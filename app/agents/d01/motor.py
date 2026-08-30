"""
app/agents/d01/motor.py — Lectura del motor (NO recálculo)
=========================================================================
Responsabilidad única: LEER las métricas de d01 del Gold Master.

REGLA 1 y 4 (inviolables): el IPE y la cobertura los calcula el Gold Master
(fórmula nativa H16b, que deriva de H12!B33 INMUTABLE). Este módulo NUNCA
recalcula — solo lee. Determinístico, sin IA.

HALLAZGO de la migración (2026-07-22): d01 calcula su IPE en H16b pero NO lo
expone en el contrato de salida H73 (fetch_gold_master_data solo trae 7
claves, ninguna de d01). Por eso aquí se lee H16b directamente. Exponer el
IPE en H73 sería una cirugía del Gold Master — se deja anotado como deuda,
NO se ejecuta en esta migración (Javo: "sin tocar el ICPI").
"""
from __future__ import annotations

import datetime as _dt
import pathlib
from typing import Any

try:
    from config import DATOS_DIR as _DATOS
except Exception:                                        # noqa: BLE001
    import os as _os
    from pathlib import Path as _P
    _DATOS = _P(_os.environ.get("QUIRA_DATOS", "."))

_GM_DEFAULT = pathlib.Path(
    str(_DATOS / "SIAP-ICPI_GOLD_MASTER_v5.5_TGI.xlsx")
)
_HOJA = "H16b_IPE"
# Celdas verificadas 2026-07-22 (PCD-D01: el IPE real curado vive en B15, no en
# el proxy B6=0.84 que quedó como nota metodológica).
_CELDA_IPE = "B15"            # IPE_Ejecutado_2026_Real
_CELDA_COBERTURA = "B12"      # Cobertura_Metas_POA_2026
_CELDA_INV_VINCULADA = "B16"  # Inversion_Vinculada_Real_USD


def _huella_del_gold_master(path: pathlib.Path) -> str:
    """SHA256 del Excel leído. Es la **evidencia** de la afirmación: identifica
    exactamente qué versión del motor produjo estos números.

    Sin esto, «el IPE es 0,84» no dice de qué Gold Master salió — y con
    correcciones sobre copia (Regla 1) eso deja de ser trivial."""
    import hashlib
    h = hashlib.sha256()
    with path.open("rb") as f:
        for bloque in iter(lambda: f.read(1 << 20), b""):
            h.update(bloque)
    return h.hexdigest()


def leer_metricas(gold_master_path: str | pathlib.Path | None = None) -> dict[str, Any]:
    """Lee (no calcula) IPE, cobertura e inversión vinculada de H16b.

    MIGRADO A AGENTE DE DOMINIO (2026-08-26 · ADR-053, piloto). La salida ya no
    es un dict de números sueltos: **lleva su procedencia dentro**, escrita aquí
    —en el generador— y sin marca de tiempo, para que el artefacto sea
    reproducible byte a byte (criterio 2 del §5).

    Lo que NO cambió: sigue leyendo H16b y sigue sin recalcular nada. La
    migración añade cadena, no aritmética."""
    path = pathlib.Path(gold_master_path) if gold_master_path else _GM_DEFAULT
    if not path.exists():
        return {"status": "failed", "error": f"Gold Master no encontrado: {path}"}

    import openpyxl
    wb = openpyxl.load_workbook(str(path), data_only=True, read_only=True)
    if _HOJA not in wb.sheetnames:
        return {"status": "failed", "error": f"Hoja {_HOJA} no existe"}
    ws = wb[_HOJA]
    return {
        "status": "ok",
        "fuente": f"{_HOJA} (leído, NO recalculado — Regla 1/4)",
        "naturaleza": "INMUTABLE",
        "ipe_ejecutado": ws[_CELDA_IPE].value,
        "cobertura_metas_poa": ws[_CELDA_COBERTURA].value,
        "inversion_vinculada_usd": ws[_CELDA_INV_VINCULADA].value,
        "evidencia_sha256": _huella_del_gold_master(path)[:16],
        "procedencia": _procedencia_de_lectura(),
    }


def _procedencia_de_lectura() -> dict[str, Any]:
    """De quién es esta lectura. **Sin reloj**, a propósito (ADR-053 §5.2)."""
    try:
        from app.agents import procedencia as P, sujeto as S
        return P.de_generacion("d01.motor",
                               f"{S.POR_DEFECTO} {S.nombre_corto()}", S.huella())
    except Exception:                                    # noqa: BLE001
        return {"etapa": "d01.motor", "estado": "sujeto_no_acreditado_por_la_cadena"}


def atender(consulta):
    """Responde a otro dominio con **evidencia, no con verdad** (ADR-053 §6-bis).

    Primera implementación real del contrato inter-dominio, y la más pequeña que
    resuelve el caso: d02 evalúa sostenibilidad presupuestaria y necesita saber
    qué parte de la inversión está vinculada a metas del POA — un dato que d01
    ya sostiene y que d02 tendría que re-derivar del mismo Gold Master.

    Lo que devuelve no es `0.9557`: es la afirmación completa de d01, con su
    sujeto, su evidencia y su grado. **Si d01 no puede sostenerla, la respuesta
    lo dice** en vez de entregar un número sin respaldo."""
    from app.agents.consulta import Respuesta

    s = sostener_ipe()
    m = leer_metricas()
    return Respuesta(
        consulta=consulta,
        sostenida=s,
        evidencia_sha256=m.get("evidencia_sha256", ""),
        motor_sha256="",          # d01 lee el Excel directo: no hay delegado
        extra={"cobertura_metas_poa": m.get("cobertura_metas_poa"),
               "inversion_vinculada_usd": m.get("inversion_vinculada_usd")}
        if m.get("status") == "ok" else {},
    )


def sostener_ipe(gold_master_path: str | pathlib.Path | None = None):
    """El IPE **como afirmación con su cadena**, no como número suelto.

    Ésta es la diferencia entre un módulo que lee un Excel y un **agente de
    dominio gobernado** (ADR-053 §4): lo que sale no es `0.84`, es *«esto puede
    sostenerse con este peso, sobre este sujeto, con esta evidencia, verificado
    por este componente, y respaldado por esta prueba»*.

    Y si algún eslabón falta, **degrada**: no calla el hueco ni lo rellena."""
    from app.agents import procedencia as P, sujeto as S

    path = pathlib.Path(gold_master_path) if gold_master_path else _GM_DEFAULT
    m = leer_metricas(path)
    if m.get("status") != "ok":
        # No se pudo leer: eso no dice nada del sujeto, dice algo del instrumento.
        return P.sostener(
            "no fue posible leer el IPE del motor",
            P.Procedencia(fuente=f"Gold Master · {_HOJA}",
                          sujeto=f"{S.POR_DEFECTO} {S.nombre_corto()}"),
            P.HALLAZGO_DE_VERIFICABILIDAD)

    return P.sostener(
        f"el IPE ejecutado del período es {m['ipe_ejecutado']}",
        P.Procedencia(
            fuente=f"Gold Master · {_HOJA} (SIAP-ICPI v5.5 · celda {_CELDA_IPE})",
            captura=_dt.date.today().isoformat(),
            estado_adquisicion="leido_del_motor",
            evidencia=m["evidencia_sha256"],
            verificador="d01.motor.leer_metricas",
            prueba_del_verificador="test_d01_lee_el_ipe_sin_recalcularlo",
            sujeto=f"{S.POR_DEFECTO} {S.nombre_corto()}",
        ))
