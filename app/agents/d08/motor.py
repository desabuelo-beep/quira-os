"""
app/agents/d08/motor.py — Lectura del Gold Master + evaluación de integridad
=============================================================================
d08 NO envuelve un enricher (a diferencia de d02/d03/d09): su motor tiene dos
lecturas distintas, coherentes con la asesoría de dirección (2026-07-24):

  1. leer_igp_diagnostico() — LEE el IGP y sus 3 componentes de H20b, SOLO para
     DIAGNÓSTICO (anotar los hallazgos de OBS-015). NO es el contrato canónico:
     el IGP actual está mal compuesto (mezcla d09 vía MFN, IGP_PP=0 pese a
     evidencia). Se reconstruirá en el Gold Master en una fase posterior (curar
     el Excel · Regla 8), y ENTONCES se expondrá por app/connectors/gold_master.py.
     Por ahora NO se toca el conector (no oficializar un valor defectuoso).

  2. evaluar_integridad() — 1ª dimensión (RO-VIII-001): evalúa la INTEGRIDAD
     NORMATIVA por la presencia/estado de la evidencia documental de cada
     instancia/mecanismo (del catálogo). Determinístico, sin IA.

  3. leer_efectividad() — 3ª dimensión (RO-VIII-003): LEE la trazabilidad
     biográfica de la demanda ya cruzada (scripts/d08/cruzar_demandas.py). Cada
     correspondencia conserva su estado epistémico; QUIRA propone, no afirma.

La dimensión 2 (vitalidad, RO-VIII-002) sigue en diseño: su índice debe sellarse
en el Gold Master (Regla 1), no en QUIRA.

LÍMITE (Regla 1/4): el IGP NO se recalcula. Aquí solo se LEE para diagnóstico.
"""
from __future__ import annotations

import json
import pathlib
from typing import Any

import openpyxl

try:
    from config import DATOS_DIR as _DATOS
except Exception:                                        # noqa: BLE001
    import os as _os
    from pathlib import Path as _P
    _DATOS = _P(_os.environ.get("QUIRA_DATOS", "."))

_EXCEL = pathlib.Path(str(_DATOS / "SIAP-ICPI_GOLD_MASTER_v5.5_TGI.xlsx"))
_HOJA_IGP = "H20b_IGP_GOBERNANZA_PARTICIPATIVA"

# Componentes del IGP y su dominio real (OBS-015)
_COMPONENTES = {
    "IGP_1_Asamblea_CPCCS": "d08",
    "IGP_2_Presupuesto_Participativo": "d08",
    "IGP_3_Fidelidad_Narrativa_MFN": "d09",   # ← pertenece a Rendición de Cuentas, no a d08
}


def leer_igp_diagnostico() -> dict[str, Any]:
    """LECTURA DIAGNÓSTICA del IGP (NO contrato canónico). Anota los hallazgos OBS-015."""
    wb = openpyxl.load_workbook(_EXCEL, read_only=True, data_only=True)
    ws = next((wb[s] for s in wb.sheetnames if s.startswith("H20b")), None)
    if ws is None:
        wb.close()
        return {"status": "error", "detalle": "hoja H20b (IGP) no encontrada"}

    etiquetas: dict[str, float] = {}
    for row in ws.iter_rows(max_row=40):
        for c in row:
            if isinstance(c.value, str) and c.value.strip() in (
                *_COMPONENTES, "IGP_Global", "Ref_2025_IGP"
            ):
                etiquetas[c.value.strip()] = ws.cell(c.row, c.column + 1).value
    wb.close()

    componentes = [
        {
            "id": k, "valor": etiquetas.get(k), "dominio_real": dom,
            "alerta": ("★ pertenece a d09 (RDC), no a d08 — OBS-015 hallazgo 1"
                       if dom == "d09" else
                       ("★ = 0 pese a evidencia de PP 2024/2025/2026 — OBS-015 hallazgo 2"
                        if k == "IGP_2_Presupuesto_Participativo" and not etiquetas.get(k) else None)),
        }
        for k, dom in _COMPONENTES.items()
    ]
    return {
        "status": "ok",
        "naturaleza": "DIAGNÓSTICO PROVISIONAL — el IGP se reconstruye en el Gold Master (fase 2). NO es contrato canónico (Regla 1/4).",
        "igp_global": etiquetas.get("IGP_Global"),
        "igp_ref_2025": etiquetas.get("Ref_2025_IGP"),
        "componentes": componentes,
        "hallazgos_obs": ["OBS-015 h1: IGP mezcla d08+d09 (MFN)", "OBS-015 h2: IGP_PP=0 con evidencia"],
    }


def leer_efectividad() -> dict[str, Any]:
    """3ª dimensión (RO-VIII-003): trazabilidad biográfica de la demanda ciudadana.

    LEE el resultado del cruce ya ejecutado (`scripts/d08/cruzar_demandas.py`), no lo
    recalcula. Cada correspondencia conserva su ESTADO EPISTÉMICO (Horizonte de Verdad):
    QUIRA propone; nunca afirma que una demanda fue satisfecha (Constitución Art. 3).
    """
    demandas = pathlib.Path("data/d08/demandas_ciudadanas.json")
    traza = pathlib.Path("data/d08/trazabilidad_demandas.json")
    if not (demandas.exists() and traza.exists()):
        return {"dimension": "efectividad_incidencia", "estado": "sin_datos",
                "nota": "ejecutar scripts/d08/extraer_demandas.py y cruzar_demandas.py"}

    d = json.loads(demandas.read_text(encoding="utf-8"))
    t = json.loads(traza.read_text(encoding="utf-8"))
    return {
        "dimension": "efectividad_incidencia",
        "ro": "RO-VIII-003",
        "postulado": "I · Trazabilidad Biográfica del Dato",
        "naturaleza": "QUIRA PROPONE correspondencias; el humano valida (Horizonte de Verdad)",
        "demandas_catalogadas": d["total"],
        "por_mecanismo": d["por_mecanismo"],
        "por_naturaleza_juridica": d["por_naturaleza"],
        "registros_poa_contrastados": t["registros_poa_contrastados"],
        "por_estado_epistemico": t["por_estado_epistemico"],
        "vinculantes_por_estado": t["vinculantes_por_estado"],
        "advertencia": "'hipotesis' NO significa 'atendida'; 'sin_correlato' NO significa 'no se atendió'",
    }


def evaluar_integridad(catalogo: dict[str, Any]) -> dict[str, Any]:
    """1ª dimensión (RO-VIII-001): estado de acreditación documental por instancia/mecanismo.
    Determinístico — refleja el catálogo (Nivel 4), no interpreta contenido (eso es Fase 4/IA)."""
    filas = []
    for inst in catalogo["instancias"]:
        ev = inst.get("evidencia", {})
        estado = ev.get("estado")
        # un campo `acreditacion` explícito gana sobre la derivación por estado: p. ej. la
        # audiencia es documentalmente 'procesable' pero su acreditación es PARCIAL por OBS-017
        # (el acto se realiza, la resolución del Art. 75 no consta).
        acreditacion = ev.get("acreditacion") or {
            "procesable": "acreditada",
            "mixto": "parcial",
            "existe_documento_no_accesible": "existe_no_verificable_aun",
            "evidencia_indirecta": "indirecta",
            "pendiente_adquisicion": "sin_evidencia_disponible",
            "no_localizada": "sin_evidencia",
        }.get(estado, "desconocido")
        filas.append({
            "cno": inst["cno"], "mecanismo": inst["mecanismo"],
            "evidencia_estado": estado, "acreditacion": acreditacion,
        })
    return {
        "dimension": "integridad_normativa",
        "ro": "RO-VIII-001",
        "naturaleza": "verificabilidad documental — nunca 'incumple/ilegal' (Regla 2)",
        "senales": filas,
    }
