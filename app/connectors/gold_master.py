"""
app/connectors/gold_master.py — QUIRA OS · Sprint 3
Conector institucional: Gold Master TGI (H73_OUTPUT_API)

Lee los valores calculados del Motor TGI desde la hoja H73_OUTPUT_API
del Gold Master v5.5 canónico. Actúa como fuente de verdad cuando los APIs
institucionales (DPE/SERCOP/CPCCS) no están disponibles.

Doctrina: Gold Master es el único cálculo canónico. El pipeline Q1
observa APIs; el Gold Master valida y certifica. Esta interfaz
SOLO LEE — nunca modifica el Excel.

DECISIÓN ARQUITECTÓNICA (2026-05-26):
  v5.5 es el Gold Master canónico activo — 123 hojas · 931 KB · audit trail completo.
  TGI Framework v6.0 fue incorporado como extensión de H73 (claves TGI_SCORE, TGI_D1-D5,
  SAT_CLASIF_RIESGO añadidas canónicamente). El archivo TGI_GOLD_MASTER_v6.0_20260525.xlsx
  es un template de referencia — no tiene los datos de soporte para auditoría.

ACTUALIZACIÓN (2026-06-15 · cirugía D.2A → v6.0 metodológico):
  El slot vivo (SIAP-ICPI_GOLD_MASTER_v5.5_TGI.xlsx) fue promovido al motor CORREGIDO:
  ICPI 17.45% → 27.46% (adscritas→cédula SERCOP · semáforo escala+corte · FactorTemporal curva real 2025).
  B33 INTACTA · 0 errores añadidos. El NOMBRE del slot se conserva a propósito (contrato con el código);
  el v6.0 se sella aparte en SIAP-ICPI_GOLD_MASTER_v6.0_FREEZE_20260615.xlsx.
  Registro celda-a-celda: docs/architecture/CIRUGIA_GOLD_MASTER_D2A.md.

Versiones soportadas:
  v5.5 (ACTIVO):    SIAP-ICPI_GOLD_MASTER_v5.5_TGI_20260518.xlsx → H73_OUTPUT_API
  v6.0 (template):  TGI_GOLD_MASTER_v6.0_20260525.xlsx → G6.1_OUTPUT_API (fallback)

El conector lee v5.5 primero; si no existe, cae a v6.0 automáticamente.
Dylus Lab © 2026
"""
# ---
# authority:
#   parent: GOVERNANCE-001
#   constitution_articles: [2]
#   type: TECNICA
# ---
from __future__ import annotations

import logging
import re
from pathlib import Path
from typing import Any

logger = logging.getLogger(__name__)

SOURCE_ID   = "gold_master"
RELIABILITY = 0.99   # Fuente máxima confiabilidad — datos certificados

# Rutas canónicas. La carpeta de los documentos se recibe de `config.DATOS_DIR`
# —declarada UNA vez y configurable con `QUIRA_DATOS`— en vez de escribirse aquí.
# Hasta el 2026-08-19 este conector, que es la puerta canónica al Gold Master,
# llevaba dentro el escritorio de una persona: en cualquier otro equipo no
# encontraba el motor (OBS-032).
try:
    from config import DATOS_DIR as _PROYECT
except Exception:                                        # noqa: BLE001
    # Sin `config` importable (script suelto, prueba aislada), la frontera sigue
    # siendo declarable por entorno. Lo que no vuelve es la ruta fija.
    import os as _os
    _PROYECT = Path(_os.environ.get("QUIRA_DATOS", "."))

_DEFAULT_GOLD_MASTER_V6 = _PROYECT / "TGI_GOLD_MASTER_v6.0_20260525.xlsx"
_DEFAULT_GOLD_MASTER_V55 = _PROYECT / "SIAP-ICPI_GOLD_MASTER_v5.5_TGI.xlsx"


def _resolver_gold_master() -> Path:
    """Localiza el Gold Master vigente SIN depender del número de versión en el nombre.

    ★ Corrección 2026-07-29 (Javo): "el ecosistema está conectado al v5.5_TGI por el
    NOMBRE; si se cambia, todo se desajusta. Ya nos pasó una vez y fue un problema."

    Tenía razón: la versión estaba **hardcodeada en la ruta**, así que subir a v5.6
    exigía tocar código — y olvidarlo rompía el sistema en silencio. La versión debe
    ser metadato del CONTENIDO, no del nombre del archivo.

    ⚠️ PRIMER INTENTO FALLIDO — se documenta para que no se repita. Resolver por
    "versión más alta" ELIGIÓ UN FREEZE: la carpeta tiene 4 archivos y tres son
    respaldos pre-cirugía (`..._v5.5_FREEZE_20260701_PRE_IPEEJEC`, etc.), todos
    v5.5. El sistema habría leído un motor congelado **en silencio**. Es el mismo
    fallo que Javo advertía, cometido al intentar corregirlo.

    Lo que distingue al canónico NO es la versión: es el **sufijo `_TGI`**.

    Reglas de resolución:
      1. Solo `SIAP-ICPI_GOLD_MASTER_v*_TGI.xlsx` — el sufijo `_TGI` marca el slot vivo.
      2. Se EXCLUYE `_FREEZE` (respaldos), `_` inicial (candidatos) y `~$` (temporales
         de Excel). Un respaldo jamás debe convertirse en canónico por accidente.
      3. Entre los válidos gana la versión más alta, numérica: v5.10 > v5.9.
      4. Sin coincidencias → nombre histórico v5.5 (el contrato documentado arriba).
         ⚠️ Ese archivo YA NO EXISTE en el slot: la ruta devuelta entonces no
         resuelve, y quien la reciba debe tratarla como «no determinable», no
         como «es la v5.5» (mismo defecto que `config` corrigió el 02-sep).
    """
    if not _PROYECT.is_dir():
        return _DEFAULT_GOLD_MASTER_V55

    validos = []
    for f in _PROYECT.glob("SIAP-ICPI_GOLD_MASTER_v*_TGI.xlsx"):
        if f.name.startswith(("_", "~$")) or "_FREEZE" in f.name.upper():
            continue
        m = re.search(r"_v(\d+)\.(\d+)_TGI", f.name)
        if m:
            validos.append(((int(m.group(1)), int(m.group(2))), f))

    if not validos:
        logger.warning(
            "[GoldMaster] Ningún 'SIAP-ICPI_GOLD_MASTER_v*_TGI.xlsx' válido en %s — "
            "se usa la ruta histórica. Candidatos descartados: %s",
            _PROYECT, [f.name for f in _PROYECT.glob("SIAP-ICPI_GOLD_MASTER*.xlsx")],
        )
        return _DEFAULT_GOLD_MASTER_V55
    return max(validos, key=lambda x: x[0])[1]


# Gold Master canónico activo — se resuelve por contenido de carpeta, no por nombre fijo
_DEFAULT_GOLD_MASTER = _resolver_gold_master()

# ⚠️ Y SI NO RESOLVIÓ, SE DICE (D-002 · 2026-09-03). El colega puso la pregunta
# exacta: *«¿puede este código producir un objeto que parezca provenir del Gold
# Master vigente cuando el Gold Master no fue resuelto?»*
#
# Medido: esta puerta NO emite versión —sólo usa el nombre para filtrar
# candidatos y para el log—, así que no puede declarar una procedencia falsa.
# Pero sí devolvía la ruta histórica v5.5 con la misma cara que una resuelta, y
# ese archivo ya no existe: quien la recibiera creería tener el slot vivo.
#
# No se cambia el retorno —romperlo afectaría a todo consumidor—: se añade la
# constancia. Es lo mismo que `config.GOLD_MASTER_RESUELTO` hizo el 02-sep, y la
# misma distinción de siempre: «no lo encontré» no es «es la v5.5».
GOLD_MASTER_RESUELTO = (
    _DEFAULT_GOLD_MASTER != _DEFAULT_GOLD_MASTER_V55
    and _DEFAULT_GOLD_MASTER.exists()
)
if not GOLD_MASTER_RESUELTO:
    logger.warning(
        "[GoldMaster] NO RESUELTO — se devuelve la ruta histórica %s, que no es "
        "el slot vivo. Quien la consuma debe tratarla como «no determinable», "
        "nunca como el Gold Master vigente.", _DEFAULT_GOLD_MASTER.name)

# Hoja de salida API — v5.5 canónico (v6.0 template como fallback)
_OUTPUT_API_SHEET     = "H73_OUTPUT_API"    # v5.5 canónico activo
_OUTPUT_API_SHEET_V6  = "G6.1_OUTPUT_API"  # v6.0 fallback (template)

# Claves que el pipeline extrae de H73
_KEYS_OF_INTEREST = {
    # ICPI / TGI global
    "ICPI_GLOBAL", "ICPI_GLOBAL_PCT", "ICPI_CLASIFICACION",
    "ICPI_2023", "ICPI_2024", "ICPI_2025", "ICPI_ACUMULADO_Q1",
    # Financiero / ejecución
    "ISP_SALUD_PRESUP", "ISP_META", "ISP_BRECHA_PP",
    "PSG_EJECUCION", "PSG_FIDELIDAD",
    # Inversión
    "IFE_INVERSION_PUBLICA", "IFE_META",
    # Contratación / SERCOP
    "PAC_PUBLICADO", "PROCESOS_ADJUDICADOS", "PROCESOS_CANCELADOS",
    "PROCESOS_CANCELADOS_PCT",
    # CPCCS / RdC
    "RDC_SCORE", "RDC_CLASIFICACION",
    # Participación ciudadana (d08) — el motor YA las publicaba en H73 filas 21-22;
    # el conector no las leía, así que la UI de d08 no tenía de dónde tomar su
    # indicador madre. No es cambio conceptual: es consumir lo que el canon ya declara.
    "IGP_2026_ACTUAL", "IGP_REF_2025",
    # SAT
    "SAT_RIESGO_TOTAL", "SAT_CLASIF_RIESGO", "SAT_ACTIVAS_COUNT",
    # TGI
    "TGI_SCORE", "TGI_D1", "TGI_D2", "TGI_D3", "TGI_D4", "TGI_D5",
}


def fetch_gold_master_data(
    gold_master_path: str | Path | None = None,
) -> dict:
    """Lee H73_OUTPUT_API del Gold Master y devuelve métricas canónicas.

    Args:
        gold_master_path: Ruta al archivo Excel. Si None, usa la ruta
                          canónica o la configurada en config.py.

    Returns:
        dict con claves:
            status:       "ok" | "partial" | "failed"
            source_id:    "gold_master"
            reliability:  float
            data:         dict con métricas de H73_OUTPUT_API
            error:        str | None
            sheet_rows:   int — número de filas leídas
    """
    result: dict[str, Any] = {
        "status":      "pending",
        "source_id":   SOURCE_ID,
        "reliability": RELIABILITY,
        "data":        {},
        "error":       None,
        "sheet_rows":  0,
    }

    # ── Resolver ruta ──────────────────────────────────────────────────────
    path = Path(gold_master_path) if gold_master_path else _resolve_path()
    if not path or not path.exists():
        result["status"]      = "failed"
        result["reliability"] = 0.0
        result["error"]       = f"Gold Master no encontrado: {path}"
        logger.warning(f"[GoldMaster] Archivo no encontrado: {path}")
        return result

    # ── Leer H73_OUTPUT_API (v5.5 canónico) o G6.1_OUTPUT_API (v6.0 fallback) ──
    try:
        import openpyxl
        wb = openpyxl.load_workbook(str(path), data_only=True, read_only=True)

        # Detectar hoja disponible (v5.5 H73 preferido, v6.0 G6.1 fallback)
        sheet_name = _OUTPUT_API_SHEET
        if sheet_name not in wb.sheetnames:
            if _OUTPUT_API_SHEET_V6 in wb.sheetnames:
                sheet_name = _OUTPUT_API_SHEET_V6
                logger.info(f"[GoldMaster] H73 no encontrado, usando G6.1_OUTPUT_API (v6.0 template)")
            else:
                result["status"]  = "failed"
                result["error"]   = (
                    f"Hoja de output no encontrada. "
                    f"Buscado: {_OUTPUT_API_SHEET} | {_OUTPUT_API_SHEET_V6}"
                )
                wb.close()
                return result

        ws    = wb[sheet_name]
        data  = {}
        rows  = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            key, value = row[0], row[1]
            if key is None:
                continue
            rows += 1
            key_str = str(key).strip()
            data[key_str] = value

        wb.close()

        result["sheet_rows"] = rows
        result["data"]       = _normalize_h73(data)
        result["status"]     = "ok" if rows > 5 else "partial"
        result["reliability"] = RELIABILITY if result["status"] == "ok" else RELIABILITY * 0.5

        logger.info(
            f"[GoldMaster] {path.name} [{sheet_name}] → {rows} métricas | "
            f"ICPI={data.get('ICPI_GLOBAL_PCT', 'N/A')} | "
            f"TGI={data.get('TGI_SCORE', 'N/A')}"
        )

    except Exception as exc:
        logger.error(f"[GoldMaster] Error leyendo {path}: {exc}")
        result["status"]      = "failed"
        result["reliability"] = 0.0
        result["error"]       = str(exc)

    return result


def _resolve_path() -> Path | None:
    """Resuelve la ruta del Gold Master desde config o usa la canónica.

    Prioridad: config.py → v5.5 canónico activo → v6.0 template (fallback).
    """
    try:
        import config as cfg
        gm_path = getattr(cfg, "GOLD_MASTER_PATH", None)
        if gm_path:
            return Path(gm_path)
    except ImportError:
        pass
    # ⚠️ v5.5 NO es el canónico activo desde el cierre de D-002: lo es el que
    # `_resolver_gold_master()` encuentre por sufijo `_TGI`, hoy v5.7. Este
    # camino sólo se toma sin `config` importable, y v5.5 ya no está en disco
    # —vive en `historial_gold_master/`—: es un respaldo histórico, no el slot.
    if _DEFAULT_GOLD_MASTER_V55.exists():
        return _DEFAULT_GOLD_MASTER_V55
    # v6.0 como fallback (template de referencia)
    if _DEFAULT_GOLD_MASTER_V6.exists():
        logger.warning("[GoldMaster] Usando v6.0 template como fallback — v5.5 no encontrado")
        return _DEFAULT_GOLD_MASTER_V6
    return None


def _normalize_h73(raw: dict) -> dict:
    """Normaliza y tipifica los valores de H73 en el namespace del pipeline.

    Devuelve un dict estructurado por dominio:
        icpi, tgi, financiero, contratacion, accountability, sat
    """
    def _float(key, default=None):
        v = raw.get(key)
        if v is None or (isinstance(v, str) and v.startswith('#')):
            return default
        try:
            return float(v)
        except (TypeError, ValueError):
            return default

    def _str(key, default=None):
        v = raw.get(key)
        return str(v).strip() if v is not None else default

    def _bool(key, default=None):
        v = raw.get(key)
        if isinstance(v, bool):
            return v
        if isinstance(v, str):
            return v.strip().upper() in ("SI", "TRUE", "1", "SÍ")
        if isinstance(v, (int, float)):
            return bool(v)
        return default

    # fallback 5D ponderado cuando TGI_SCORE falla en Excel (#VALOR!)
    # pesos canónicos: D1×20 + D2×20 + D3×25 + D4×25 + D5×10 = 100
    _tgi_d = [_float(f"TGI_D{i}") for i in range(1, 6)]
    _tgi_score = _float("TGI_SCORE")
    if _tgi_score is None and all(d is not None for d in _tgi_d):
        _tgi_score = round(
            _tgi_d[0]*0.20 + _tgi_d[1]*0.20 + _tgi_d[2]*0.25 + _tgi_d[3]*0.25 + _tgi_d[4]*0.10,
            2,
        )

    return {
        "icpi": {
            "global":          _float("ICPI_GLOBAL"),
            "global_pct":      _float("ICPI_GLOBAL_PCT"),
            "clasificacion":   _str("ICPI_CLASIFICACION"),
            "historico": {
                "2023": _float("ICPI_2023"),
                "2024": _float("ICPI_2024"),
                "2025": _float("ICPI_2025"),
            },
            "acumulado_q1":    _float("ICPI_ACUMULADO_Q1"),
        },
        "tgi": {
            "score": _tgi_score,
            "d1":    _tgi_d[0],
            "d2":    _tgi_d[1],
            "d3":    _tgi_d[2],
            "d4":    _tgi_d[3],
            "d5":    _tgi_d[4],
        },
        "financiero": {
            "isp_salud_presup":   _float("ISP_SALUD_PRESUP"),
            "isp_meta":           _float("ISP_META"),
            "isp_brecha_pp":      _float("ISP_BRECHA_PP"),
            "psg_ejecucion":      _float("PSG_EJECUCION"),
            "psg_fidelidad":      _float("PSG_FIDELIDAD"),
            "ife_inversion":      _float("IFE_INVERSION_PUBLICA"),
            "ife_meta":           _float("IFE_META"),
        },
        "contratacion": {
            "pac_publicado":         _bool("PAC_PUBLICADO"),
            "procesos_adjudicados":  _float("PROCESOS_ADJUDICADOS"),
            "procesos_cancelados":   _float("PROCESOS_CANCELADOS"),
            "cancelados_pct":        _float("PROCESOS_CANCELADOS_PCT"),
        },
        "accountability": {
            "rdc_score":         _float("RDC_SCORE"),
            "rdc_clasificacion": _str("RDC_CLASIFICACION"),
        },
        # Participación ciudadana (d08). El IGP se compone hoy de IGP_1 (Asamblea
        # CPCCS) + IGP_2 (Presupuesto Participativo) + IGP_3 (Fidelidad Narrativa).
        # ⚠️ IGP_3 pertenece a d09 (control social) y cruza la frontera DEC-0004;
        # IGP_2 vale 0 porque H10b no tiene montos por proyecto PP, no por un bug.
        # Ambas correcciones se sellan en el Excel — ver ESPECIFICACION_GOLD_MASTER_D08.
        "participacion": {
            "igp_actual": _float("IGP_2026_ACTUAL"),
            "igp_ref_2025": _float("IGP_REF_2025"),
        },
        "sat_engine": {
            "riesgo_total":   _float("SAT_RIESGO_TOTAL"),
            "clasif_riesgo":  _str("SAT_CLASIF_RIESGO"),
            "activas_count":  _float("SAT_ACTIVAS_COUNT"),
        },
        "_raw_h73": raw,   # preservar todos los datos crudos
    }
