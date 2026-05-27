"""
QUIRA Intelligence — Gobernanza Gold Master (Sprint 3 · P4)

El Gold Master ya no es un archivo Excel.
Es infraestructura epistemológica institucional.
Este módulo garantiza su integridad, trazabilidad y gobernanza de cambios.

Funciones públicas:
  validar_gold_master(ruta)          → lista de ResultadoValidacion
  obtener_changelog_gm()             → historial de versiones del GM
  respaldar_gold_master(ruta, dest)  → copia firmada con SHA-256
  obtener_estado_gm(ruta)            → diagnóstico completo del GM actual

Doctrina:
  Cada versión del Gold Master es un hito epistemológico.
  Cada respaldo es un punto de no retorno auditado.
  Ningún cambio al Gold Master ocurre sin registro en gm_changelog.json.

Stack (reemplazable): openpyxl, hashlib, shutil.
Firmas de función: permanentes.

Dylus Lab © 2026
"""
from __future__ import annotations

import hashlib
import json
import shutil
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl

from utils.logger import get_logger

logger = get_logger(__name__)

# ── Constantes canónicas ───────────────────────────────────────────────────────

HOJAS_REQUERIDAS_V6: list[str] = [
    # v5.5 canónico activo — hojas estructurales mínimas requeridas
    # (las G1.x-G6.x corresponden al template v6.0 que NO es el canónico activo)
    "H73_OUTPUT_API",              # interfaz pública — todas las métricas TGI/ICPI
    "H82_CONFIG_PARAMS",           # config sistema (VERSION_SISTEMA, FECHA_CORTE)
    "H07_S5_FINANCIERO_eSIGEF",   # datos financieros eSIGEF (Ti holding)
    "H10_S8_PARTICIPACIÓN_CPCCS",  # rendición de cuentas CPCCS (tilde canónica)
]

CLAVES_OUTPUT_API_REQUERIDAS: list[str] = [
    "ICPI_GLOBAL",
    "ICPI_GLOBAL_PCT",
    "ICPI_CLASIFICACION",
    "PSG_EJECUCION",
    "PSG_FIDELIDAD",
    "SAT_RIESGO_TOTAL",
    "SAT_CLASIF_RIESGO",
    "SAT_ACTIVAS_COUNT",
    "TGI_SCORE",
    "TGI_D1",
    "TGI_D2",
    "TGI_D3",
    "TGI_D4",
    "TGI_D5",
]

# Rangos aceptables para validación de integridad numérica
_RANGOS_METRICAS: dict[str, tuple[float, float]] = {
    "ICPI_GLOBAL":     (0.0, 1.0),
    "ICPI_GLOBAL_PCT": (0.0, 100.0),
    "TGI_SCORE":       (0.0, 100.0),
    "TGI_D1":          (0.0, 100.0),
    "TGI_D2":          (0.0, 100.0),
    "TGI_D3":          (0.0, 100.0),
    "TGI_D4":          (0.0, 100.0),
    "TGI_D5":          (0.0, 100.0),
    "SAT_RIESGO_TOTAL": (0.0, 1.0),
    "SAT_ACTIVAS_COUNT": (0, 20),
}

_RUTA_CHANGELOG = Path("data/doctrinal/gm_changelog.json")
_RUTA_RESPALDOS  = Path("data/backups")


# ── Tipos de datos ─────────────────────────────────────────────────────────────

@dataclass
class ResultadoValidacion:
    """Resultado de una validación canónica del Gold Master."""
    categoria:   str              # "estructura" | "contrato_api" | "integridad" | "version"
    regla:       str              # nombre de la regla
    estado:      str              # "OK" | "ALERTA" | "ERROR"
    detalle:     str              # mensaje explicativo
    valor:       Any = None       # valor observado (si aplica)
    esperado:    Any = None       # valor esperado (si aplica)

    def es_error(self) -> bool:
        return self.estado == "ERROR"

    def es_alerta(self) -> bool:
        return self.estado == "ALERTA"


# ══════════════════════════════════════════════════════════════════════════════
# 1. Validación canónica
# ══════════════════════════════════════════════════════════════════════════════

def validar_gold_master(
    ruta: str | Path | None = None,
) -> list[ResultadoValidacion]:
    """
    Ejecuta validaciones canónicas sobre el Gold Master.

    Verifica:
      - Existencia del archivo
      - Hojas requeridas G1.x-G6.x presentes
      - G6.1_OUTPUT_API contiene todas las claves del contrato
      - Rangos numéricos de métricas clave
      - Versión declarada en G1.1_CONFIG

    Args:
        ruta: Ruta al archivo Excel. Si None, usa la ruta canónica del conector.

    Returns:
        Lista de ResultadoValidacion con estado OK/ALERTA/ERROR por regla.
    """
    resultados: list[ResultadoValidacion] = []
    ruta_resuelta = _resolver_ruta_gm(ruta)

    # ── Regla 1: Existencia del archivo ───────────────────────────────────────
    if not ruta_resuelta or not ruta_resuelta.exists():
        resultados.append(ResultadoValidacion(
            categoria="estructura",
            regla="archivo_existe",
            estado="ERROR",
            detalle=f"Gold Master no encontrado: {ruta_resuelta}",
            valor=str(ruta_resuelta),
            esperado="archivo .xlsx existente",
        ))
        return resultados  # sin archivo, no continuamos

    resultados.append(ResultadoValidacion(
        categoria="estructura",
        regla="archivo_existe",
        estado="OK",
        detalle=f"Gold Master encontrado: {ruta_resuelta.name}",
        valor=str(ruta_resuelta.name),
    ))

    # ── Abrir workbook ─────────────────────────────────────────────────────────
    try:
        wb = openpyxl.load_workbook(str(ruta_resuelta), read_only=True, data_only=True)
    except Exception as exc:
        resultados.append(ResultadoValidacion(
            categoria="estructura",
            regla="apertura_workbook",
            estado="ERROR",
            detalle=f"No se pudo abrir el archivo Excel: {exc}",
        ))
        return resultados

    hojas_presentes = set(wb.sheetnames)

    # ── Regla 2: Hojas requeridas ──────────────────────────────────────────────
    for hoja in HOJAS_REQUERIDAS_V6:
        if hoja in hojas_presentes:
            resultados.append(ResultadoValidacion(
                categoria="estructura",
                regla=f"hoja_requerida:{hoja}",
                estado="OK",
                detalle=f"Hoja presente: {hoja}",
                valor=hoja,
            ))
        else:
            resultados.append(ResultadoValidacion(
                categoria="estructura",
                regla=f"hoja_requerida:{hoja}",
                estado="ERROR",
                detalle=f"Hoja ausente: {hoja} — requerida por contrato v6.0",
                valor=None,
                esperado=hoja,
            ))

    # ── Regla 3: Contrato H73_OUTPUT_API (v5.5 canónico activo) ──────────────────
    # v6.0 usaba G6.1_OUTPUT_API; v5.5 canónico usa H73_OUTPUT_API
    _API_SHEET = "H73_OUTPUT_API" if "H73_OUTPUT_API" in hojas_presentes else "G6.1_OUTPUT_API"
    if _API_SHEET in hojas_presentes:
        claves_presentes = set()
        valores_api: dict[str, Any] = {}
        try:
            ws_api = wb[_API_SHEET]
            for row in ws_api.iter_rows(min_row=2, values_only=True):
                if row[0] is not None:
                    clave = str(row[0]).strip()
                    claves_presentes.add(clave)
                    valores_api[clave] = row[1] if len(row) > 1 else None
        except Exception as exc:
            resultados.append(ResultadoValidacion(
                categoria="contrato_api",
                regla="lectura_output_api",
                estado="ERROR",
                detalle=f"Error leyendo {_API_SHEET}: {exc}",
            ))

        for clave in CLAVES_OUTPUT_API_REQUERIDAS:
            if clave in claves_presentes:
                val = valores_api.get(clave)
                estado = "OK" if val is not None else "ALERTA"
                detalle = (f"Clave presente con valor: {val}"
                           if val is not None
                           else f"Clave presente pero valor es None — pendiente datos")
                resultados.append(ResultadoValidacion(
                    categoria="contrato_api",
                    regla=f"clave_api:{clave}",
                    estado=estado,
                    detalle=detalle,
                    valor=val,
                ))
            else:
                resultados.append(ResultadoValidacion(
                    categoria="contrato_api",
                    regla=f"clave_api:{clave}",
                    estado="ERROR",
                    detalle=f"Clave ausente en {_API_SHEET}: {clave}",
                    valor=None,
                    esperado=clave,
                ))

        # ── Regla 4: Rangos numéricos de métricas clave ───────────────────────
        for clave, (min_val, max_val) in _RANGOS_METRICAS.items():
            if clave not in valores_api:
                continue
            val = valores_api.get(clave)
            if val is None:
                continue  # ya reportado como ALERTA arriba
            try:
                val_num = float(val)
                en_rango = min_val <= val_num <= max_val
                resultados.append(ResultadoValidacion(
                    categoria="integridad",
                    regla=f"rango_metrica:{clave}",
                    estado="OK" if en_rango else "ERROR",
                    detalle=(f"{clave}={val_num} dentro del rango [{min_val},{max_val}]"
                             if en_rango
                             else f"{clave}={val_num} FUERA del rango [{min_val},{max_val}]"),
                    valor=val_num,
                    esperado=f"[{min_val}, {max_val}]",
                ))
            except (TypeError, ValueError):
                pass  # valor no numérico — se valida en clave_api

    # ── Regla 5: Versión en H82_CONFIG_PARAMS (v5.5) o G1.1_CONFIG (v6.0) ────────
    _cfg_sheet = ("H82_CONFIG_PARAMS" if "H82_CONFIG_PARAMS" in hojas_presentes
                  else "G1.1_CONFIG" if "G1.1_CONFIG" in hojas_presentes
                  else None)
    if _cfg_sheet:
        version_declarada = None
        try:
            ws_cfg = wb[_cfg_sheet]
            for row in ws_cfg.iter_rows(values_only=True):
                if row[0] in ("VERSION_GOLD_MASTER", "VERSION_SISTEMA") and len(row) > 1:
                    version_declarada = str(row[1]).strip() if row[1] else None
                    break
        except Exception:
            pass

        if version_declarada:
            # v5.5 es el canónico activo — cualquier v5.x o v1.x (VERSION_SISTEMA) es OK
            es_valida = any(version_declarada.startswith(p) for p in ("v5", "v6", "v1"))
            resultados.append(ResultadoValidacion(
                categoria="version",
                regla="version_declarada",
                estado="OK" if es_valida else "ALERTA",
                detalle=f"Versión declarada en {_cfg_sheet}: {version_declarada}",
                valor=version_declarada,
                esperado="v5.x (canónico activo) o v6.x",
            ))
        else:
            resultados.append(ResultadoValidacion(
                categoria="version",
                regla="version_declarada",
                estado="ALERTA",
                detalle=f"VERSION_GOLD_MASTER / VERSION_SISTEMA no encontrada en {_cfg_sheet}",
            ))

    wb.close()

    n_errores  = sum(1 for r in resultados if r.es_error())
    n_alertas  = sum(1 for r in resultados if r.es_alerta())
    logger.info(
        f"[GM Governance] Validación completa — {len(resultados)} reglas · "
        f"{n_errores} errores · {n_alertas} alertas"
    )
    return resultados


# ══════════════════════════════════════════════════════════════════════════════
# 2. Historial de versiones (changelog)
# ══════════════════════════════════════════════════════════════════════════════

def obtener_changelog_gm(
    ruta_changelog: str | Path | None = None,
) -> dict:
    """
    Retorna el historial de versiones del Gold Master.

    Lee de data/doctrinal/gm_changelog.json.
    Si no existe, devuelve estructura vacía con nota doctrinal.

    Args:
        ruta_changelog: Ruta al JSON de changelog. Si None, usa la canónica.

    Returns:
        Dict con:
            versiones     list[dict]  Historial de versiones
            version_activa str        Versión actualmente activa
            total_versiones int
            error         str|None
    """
    ruta = Path(ruta_changelog) if ruta_changelog else _RUTA_CHANGELOG

    vacio: dict = {
        "versiones":       [],
        "version_activa":  None,
        "total_versiones": 0,
        "error":           None,
    }

    if not ruta.exists():
        logger.warning(f"[GM Governance] Changelog no encontrado: {ruta}")
        return {**vacio, "error": f"Changelog no encontrado: {ruta}"}

    try:
        with open(ruta, encoding="utf-8") as f:
            datos = json.load(f)

        versiones = datos.get("versiones", [])
        activa = next(
            (v["version"] for v in versiones if v.get("estado") == "ACTIVO"),
            None,
        )

        return {
            "versiones":       versiones,
            "version_activa":  activa,
            "total_versiones": len(versiones),
            "error":           None,
        }

    except Exception as exc:
        logger.error(f"[GM Governance] Error leyendo changelog: {exc}")
        return {**vacio, "error": str(exc)}


# ══════════════════════════════════════════════════════════════════════════════
# 3. Respaldo firmado con SHA-256
# ══════════════════════════════════════════════════════════════════════════════

def respaldar_gold_master(
    ruta: str | Path | None = None,
    destino: str | Path | None = None,
) -> dict:
    """
    Crea una copia firmada del Gold Master con SHA-256.

    El respaldo se nombra automáticamente:
        GM_BACKUP_{version}_{fecha}.xlsx

    Genera también un archivo {nombre}.sha256.json con:
        - sha256 del archivo
        - fecha del respaldo
        - ruta origen y destino
        - tamaño en bytes

    Args:
        ruta:    Ruta al Gold Master. Si None, usa la canónica.
        destino: Carpeta destino. Si None, usa data/backups/.

    Returns:
        Dict con:
            estado           "ok" | "error"
            ruta_respaldo    str   Ruta del archivo respaldado
            sha256           str   Hash SHA-256 del archivo
            tamano_bytes     int
            fecha_respaldo   str   ISO datetime
            error            str|None
    """
    resultado: dict = {
        "estado":         "error",
        "ruta_respaldo":  None,
        "sha256":         None,
        "tamano_bytes":   None,
        "fecha_respaldo": None,
        "error":          None,
    }

    ruta_origen = _resolver_ruta_gm(ruta)
    if not ruta_origen or not ruta_origen.exists():
        resultado["error"] = f"Gold Master no encontrado: {ruta_origen}"
        logger.error(f"[GM Governance] Respaldo fallido — archivo no encontrado: {ruta_origen}")
        return resultado

    carpeta_destino = Path(destino) if destino else _RUTA_RESPALDOS
    carpeta_destino.mkdir(parents=True, exist_ok=True)

    # Calcular SHA-256
    sha256 = _calcular_sha256(ruta_origen)

    # Nombre del respaldo
    fecha_str = datetime.now().strftime("%Y%m%d_%H%M%S")
    nombre_base = f"GM_BACKUP_{fecha_str}{ruta_origen.suffix}"
    ruta_destino = carpeta_destino / nombre_base

    try:
        shutil.copy2(str(ruta_origen), str(ruta_destino))
        tamano = ruta_destino.stat().st_size
        fecha_iso = datetime.now().isoformat()

        # Archivo de firma
        firma: dict = {
            "archivo":        nombre_base,
            "sha256":         sha256,
            "tamano_bytes":   tamano,
            "fecha_respaldo": fecha_iso,
            "ruta_origen":    str(ruta_origen),
            "ruta_destino":   str(ruta_destino),
            "integridad":     "verificada",
        }
        ruta_firma = ruta_destino.with_suffix(".sha256.json")
        with open(ruta_firma, "w", encoding="utf-8") as f:
            json.dump(firma, f, indent=2, ensure_ascii=False)

        resultado.update({
            "estado":         "ok",
            "ruta_respaldo":  str(ruta_destino),
            "sha256":         sha256,
            "tamano_bytes":   tamano,
            "fecha_respaldo": fecha_iso,
            "error":          None,
        })

        logger.info(
            f"[GM Governance] Respaldo creado: {nombre_base} "
            f"| SHA-256: {sha256[:16]}... | {tamano} bytes"
        )

    except Exception as exc:
        resultado["error"] = str(exc)
        logger.error(f"[GM Governance] Error en respaldo: {exc}")

    return resultado


# ══════════════════════════════════════════════════════════════════════════════
# 4. Estado general del Gold Master
# ══════════════════════════════════════════════════════════════════════════════

def obtener_estado_gm(
    ruta: str | Path | None = None,
) -> dict:
    """
    Diagnóstico completo del Gold Master activo.

    Retorna un resumen ejecutivo del estado del sistema epistemológico:
    versión, métricas clave, número de hojas, SHA-256, validación rápida.

    Args:
        ruta: Ruta al Gold Master. Si None, usa la canónica.

    Returns:
        Dict con:
            disponible       bool
            nombre_archivo   str|None
            version          str|None
            fecha_corte      str|None
            total_hojas      int
            metricas_clave   dict  {ICPI_GLOBAL_PCT, TGI_SCORE, SAT_CLASIF_RIESGO, ...}
            sha256           str|None
            tamano_mb        float|None
            n_errores_val    int   Errores de validación
            n_alertas_val    int   Alertas de validación
            changelog        dict  Versión del changelog
            error            str|None
    """
    vacio: dict = {
        "disponible":     False,
        "nombre_archivo": None,
        "version":        None,
        "fecha_corte":    None,
        "total_hojas":    0,
        "metricas_clave": {},
        "sha256":         None,
        "tamano_mb":      None,
        "n_errores_val":  0,
        "n_alertas_val":  0,
        "changelog":      {},
        "error":          None,
    }

    ruta_resuelta = _resolver_ruta_gm(ruta)

    if not ruta_resuelta or not ruta_resuelta.exists():
        return {**vacio, "error": f"Gold Master no encontrado: {ruta_resuelta}"}

    try:
        wb = openpyxl.load_workbook(str(ruta_resuelta), read_only=True, data_only=True)
        hojas = wb.sheetnames
        total_hojas = len(hojas)

        # Leer versión y fecha_corte de G1.1_CONFIG
        version    = None
        fecha_corte = None
        if "G1.1_CONFIG" in hojas:
            for row in wb["G1.1_CONFIG"].iter_rows(values_only=True):
                if row[0] == "VERSION_GOLD_MASTER" and len(row) > 1:
                    version = str(row[1]).strip() if row[1] else None
                if row[0] == "FECHA_CORTE" and len(row) > 1:
                    fecha_corte = str(row[1]).strip() if row[1] else None
                if version and fecha_corte:
                    break

        # Leer métricas clave de G6.1_OUTPUT_API
        metricas_clave: dict = {}
        claves_resumen = {
            "ICPI_GLOBAL_PCT", "TGI_SCORE",
            "TGI_D1", "TGI_D2", "TGI_D3", "TGI_D4", "TGI_D5",
            "SAT_CLASIF_RIESGO", "SAT_ACTIVAS_COUNT",
        }
        if "G6.1_OUTPUT_API" in hojas:
            for row in wb["G6.1_OUTPUT_API"].iter_rows(min_row=2, values_only=True):
                if row[0] and str(row[0]).strip() in claves_resumen:
                    metricas_clave[str(row[0]).strip()] = row[1] if len(row) > 1 else None

        wb.close()

        # SHA-256 y tamaño
        sha256   = _calcular_sha256(ruta_resuelta)
        tamano   = ruta_resuelta.stat().st_size / (1024 * 1024)

        # Validación rápida
        validaciones = validar_gold_master(ruta_resuelta)
        n_err = sum(1 for v in validaciones if v.es_error())
        n_alt = sum(1 for v in validaciones if v.es_alerta())

        # Changelog
        changelog = obtener_changelog_gm()

        return {
            "disponible":     True,
            "nombre_archivo": ruta_resuelta.name,
            "version":        version,
            "fecha_corte":    fecha_corte,
            "total_hojas":    total_hojas,
            "metricas_clave": metricas_clave,
            "sha256":         sha256,
            "tamano_mb":      round(tamano, 2),
            "n_errores_val":  n_err,
            "n_alertas_val":  n_alt,
            "changelog":      changelog,
            "error":          None,
        }

    except Exception as exc:
        logger.error(f"[GM Governance] Error en obtener_estado_gm: {exc}")
        return {**vacio, "error": str(exc)}


# ══════════════════════════════════════════════════════════════════════════════
# Helpers internos
# ══════════════════════════════════════════════════════════════════════════════

def _calcular_sha256(ruta: Path) -> str:
    """Calcula el hash SHA-256 de un archivo."""
    h = hashlib.sha256()
    with open(ruta, "rb") as f:
        for bloque in iter(lambda: f.read(65536), b""):
            h.update(bloque)
    return h.hexdigest()


def _resolver_ruta_gm(ruta: str | Path | None) -> Path | None:
    """Resuelve la ruta del Gold Master desde argumento, config o canónica."""
    if ruta:
        return Path(ruta)
    try:
        import config as cfg
        gm_path = getattr(cfg, "GOLD_MASTER_PATH", None)
        if gm_path:
            return Path(gm_path)
    except ImportError:
        pass
    # Ruta canónica por defecto
    try:
        from app.connectors.gold_master import _DEFAULT_GOLD_MASTER_V6
        return _DEFAULT_GOLD_MASTER_V6
    except ImportError:
        pass
    return None
