"""
QUIRA OS · p_seguimiento.py
Seguimiento Institucional — Sprint 2.6.2

Capacidad de aprendizaje operativo:
  • Aging de alertas — backlog, promedio, máximo días sin resolver
  • Reincidencia por entidad — patrones estructurales repetidos
  • Ranking de riesgo institucional — score compuesto auditado
  • Alertas sin resolver — señal de abandono o cuello de botella

Lenguaje institucional — sin terminología técnica.
Apto para: alcalde, dirección financiera, auditoría, Contraloría.

Dylus Lab © 2026
"""
from __future__ import annotations

import io
import streamlit as st
import streamlit.components.v1 as components

from quira_pages.html_engine import page_frame, page_header
from sentinel.alert_engine import (
    get_aging_report, get_reincidencia, get_ranking_riesgo, get_alertas_sin_resolver,
    ENTITY_LABELS, OK, WARNING, ERROR,
)

# ── CSS ───────────────────────────────────────────────────────────────────────
EXTRA_CSS = """
.seg-bloque-label {
  font-size: 10px; font-weight: 700; color: var(--muted);
  text-transform: uppercase; letter-spacing: .08em;
  margin: 18px 0 10px;
  padding-top: 14px; border-top: 1px solid var(--divider);
}
.seg-bloque-label:first-child { margin-top: 0; border-top: none; }

/* ── Ranking ── */
.rank-row {
  display: grid;
  grid-template-columns: 24px 1fr 54px 54px 54px 54px 80px;
  gap: 8px; align-items: center;
  padding: 8px 4px; border-bottom: 1px solid rgba(255,255,255,.04);
  font-size: 11px;
}
.rank-row:last-child { border-bottom: none; }
.rank-head {
  font-size: 9px; font-weight: 700; color: var(--muted);
  text-transform: uppercase; letter-spacing: .06em;
}
.rank-pos { font-family: var(--mono); font-size: 13px; font-weight: 700; }
.rank-ent { font-weight: 600; }
.rank-val { text-align: right; font-family: var(--mono); font-size: 12px; }
.nivel-badge {
  font-size: 9px; font-weight: 800; text-transform: uppercase;
  padding: 3px 9px; border-radius: 10px; text-align: center;
}

/* ── Reincidencia ── */
.reinc-grid {
  display: grid; grid-template-columns: repeat(2, 1fr);
  gap: 10px; margin-bottom: 6px;
}
@media (max-width: 600px) { .reinc-grid { grid-template-columns: 1fr; } }
.reinc-card {
  border-radius: 10px; padding: 12px 14px;
  border: 1px solid var(--divider); font-size: 11px;
}
.reinc-ent { font-weight: 700; font-size: 12px; margin-bottom: 6px; }
.reinc-meta { color: var(--muted); font-size: 10px; line-height: 1.8; }
.patron-tag {
  display: inline-block; font-size: 9px; font-weight: 800;
  padding: 2px 8px; border-radius: 8px; margin-top: 6px;
  background: rgba(255,77,109,.12); color: var(--red);
  text-transform: uppercase; letter-spacing: .05em;
}

/* ── Alertas sin resolver ── */
.cronico-row {
  display: grid;
  grid-template-columns: 1fr 90px 90px 70px;
  gap: 8px; align-items: center;
  padding: 8px 4px; border-bottom: 1px solid rgba(255,255,255,.04);
  font-size: 11px;
}
.cronico-row:last-child { border-bottom: none; }
.cronico-head {
  font-size: 9px; font-weight: 700; color: var(--muted);
  text-transform: uppercase; letter-spacing: .06em;
}
.dias-badge {
  font-size: 11px; font-weight: 800;
  padding: 3px 10px; border-radius: 8px;
  text-align: center;
}
.no-cronicos {
  text-align: center; padding: 24px 16px;
  font-size: 13px; font-weight: 700; color: var(--green);
}

.seg-footer {
  text-align: center; font-size: 9px; color: var(--muted);
  margin-top: 18px;
}
"""

# ── COLORES ───────────────────────────────────────────────────────────────────
_CLR    = {OK: "var(--green)", WARNING: "var(--amber)", ERROR: "var(--red)"}
_BG     = {OK: "rgba(0,224,150,.08)", WARNING: "rgba(255,184,0,.08)", ERROR: "rgba(255,77,109,.08)"}
_ACCENT = {OK: "#00E096", WARNING: "#FFB800", ERROR: "#FF4D6D"}
_TIPO_LABEL = {"ejecucion": "Ejecución", "cobertura": "Cobertura"}


# ── HTML HELPERS ──────────────────────────────────────────────────────────────
def _ranking_html(ranking: list[dict]) -> str:
    if not ranking:
        return '<div style="color:var(--muted);font-size:12px;padding:16px 0">Sin datos disponibles.</div>'

    header = """
<div class="rank-row rank-head">
  <span>#</span>
  <span>Institución</span>
  <span style="text-align:right">Críticas</span>
  <span style="text-align:right">Advert.</span>
  <span style="text-align:right">Días máx</span>
  <span style="text-align:right">Reincid.</span>
  <span style="text-align:center">Nivel</span>
</div>"""

    rows = ""
    for i, r in enumerate(ranking, 1):
        niv   = r["nivel"]
        clr   = _CLR.get(niv, "var(--white)")
        bg    = _BG.get(niv, "transparent")
        bdr   = _ACCENT.get(niv, "#666")
        label_map = {OK: "Controlado", WARNING: "Atención", ERROR: "Alto Riesgo"}
        niv_lbl = label_map.get(niv, niv)

        pos_clr = "var(--red)" if i == 1 and niv == ERROR else (
            "var(--amber)" if i <= 2 and niv != OK else "var(--muted)"
        )

        rows += f"""
<div class="rank-row" style="background:{bg if i % 2 == 0 else 'transparent'}">
  <span class="rank-pos" style="color:{pos_clr}">{i}</span>
  <span class="rank-ent">{r['entidad_label']}</span>
  <span class="rank-val" style="color:{_CLR.get(ERROR,'var(--red)') if r['n_criticas'] > 0 else 'var(--muted)'}">
    {r['n_criticas']}
  </span>
  <span class="rank-val" style="color:{_CLR.get(WARNING,'var(--amber)') if r['n_advertencias'] > 0 else 'var(--muted)'}">
    {r['n_advertencias']}
  </span>
  <span class="rank-val">{r['dias_max']}d</span>
  <span class="rank-val">{r['meses_reincidencia']} meses</span>
  <span>
    <span class="nivel-badge" style="background:{bg};color:{clr};border:1px solid {bdr}40">
      {niv_lbl}
    </span>
  </span>
</div>"""

    return header + rows


def _reincidencia_html(reincidencias: list[dict]) -> str:
    if not reincidencias:
        return '<div style="color:var(--green);font-size:12px;padding:16px 0">✓ Sin patrones de reincidencia detectados.</div>'

    cards = ""
    for r in reincidencias[:8]:
        niv = r.get("status", WARNING)
        clr = _CLR.get(niv, "var(--white)")
        bg  = _BG.get(niv, "transparent")
        bdr = _ACCENT.get(niv, "#666")
        tipo_lbl = _TIPO_LABEL.get(r["tipo"], r["tipo"].title())
        patron_tag = '<span class="patron-tag">⚠ Patrón Estructural</span>' if r["es_patron"] else ""

        cards += f"""
<div class="reinc-card" style="background:{bg};border-color:{bdr}40">
  <div class="reinc-ent" style="color:{clr}">{r['entidad_label']}</div>
  <div class="reinc-meta">
    Tipo: {tipo_lbl}<br>
    Períodos afectados: <b style="color:{clr}">{r['n_periodos']}</b><br>
    Alertas críticas: <b>{r['n_criticas']}</b><br>
    Primera vez: {r['primera_vez']} · Última: {r['ultima_vez']}
  </div>
  {patron_tag}
</div>"""

    return f'<div class="reinc-grid">{cards}</div>'


def _cronicos_html(cronicos: list[dict]) -> str:
    if not cronicos:
        return '<div class="no-cronicos">✓ Sin alertas crónicas</div>'

    header = """
<div class="cronico-row cronico-head">
  <span>Alerta</span>
  <span>Institución</span>
  <span>Tipo</span>
  <span style="text-align:right">Días abierta</span>
</div>"""

    rows = ""
    for a in cronicos:
        dias  = a.get("dias_abierta", 0)
        niv   = a.get("status", WARNING)
        clr   = _CLR.get(niv, "var(--white)")
        bg    = _BG.get(niv, "transparent")
        tipo_lbl = _TIPO_LABEL.get(a.get("tipo", ""), a.get("tipo", ""))

        dias_clr = "var(--red)" if dias >= 90 else ("var(--amber)" if dias >= 60 else clr)
        dias_bg  = ("rgba(255,77,109,.15)" if dias >= 90
                    else "rgba(255,184,0,.12)" if dias >= 60
                    else "rgba(255,184,0,.08)")

        rows += f"""
<div class="cronico-row">
  <span style="color:var(--white)">{(a.get('titulo') or '')[:60]}</span>
  <span style="color:var(--muted)">{a.get('entidad','–')}</span>
  <span style="color:var(--muted)">{tipo_lbl}</span>
  <span style="text-align:right">
    <span class="dias-badge" style="background:{dias_bg};color:{dias_clr}">
      {dias}d
    </span>
  </span>
</div>"""

    return header + rows


def _build_html(
    aging: dict,
    ranking: list[dict],
    reincidencias: list[dict],
    cronicos: list[dict],
) -> str:
    hdr = page_header(
        "📊 SEGUIMIENTO",
        "Seguimiento Institucional",
        "Aging · Reincidencia · Ranking de Riesgo · Alertas Crónicas",
        '<span class="badge badge-amber">Sprint 2.6.2</span>',
    )

    # ── Aging table ──────────────────────────────────────────────────────────
    aging_ent_rows = ""
    for ed in aging["por_entidad"]:
        clr = "var(--red)" if ed["dias_max"] >= 60 else (
            "var(--amber)" if ed["dias_max"] >= 30 else "var(--green)")
        aging_ent_rows += f"""
<div style="display:grid;grid-template-columns:1fr 60px 60px 60px;
            gap:8px;padding:6px 4px;border-bottom:1px solid rgba(255,255,255,.04);
            font-size:11px;align-items:center">
  <span style="font-weight:600">{ENTITY_LABELS.get(ed['entidad'], ed['entidad'])}</span>
  <span style="text-align:right;color:var(--muted)">{ed['n_abiertas']}</span>
  <span style="text-align:right;color:{'var(--red)' if ed['n_criticas'] > 0 else 'var(--muted)'}">{ed['n_criticas']}</span>
  <span style="text-align:right;color:{clr};font-weight:700;font-family:var(--mono)">{ed['dias_max']}d</span>
</div>"""

    if not aging_ent_rows:
        aging_ent_rows = '<div style="color:var(--green);font-size:12px;padding:12px 0">✓ Sin backlog activo</div>'

    aging_header = """
<div style="display:grid;grid-template-columns:1fr 60px 60px 60px;
            gap:8px;padding:4px 4px 6px;
            font-size:9px;font-weight:700;color:var(--muted);
            text-transform:uppercase;letter-spacing:.06em;
            border-bottom:1px solid var(--divider)">
  <span>Institución</span>
  <span style="text-align:right">Abiertas</span>
  <span style="text-align:right">Críticas</span>
  <span style="text-align:right">Días máx</span>
</div>"""

    # ── Contenido ─────────────────────────────────────────────────────────────
    n_cronicos = len(cronicos)
    cron_label = (
        f'<span style="color:var(--red);font-weight:800">{n_cronicos} alerta{"s" if n_cronicos != 1 else ""} crónica{"s" if n_cronicos != 1 else ""}</span>'
        if n_cronicos > 0 else
        '<span style="color:var(--green)">Sin alertas crónicas</span>'
    )

    content = f"""
<div class="seg-bloque-label" style="margin-top:0;border-top:none">
  Backlog de Alertas · Distribución por Institución
</div>
{aging_header}
{aging_ent_rows}

<div class="seg-bloque-label">Ranking de Riesgo Institucional</div>
{_ranking_html(ranking)}

<div class="seg-bloque-label">Reincidencia · Patrones Detectados</div>
{_reincidencia_html(reincidencias)}

<div class="seg-bloque-label">
  Alertas sin Resolver · Más de 45 días · {cron_label}
</div>
{_cronicos_html(cronicos)}

<div class="seg-footer">
  Seguimiento Institucional · Sprint 2.6.2 · Dylus Lab © 2026<br>
  <span style="font-size:8px">
    Score riesgo = (críticas×3) + (advertencias×1) + (días//15) + (reincidencias×2) ·
    Metodología auditada — sin aprendizaje automático
  </span>
</div>"""

    return hdr + content


# ── EXPORT XLSX ───────────────────────────────────────────────────────────────
def _export_xlsx(aging: dict, ranking: list[dict],
                 reincidencias: list[dict], cronicos: list[dict]) -> bytes:
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from datetime import datetime as _dt

        wb  = openpyxl.Workbook()
        _HF = PatternFill("solid", fgColor="0D1F3C")
        _HT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        _NM = Font(name="Arial", size=10)
        _BL = Font(name="Arial", bold=True, size=10)
        _BD = Border(bottom=Side(style="thin", color="DDDDDD"))
        _CT = Alignment(horizontal="center", vertical="center")

        def _hrow(ws, row, cols, widths=None):
            for c, v in enumerate(cols, 1):
                cell = ws.cell(row=row, column=c, value=v)
                cell.font = _HT; cell.fill = _HF; cell.alignment = _CT
            if widths:
                for c, w in enumerate(widths, 1):
                    ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = w

        # ── Hoja 1: Ranking ──────────────────────────────────────────────────
        ws1 = wb.active
        ws1.title = "Ranking de Riesgo"
        ws1["A1"] = "QUIRA OS · Seguimiento Institucional — Ranking de Riesgo"
        ws1["A1"].font = Font(name="Arial", bold=True, size=13)
        ws1["A2"] = f"Generado: {_dt.now().strftime('%d %b %Y %H:%M')} · Dylus Lab · GAD Montecristi"
        ws1["A2"].font = Font(name="Arial", size=9, color="888888")
        _hrow(ws1, 4, ["#", "Institución", "Score", "Críticas", "Advertencias",
                       "Días máx", "Meses reinc.", "Nivel"],
              [4, 38, 10, 12, 14, 12, 14, 16])
        for i, r in enumerate(ranking, 5):
            label_map = {"ok": "Controlado", "warning": "Atención", "error": "Alto Riesgo"}
            for col, val in enumerate([
                i - 4, r["entidad_label"], r["score"], r["n_criticas"],
                r["n_advertencias"], r["dias_max"], r["meses_reincidencia"],
                label_map.get(r["nivel"], r["nivel"]),
            ], 1):
                c = ws1.cell(row=i, column=col, value=val)
                c.font = _NM; c.border = _BD

        # ── Hoja 2: Reincidencia ─────────────────────────────────────────────
        ws2 = wb.create_sheet("Reincidencia")
        _hrow(ws2, 1, ["Institución", "Tipo", "Períodos", "Críticas",
                       "Primera vez", "Última vez", "Patrón estructural"],
              [38, 16, 12, 12, 14, 14, 20])
        for i, r in enumerate(reincidencias, 2):
            for col, val in enumerate([
                r["entidad_label"], _TIPO_LABEL.get(r["tipo"], r["tipo"].title()),
                r["n_periodos"], r["n_criticas"],
                r["primera_vez"], r["ultima_vez"],
                "Sí" if r["es_patron"] else "No",
            ], 1):
                ws2.cell(row=i, column=col, value=val).font = _NM

        # ── Hoja 3: Alertas crónicas ─────────────────────────────────────────
        ws3 = wb.create_sheet("Alertas Crónicas (+45d)")
        _hrow(ws3, 1, ["Alerta", "Institución", "Tipo", "Estado", "Días abierta"],
              [42, 20, 16, 16, 14])
        _REF = {"pendiente": "Abierta", "en_revision": "En Revisión"}
        for i, a in enumerate(cronicos, 2):
            for col, val in enumerate([
                a.get("titulo", ""),
                a.get("entidad", ""),
                _TIPO_LABEL.get(a.get("tipo", ""), a.get("tipo", "")),
                _REF.get(a.get("estado", ""), a.get("estado", "")),
                a.get("dias_abierta", 0),
            ], 1):
                ws3.cell(row=i, column=col, value=val).font = _NM

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()
    except Exception:
        return b""


# ── RENDER ────────────────────────────────────────────────────────────────────
def render() -> None:
    col_t, col_btn = st.columns([5, 1])
    with col_t:
        st.caption("**Seguimiento Institucional** · Aging · Reincidencia · Riesgo")
    with col_btn:
        refresh = st.button("↻ Actualizar", use_container_width=True,
                            help="Recalcular métricas de seguimiento")

    cache_key = "seguimiento_result"
    if refresh or cache_key not in st.session_state:
        with st.spinner("Analizando patrones institucionales…"):
            aging_data    = get_aging_report()
            ranking_data  = get_ranking_riesgo()
            reinc_data    = get_reincidencia(min_meses=2)
            cronicos_data = get_alertas_sin_resolver(umbral_dias=45)
            st.session_state[cache_key] = {
                "aging":    aging_data,
                "ranking":  ranking_data,
                "reinc":    reinc_data,
                "cronicos": cronicos_data,
            }

    d = st.session_state[cache_key]
    aging    = d["aging"]
    ranking  = d["ranking"]
    reinc    = d["reinc"]
    cronicos = d["cronicos"]

    # ── KPIs nativos ──────────────────────────────────────────────────────────
    k1, k2, k3, k4 = st.columns(4)
    k1.metric("📋 Backlog activo",       aging["backlog_total"],
              help="Alertas sin resolver actualmente")
    k2.metric("📅 Días sin resolver (máx.)", f"{aging['dias_max']}d")
    k3.metric("⏱ Días promedio resolución",  f"{aging['resolucion_promedio']}d",
              help="Tiempo medio hasta cierre (alertas resueltas)")
    k4.metric("⚠ Alertas crónicas (+45d)",   len(cronicos),
              delta=len(cronicos) if len(cronicos) > 0 else None,
              delta_color="inverse")

    # ── Iframe principal ──────────────────────────────────────────────────────
    html  = _build_html(aging, ranking, reinc, cronicos)
    full  = page_frame(html, show_tech=False, extra_css=EXTRA_CSS)

    n_reinc   = len(reinc)
    n_cronicos = len(cronicos)
    height = max(900, 420 + len(ranking) * 44 + n_reinc * 70 + n_cronicos * 36)
    components.html(full, height=height, scrolling=False)

    # ── Export ────────────────────────────────────────────────────────────────
    st.markdown("---")
    xlsx = _export_xlsx(aging, ranking, reinc, cronicos)
    if xlsx:
        from datetime import date as _date
        st.download_button(
            label="📊 Exportar Seguimiento a Excel (Auditoría / Contraloría)",
            data=xlsx,
            file_name=f"seguimiento_institucional_{_date.today().strftime('%Y%m%d')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

    # ── Navegación ────────────────────────────────────────────────────────────
    c1, c2 = st.columns(2)
    with c1:
        if st.button("🔔 Ver Alertas de Cumplimiento", use_container_width=True):
            st.session_state["page"] = "alertas"
            st.rerun()
    with c2:
        if st.button("⬡ Centro de Control", use_container_width=True):
            st.session_state["page"] = "sentinel_hub"
            st.rerun()
