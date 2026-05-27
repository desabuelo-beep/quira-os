
import openpyxl, sys
sys.stdout.reconfigure(encoding='utf-8', errors='replace')

GM = r'C:\Users\DELL\Desktop\Javo\Dylus Lab\ProyecT\SIAP-ICPI_GOLD_MASTER_v5.5_TGI.xlsx'
wb = openpyxl.load_workbook(GM, data_only=True)

def show(name, head=5, tail=5):
    try:
        ws = wb[name]
        rows = list(ws.iter_rows(values_only=True))
        data = [r for r in rows if any(c for c in r if c is not None)]
        print(f"\n=== {name} === ({len(data)} filas con datos)")
        for r in data[:head]:
            print(f"  {r}")
        if len(data) > head + tail:
            print(f"  ... ({len(data)-head-tail} filas intermedias) ...")
        for r in data[-tail:]:
            print(f"  {r}")
    except Exception as e:
        print(f"\n=== {name} === ERROR: {e}")

# Sheets relevantes para actualizar
show('H84_SNAPSHOT_REGISTRY', head=10, tail=3)
show('H85_ALERTS_LOG', head=4, tail=8)
show('H36c_OBSIDIAN_MAP', head=3, tail=10)
show('H80_MODEL_REGISTRY', head=10, tail=3)

# H07b — leer todas las filas para ver estructura Ti
name07b = [s for s in wb.sheetnames if 'H07b' in s][0]
show(name07b, head=30, tail=3)

# H90 — ver si existe
name90 = [s for s in wb.sheetnames if 'H90' in s]
if name90:
    show(name90[0], head=8, tail=5)

# H_HOLDING
nameH = [s for s in wb.sheetnames if 'HOLDING' in s]
if nameH:
    show(nameH[0], head=8, tail=5)
