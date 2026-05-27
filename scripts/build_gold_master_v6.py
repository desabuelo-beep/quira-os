"""
scripts/build_gold_master_v6.py — QUIRA OS
Construye el Gold Master TGI v6.0 desde cero con terminología canónica TGI.

34 hojas / 7 grupos:
  G1 Configuración (5) · G2 Fuentes (4) · G3 Dimensiones (8)
  G4 Índices (5) · G5 SAT (5) · G6 Outputs (5) · G7 Gobernanza (2)

El Gold Master v5.5 permanece intacto como referencia histórica.
Conector Python: gold_master.py → leer G6.1_OUTPUT_API (mismo contrato que H73).

Dylus Lab © 2026
"""
from __future__ import annotations

import sys
from pathlib import Path
from datetime import date

try:
    from openpyxl import Workbook
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, numbers
    )
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl no instalado. Ejecutar: pip install openpyxl")
    sys.exit(1)

# ── Rutas ──────────────────────────────────────────────────────────────────────
OUTPUT_PATH = Path(
    r"C:\Users\DELL\Desktop\Javo\Dylus Lab\ProyecT"
    r"\TGI_GOLD_MASTER_v6.0_20260525.xlsx"
)

# ── Paleta de colores QUIRA ────────────────────────────────────────────────────
C = {
    # Headers de grupo
    "g1": "1A3A5C",   # Azul marino — Configuración
    "g2": "1E4D6B",   # Azul petróleo — Fuentes
    "g3": "0D4F3C",   # Verde oscuro — Dimensiones
    "g4": "4A235A",   # Púrpura — Índices
    "g5": "7B1C1C",   # Rojo oscuro — SAT
    "g6": "1C4B7B",   # Azul — Outputs
    "g7": "3D3D3D",   # Gris — Gobernanza
    # Semáforo
    "verde":    "00C896",
    "amarillo": "FFB700",
    "naranja":  "FF8C00",
    "rojo":     "FF4D6D",
    # Fondo alternado
    "fila_par":  "F0F4F8",
    "fila_impar":"FFFFFF",
    # Texto IP (input hardcoded) — azul estándar industria
    "input_blue": "0000FF",
    # Header texto
    "header_text": "FFFFFF",
}

# ── Estilos base ───────────────────────────────────────────────────────────────
FONT_BASE    = Font(name="Arial", size=10)
FONT_BOLD    = Font(name="Arial", size=10, bold=True)
FONT_HEADER  = Font(name="Arial", size=10, bold=True, color=C["header_text"])
FONT_TITLE   = Font(name="Arial", size=12, bold=True)
FONT_INPUT   = Font(name="Arial", size=10, color=C["input_blue"])
FONT_FORMULA = Font(name="Arial", size=10, color="000000")

ALIGN_LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=False)
ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
ALIGN_RIGHT  = Alignment(horizontal="right",  vertical="center")
ALIGN_WRAP   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

THIN  = Side(style="thin",   color="CCCCCC")
MED   = Side(style="medium", color="999999")
BORDER_THIN = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _group_header(ws, title: str, subtitle: str, color: str,
                  row: int = 1, max_col: int = 8) -> None:
    """Bloque de título de grupo en fila 1-2."""
    ws.merge_cells(start_row=row,   start_column=1, end_row=row,   end_column=max_col)
    ws.merge_cells(start_row=row+1, start_column=1, end_row=row+1, end_column=max_col)
    c1 = ws.cell(row=row,   column=1, value=title)
    c2 = ws.cell(row=row+1, column=1, value=subtitle)
    c1.font      = Font(name="Arial", size=13, bold=True, color="FFFFFF")
    c1.fill      = _fill(color)
    c1.alignment = ALIGN_CENTER
    c2.font      = Font(name="Arial", size=9, italic=True, color="FFFFFF")
    c2.fill      = _fill(color)
    c2.alignment = ALIGN_CENTER


def _col_header(ws, row: int, headers: list[tuple[str, int]],
                color: str = "374151") -> None:
    """Encabezados de columnas con ancho automático."""
    for col_idx, (label, width) in enumerate(headers, start=1):
        c = ws.cell(row=row, column=col_idx, value=label)
        c.font      = FONT_HEADER
        c.fill      = _fill(color)
        c.alignment = ALIGN_CENTER
        c.border    = BORDER_THIN
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _data_row(ws, row: int, values: list, bold: bool = False,
              fill_color: str | None = None) -> None:
    """Fila de datos con borde fino."""
    bg = fill_color or (C["fila_par"] if row % 2 == 0 else C["fila_impar"])
    for col_idx, val in enumerate(values, start=1):
        c = ws.cell(row=row, column=col_idx, value=val)
        c.font   = Font(name="Arial", size=10, bold=bold)
        c.fill   = _fill(bg)
        c.border = BORDER_THIN
        c.alignment = ALIGN_LEFT


def _input_cell(ws, row: int, col: int, value,
                comment: str = "") -> None:
    """Celda de input (azul — hardcoded)."""
    c = ws.cell(row=row, column=col, value=value)
    c.font      = FONT_INPUT
    c.alignment = ALIGN_LEFT
    c.border    = BORDER_THIN


# ══════════════════════════════════════════════════════════════════════════════
# CONSTRUCTORES POR HOJA
# ══════════════════════════════════════════════════════════════════════════════

def build_g1_1_config(ws) -> None:
    _group_header(ws,
        "G1.1 — CONFIG | Parámetros Globales Gold Master TGI v6.0",
        "Versión · Municipio · Pesos TGI · Umbrales · Fecha de corte",
        C["g1"])
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 16

    # Sección: Identificación
    r = 4
    ws.cell(r, 1, "── IDENTIFICACIÓN ──────────────────────────────────────").font = FONT_BOLD
    params_id = [
        ("VERSION_GOLD_MASTER",  "v6.0",                   "Versión canónica del Gold Master"),
        ("FECHA_CONSTRUCCION",   "2026-05-25",              "Fecha de construcción v6.0"),
        ("FECHA_CORTE",          "2026-05-18",              "Fecha de corte datos actuales"),
        ("MUNICIPIO_NOMBRE",     "Montecristi",             "Nombre del municipio"),
        ("MUNICIPIO_CODIGO",     "130801",                  "Código INEC del municipio"),
        ("MUNICIPIO_PROVINCIA",  "Manabí",                  "Provincia"),
        ("PERIODO_MANDATO",      "2023-2027",               "Período de mandato actual"),
        ("VERSION_ANTERIOR",     "v5.5 (congelado)",        "Gold Master de referencia histórica"),
    ]
    _col_header(ws, r+1, [("Parámetro", 32), ("Valor", 28), ("Descripción", 52)], C["g1"])
    for i, (k, v, d) in enumerate(params_id, start=r+2):
        _input_cell(ws, i, 1, k)
        _input_cell(ws, i, 2, v)
        ws.cell(i, 3, d).font = FONT_BASE

    # Sección: Pesos TGI
    r2 = r + len(params_id) + 4
    ws.cell(r2, 1, "── PESOS TGI D1-D5 ─────────────────────────────────────").font = FONT_BOLD
    _col_header(ws, r2+1, [("Dimensión", 32), ("Peso (%)", 14), ("Justificación", 52)], C["g3"])
    pesos = [
        ("D1 — Legalidad y Coherencia Normativa",    0.20, "Condición necesaria. Sin legalidad, el resto colapsa"),
        ("D2 — Fidelidad a la Planificación",        0.20, "Compromiso ejecutado vs. prometido (PDOT/POA)"),
        ("D3 — Ejecución Presupuestaria",            0.25, "Mayor peso: sub-ejecución es el problema crónico"),
        ("D4 — Equidad Territorial",                 0.25, "Mayor peso: regresividad es el otro problema crónico"),
        ("D5 — Capacidad Institucional",             0.10, "Piso mínimo: sin capacidad, los 4 anteriores colapsan"),
        ("TOTAL",                                    1.00, "Suma = 1.00"),
    ]
    for i, (dim, peso, just) in enumerate(pesos, start=r2+2):
        bold = dim == "TOTAL"
        _data_row(ws, i, [dim, f"{peso:.0%}", just], bold=bold)

    # Sección: Umbrales AVEP
    r3 = r2 + len(pesos) + 4
    ws.cell(r3, 1, "── UMBRALES AVEP ────────────────────────────────────────").font = FONT_BOLD
    _col_header(ws, r3+1, [("Nivel", 26), ("ICPI Mín (%)", 14), ("ICPI Máx (%)", 14), ("Color", 14), ("Descripción", 40)], C["g5"])
    avep = [
        ("🟢 Gestión por Mandato",   75, 100, "#00C896", "Ejecución alineada al PDOT"),
        ("🟡 Transición Crítica",    60,  74, "#FFB700", "SAT-I/II activa — decisiones correctivas"),
        ("🟠 Gestión por Ocurrencia", 50,  59, "#FF8C00", "SAT-III reincidencia — reorientación urgente"),
        ("🔴 Nivel de Atención Alta",  0,  49, "#FF4D6D", "SAT-IV/V/VI — protocolo de emergencia"),
    ]
    for i, row_data in enumerate(avep, start=r3+2):
        _data_row(ws, i, list(row_data))

    # Ancho columna A
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 54


def build_g1_2_municipio(ws) -> None:
    _group_header(ws, "G1.2 — MUNICIPIO_PROFILE | Perfil GADM Montecristi",
        "Datos oficiales del municipio: área, población, parroquias, holding", C["g1"])
    _col_header(ws, 4, [("Campo", 36), ("Valor", 22), ("Fuente", 28), ("Año", 10)], C["g1"])
    perfil = [
        ("Nombre oficial",            "GAD Municipal de Montecristi",      "COOTAD Art. 53",     2023),
        ("Código INEC",               "130801",                             "INEC",               2023),
        ("Área total (km²)",          335.8,                                "PDOT 2023-2027",     2023),
        ("Población total",           100947,                               "Censo INEC 2022",    2022),
        ("Población urbana",          69582,                                "Censo INEC 2022",    2022),
        ("Población rural",           31365,                                "Censo INEC 2022",    2022),
        ("Parroquias totales",        7,                                    "COOTAD",             2023),
        ("Parroquia urbana",          "Montecristi (cabecera cantonal)",    "GADM",               2023),
        ("Parroquias rurales",        6,                                    "COOTAD",             2023),
        ("Presupuesto codificado GAD","$8,245,612",                         "eSIGEF 2026",        2026),
        ("Entidades Holding",        4,                                    "Resolución 040-2025", 2025),
        ("Período de mandato",        "2023-2027",                          "CNE",                2023),
        ("Alcalde",                   "[Nombre del Alcalde]",              "CNE",                2023),
        ("NBI cantonal (%)",          35.2,                                 "INEC / PDOT",        2022),
        ("NBI rural promedio (%)",    55.7,                                 "INEC / PDOT",        2022),
        ("IRS global",                79.7,                                 "Motor TGI v5.5",     2026),
        ("Brecha rural USD",          1791935,                              "Motor TGI v5.5",     2026),
    ]
    for i, row_data in enumerate(perfil, start=5):
        _data_row(ws, i, list(row_data))
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 34
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 10


def build_g1_3_holding(ws) -> None:
    _group_header(ws, "G1.3 — HOLDING_ENTIDADES | Registro de Entidades Municipales",
        "GAD Central · EMAI · Bomberos · Patronato — códigos, responsables, tipo", C["g1"])
    _col_header(ws, 4, [
        ("Código", 10), ("Entidad", 34), ("Tipo", 22),
        ("Responsable", 22), ("RUC", 16), ("Fuente Ti", 22), ("Reliability", 12)
    ], C["g1"])
    entidades = [
        ("GAD-01", "GAD Municipal de Montecristi",       "Gobierno Autónomo Descentralizado", "Alcalde Municipal",  "1360000410001", "eSIGEF / DPE", 0.95),
        ("GAD-02", "EMAI (EP Aseo Integral Montecristi)", "Empresa Pública Municipal",          "Gerente EP Aseo",    "1368000830001", "Presupuesto EP", 0.85),
        ("GAD-03", "Cuerpo de Bomberos Montecristi",     "Entidad Adscrita",                   "Jefe de Bomberos",   "1360004100001", "Presupuesto CB", 0.85),
        ("GAD-04", "Patronato Municipal Montecristi",    "Entidad Adscrita",                   "Presidente Patronato","1360004100001", "Presupuesto PM", 0.85),
    ]
    for i, row_data in enumerate(entidades, start=5):
        _data_row(ws, i, list(row_data))
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 36
    ws.column_dimensions["C"].width = 24
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 22
    ws.column_dimensions["G"].width = 12


def build_g1_5_changelog(ws) -> None:
    _group_header(ws, "G1.5 — CHANGELOG | Historial de Versiones Gold Master",
        "Trazabilidad documental del Gold Master — cada versión preservada", C["g1"])
    _col_header(ws, 4, [
        ("Versión", 12), ("Fecha", 14), ("Cambio Principal", 60), ("SHA-256 (archivo)", 68)
    ], C["g1"])
    log = [
        ("v5.0", "2026-05-01", "Base TGI 3D inicial — motor SIAP-ICPI primigenio",                        "pendiente"),
        ("v5.1", "2026-05-08", "Agrega D4 Equidad territorial — IET por parroquia",                        "pendiente"),
        ("v5.2", "2026-05-10", "IRS Global + alertas cantonales",                                           "pendiente"),
        ("v5.3", "2026-05-14", "TGI 5D completo, 7 parroquias validadas",                                  "pendiente"),
        ("v5.4", "2026-05-16", "Brecha USD · Prioridad Reequil · Clasificación Equidad",                   "pendiente"),
        ("v5.5", "2026-05-18", "TGI terminology · SAT_Catalogo · ICPI 53.56% validado — CONGELADO",       "pendiente"),
        ("v6.0", "2026-05-25", "Refactorización completa: 34 hojas/7 grupos, nomenclatura TGI canónica",  "pendiente"),
    ]
    for i, row_data in enumerate(log, start=5):
        _data_row(ws, i, list(row_data), bold=(row_data[0] in ("v5.5", "v6.0")))
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 62
    ws.column_dimensions["D"].width = 70


def build_g2_1_fuentes(ws) -> None:
    _group_header(ws, "G2.1 — FUENTES_INSTITUCIONALES | Catálogo de Fuentes",
        "Todas las fuentes que alimentan el Motor TGI con reliability y frecuencia", C["g2"])
    _col_header(ws, 4, [
        ("ID Fuente", 14), ("Nombre", 36), ("Tipo", 16),
        ("URL / Acceso", 40), ("Frecuencia", 14), ("Reliability", 12), ("Dimensión TGI", 16)
    ], C["g2"])
    fuentes = [
        ("GOLD_MASTER", "Gold Master TGI (este archivo)",          "Doctrinal",    "Interno",                             "Cada versión", 0.99, "D1-D5"),
        ("DPE",         "Datos Públicos Ecuador / eSIGEF",         "API Oficial",  "dpe.fin.gob.ec",                      "Mensual",       0.95, "D3"),
        ("SERCOP",      "SERCOP OCDS — Contratación Pública",      "API Oficial",  "ocds.sercop.gob.ec",                  "Mensual",       0.95, "D3"),
        ("CPCCS",       "CPCCS — Rendición de Cuentas",            "Web Scraping", "cpccs.gob.ec/rendicion-cuentas",      "Anual",         0.80, "D1/D5"),
        ("SIGAD",       "SIGAD SENPLADES — Planificación",         "Web",          "sigad.senplades.gob.ec",              "Trimestral",    0.85, "D2"),
        ("INEC",        "INEC — Censos y estadísticas",            "Descarga",     "ecuadorencifras.gob.ec",              "Censal",        0.95, "D4"),
        ("PDOT",        "PDOT Montecristi 2023-2027",              "Doctrinal",    "GAD Montecristi / Resolución Concejo", "Cuatrianual",   0.99, "D1/D2"),
        ("CNE",         "CNE — Plan de Trabajo Alcalde",           "Doctrinal",    "cne.gob.ec",                          "Cuatrianual",   0.95, "D1/D2"),
        ("EMAI",        "EP Aseo — Presupuesto mensual",           "Doctrinal",    "Ingesta manual H07c",                 "Mensual",       0.85, "D3"),
        ("BOMBEROS",    "Cuerpo Bomberos — Presupuesto mensual",   "Doctrinal",    "Ingesta manual H07c",                 "Mensual",       0.85, "D3"),
        ("PATRONATO",   "Patronato Municipal — Presupuesto mensual","Doctrinal",   "Ingesta manual H07c",                 "Mensual",       0.85, "D3"),
        ("SOCIAL",      "Evidencia Social / Comunidades",          "Manual",       "Trabajo de campo",                    "Variable",      0.45, "D4"),
    ]
    for i, row_data in enumerate(fuentes, start=5):
        _data_row(ws, i, list(row_data))


def build_g2_2_reliability(ws) -> None:
    _group_header(ws, "G2.2 — RELIABILITY_SCORE | Scores de Confiabilidad por Fuente",
        "Escala 0.00-1.00 · 0.99=Canónica · 0.95=Operativa · 0.80=Funcional · 0.45=Manual", C["g2"])
    _col_header(ws, 4, [
        ("Fuente", 22), ("Reliability", 12), ("Nivel", 16),
        ("Razón", 50), ("Última verificación", 20)
    ], C["g2"])
    datos = [
        ("Gold Master (propio)",     0.99, "Canónico",  "Motor propietario validado — source of truth absoluto",    "2026-05-25"),
        ("DPE API / eSIGEF",         0.95, "Operativo", "API oficial del Estado, datos devengados verificados",     "2026-05-18"),
        ("SERCOP OCDS",              0.95, "Operativo", "Estándar OCDS internacional, contratos públicos trazables","2026-05-18"),
        ("SIGAD / SENPLADES",        0.85, "Funcional", "Auto-reporte GAD — sujeto a subjetividad institucional",   "2026-05-18"),
        ("CPCCS RdC",                0.80, "Funcional", "Requiere procesamiento manual de PDF/web scraping",        "2026-05-18"),
        ("INEC",                     0.95, "Operativo", "Censos oficiales, alta confiabilidad estadística",         "2022-11-01"),
        ("Ingesta mensual (H07c)",   0.85, "Funcional", "Firmada por Director responsable — SHA-256 validado",     "Mensual"),
        ("Evidencia social / campo", 0.45, "Manual",    "No verificable sistémicamente — solo contexto cualitativo","Variable"),
    ]
    for i, row_data in enumerate(datos, start=5):
        _data_row(ws, i, list(row_data))


def build_g3_dimensiones(ws, dim_code: str, dim_nombre: str,
                          peso: float, color: str, datos_tabla: list) -> None:
    """Template genérico para hojas G3.x_D*."""
    titulo = f"G3.{dim_code} — D{dim_code[-1]}_{dim_nombre.upper().replace(' ', '_')}"
    _group_header(ws, titulo,
        f"Dimensión TGI: {dim_nombre} · Peso {peso:.0%} en ICPI Global", color)
    # Ficha dimensional
    _col_header(ws, 4, [("Campo", 32), ("Valor", 22), ("Fuente / Nota", 48)], color)
    for i, row_data in enumerate(datos_tabla, start=5):
        _data_row(ws, i, list(row_data))
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 50


def build_g4_1_icpi(ws) -> None:
    _group_header(ws, "G4.1 — ICPI_GLOBAL | Índice Compuesto de Progreso Institucional",
        "ICPI = D1×0.20 + D2×0.20 + D3×0.25 + D4×0.25 + D5×0.10 | Valor actual: 53.56%",
        C["g4"])
    _col_header(ws, 4, [
        ("Componente", 38), ("Valor", 14), ("Peso", 10),
        ("Aporte", 12), ("Estado", 16), ("Fuente", 30)
    ], C["g4"])
    componentes = [
        ("D1 — Legalidad y Coherencia Normativa",  0.835, 0.20, "=B5*C5", "OK",           "Gold Master G3.1"),
        ("D2 — Fidelidad a la Planificación",       0.699, 0.20, "=B6*C6", "⚠ Alerta",    "Gold Master G3.2"),
        ("D3 — Ejecución Presupuestaria (Holding)", 0.146, 0.25, "=B7*C7", "🔴 Crítico",  "Gold Master G3.5"),
        ("D4 — Equidad Territorial (cantonal avg)", 0.669, 0.25, "=B8*C8", "⚠ Alerta",    "Gold Master G3.6"),
        ("D5 — Capacidad Institucional",            1.000, 0.10, "=B9*C9", "✅ OK",         "Gold Master G3.7"),
    ]
    for i, (comp, val, peso, formula, estado, fuente) in enumerate(componentes, start=5):
        _data_row(ws, i, [comp, val, f"{peso:.0%}", None, estado, fuente])
        ws.cell(i, 2).font = FONT_INPUT
        ws.cell(i, 4).value = formula
        ws.cell(i, 4).font  = FONT_FORMULA
    # Fila TOTAL
    r = 10
    _data_row(ws, r, ["ICPI GLOBAL", None, "100%", None, "🔴 Ruptura Sistémica", "D1-D5 ponderado"], bold=True)
    ws.cell(r, 2).value = "=SUM(D5:D9)"
    ws.cell(r, 2).font  = FONT_FORMULA
    ws.cell(r, 4).value = "=SUM(D5:D9)"
    ws.cell(r, 4).font  = FONT_FORMULA

    # Nota doctrinal
    r2 = 12
    ws.cell(r2, 1, "NOTA DOCTRINAL").font = FONT_BOLD
    ws.cell(r2+1, 1,
        "ICPI < 50% = Ruptura Sistémica · ICPI 50-64% = Gestión por Ocurrencia · "
        "ICPI 65-74% = Transición Crítica · ICPI ≥ 75% = Gestión por Mandato"
    ).font = Font(name="Arial", size=9, italic=True, color="666666")
    ws.cell(r2+1, 1).alignment = ALIGN_WRAP
    ws.merge_cells(f"A{r2+1}:F{r2+1}")
    ws.row_dimensions[r2+1].height = 30

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 30


def build_g4_5_rcm(ws) -> None:
    _group_header(ws, "G4.5 — RC_M | Tabla de Memoria Longitudinal",
        "RC-M: Período | ICPI (%) | D3 Ti (%) | SAT-IV | Riesgo Ponderado",
        C["g4"])
    _col_header(ws, 4, [
        ("Período", 14), ("Fecha Corte", 14), ("ICPI (%)", 12),
        ("D3 Ti (%)", 12), ("SAT-IV", 14), ("Riesgo", 14),
        ("Tendencia", 14), ("Snapshot ID", 36)
    ], C["g4"])
    # Datos reales del primer snapshot
    _data_row(ws, 5, [
        "Mayo 2026", "2026-05-18", 53.56, 14.58,
        "ACTIVA", "ALTO", "DETERIORO (base)", "snapshot_130801_20260518"
    ], bold=True)
    # Filas vacías para próximos períodos
    for r in range(6, 18):
        _data_row(ws, r, ["", "", None, None, "", "", "", ""])

    # Nota doctrinal RC-M
    ws.cell(20, 1, "REGLA SAT-III REINCIDENTE").font = FONT_BOLD
    ws.cell(21, 1,
        "Si D3 Ti < 60% por 3 períodos consecutivos → SAT-III REINCIDENTE (COPFP Art. 113). "
        "Si ICPI < 65% sostenido → riesgo pérdida elegibilidad fondos externos (CAF/BID/GIZ)."
    ).font = Font(name="Arial", size=9, italic=True, color="666666")
    ws.cell(21, 1).alignment = ALIGN_WRAP
    ws.merge_cells("A21:H21")
    ws.row_dimensions[21].height = 30


def build_g5_1_sat_catalogo(ws) -> None:
    _group_header(ws, "G5.1 — SAT_CATALOGO | Sistema de Alertas y Trazabilidad",
        "SAT-0 a SAT-VI: Triple ancla Legal + Operativa + Doctrinal", C["g5"])
    _col_header(ws, 4, [
        ("Nivel SAT", 12), ("Nombre", 30), ("Activador Técnico", 40),
        ("Base Legal", 36), ("Base Operativa", 28),
        ("Dimensión", 14), ("Umbral", 16)
    ], C["g5"])
    catalogo = [
        ("SAT-0", "Observación Inicial",   "Cualquier métrica en zona amarilla (primera vez)",
         "—",                            "Primer período con alerta",           "D1-D5",  "Primera vez"),
        ("SAT-I", "Alerta Operativa",     "Ti < 60% en 1 período",
         "COPFP Art. 113",              "Ti observado < 60%",                  "D3",     "< 60%"),
        ("SAT-II","Alerta Planificación", "D2_Score < 65% en 1 período",
         "COPFP Art. 41-44",            "Metas PDOT cumplimiento < 65%",       "D2",     "< 65%"),
        ("SAT-III","Reincidencia Crítica","Ti < 60% por 3 períodos consecutivos",
         "COPFP Art. 113 — persistencia","Ti < 60% × ≥ 3 períodos",            "D3",     "≥ 3 períodos"),
        ("SAT-IV", "Brecha Territorial",  "IRS > 70 o IET_parroquia < 50",
         "COOTAD Art. 249 CRE",         "IRS=79.7 — inversión concentrada",    "D4",     "IRS > 70"),
        ("SAT-V",  "Opacidad Directiva",  "Director no entrega informe firmado",
         "LOSEP Art. 76-82",            "Ti_V = 0 por ausencia de SHA-256",    "D5",     "Ti_V = 0"),
        ("SAT-VI", "Riesgo Sistémico",    "ICPI Global < 50%",
         "COPFP Art. 115-116",          "ICPI compuesto < 0.50",               "D1-D5",  "< 50%"),
    ]
    for i, row_data in enumerate(catalogo, start=5):
        _data_row(ws, i, list(row_data))
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 44
    ws.column_dimensions["D"].width = 38
    ws.column_dimensions["E"].width = 30
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 16


def build_g5_5_avep(ws) -> None:
    _group_header(ws, "G5.5 — AVEP_LENGUAJE | Sistema de Comunicación Política",
        "4 niveles: Gestión por Mandato · Transición Crítica · Gestión por Ocurrencia · Atención Alta",
        C["g5"])
    _col_header(ws, 4, [
        ("Nivel", 10), ("Nombre AVEP", 30), ("ICPI Rango", 16),
        ("SAT Activas", 18), ("Mensaje al Alcalde", 54), ("Acción", 44)
    ], C["g5"])
    avep = [
        ("🟢", "Gestión por Mandato",    "≥ 75%",   "Ninguna",     "La institución gobierna como se comprometió. Mantener ritmo.", "Monitoreo rutinario"),
        ("🟡", "Transición Crítica",     "60-74%",  "SAT-I, SAT-II","Señales de desaceleración. Decisiones correctivas esta semana.", "Reunión seguimiento con directores afectados"),
        ("🟠", "Gestión por Ocurrencia", "50-59%",  "SAT-III",     "El mandato está siendo sustituido por la ocurrencia. Reorientación urgente.", "Sesión extraordinaria · Revisión prioridades · SAT-III activado"),
        ("🔴", "Nivel de Atención Alta",  "< 50%",   "SAT-IV a VI", "Riesgo sistémico. Protocolo de emergencia institucional.", "Notificación Concejo · Revisión presupuestaria · Órganos control"),
    ]
    for i, row_data in enumerate(avep, start=5):
        _data_row(ws, i, list(row_data))
    ws.cell(10, 1, "PRINCIPIO FUNDACIONAL").font = FONT_BOLD
    ws.cell(11, 1,
        "AVEP no es calificación ni sanción. Es detección temprana. El lenguaje es "
        "100% preventivo y prospectivo. Origen doctrinal: IP Dylus Lab — "
        "'Gestión por Mandato vs. Gestión por Ocurrencia'."
    ).font = Font(name="Arial", size=9, italic=True, color="444444")
    ws.cell(11, 1).alignment = ALIGN_WRAP
    ws.merge_cells("A11:F11")
    ws.row_dimensions[11].height = 36


def build_g6_1_output_api(ws) -> None:
    """G6.1_OUTPUT_API — mismo contrato que H73_OUTPUT_API del v5.5.
    El conector gold_master.py lee esta hoja con data_only=True.
    Columna A = clave (str), Columna B = valor calculado.
    """
    _group_header(ws, "G6.1 — OUTPUT_API | Interfaz de Salida para Conector Python",
        "Contrato H73 preservado. Columna A = clave · Columna B = valor · Columna C = descripción",
        C["g6"])

    # Encabezado de columnas
    _col_header(ws, 4, [
        ("CLAVE", 34), ("VALOR", 18), ("DESCRIPCIÓN", 52),
        ("DIMENSIÓN TGI", 16), ("FUENTE HOJA", 18)
    ], C["g6"])

    # Todas las métricas — MISMO CONTRATO QUE H73_OUTPUT_API
    metricas = [
        # ── ICPI Global ──────────────────────────────────────────────────────
        ("ICPI_GLOBAL",           0.5356,  "ICPI ratio (0-1) — Índice Compuesto de Progreso Institucional", "D1-D5",    "G4.1"),
        ("ICPI_GLOBAL_PCT",       53.56,   "ICPI en porcentaje — 53.56%",                                   "D1-D5",    "G4.1"),
        ("ICPI_CLASIFICACION",    "Ruptura Sistémica", "Nivel AVEP: Gestión por Ocurrencia/Ruptura",         "D1-D5",    "G4.1"),
        ("ICPI_2023",             None,    "ICPI período 2023 — pendiente datos históricos",                "Histórico", "G4.5"),
        ("ICPI_2024",             None,    "ICPI período 2024 — pendiente datos históricos",                "Histórico", "G4.5"),
        ("ICPI_2025",             None,    "ICPI período 2025 — pendiente datos históricos",                "Histórico", "G4.5"),
        ("ICPI_ACUMULADO_Q1",     0.5356,  "ICPI acumulado Q1 2026",                                        "D1-D5",    "G4.1"),
        # ── Financiero ───────────────────────────────────────────────────────
        ("ISP_SALUD_PRESUP",      None,    "Inversión en Salud sobre Presupuesto total",                    "D3",       "G3.3"),
        ("ISP_META",              0.75,    "Meta ISP: 75% mínimo",                                          "D3",       "G1.1"),
        ("ISP_BRECHA_PP",         None,    "Brecha en pp: ISP_META - ISP_SALUD_PRESUP",                    "D3",       "G3.3"),
        ("PSG_EJECUCION",         0.5985,  "Ejecución presupuestaria GAD central (devengado/codificado)",   "D3",       "G3.3"),
        ("PSG_FIDELIDAD",         0.6993,  "D2_Score: fidelidad a planificación — % metas PDOT cumplidas", "D2",       "G3.2"),
        # ── Inversión ────────────────────────────────────────────────────────
        ("IFE_INVERSION_PUBLICA", None,    "Inversión pública efectiva — grupos 7+8 devengado",            "D3",       "G3.3"),
        ("IFE_META",              None,    "Meta de inversión pública para el período",                     "D3",       "G3.3"),
        # ── Contratación ─────────────────────────────────────────────────────
        ("PAC_PUBLICADO",         True,    "PAC publicado en SERCOP (True/False)",                          "D3",       "G2.1"),
        ("PROCESOS_ADJUDICADOS",  None,    "Número de procesos adjudicados SERCOP",                         "D3",       "G3.3"),
        ("PROCESOS_CANCELADOS",   None,    "Número de procesos cancelados SERCOP",                          "D3",       "G3.3"),
        ("PROCESOS_CANCELADOS_PCT",None,   "% procesos cancelados / total",                                 "D3",       "G3.3"),
        # ── CPCCS / Accountability ───────────────────────────────────────────
        ("RDC_SCORE",             None,    "Score Rendición de Cuentas CPCCS (0-100)",                      "D1",       "G3.1"),
        ("RDC_CLASIFICACION",     None,    "Clasificación CPCCS RdC",                                       "D1",       "G3.1"),
        # ── SAT ──────────────────────────────────────────────────────────────
        ("SAT_RIESGO_TOTAL",      0.72,    "Score de riesgo SAT ponderado (0-1)",                           "D1-D5",    "G5.2"),
        ("SAT_CLASIF_RIESGO",     "ALTO",  "Clasificación riesgo: BAJO/MEDIO/ALTO/CRÍTICO",                "D1-D5",    "G5.2"),
        ("SAT_ACTIVAS_COUNT",     3,       "Número de alertas SAT activas en el período",                   "D1-D5",    "G5.2"),
        # ── TGI 5D ───────────────────────────────────────────────────────────
        ("TGI_SCORE",             68.82,   "TGI Score cantonal ponderado 5D",                               "D1-D5",    "G4.2"),
        ("TGI_D1",                83.50,   "D1 Legalidad y Coherencia Normativa (%)",                       "D1",       "G3.1"),
        ("TGI_D2",                69.93,   "D2_Score Fidelidad a la Planificación (%)",                     "D2",       "G3.2"),
        ("TGI_D3",                14.58,   "D3 Ti Holding Municipal — ejecución (%) — ratio estricto",      "D3",       "G3.5"),
        ("TGI_D4",                66.85,   "D4 Equidad Territorial — TGI rural promedio (%)",               "D4",       "G3.6"),
        ("TGI_D5",               100.00,   "D5 Capacidad Institucional (%)",                                "D5",       "G3.7"),
        # ── IED y MMP (extras v6.0) ─────────────────────────────────────────
        ("IED_GLOBAL",            31.14,   "Índice de Eficiencia Directiva global (%)",                     "D5",       "G4.4"),
        ("IED_DIRECCIONES_N",     11,      "Número de direcciones municipales activas",                     "D5",       "G4.4"),
        ("MMP_AVANCE_PCT",        None,    "Avance MMP: % metas con progreso documentado",                  "D2",       "G4.3"),
        ("IRS_GLOBAL",            79.7,    "Índice de Regresividad Social cantonal",                        "D4",       "G3.6"),
        ("BRECHA_RURAL_USD",      1791935, "Brecha de subinversión rural acumulada (USD)",                  "D4",       "G3.6"),
    ]

    for i, (clave, valor, desc, dim, fuente) in enumerate(metricas, start=5):
        bg = C["fila_par"] if i % 2 == 0 else C["fila_impar"]
        # Clave
        c_k = ws.cell(i, 1, clave)
        c_k.font = FONT_BOLD
        c_k.fill = _fill(bg)
        c_k.border = BORDER_THIN
        # Valor (input — azul)
        c_v = ws.cell(i, 2, valor)
        c_v.font = FONT_INPUT
        c_v.fill = _fill(bg)
        c_v.border = BORDER_THIN
        c_v.alignment = ALIGN_RIGHT
        # Descripción, dimensión, fuente
        for col_idx, val in enumerate([desc, dim, fuente], start=3):
            c = ws.cell(i, col_idx, val)
            c.font = FONT_BASE
            c.fill = _fill(bg)
            c.border = BORDER_THIN

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 54
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 16

    # Nota importante
    last = 5 + len(metricas)
    ws.cell(last + 2, 1,
        "CONTRATO API: El conector Python gold_master.py lee esta hoja con data_only=True. "
        "Columna A = clave, Columna B = valor calculado. NO modificar claves sin actualizar el conector."
    ).font = Font(name="Arial", size=9, italic=True, color="444444")
    ws.cell(last + 2, 1).alignment = ALIGN_WRAP
    ws.merge_cells(f"A{last+2}:E{last+2}")
    ws.row_dimensions[last+2].height = 36


def build_g6_3_reporte_ejecutivo(ws) -> None:
    _group_header(ws, "G6.3 — REPORTE_EJECUTIVO | Las 5 Preguntas del Alcalde",
        "Respuestas en lenguaje AVEP a las 5 preguntas estratégicas (H29 canónicas)", C["g6"])
    _col_header(ws, 4, [
        ("#", 6), ("Pregunta", 38), ("Respuesta", 40),
        ("Nivel AVEP", 20), ("Acción Recomendada", 44)
    ], C["g6"])
    preguntas = [
        (1, "¿Cómo vamos?",
         f"ICPI Global: 53.56% — Ruptura Sistémica",
         "🔴 Atención Alta",
         "Protocolo de emergencia. Revisión urgente de prioridades."),
        (2, "¿Dónde enfocar la atención?",
         "D3 Ti Holding: 14.58% — ejecución presupuestaria crítica en entidades adscritas",
         "🔴 Atención Alta",
         "Reunión inmediata con EMAI, Bomberos, Patronato."),
        (3, "¿Qué alertas SAT están activas?",
         "SAT-I (Ti < 60%), SAT-IV (IRS 79.7 — brecha rural $1.79M), SAT-VI (ICPI < 50%)",
         "🔴 Múltiples",
         "Activar protocolo SAT por nivel. Ver G5.2_SAT_ACTIVAS."),
        (4, "¿Cuál es la tendencia?",
         "Período base (1 snapshot). RC-M activo desde Sprint 3. Siguiente período: Junio 2026.",
         "— Base",
         "Ejecutar pipeline en Junio para ver trayectoria."),
        (5, "¿Están en riesgo los fondos externos?",
         "IRS 79.7 > umbral 70. ICPI < 65% implica riesgo elegibilidad CAF/BID/GIZ.",
         "🟠 Riesgo",
         "Revisar compromisos con cooperantes. Plan reequilibrio territorial."),
    ]
    for i, row_data in enumerate(preguntas, start=5):
        _data_row(ws, i, list(row_data))
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 42
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 46


def build_g7_sod(ws) -> None:
    _group_header(ws, "G7.1 — GOBERNANZA_SOD | Segregación de Funciones",
        "Marco de segregación: quién puede leer / editar / validar cada hoja", C["g7"])
    _col_header(ws, 4, [
        ("Grupo", 12), ("Función", 36), ("Puede leer", 20),
        ("Puede editar", 20), ("Puede validar", 20)
    ], C["g7"])
    sod = [
        ("G1", "Configuración y parámetros",        "Todos",         "Director QUIRA",   "Dylus Lab"),
        ("G2", "Fuentes y trazabilidad",             "Todos",         "Técnico QUIRA",    "Director QUIRA"),
        ("G3", "Dimensiones TGI",                    "Todos",         "Motor (automático)","Director QUIRA"),
        ("G4", "Índices compuestos",                 "Todos",         "Motor (automático)","Director QUIRA"),
        ("G5", "SAT y alertas",                      "Todos",         "Técnico QUIRA",    "Director QUIRA"),
        ("G6", "Outputs y reportes",                 "Todos",         "Motor (automático)","Dylus Lab"),
        ("G7", "Gobernanza",                         "Director QUIRA","Dylus Lab",         "Dylus Lab"),
    ]
    for i, row_data in enumerate(sod, start=5):
        _data_row(ws, i, list(row_data))


def build_placeholder(ws, nombre: str, desc: str, color: str) -> None:
    """Hoja con estructura básica para completar."""
    _group_header(ws, nombre, desc, color)
    ws.cell(5, 1, "🔨 En construcción — datos se incorporan en fase de implementación").font = Font(
        name="Arial", size=10, italic=True, color="888888"
    )


# ══════════════════════════════════════════════════════════════════════════════
# ORQUESTADOR PRINCIPAL
# ══════════════════════════════════════════════════════════════════════════════

def build_workbook() -> Workbook:
    wb = Workbook()

    # Eliminar hoja default
    default = wb.active
    wb.remove(default)

    sheets = [
        # G1 — CONFIGURACIÓN
        ("G1.1_CONFIG",              C["g1"], build_g1_1_config),
        ("G1.2_MUNICIPIO_PROFILE",   C["g1"], build_g1_2_municipio),
        ("G1.3_HOLDING_ENTIDADES",   C["g1"], build_g1_3_holding),
        ("G1.4_PESOS_PONDERACIONES", C["g1"], None),  # placeholder
        ("G1.5_CHANGELOG",           C["g1"], build_g1_5_changelog),
        # G2 — FUENTES
        ("G2.1_FUENTES_INSTITUCIONALES", C["g2"], build_g2_1_fuentes),
        ("G2.2_RELIABILITY_SCORE",       C["g2"], build_g2_2_reliability),
        ("G2.3_INGESTA_CALENDARIO",      C["g2"], None),
        ("G2.4_TRAZABILIDAD",            C["g2"], None),
        # G3 — DIMENSIONES
        ("G3.1_D1_LEGALIDAD",       C["g3"], None),
        ("G3.2_D2_PLANIFICACION",   C["g3"], None),
        ("G3.3_D3_EJECUCION_GAD",   C["g3"], None),
        ("G3.4_D3_HOLDING",         C["g3"], None),
        ("G3.5_D3_CONSOLIDADO",     C["g3"], None),
        ("G3.6_D4_EQUIDAD",         C["g3"], None),
        ("G3.7_D5_CAPACIDAD",       C["g3"], None),
        ("G3.8_D3_HISTORICO",       C["g3"], None),
        # G4 — ÍNDICES
        ("G4.1_ICPI_GLOBAL",        C["g4"], build_g4_1_icpi),
        ("G4.2_TGI_SCORE",          C["g4"], None),
        ("G4.3_MMP",                C["g4"], None),
        ("G4.4_IED",                C["g4"], None),
        ("G4.5_RC_M",               C["g4"], build_g4_5_rcm),
        # G5 — SAT
        ("G5.1_SAT_CATALOGO",       C["g5"], build_g5_1_sat_catalogo),
        ("G5.2_SAT_ACTIVAS",        C["g5"], None),
        ("G5.3_SAT_LONGITUDINAL",   C["g5"], None),
        ("G5.4_PROTOCOLO_ACCION",   C["g5"], None),
        ("G5.5_AVEP_LENGUAJE",      C["g5"], build_g5_5_avep),
        # G6 — OUTPUTS
        ("G6.1_OUTPUT_API",         C["g6"], build_g6_1_output_api),
        ("G6.2_SNAPSHOT_SCHEMA",    C["g6"], None),
        ("G6.3_REPORTE_EJECUTIVO",  C["g6"], build_g6_3_reporte_ejecutivo),
        ("G6.4_REPORTE_DIRECTIVO",  C["g6"], None),
        ("G6.5_DASHBOARD_DATA",     C["g6"], None),
        # G7 — GOBERNANZA
        ("G7.1_GOBERNANZA_SOD",     C["g7"], build_g7_sod),
        ("G7.2_POLITICA_INTEGRIDAD",C["g7"], None),
    ]

    desc_map = {
        "G1.4_PESOS_PONDERACIONES": "Tabla de pesos TGI y variantes históricas por sub-indicador",
        "G2.3_INGESTA_CALENDARIO":  "Calendario: fuente, frecuencia, responsable, estado mensual",
        "G2.4_TRAZABILIDAD":        "Cadena: dato bruto → campo → fórmula → score → output API",
        "G3.1_D1_LEGALIDAD":        "D1 Legalidad: Trust Score, PDOT-POA-PAC, LOTAIP · Peso 20%",
        "G3.2_D2_PLANIFICACION":    "D2 Planificación: D2_Score (ex-ICPI), 25 metas PDOT · Peso 20%",
        "G3.3_D3_EJECUCION_GAD":    "D3 Ejecución GAD central: devengado/codificado · Peso 25%",
        "G3.4_D3_HOLDING":          "D3 Ti por entidad: EMAI / Bomberos / Patronato · Peso 25%",
        "G3.5_D3_CONSOLIDADO":      "D3 Consolidado Holding: ponderado por presupuesto relativo",
        "G3.6_D4_EQUIDAD":          "D4 Equidad: IET por parroquia, IRS 79.7, Brecha $1.79M · Peso 25%",
        "G3.7_D5_CAPACIDAD":        "D5 Capacidad: IED 31.14%, ICM-SNP, Orgánico 040 · Peso 10%",
        "G3.8_D3_HISTORICO":        "D3 histórico por período: base para RC-M y Longitudinal Engine",
        "G4.2_TGI_SCORE":           "TGI Score cantonal 5D ponderado: D1×0.20+D2×0.20+D3×0.25+D4×0.25+D5×0.10",
        "G4.3_MMP":                 "Monitor Progreso Mensual: 25 metas PDOT × 12 meses",
        "G4.4_IED":                 "IED por dirección: 11 direcciones municipales · LOSEP Art.76-82",
        "G5.2_SAT_ACTIVAS":         "Alertas SAT activas período actual con estado y prioridad",
        "G5.3_SAT_LONGITUDINAL":    "Evolución temporal de alertas: activación, mitigación, reincidencia",
        "G5.4_PROTOCOLO_ACCION":    "Protocolo preventivo por nivel SAT: qué hacer antes de que escale",
        "G6.2_SNAPSHOT_SCHEMA":     "Esquema canónico del snapshot JSON Supabase (campos, tipos, defaults)",
        "G6.4_REPORTE_DIRECTIVO":   "Reporte técnico para directores: IED por dirección + SAT activas",
        "G6.5_DASHBOARD_DATA":      "Datos procesados listos para Streamlit (evita re-cálculo en Python)",
        "G7.2_POLITICA_INTEGRIDAD": "Política de integridad, backup SHA-256, recuperación de desastres",
    }

    for sheet_name, color, builder in sheets:
        ws = wb.create_sheet(title=sheet_name)
        ws.sheet_view.showGridLines = False
        ws.row_dimensions[1].height = 24
        ws.row_dimensions[2].height = 18

        if builder:
            builder(ws)
        else:
            desc = desc_map.get(sheet_name, "En construcción")
            build_placeholder(ws, sheet_name, desc, color)

    return wb


def main() -> None:
    print("-" * 60)
    print("  QUIRA OS - Gold Master TGI v6.0 Builder")
    print("-" * 60)
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    print("[>>] Construyendo workbook...")
    wb = build_workbook()
    outpath = str(OUTPUT_PATH)
    print("[>>] Guardando en: " + outpath)
    wb.save(outpath)
    n = len(wb.sheetnames)
    print("[OK] Gold Master TGI v6.0 creado. Hojas: " + str(n))
    for name in wb.sheetnames:
        print("   - " + name)
    print("-" * 60)

if __name__ == "__main__":
    main()
