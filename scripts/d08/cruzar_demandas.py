# -*- coding: utf-8 -*-
"""
scripts/d08/cruzar_demandas.py — Fase 2 · Trazabilidad Biográfica de la demanda
═══════════════════════════════════════════════════════════════════════════════
authority:
  parent: MARCO-TEORICO-001
  constitution_articles: [1, 2, 3, 9]
  type: TECNICA

Implementa el POSTULADO I (Trazabilidad Biográfica del Dato) sobre las demandas
ciudadanas: reconstruye la historia de vida de cada demanda a través de la cadena

    demanda ciudadana → POA → PAC → presupuesto → ejecución

★ HORIZONTE DE VERDAD (marco_teorico · concepto 6) — TODO resultado declara su
  estado epistémico. Ésta es la frontera obligatoria de la Fase 2:

  · `confirmado`  → HECHO OBSERVABLE: existe la demanda · existe el proyecto ·
                    hay coincidencia semántica medible. QUIRA lo certifica.
  · `hipotesis`   → INFERENCIA ANALÍTICA: "esta demanda fue atendida". QUIRA la
                    PROPONE; el humano la valida (Constitución Art. 3).
  · `sin_correlato` → no se halló proyecto asociable en la evidencia disponible.
                    NO significa "no se atendió": significa que el expediente no
                    lo acredita (Principio de No-Inferencia).

NUNCA se afirma "la demanda fue satisfecha" como hecho. La prohibición de
alucinación es arquitectónica, no operacional.

MOTOR: embeddings locales (mismo modelo del corpus), SIN API y sin costo. Se reusa
la TÉCNICA de d09, no sus datos — la frontera d08/d09 es sobre el ORIGEN de los
aportes (DEC-0004), no sobre el método de cruce.

Uso:  python scripts/d08/cruzar_demandas.py [--anio 2025]
Salida: data/d08/trazabilidad_demandas.json
Dylus Lab © 2026
"""
from __future__ import annotations

import json
import sys
import tomllib
from collections import Counter
from datetime import date
from pathlib import Path

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

REPO = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(REPO / "scripts" / "normativa"))
sys.path.insert(0, str(REPO / "scripts"))

DEMANDAS = REPO / "data" / "d08" / "demandas_ciudadanas.json"
SALIDA = REPO / "data" / "d08" / "trazabilidad_demandas.json"

# Umbrales calibrados y ya validados por Javo en el cruce de d09 (METODOLOGIA_TRAZABILIDAD_APORTES)
TH_FUERTE = 0.62      # candidato fuerte → hipótesis de correspondencia
TH_REVISAR = 0.52     # banda de validación experta obligatoria
# Piso de evaluación: por debajo es ruido puro y no vale ni pasar por el filtro.
# Más bajo que TH_REVISAR a propósito: el filtro ontológico puede rescatar una
# relación legítima que el embedding rankeó bajo por ruido OCR en la demanda.
PISO_EVALUACION = 0.42


POA_XLSX = Path(r"C:\Users\DELL\Desktop\Javo\Dylus Lab\ProyecT\Holding_Municipal_Montecristi"
                r"\POA 2023-2026\GAD Montecristi")

# Membretes, títulos y filas de encabezado — NO son proyectos. Los tokens provienen del
# extractor ya existente (scripts/extract_poa_pdf.py::_HDR_TOKENS): la solución estaba en
# el canon. Sin este filtro, el cruce empareja demandas contra el membrete institucional.
_HDR = ("ACTIVIDAD", "ACTVIDAD", "DESCRIPCIÓN", "DESCRIPCION", "PROYECTO", "PARTIDA",
        "MONTO", "NO. DE", "RESPONSABLE", "META", "GOBIERNO AUTONOMO", "GOBIERNO AUTÓNOMO",
        "MISIÓN INSTITUCIONAL", "VISIÓN INSTITUCIONAL", "PLAN OPERATIVO ANUAL",
        "ALINEACIÓN", "OBJETIVO DE DESARROLLO SOS", "SEGUIMIENTO EJECUCIÓN",
        "PROGRAMACIÓN DE LA META", "TIPO DE FINANCIAMIENTO", "UNIDAD ADMINISTRATIVA")


def es_encabezado(texto: str) -> bool:
    """True si la fila es membrete/título/cabecera de tabla, no un proyecto."""
    up = texto.upper()
    hits = sum(1 for t in _HDR if t in up)
    return hits >= 2 or up.startswith(("GOBIERNO AUTONOMO", "GOBIERNO AUTÓNOMO"))


def cargar_poa(anios: tuple[str, ...]) -> list[dict]:
    """Lee los proyectos POA de los XLSX oficiales del GAD. HECHO OBSERVABLE.

    ★ NO se usa el corpus vectorizado (POA-GAD-*): el canon lo PROHÍBE expresamente
    (`docs/architecture/METODOLOGIA_TRAZABILIDAD_APORTES.md` §3): "la vectorización de
    esos PDFs quedó CORRUPTA (OCR fallido — chunks de caracteres sueltos)". Un primer
    intento ignoró esa advertencia y produjo correspondencias contra texto ilegible
    ("bsta idl s iae d r li eao s INDICADOR OPERATIVO...") — resultado descartado (OBS-018).

    El XLSX es la fuente estructurada y limpia: sin OCR, sin corrupción.
    """
    import openpyxl
    out = []
    for anio in anios:
        f = POA_XLSX / f"GAD Montecristi POA {anio}.xlsx"
        if not f.exists():
            f = POA_XLSX / f"GAD Monteristi POA {anio}.xlsx"     # typo en el archivo oficial
        if not f.exists():
            print(f"  [skip] POA {anio}: no existe XLSX")
            continue
        wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        n = 0
        for row in ws.iter_rows(values_only=True):
            celdas = [" ".join(str(c).split()) for c in row if c not in (None, "")]
            # se conserva el texto sustantivo de la fila (proyecto · actividad · meta)
            texto = " · ".join(c for c in celdas if len(c) > 12)[:400]
            if len(texto) < 40 or es_encabezado(texto):
                continue
            out.append({"fuente": f"POA {anio} (XLSX oficial)", "sha": "", "texto": texto})
            n += 1
        wb.close()
        print(f"  POA {anio}: {n} filas con contenido sustantivo (XLSX limpio)")
    return out


def main() -> int:
    if not DEMANDAS.exists():
        print("ERROR: falta data/d08/demandas_ciudadanas.json — ejecutar extraer_demandas.py")
        return 1
    d = json.loads(DEMANDAS.read_text(encoding="utf-8"))
    demandas = d["demandas"]
    print(f"=== FASE 2 · Trazabilidad Biográfica de {len(demandas)} demandas ===\n")

    poa = cargar_poa(("2023", "2024", "2025", "2026"))
    print(f"  TOTAL POA (fuente XLSX oficial, sin OCR): {len(poa)} registros\n")
    if not poa:
        print("  ERROR: sin POA — no se puede cruzar")
        return 1

    import numpy as np
    import ingest as ing
    model = ing._get_model()
    print("  modelo de embeddings local cargado (sin API)\n")

    def emb(txts):
        v = model.encode(txts, convert_to_numpy=True, show_progress_bar=False, batch_size=64)
        return v / (np.linalg.norm(v, axis=1, keepdims=True) + 1e-9)

    D = emb([x["demanda"][:300] for x in demandas])
    P = emb([x["texto"] for x in poa])

    import filtro_ontologico as fo
    print("  MRSPP v3 cargado — el embedding PROPONE, el modelo de satisfacción DECIDE\n")

    resultados = []
    descartes = Counter()
    for i, dem in enumerate(demandas):
        sims = D[i] @ P.T
        # ★ EL EMBEDDING YA NO DECIDE: propone candidatos; el filtro elige.
        # TOP-30 (no 10): ahora que el filtro garantiza precisión, se prioriza el
        # RECALL. Con top-10 se perdían correspondencias legítimas que el embedding
        # rankeaba bajo por ruido OCR ("REALIZAR AREAS VERDES-EL SECAL" quedaba sin
        # correlato aunque el POA sí tenía proyectos de áreas verdes).
        top = sims.argsort()[-30:][::-1]

        # ★ DIVISIÓN DE RESPONSABILIDADES (corrección 2026-07-29):
        #   el FILTRO ONTOLÓGICO decide SI existe relación y de qué TIPO;
        #   el SCORE decide CUÁNTA CONFIANZA merece esa relación.
        # Antes el piso de score cortaba ANTES de evaluar el filtro, y demandas
        # legítimas ("REALIZAR AREAS VERDES-EL SECAL") quedaban sin correlato pese a
        # existir proyectos de áreas verdes en el POA. El umbral gobernaba el recall
        # cuando debía gobernarlo el conocimiento institucional.
        j, score, motivo, tipo = None, 0.0, "sin_candidato_que_pase_el_filtro", "nula"
        for k in top:
            cand_score = float(sims[k])
            if cand_score < PISO_EVALUACION:
                break                                   # ruido puro, ni evaluar
            tipo_rel, razon = fo.evaluar_relacion(dem["demanda"], poa[k]["texto"])
            if tipo_rel != "nula":
                j, score, motivo, tipo = int(k), cand_score, f"{tipo_rel}: {razon}", tipo_rel
                break
            descartes[razon.split(":")[0]] += 1

        # ★ ESTADO EPISTÉMICO = tipo de relación × confianza del score
        fuerte = tipo in ("directa", "funcional")
        if j is not None and fuerte and score >= TH_FUERTE:
            estado, naturaleza = "hipotesis", f"INFERENCIA ANALÍTICA — relación {tipo} con alta similitud; requiere validación experta"
        elif j is not None and fuerte:
            estado, naturaleza = "pendiente_validacion", f"relación {tipo} verificada ontológicamente, similitud media — el analista confirma"
        elif j is not None:
            # instrumental/complementaria son débiles por naturaleza: nunca ascienden solas
            estado, naturaleza = "pendiente_validacion", f"relación {tipo} — débil por naturaleza, requiere validación experta"
        else:
            j = int(sims.argmax())                       # se conserva el más próximo solo como referencia
            score = float(sims[j])
            estado = "sin_correlato"
            naturaleza = ("SIN CORRELATO PRESUPUESTARIO VERIFICABLE — ningún candidato superó el "
                          "filtro ontológico. NO significa que no se atendió: significa que el "
                          "expediente no acredita correspondencia (alimenta la brecha de atención)")

        resultados.append({
            # HECHOS OBSERVABLES (confirmados)
            "demanda": dem["demanda"][:220],
            "mecanismo": dem["mecanismo"],
            "naturaleza_juridica": dem["naturaleza_juridica"],
            "anio_demanda": dem["anio"],
            "fuente_demanda": dem["fuente"],
            "similitud": round(score, 3),
            # ★ TEXTO ÍNTEGRO, no recortado: se guarda EXACTAMENTE lo que el filtro
            #   juzgó. Con el recorte a 200 chars el token que disparaba el veredicto
            #   quedaba fuera del expediente y la correspondencia era inauditable —
            #   así se ocultó durante una corrida el falso positivo de REGLA 0.
            "proyecto_poa_mas_proximo": poa[j]["texto"] if estado != "sin_correlato" else "",
            "fuente_poa": poa[j]["fuente"] if estado != "sin_correlato" else "",
            "sha_poa": poa[j]["sha"] if estado != "sin_correlato" else "",
            "filtro_ontologico": motivo,
            # INFERENCIA (declarada como tal)
            "estado_epistemico": estado,
            "naturaleza_del_juicio": naturaleza,
        })

    est = Counter(r["estado_epistemico"] for r in resultados)
    vinc = [r for r in resultados if r["naturaleza_juridica"] == "vinculante"]
    est_vinc = Counter(r["estado_epistemico"] for r in vinc)

    salida = {
        "_fuente": "GENERADO por scripts/d08/cruzar_demandas.py — Fase 2",
        "_postulado": "I · Trazabilidad Biográfica del Dato (marco_teorico/MARCO_TEORICO_QUIRA.md)",
        "_horizonte_de_verdad": {
            "confirmado": "hechos observables: la demanda existe, el proyecto existe, hay similitud medible",
            "hipotesis": "INFERENCIA: correspondencia propuesta por la máquina — requiere validación experta",
            "pendiente_validacion": "banda de revisión obligatoria (0.52-0.62)",
            "sin_correlato": "el expediente no acredita correspondencia — NO afirma que no se atendió",
        },
        "_advertencia": "QUIRA NUNCA afirma 'la demanda fue satisfecha' como hecho. Propone; el humano valida (Constitución Art. 3).",
        "generado": date.today().isoformat(),
        "umbrales": {"fuerte": TH_FUERTE, "revisar": TH_REVISAR},
        "total_demandas": len(resultados),
        "registros_poa_contrastados": len(poa),
        "por_estado_epistemico": dict(est),
        "vinculantes_por_estado": dict(est_vinc),
        "trazabilidad": resultados,
    }
    SALIDA.write_text(json.dumps(salida, ensure_ascii=False, indent=2), encoding="utf-8")

    print(f"  DESCARTES DEL FILTRO ONTOLÓGICO: {sum(descartes.values())} candidatos rechazados")
    for k, v in descartes.most_common():
        print(f"     {k:32} {v:>5}")
    print()
    print(f"  {'ESTADO EPISTÉMICO':<24} {'TODAS':>7}   {'VINCULANTES (COOTAD 238)':>26}")
    for e in ("hipotesis", "pendiente_validacion", "sin_correlato"):
        print(f"  {e:<24} {est.get(e,0):>7}   {est_vinc.get(e,0):>26}")
    print(f"\nOK — {SALIDA.relative_to(REPO)}")
    print("\n  RECORDATORIO: 'hipotesis' NO es 'atendida'. Requiere validación experta")
    print("  antes de cualquier afirmación pública (Horizonte de Verdad).")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
