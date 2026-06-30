# -*- coding: utf-8 -*-
"""
scripts/dev/gm_diff2.py — DIFF quirúrgico A vs B (data_only · solo lectura · $0)
Verifica que una ingesta tocó SOLO lo que debía. Imprime celdas distintas por hoja,
muestra, y el centinela ICPI (H12!B33) en ambos. NO modifica nada.
Uso: python scripts/dev/gm_diff2.py "<A work>" "<B baseline>"
Dylus Lab © 2026
"""
import sys
from pathlib import Path

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

# Celdas/fórmulas que cambian solas por TODAY() — ruido esperado, no señal
RUIDO_HOJAS = ()  # (se filtra por contenido tipo fecha abajo si hiciera falta)


def _cellmap(ws):
    out = {}
    for row in ws.iter_rows():
        for cell in row:
            v = cell.value
            if v is not None and str(v).strip() != "":
                out[cell.coordinate] = v
    return out


def _differ(a, b):
    if isinstance(a, (int, float)) and isinstance(b, (int, float)):
        return abs(float(a) - float(b)) > 1e-6
    return str(a) != str(b)


def _icpi(wb):
    for name in wb.sheetnames:
        if name.startswith("H12"):
            return wb[name]["B33"].value
    return "<sin H12>"


def main():
    import openpyxl
    A, B = Path(sys.argv[1]), Path(sys.argv[2])
    wa = openpyxl.load_workbook(A, read_only=True, data_only=True)
    wb = openpyxl.load_workbook(B, read_only=True, data_only=True)
    sa, sb = set(wa.sheetnames), set(wb.sheetnames)

    print(f"A={A.name}  B={B.name}")
    print(f"ICPI(H12!B33)  A={_icpi(wa)!r}  B={_icpi(wb)!r}")
    print(f"hojas solo-A={sorted(sa-sb)}  solo-B={sorted(sb-sa)}")
    print("-" * 64)

    total = 0
    for name in wa.sheetnames:
        if name not in sb:
            continue
        ma, mb = _cellmap(wa[name]), _cellmap(wb[name])
        keys = set(ma) | set(mb)
        diffs = []
        for k in keys:
            x, y = ma.get(k), mb.get(k)
            if x is None or y is None or _differ(x, y):
                diffs.append((k, x, y))
        if diffs:
            total += len(diffs)
            sample = "; ".join(
                f"{k}:{str(x)[:14]}->{str(y)[:14]}"
                for k, x, y in sorted(diffs)[:8]
            )
            print(f"[{name}] {len(diffs)} celdas | {sample[:150]}")
    wa.close(); wb.close()
    print("-" * 64)
    print(f"TOTAL celdas distintas: {total}")


if __name__ == "__main__":
    main()
