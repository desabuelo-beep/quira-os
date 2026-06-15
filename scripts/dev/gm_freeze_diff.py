"""
scripts/dev/gm_freeze_diff.py — DIFF Canónico-vivo vs FREEZE (determinista · solo lectura · $0)

Compara valores cacheados (data_only=True) del Gold Master VIVO contra la copia FREEZE.
Reporta: hojas solo-en-uno, y por hoja común el #celdas con valor distinto + muestra.
NO modifica nada. El árbitro es la celda.

Uso: python scripts/dev/gm_freeze_diff.py
Dylus Lab © 2026
"""
from __future__ import annotations

import sys
from pathlib import Path

REPO = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(REPO))
try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

PROYECT = Path(r"C:\Users\DELL\Desktop\Javo\Dylus Lab\ProyecT")
LIVE = PROYECT / "SIAP-ICPI_GOLD_MASTER_v5.5_TGI.xlsx"
FREEZE = PROYECT / "SIAP-ICPI_GOLD_MASTER_v5.5_FREEZE_20260526.xlsx"
SAMPLE = 10  # muestra de celdas distintas por hoja


def _cellmap(ws) -> dict[str, object]:
    out: dict[str, object] = {}
    for row in ws.iter_rows():
        for cell in row:
            v = cell.value
            if v is not None and str(v).strip() != "":
                out[cell.coordinate] = v
    return out


def _differ(a, b) -> bool:
    if isinstance(a, (int, float)) and isinstance(b, (int, float)):
        return abs(float(a) - float(b)) > 1e-6
    return str(a) != str(b)


def main() -> int:
    import openpyxl
    if not LIVE.exists() or not FREEZE.exists():
        print(f"Falta archivo. live={LIVE.exists()} freeze={FREEZE.exists()}"); return 1

    wl = openpyxl.load_workbook(LIVE, read_only=True, data_only=True)
    wf = openpyxl.load_workbook(FREEZE, read_only=True, data_only=True)
    sl, sf = set(wl.sheetnames), set(wf.sheetnames)

    lines = [
        "# GM _FREEZE_DIFF — vivo vs FREEZE (determinista)",
        f"**vivo:** `{LIVE.name}` ({len(sl)} hojas) · **freeze:** `{FREEZE.name}` ({len(sf)} hojas)",
        "",
        f"## Hojas solo en VIVO: {sorted(sl - sf) or '—'}",
        f"## Hojas solo en FREEZE: {sorted(sf - sl) or '—'}",
        "",
        "## Hojas con valores distintos",
        "| Hoja | celdas dif | solo-vivo | solo-freeze | muestra |",
        "|---|---|---|---|---|",
    ]
    total_diff_sheets = 0
    detail = []
    for name in wl.sheetnames:
        if name not in sf:
            continue
        ml = _cellmap(wl[name])
        mf = _cellmap(wf[name])
        keys = set(ml) | set(mf)
        diffs = []
        only_l = only_f = 0
        for k in keys:
            a, b = ml.get(k), mf.get(k)
            if a is None:
                only_f += 1; diffs.append((k, None, b))
            elif b is None:
                only_l += 1; diffs.append((k, a, None))
            elif _differ(a, b):
                diffs.append((k, a, b))
        changed = [d for d in diffs if d[1] is not None and d[2] is not None]
        if diffs:
            total_diff_sheets += 1
            sample = "; ".join(
                f"{k}: {str(a)[:18]}→{str(b)[:18]}" for k, a, b in
                sorted(changed, key=lambda x: x[0])[:SAMPLE]
            ) or "(solo altas/bajas de celdas)"
            lines.append(f"| {name} | {len(changed)} | {only_l} | {only_f} | {sample[:160]} |")
            # detalle ampliado para hojas motor
            if name.startswith(("H12", "H07", "H73", "H90", "H98", "H99", "H15", "H39", "H85", "H97")):
                detail.append(f"\n### {name} — {len(changed)} celdas cambiadas")
                for k, a, b in sorted(changed, key=lambda x: x[0])[:30]:
                    detail.append(f"- {k}: `{str(a)[:40]}` → `{str(b)[:40]}`")
    wl.close(); wf.close()

    lines += ["", f"**hojas con diferencias:** {total_diff_sheets}", ""]
    if detail:
        lines += ["## Detalle hojas-motor"] + detail

    dest = REPO / "docs" / "architecture" / "gm_dumps" / "_FREEZE_DIFF.md"
    dest.write_text("\n".join(lines), encoding="utf-8")
    print(f"vivo {len(sl)} · freeze {len(sf)} · solo-vivo={sorted(sl-sf)} · solo-freeze={sorted(sf-sl)}")
    print(f"hojas con diferencias: {total_diff_sheets} -> {dest.relative_to(REPO)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
