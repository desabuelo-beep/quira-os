"""
scripts/dev/gm_full_audit.py — AUDITOR INTEGRAL del Gold Master (determinista · solo lectura · $0)

Extiende gm_sheet_dump.py a las 123 hojas. NO modifica el Excel · NO recalcula el motor ·
NO usa LLM. Para CADA hoja:
  - vuelca TODAS las fórmulas (data_only=False) + celdas etiqueta a docs/architecture/gm_dumps/<hoja>.md
  - parsea referencias cruzadas `Hoja!Celda` → grafo de dependencias (inputs/outputs)
  - cuenta llenado (data_only=True), detecta errores (#REF!/#DIV/0!/…) y marcadores (PENDIENTE/MISSING/…)
Emite:
  - docs/architecture/gm_dumps/<hoja>.md   (123 volcados durables)
  - docs/architecture/gm_dumps/_INDEX.md   (índice: hoja → llenado · #fórmulas · #in · #out · marcadores)
  - docs/architecture/gm_dumps/_ANALYSIS.md (análisis compacto · fuente para el REGISTRO INTEGRAL)

El árbitro es SIEMPRE la celda. Uso: python scripts/dev/gm_full_audit.py
Dylus Lab © 2026
"""
from __future__ import annotations

import re
import sys
from pathlib import Path

REPO = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(REPO))
try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")  # consola Windows = cp1252
except Exception:
    pass

# Referencia a otra hoja: 'Nombre Hoja'!  ó  NombreHoja!  (acentos permitidos, sin operadores)
REF_RE = re.compile(r"(?:'((?:[^']|'')+)'|([^\s'\"!=+\-*/(),;:<>&^%$\[\]{}]+))!")
ERROR_TOKENS = ("#REF!", "#DIV/0!", "#NAME?", "#VALUE!", "#N/A", "#NULL!", "#NUM!", "#SPILL!", "#CALC!")
MARKERS = (
    "PENDIENTE", "MISSING", "HILO ROTO", "HILO_ROTO", "TBL_CALIBRACION", "GAD_SIN_ESIGEF",
    "HARDCODE", "POR DEFINIR", "A PRECISAR", "NO DISPONIBLE", "PLACEHOLDER", "SIN FUENTE",
    "SIN DATO", "FALTA ", "POR CALCULAR", "TODO:", "FIXME", "DUMMY",
)
NONFORMULA_CAP = 600  # tope de celdas-dato no-fórmula por hoja (las fórmulas SIEMPRE van completas)


def _resolve_excel() -> Path | None:
    try:
        from app.connectors.gold_master import _resolve_path
        p = _resolve_path()
        if p and Path(p).exists():
            return Path(p)
    except Exception as e:  # noqa: BLE001
        print(f"<!-- resolve fail: {e} -->")
    backups = sorted((REPO / "data" / "backups").glob("GM_BACKUP_*.xlsx"))
    return backups[-1] if backups else None


def _safe(name: str) -> str:
    out = re.sub(r'[\\/:*?"<>|]', "_", name).strip().strip(".")
    return out or "_"


def main() -> int:
    import openpyxl

    path = _resolve_excel()
    if not path:
        print("GM_NOT_FOUND"); return 1
    outdir = REPO / "docs" / "architecture" / "gm_dumps"
    outdir.mkdir(parents=True, exist_ok=True)

    print(f"fuente: {path.name}")
    wb_f = openpyxl.load_workbook(path, read_only=True, data_only=False)  # fórmulas
    wb_v = openpyxl.load_workbook(path, read_only=True, data_only=True)   # valores cacheados
    names = list(wb_f.sheetnames)
    real = {n.upper(): n for n in names}

    sheets: dict[str, dict] = {}
    global_markers: list[str] = []
    global_errors: list[str] = []

    for name in names:
        ws_f = wb_f[name]
        ws_v = wb_v[name]

        formulas: list[tuple[str, str]] = []   # (coord, formula)
        labels: list[tuple[str, str]] = []      # (coord, texto) no-fórmula
        inputs: set[str] = set()
        unresolved: set[str] = set()
        markers: list[str] = []
        errors: list[str] = []
        total = 0

        for row in ws_f.iter_rows():
            for cell in row:
                v = cell.value
                if v is None:
                    continue
                s = str(v)
                if s.strip() == "":
                    continue
                total += 1
                is_formula = isinstance(v, str) and v.startswith("=")
                if is_formula:
                    formulas.append((cell.coordinate, s))
                    for m in REF_RE.finditer(s):
                        ref = (m.group(1).replace("''", "'") if m.group(1) else m.group(2))
                        key = ref.upper()
                        if key in real and real[key] != name:
                            inputs.add(real[key])
                        elif key not in real:
                            unresolved.add(ref)
                else:
                    if len(labels) < NONFORMULA_CAP:
                        labels.append((cell.coordinate, s))
                # errores y marcadores (en fórmula o texto)
                su = s.upper()
                for tok in ERROR_TOKENS:
                    if tok in s:
                        errors.append(f"{cell.coordinate}: {tok}")
                        global_errors.append(f"{name}!{cell.coordinate}: {tok}")
                        break
                for mk in MARKERS:
                    if mk in su:
                        snip = s.strip()[:70]
                        markers.append(f"{cell.coordinate}: {snip}")
                        global_markers.append(f"{name}!{cell.coordinate}: {snip}")
                        break

        # llenado real (valores cacheados, semántica del surface dump)
        pop = rows_t = 0
        for vr in ws_v.iter_rows(values_only=True):
            rows_t += 1
            if any(c is not None and str(c).strip() != "" for c in vr):
                pop += 1
            if rows_t >= 3000:
                break

        sheets[name] = dict(
            formulas=formulas, labels=labels, inputs=inputs, unresolved=unresolved,
            markers=markers, errors=errors, total=total, pop=pop, rows_t=rows_t,
        )

    wb_f.close(); wb_v.close()

    # outputs = inverso de inputs
    outputs: dict[str, set[str]] = {n: set() for n in names}
    for n in names:
        for inp in sheets[n]["inputs"]:
            outputs.setdefault(inp, set()).add(n)

    # ── per-hoja dumps ──
    for name in names:
        d = sheets[name]
        lines = [
            f"# {name} — volcado determinista (fórmulas + etiquetas)",
            f"fuente: `{path.name}` · filas={d['rows_t']} · pobladas={d['pop']} · fórmulas={len(d['formulas'])}",
            f"inputs(lee de): {', '.join(sorted(d['inputs'])) or '—'}",
            f"outputs(alimenta a): {', '.join(sorted(outputs[name])) or '—'}",
        ]
        if d["unresolved"]:
            lines.append(f"refs no resueltas: {', '.join(sorted(d['unresolved']))}")
        if d["errors"]:
            lines.append(f"ERRORES: {', '.join(d['errors'])}")
        if d["markers"]:
            lines.append("MARCADORES: " + " · ".join(d["markers"]))
        lines += ["", "## FÓRMULAS", "```"]
        lines += [f"{c}\t{f}" for c, f in d["formulas"]] or ["(sin fórmulas)"]
        lines += ["```", "", f"## ETIQUETAS / DATOS (tope {NONFORMULA_CAP})", "```"]
        lines += [f"{c}\t{t}" for c, t in d["labels"]]
        if d["total"] - len(d["formulas"]) > len(d["labels"]):
            lines.append(f"[TRUNCADO: {d['total'] - len(d['formulas']) - len(d['labels'])} celdas-dato más]")
        lines += ["```"]
        (outdir / f"{_safe(name)}.md").write_text("\n".join(lines), encoding="utf-8")

    # ── _INDEX.md ──
    idx = [
        "# GM gm_dumps — ÍNDICE (volcado integral determinista)",
        f"**fuente:** `{path.name}` · **hojas:** {len(names)} · `scripts/dev/gm_full_audit.py`",
        "",
        "| # | Hoja | filas | pobl. | fórmulas | in | out | err | marc |",
        "|---|---|---|---|---|---|---|---|---|",
    ]
    for i, name in enumerate(names, 1):
        d = sheets[name]
        idx.append(f"| {i} | [{name}]({_safe(name)}.md) | {d['rows_t']} | {d['pop']} | "
                   f"{len(d['formulas'])} | {len(d['inputs'])} | {len(outputs[name])} | "
                   f"{len(d['errors'])} | {len(d['markers'])} |")
    (outdir / "_INDEX.md").write_text("\n".join(idx), encoding="utf-8")

    # ── _ANALYSIS.md (compacto · fuente del registro) ──
    an = [
        "# GM _ANALYSIS — análisis compacto (fuente del REGISTRO INTEGRAL)",
        f"**fuente:** `{path.name}` · hojas={len(names)} · errores_totales={len(global_errors)} · "
        f"marcadores_totales={len(global_markers)}",
        "",
        "## GLOBAL — ERRORES (#REF!/#DIV/0!/…)",
    ]
    an += [f"- {e}" for e in global_errors] or ["- (ninguno)"]
    an += ["", "## GLOBAL — MARCADORES (PENDIENTE/MISSING/Ti=0/…)"]
    an += [f"- {m}" for m in global_markers] or ["- (ninguno)"]
    # dependencias clave
    an += ["", "## DEPENDENCIAS CLAVE (inputs directos por fórmula)"]
    for key in ("H73_OUTPUT_API", "H12_MOTOR_ICPI_CANÓNICO", "H98_TGI_FRAMEWORK",
                "H99_ENGINE_CORE", "H15_ICPI_GLOBAL"):
        if key in sheets:
            an.append(f"- **{key}** ← {', '.join(sorted(sheets[key]['inputs'])) or '(sin refs de fórmula · snapshot)'}")
    an += ["", "## FICHAS POR HOJA"]
    for i, name in enumerate(names, 1):
        d = sheets[name]
        longest = max((f for _, f in d["formulas"]), key=len, default="")
        coord = next((c for c, f in d["formulas"] if f == longest), "")
        pct = round(100 * d["pop"] / d["rows_t"]) if d["rows_t"] else 0
        an.append(f"\n### {i}. {name}  [pobl {d['pop']}/{d['rows_t']} ~{pct}% · fórmulas={len(d['formulas'])}]")
        an.append(f"IN: {', '.join(sorted(d['inputs'])) or '—'}")
        an.append(f"OUT: {', '.join(sorted(outputs[name])) or '—'}")
        an.append(f"KEYF {coord}: {longest[:240] if longest else '(sin fórmulas · hoja de datos/snapshot)'}")
        if d["errors"]:
            an.append("ERR: " + " · ".join(d["errors"]))
        if d["markers"]:
            an.append("MARK: " + " · ".join(d["markers"][:8]))
    (outdir / "_ANALYSIS.md").write_text("\n".join(an), encoding="utf-8")

    print(f"hojas procesadas: {len(names)} · errores: {len(global_errors)} · marcadores: {len(global_markers)}")
    print(f"-> {outdir.relative_to(REPO)}/  (_INDEX.md · _ANALYSIS.md · {len(names)} volcados)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
