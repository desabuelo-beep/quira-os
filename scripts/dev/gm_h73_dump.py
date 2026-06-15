"""
scripts/dev/gm_h73_dump.py — Volcado del contrato de salida del Gold Master (determinista · $0)

Dumpea H73_OUTPUT_API (la hoja que el conector LEE como fuente de verdad) — todas las
parejas clave→valor pobladas. Es "sincerar el Excel": ver QUÉ entrega realmente el motor,
para cablear el proyecto a ese contrato (no a demo_data).

Escribe el volcado completo a `docs/architecture/GM_H73_DUMP.md` (UTF-8, durable) y
imprime un resumen seguro a stdout. NO lee por LLM · NO modifica el Excel.
Uso: python scripts/dev/gm_h73_dump.py
Dylus Lab © 2026
"""
from __future__ import annotations

import sys
from pathlib import Path

REPO = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(REPO))

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")  # consola Windows = cp1252
except Exception:
    pass


def _resolve_excel() -> Path | None:
    try:
        from app.connectors.gold_master import _resolve_path
        p = _resolve_path()
        if p and Path(p).exists():
            return Path(p)
    except Exception as e:  # noqa: BLE001
        print(f"<!-- resolve fail: {e} -->")
    backups = sorted((REPO / "data" / "backups").glob("GM_BACKUP_*.xlsx"))
    return backups[-1] if backups else None


def main() -> int:
    try:
        import openpyxl
    except ImportError:
        print("openpyxl no instalado"); return 1

    path = _resolve_excel()
    if not path:
        print("GM_NOT_FOUND"); return 1

    try:
        from app.connectors.gold_master import _OUTPUT_API_SHEET, _OUTPUT_API_SHEET_V6
    except Exception:
        _OUTPUT_API_SHEET, _OUTPUT_API_SHEET_V6 = "H73_OUTPUT_API", "G6.1_OUTPUT_API"

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet = _OUTPUT_API_SHEET if _OUTPUT_API_SHEET in wb.sheetnames else (
        _OUTPUT_API_SHEET_V6 if _OUTPUT_API_SHEET_V6 in wb.sheetnames else None)
    if sheet is None:
        print(f"Hoja de salida no encontrada. Hojas: {len(wb.sheetnames)}")
        return 1

    ws = wb[sheet]
    rows = []
    for row in ws.iter_rows(values_only=True):
        cells = [str(c).strip() for c in row if c is not None and str(c).strip() != ""]
        if cells:
            rows.append(" | ".join(cells))
    wb.close()

    # ── artefacto durable (UTF-8) ──
    out = REPO / "docs" / "architecture" / "GM_H73_DUMP.md"
    lines = [
        f"# GM H73_OUTPUT_API — contrato de salida del motor (volcado determinista)",
        f"**fuente:** `{path.name}` · **hoja:** {sheet} · **filas con dato:** {len(rows)}",
        "", "```",
    ] + rows + ["```"]
    out.write_text("\n".join(lines), encoding="utf-8")

    # ── stdout (utf-8 reconfigurado) ──
    print(f"fuente: {path.name} · hoja: {sheet} · filas con dato: {len(rows)}")
    print(f"dump completo -> {out.relative_to(REPO)}")
    print("=" * 60)
    for r in rows:
        print(r)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
