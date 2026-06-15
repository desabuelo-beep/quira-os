"""
scripts/dev/gm_probe.py — Sonda de VALORES cacheados del Gold Master (determinista · solo lectura · $0)

Imprime el valor CACHEADO (data_only=True) de los rangos pedidos — para confirmar el estado VIVO
de auto-chequeos/validaciones SIN recalcular y SIN adivinar. El árbitro es la celda.

Uso: python scripts/dev/gm_probe.py "Hoja!A1:D40" "Hoja2!A1:B14" ...
Sin args: usa el set de auditoría (motor/validaciones). Escribe a gm_dumps/_PROBE_VALUES.md.
Dylus Lab © 2026
"""
from __future__ import annotations

import sys
from pathlib import Path

REPO = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(REPO))
try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

DEFAULT = [
    "H12_MOTOR_ICPI_CANÓNICO!A6:I34",
    "H39_AUTOCONTROL_ECOSISTEMA!A23:D39",
    "H97_VALIDACIONES!A1:F27",
    "H85_ALERTS_LOG!A20:G40",
    "H90_PRESUPUESTO_CONSOLIDADO_202!A1:F22",
    "H07b_Ti_INVERSIÓN_eSIGEF!A1:C30",
    "H07_S5_FINANCIERO_eSIGEF!A40:E60",
    "H75_SAT_ENGINE!A1:B14",
    "H71_EP_ADSCRITAS!A6:I26",
]


def _resolve_excel() -> Path | None:
    try:
        from app.connectors.gold_master import _resolve_path
        p = _resolve_path()
        if p and Path(p).exists():
            return Path(p)
    except Exception:
        pass
    return None


def main() -> int:
    import openpyxl
    from openpyxl.utils.cell import range_boundaries, get_column_letter

    path = _resolve_excel()
    if not path:
        print("GM_NOT_FOUND"); return 1
    ranges = sys.argv[1:] or DEFAULT
    wb = openpyxl.load_workbook(path, read_only=False, data_only=True)

    out = [f"# GM _PROBE_VALUES — valores cacheados (determinista)", f"fuente: `{path.name}`", ""]
    for spec in ranges:
        if "!" not in spec:
            continue
        sheet, rng = spec.rsplit("!", 1)
        if sheet not in wb.sheetnames:
            out.append(f"## {spec}\n(HOJA NO EXISTE)\n"); continue
        ws = wb[sheet]
        out.append(f"## {sheet}!{rng}")
        out.append("```")
        min_c, min_r, max_c, max_r = range_boundaries(rng)
        for r in range(min_r, max_r + 1):
            cells = []
            for c in range(min_c, max_c + 1):
                v = ws.cell(row=r, column=c).value
                if v is not None and str(v).strip() != "":
                    cells.append(f"{get_column_letter(c)}{r}={v}")
            if cells:
                out.append(" · ".join(cells))
        out.append("```")
        out.append("")
    wb.close()

    dest = REPO / "docs" / "architecture" / "gm_dumps" / "_PROBE_VALUES.md"
    dest.write_text("\n".join(out), encoding="utf-8")
    print("\n".join(out))
    print(f"\n-> {dest.relative_to(REPO)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
