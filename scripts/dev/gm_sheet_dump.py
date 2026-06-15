"""
scripts/dev/gm_sheet_dump.py — Auditor de hoja del Gold Master (FÓRMULAS · solo lectura · $0)

Dumpea las celdas pobladas de una hoja con su FÓRMULA (no solo el valor), para auditar la
lógica del motor sin tocarlo. Determinista · NO modifica el Excel · NO lee por LLM.

Uso: python scripts/dev/gm_sheet_dump.py [prefijo_hoja]   (default: H12_MOTOR_ICPI)
Dylus Lab © 2026
"""
from __future__ import annotations

import sys
from pathlib import Path

REPO = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(REPO))
try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass


def _resolve_excel() -> Path | None:
    try:
        from app.connectors.gold_master import _resolve_path
        p = _resolve_path()
        if p and Path(p).exists():
            return Path(p)
    except Exception:
        pass
    backups = sorted((REPO / "data" / "backups").glob("GM_BACKUP_*.xlsx"))
    return backups[-1] if backups else None


def main() -> int:
    import openpyxl
    prefix = (sys.argv[1] if len(sys.argv) > 1 else "H12_MOTOR_ICPI").upper()
    path = _resolve_excel()
    if not path:
        print("GM_NOT_FOUND"); return 1

    # data_only=False → fórmulas; read_only para eficiencia
    wb = openpyxl.load_workbook(path, read_only=True, data_only=False)
    match = [s for s in wb.sheetnames if s.upper().startswith(prefix)]
    if not match:
        print(f"Sin hoja con prefijo '{prefix}'. Candidatas H12: "
              f"{[s for s in wb.sheetnames if s.upper().startswith('H12')]}")
        return 1
    sheet = match[0]
    ws = wb[sheet]

    rows = []
    for row in ws.iter_rows():
        for cell in row:
            v = cell.value
            if v is not None and str(v).strip() != "":
                rows.append(f"{cell.coordinate}\t{v}")
    wb.close()

    out = REPO / "docs" / "architecture" / f"GM_SHEET_{prefix}.md"
    out.write_text(
        f"# {sheet} — fórmulas (volcado determinista)\nfuente: `{path.name}`\n\n```\n"
        + "\n".join(rows) + "\n```\n", encoding="utf-8")

    print(f"hoja: {sheet} · celdas con dato: {len(rows)} · -> {out.relative_to(REPO)}")
    print("=" * 60)
    for r in rows:
        print(r)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
