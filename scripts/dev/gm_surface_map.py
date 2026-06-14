"""
scripts/dev/gm_surface_map.py — Mapa de Superficie del Gold Master (determinista · $0)

Enumera las hojas del Gold Master v5.5 vivo, su llenado y un estado tentativo.
NO lee por LLM · NO modifica el Excel · solo DESCRIBE su superficie.

- Escribe la tabla completa a `docs/architecture/GM_SURFACE_DUMP.md` (artefacto durable).
- Imprime a stdout solo lo barato: total, resumen y las hojas NO llenas + una watchlist
  de hojas clave del motor (las que anclan a cajones con huecos conocidos).

Insumo del "Mapa de Anclaje del Motor". Auditoría de fórmulas = aparte (diferida).
Uso: python scripts/dev/gm_surface_map.py
Dylus Lab © 2026
"""
from __future__ import annotations

import sys
from pathlib import Path

REPO = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(REPO))

# Hojas clave del motor que anclan a cajones (watchlist — siempre se reportan)
WATCH = ("H73", "H12", "H07", "H98", "H99", "H11b", "H26", "H31", "H75", "H24", "H10")


def _resolve_excel() -> Path | None:
    """Usa la misma resolución que la app; si no, el backup más reciente."""
    try:
        from app.connectors.gold_master import _resolve_path
        p = _resolve_path()
        if p and Path(p).exists():
            return Path(p)
    except Exception as e:  # noqa: BLE001
        print(f"<!-- resolve fail: {e} -->")
    backups = sorted((REPO / "data" / "backups").glob("GM_BACKUP_*.xlsx"))
    return backups[-1] if backups else None


def _classify(nonempty: int, total: int) -> str:
    if nonempty <= 1:
        return "MUERTA"
    if nonempty < 25 and total and (nonempty / total) < 0.5:
        return "INCOMPLETA?"
    return "LLENA"


def main() -> int:
    try:
        import openpyxl
    except ImportError:
        print("openpyxl no instalado"); return 1

    path = _resolve_excel()
    if not path:
        print("GM_NOT_FOUND — Excel vivo no está en la ruta del conector ni hay backup.")
        return 1

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    names = wb.sheetnames
    CAP = 3000
    rows = []   # (name, total, nonempty, estado)
    summ = {"LLENA": 0, "INCOMPLETA?": 0, "MUERTA": 0}
    for name in names:
        ws = wb[name]
        nonempty = total = 0
        for row in ws.iter_rows(values_only=True):
            total += 1
            if any(c is not None and str(c).strip() != "" for c in row):
                nonempty += 1
            if total >= CAP:
                break
        est = _classify(nonempty, total)
        summ[est] = summ.get(est, 0) + 1
        rows.append((name, total, nonempty, est))
    wb.close()

    # --- artefacto durable: tabla completa a disco ---
    out = REPO / "docs" / "architecture" / "GM_SURFACE_DUMP.md"
    lines = [
        "# GM SURFACE DUMP — superficie viva del Gold Master (determinista)",
        f"**fuente:** `{path.name}` · **hojas:** {len(names)} · generado por `scripts/dev/gm_surface_map.py`",
        "",
        "| Hoja | filas | filas c/dato | estado |",
        "|---|---|---|---|",
    ]
    for name, total, nonempty, est in rows:
        cap = "+" if total >= CAP else ""
        lines.append(f"| {name} | {total}{cap} | {nonempty} | {est} |")
    lines += ["", f"resumen: LLENA={summ['LLENA']} · INCOMPLETA?={summ['INCOMPLETA?']} · MUERTA={summ['MUERTA']}", ""]
    out.write_text("\n".join(lines), encoding="utf-8")

    # --- stdout barato: resumen + no-llenas + watchlist ---
    print(f"fuente: {path.name} | hojas: {len(names)}")
    print(f"resumen: LLENA={summ['LLENA']} · INCOMPLETA?={summ['INCOMPLETA?']} · MUERTA={summ['MUERTA']}")
    print(f"dump completo -> {out.relative_to(REPO)}")
    print("\n-- NO LLENAS (huecos candidatos) --")
    any_hole = False
    for name, total, nonempty, est in rows:
        if est != "LLENA":
            print(f"  {est:12} {name}  (filas={total} c/dato={nonempty})")
            any_hole = True
    if not any_hole:
        print("  (ninguna)")
    print("\n-- WATCHLIST motor (anclan a cajones) --")
    for name, total, nonempty, est in rows:
        if name.startswith(WATCH):
            print(f"  {est:12} {name}  (filas={total} c/dato={nonempty})")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
