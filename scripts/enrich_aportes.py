# -*- coding: utf-8 -*-
"""
enrich_aportes.py — Cable de Trazabilidad de Aportes Ciudadanos (d09)
QUIRA Gov · Dylus Lab © 2026

Implementa METODOLOGIA_TRAZABILIDAD_APORTES.md v0.2 (Regla 9: deriva del canon
conceptual, no lo inventa). Cruza cada aporte ciudadano (H10c) contra la ejecución
del PERIODO DE GOBIERNO (POA de años > año-del-aporte, extraídos de los PDFs
oficiales vía extract_poa_pdf) y clasifica:

  Nivel_Atención  : Atendido (≥0.62) · Por validar (0.52–0.62) · Sin correlato (<0.52)
  Tiempo_Respuesta: A tiempo (año+1) · Tarde (dentro del periodo) · Olvidado

Método semiautomático (evaluación experta trazable, como el IF_n): la banda
0.52–0.62 se marca «por validar» — la máquina propone, el analista confirma.
Escribe snap["rendicion"]["aportes"] (merge, preserva fidelidad/cpccs/serie).

Uso:  python scripts/enrich_aportes.py
"""
from __future__ import annotations

import json
import sys
from collections import Counter
from pathlib import Path

import numpy as np
import openpyxl

_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(_ROOT / "scripts" / "normativa"))
sys.path.insert(0, str(_ROOT / "scripts"))
import ingest as ing                       # modelo local de embeddings (offline)
from extract_poa_pdf import extract_poa

try:
    from config import DATOS_DIR as _DATOS
except Exception:                                        # noqa: BLE001
    import os as _os
    from pathlib import Path as _P
    _DATOS = _P(_os.environ.get("QUIRA_DATOS", "."))
def _gold_master_vigente() -> str:
    """El Gold Master VIGENTE, resuelto por `config` (sufijo `_TGI`,
    versión más alta). Antes aquí había un literal `v5.5` y por eso este
    módulo leía una versión que el canon ya había superado — D-002."""
    try:
        from config import SIAP_PATH
        return str(SIAP_PATH)
    except Exception:                                    # noqa: BLE001
        return str(_DATOS / "SIAP-ICPI_GOLD_MASTER_v5.5_TGI.xlsx")


EXCEL = _gold_master_vigente()
SNAP = _ROOT / "data" / "gm_snapshot.json"

TH_ATENDIDO = 0.62     # candidato fuerte
TH_VALIDAR = 0.52      # banda de validación experta (0.52–0.62)
AÑOS_POA = (2024, 2025, 2026)   # POA disponibles en PDF (2023 no aplica: es el año base)

# Eje PDOT (col I de H10c) → etiqueta legible
EJE = {"SC": "Sociocultural", "AH": "Asentamientos Humanos", "FA": "Físico-Ambiental",
       "EP": "Económico Productivo", "PI": "Político Institucional"}


def _cargar_aportes() -> list[dict]:
    wb = openpyxl.load_workbook(EXCEL, data_only=True, read_only=True)
    ws = wb["H10c_RDC_APORTES"]
    out = []
    for r in range(5, ws.max_row + 1):
        a = str(ws.cell(r, 1).value or "").strip()
        dem = ws.cell(r, 6).value
        if a not in ("2023", "2024") or not dem:
            continue
        d = str(dem).strip()
        if len(d) < 10 or d.upper() in ("N/A", "NA", "CUMPLIDO"):
            continue
        sec = str(ws.cell(r, 5).value or "").strip()
        eje = str(ws.cell(r, 9).value or "").strip()
        out.append({"anio": int(a), "demanda": d, "sector": sec,
                    "eje": EJE.get(eje, eje or "—"), "texto": f"{d} ({sec})"})
    wb.close()
    return out


def _embed(model, txts):
    v = model.encode(txts, convert_to_numpy=True, show_progress_bar=False, batch_size=64)
    return v / (np.linalg.norm(v, axis=1, keepdims=True) + 1e-9)


def build() -> dict:
    aportes = _cargar_aportes()
    valid_path = _ROOT / "data" / "aportes_validacion.json"
    _VALID = json.loads(valid_path.read_text(encoding="utf-8")) if valid_path.exists() else {}
    poa = {y: extract_poa(str(y)) for y in AÑOS_POA}
    print(f"Aportes: {len(aportes)} · POA: {{{', '.join(f'{y}:{len(poa[y])}' for y in poa)}}}")

    model = ing._get_model()
    A = _embed(model, [a["texto"] for a in aportes])
    P = {y: _embed(model, [p["desc"] for p in poa[y]]) for y in poa if poa[y]}

    detalle = []
    for i, ap in enumerate(aportes):
        # score máximo por año (solo años posteriores al aporte)
        by_year = {}
        for y in P:
            if y <= ap["anio"]:
                continue
            s = A[i] @ P[y].T
            j = int(s.argmax())
            by_year[y] = (float(s[j]), poa[y][j]["desc"], poa[y][j].get("monto", 0.0))
        if not by_year:
            continue
        best_y = max(by_year, key=lambda y: by_year[y][0])
        best_score, best_desc, best_monto = by_year[best_y]
        # año de ejecución = primer año (cronológico) que supera el umbral de validación
        ejec_y = next((y for y in sorted(by_year) if by_year[y][0] >= TH_VALIDAR), None)

        if best_score >= TH_ATENDIDO:
            estado = "atendido"
        elif best_score >= TH_VALIDAR:
            estado = "por_validar"
        else:
            estado = "sin_correlato"

        # ── validación experta (overrides trazables · data/aportes_validacion.json) ──
        validado = False
        dem = ap["demanda"]
        resc = next((val for pref, val in _VALID.get("rescatar", {}).items() if dem.startswith(pref)), None)
        if resc is not None:                       # falso negativo rescatado → atendido
            estado, validado = "atendido", True
            if isinstance(resc, dict):             # {anio, obra}: apunta a la obra correcta del POA
                ejec_y = resc.get("anio")
                obra = (resc.get("obra") or "").lower()
                cand = []
                if ejec_y in P:
                    sc = A[i] @ P[ejec_y].T
                    cand = [(float(sc[k]), poa[ejec_y][k]["desc"], poa[ejec_y][k].get("monto", 0.0))
                            for k in range(len(poa[ejec_y])) if obra in poa[ejec_y][k]["desc"].lower()]
                if cand:
                    best_score, best_desc, best_monto = max(cand)
                elif ejec_y in by_year:
                    best_score, best_desc, best_monto = by_year[ejec_y]
            else:                                   # año simple: mejor match del año
                ejec_y = resc
                if resc in by_year:
                    best_score, best_desc, best_monto = by_year[resc]
        elif any(dem.startswith(p) for p in _VALID.get("rechazar", [])):   # falso positivo → sin correlato
            estado, validado = "sin_correlato", True
        elif estado == "por_validar":              # banda 0.52-0.62 no confirmada → sin correlato
            estado, validado = "sin_correlato", True

        if estado == "sin_correlato":
            tiempo, ejec_y = "olvidado", None
        else:
            tiempo = "a_tiempo" if ejec_y == ap["anio"] + 1 else "tarde"

        detalle.append({
            "anio_aporte": ap["anio"], "demanda": ap["demanda"], "sector": ap["sector"],
            "eje": ap["eje"], "estado": estado, "tiempo": tiempo,
            "anio_ejecucion": ejec_y, "validado": validado,
            "evidencia": best_desc[:120] if estado != "sin_correlato" else "",
            "monto": round(best_monto, 2) if best_monto else 0.0,
            "score": round(best_score, 3),
        })

    est = Counter(d["estado"] for d in detalle)
    tmp = Counter(d["tiempo"] for d in detalle)
    eje = Counter(d["eje"] for d in detalle)
    eje_atend = Counter(d["eje"] for d in detalle if d["estado"] != "sin_correlato")
    return {
        "_fuente": "H10c aportes (2023-24) × POA PDFs 2024-2026 · cruce semiautomático",
        "_nota": "Método semiautomático + PRE-VALIDACIÓN experta (data/aportes_validacion.json, "
                 "pendiente ratificación humana). Ventana = periodo de gobierno. La banda dudosa "
                 "0.52-0.62 se degrada a sin_correlato salvo confirmación (fue ~74% falsos positivos).",
        "total": len(detalle),
        "n_validados": sum(1 for d in detalle if d.get("validado")),
        "umbral": {"atendido": TH_ATENDIDO, "validar": TH_VALIDAR},
        "por_estado": dict(est),
        "por_tiempo": dict(tmp),
        "por_eje": {k: {"total": eje[k], "atendidos": eje_atend[k]} for k in eje},
        "detalle": detalle,
    }


def main() -> None:
    block = build()
    with open(SNAP, encoding="utf-8") as f:
        snap = json.load(f)
    snap.setdefault("rendicion", {})["aportes"] = block   # merge, no sobrescribe fidelidad/cpccs
    with open(SNAP, "w", encoding="utf-8") as f:
        json.dump(snap, f, ensure_ascii=False, indent=2)
    e = block["por_estado"]; t = block["por_tiempo"]
    print("OK - 'rendicion.aportes' escrito en gm_snapshot.json")
    print(f"   Estado:  atendido={e.get('atendido',0)} · por_validar={e.get('por_validar',0)} "
          f"· sin_correlato={e.get('sin_correlato',0)}")
    print(f"   Tiempo:  a_tiempo={t.get('a_tiempo',0)} · tarde={t.get('tarde',0)} · olvidado={t.get('olvidado',0)}")


if __name__ == "__main__":
    main()
