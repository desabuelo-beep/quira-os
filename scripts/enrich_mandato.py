"""
QUIRA OS — Enriquecedor del snapshot · bloque `mandato_dom` (DOM d03 · Gobernanza del Mandato)
═══════════════════════════════════════════════════════════════════════════════
Puente Excel→snapshot (Regla 1). Lee el Gold Master LOCAL (solo lectura · openpyxl
data_only · NO corrompe) y escribe `mandato_dom` en data/gm_snapshot.json.

VISIÓN d03: **la palabra empeñada.** Qué proporción de lo que el candidato prometió ante
el CNE se convirtió en meta del plan de desarrollo. No mide si la obra se hizo (eso es
ejecución): mide el **origen democrático** de lo que hoy se planifica.

ROSTER (auditoría del canon · 2026-07-15):
  ✅ H03_S1_ELECTORAL_CNE — promesas + fidelidad. FUENTE CANÓNICA (ingesta curada y
     corroborada por Javo · validación humana, ADR-035 §5).
  ✅ H16_IFE            — el índice, su clasificación y su escala.
  ✅ SCHEMA_CNE         — autoridades electas (con ausencias declaradas [VER_CNE]).
  ❌ H63_S0_CNE_TRAZABILIDAD — NO CANÓNICA: contiene 66 promesas genéricas de plantilla
     que NO existen en el Plan de Trabajo real (quedó a medias · Javo 2026-07-15).
  ❌ H20b_IGP           — es Participación (d08): Asamblea CPCCS + PP + Fidelidad Narrativa.

HONESTIDAD EPISTEMOLÓGICA (Principio Rector): el registro publica 20 promesas trazadas
de 66 declaradas. Esa brecha se DECLARA (`cobertura_registro`), no se disimula.

Uso:  python scripts/enrich_mandato.py
Dylus Lab © 2026
"""
from __future__ import annotations

import hashlib
import json
import os
import re

import openpyxl

EXCEL = r"C:\Users\DELL\Desktop\Javo\Dylus Lab\ProyecT\SIAP-ICPI_GOLD_MASTER_v5.5_TGI.xlsx"
PLAN_CNE = (r"C:\Users\DELL\Desktop\Javo\Dylus Lab\ProyecT\Holding_Municipal_Montecristi"
            r"\Plan CNE ALcalde Montecristi.docx")
SNAP = os.path.join(os.path.dirname(__file__), "..", "data", "gm_snapshot.json")

_HCODE = re.compile(r"\bH\d{1,2}[a-z]?\b")

# Score de vinculación → lenguaje público (Firewall: jamás "Score_IFE" hacia afuera)
_NIVEL = {
    1.0:  ("Directa", "#1E8E3E", "la promesa se convirtió en meta del plan"),
    0.75: ("Sustancial", "#1E8E3E", "la meta recoge la promesa casi por completo"),
    0.5:  ("Parcial", "#F9AB00", "la meta la recoge solo en parte"),
    0.0:  ("Sin correspondencia", "#D93025", "no hay meta del plan que la recoja"),
}


def _fw(s) -> bool:
    """Seguro para público: sin nomenclatura canónica (H01-H99) ni fila-nota."""
    s = str(s or "")
    return not _HCODE.search(s) and not s.strip().upper().startswith(("NOTA", "INSTRUCCIÓN", "ESCALA"))


def _num(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _sha(path: str) -> str:
    try:
        with open(path, "rb") as f:
            return hashlib.sha256(f.read()).hexdigest()
    except OSError:
        return ""


def build_block() -> dict:
    wb = openpyxl.load_workbook(EXCEL, read_only=True, data_only=True)

    def sh(prefix: str):
        return wb[next(s for s in wb.sheetnames if s.startswith(prefix))]

    def find(ws, etiqueta: str, col_val: int = 2, maxr: int = 90):
        for r in ws.iter_rows(min_row=1, max_row=maxr, values_only=True):
            if r and r[0] and etiqueta.lower() in str(r[0]).lower():
                return r[col_val - 1] if len(r) >= col_val else None
        return None

    # ── ① EL MANDATO: promesas registradas (H03 · fuente canónica validada) ───
    ws = sh("H03_S1")
    total = int(_num(find(ws, "Total_Promesas_CNE")) or 0)
    con_meta = int(_num(find(ws, "Promesas_Con_Meta_PDOT")) or 0)
    ife = _num(find(ws, "IFE_Global")) or 0
    clasif = re.sub(r"[🔴🟡🟢🟠⚠️✅]", "", str(find(ws, "Clasificación_IFE") or "")).strip(" —·")
    anio = int(_num(find(ws, "Año_Elección")) or 2023)

    # OJO (lección 2026-07-15): el filtro previo exigía len(id)>=6 y perdía las 46 promesas
    # "PR-21".."PR-66" (5 chars) → se auditaba sobre una muestra mutilada. Filtro correcto:
    _PID = re.compile(r"^(EC|IN|SC|PR)-\d{2,3}$")
    promesas, ejes = [], {}
    for r in ws.iter_rows(min_row=1, max_row=90, values_only=True):
        if not (r and r[0] and isinstance(r[0], str) and _PID.match(r[0].strip())):
            continue
        pid = r[0].strip()
        desc = str(r[2] or "").strip()
        meta = str(r[3] or "").strip()
        sc = _num(r[4])
        tipo = str(r[5] or "").strip()
        if not _fw(pid) or not _fw(desc):
            continue
        nivel, col, gloss = _NIVEL.get(sc, ("Sin correspondencia", "#D93025", ""))
        eje = str(r[1] or "").strip()
        ejes[eje] = ejes.get(eje, 0) + 1
        # NIVEL DE VERIFICACIÓN — se LEE del canon (col G · curada por Javo 2026-07-15).
        # Antes se INFERÍA del prefijo del ID ("PR-" = pendiente): eso nacía en Python y violaba
        # la Regla 9. El DOM reveló el defecto y el canon lo absorbió: ahora el estado es dato.
        estado_canon = str(r[6] or "").strip() if len(r) > 6 else ""
        verificada = estado_canon.lower().startswith("verific")
        promesas.append({
            "id": pid, "eje": eje, "promesa": desc[:150],
            "meta": meta if meta and meta not in ("—", "-") else "",
            "nivel": nivel, "color": col, "glosa": gloss,
            "tipo": tipo if _fw(tipo) else "", "score": sc if sc is not None else 0,
            "verificada": verificada,
            "estado": estado_canon if _fw(estado_canon) else "",
        })

    # LAS DOS VERDADES (familia de fidelidad · aporte del colega 2026-07-15):
    #  · INCORPORACIÓN = HECHO documental: cuántas promesas tienen meta (contar ≠ recalcular el motor).
    #  · CALIDAD       = ÍNDICE del motor: suma de scores / total (pondera cómo se incorporó).
    # No se inventan los demás eslabones (POA/presupuesto/ejecución): sin dato, no hay índice (Regla 3),
    # y esas transiciones ya se miden en d01/d02 — aquí se REFERENCIAN, no se recalculan (Regla 4 · ADR-032).
    reg_con_meta = sum(1 for p in promesas if p["meta"])
    n_verif = sum(1 for p in promesas if p["verificada"] and p["meta"])
    suma_score = round(sum(p["score"] for p in promesas), 2)
    n_reg = len(promesas)

    # ── ② EL ÍNDICE (H16) — panel, interpretación y escala ────────────────────
    ws16 = sh("H16_IFE")
    interp = str(find(ws16, "Interpretación") or "").strip()
    escala = []
    for r in ws16.iter_rows(min_row=1, max_row=40, values_only=True):
        if r and r[0] and isinstance(r[0], str) and "%" in r[0] and r[1]:
            niv = re.sub(r"[🔴🟡🟢🟠⚠️✅]", "", str(r[1])).strip()
            if _fw(niv):
                escala.append({"umbral": r[0].strip(), "nivel": niv})

    # ── ③ AUTORIDADES ELECTAS (SCHEMA_CNE) — con ausencias declaradas ─────────
    wsa = wb["SCHEMA_CNE"]
    autoridades, sin_dato = [], 0
    for r in wsa.iter_rows(min_row=1, max_row=22, values_only=True):
        if not (r and r[0] and isinstance(r[0], str)):
            continue
        cargo = r[0].strip()
        if not any(k in cargo.lower() for k in ("alcalde", "concejal", "junta")):
            continue
        nombre = str(r[1] or "").strip()
        ausente = "VER_CNE" in nombre or not nombre
        if ausente:
            sin_dato += 1
        autoridades.append({
            "cargo": cargo,
            "nombre": "" if ausente else nombre,
            "movimiento": "" if "VER_CNE" in str(r[2] or "") else str(r[2] or "").strip(),
            "verificado": not ausente,
        })

    return {
        "_fuente": f"Plan de Trabajo del CNE {anio} (documento verificado por huella SHA256) · "
                   f"metas del plan de desarrollo · corte 2026",
        "vision": "La palabra empeñada: qué proporción de lo prometido en campaña se convirtió en "
                  "meta del plan de desarrollo. Mide el origen democrático de lo que hoy se planifica.",
        "eleccion": {
            "anio": anio, "periodo": f"{anio}-{anio + 4}",
            "documento": "Plan de Trabajo presentado ante el Consejo Nacional Electoral",
            "sha256": _sha(PLAN_CNE)[:16],
        },
        # ── INCORPORACIÓN · con su NIVEL DE VERIFICABILIDAD (Principio Rector) ──
        # Tener meta asignada ≠ tener la vinculación verificada. Publicar "97%" a secas
        # sería inferir: solo 20 de las 66 están contrastadas contra el documento del CNE.
        "incorporacion": {   # 🟢 HECHO documental (contar el registro)
            "con_meta": reg_con_meta, "total": n_reg,
            "sin_meta": n_reg - reg_con_meta,
            "pct": round(reg_con_meta / n_reg * 100, 1) if n_reg else 0,
            "pregunta": "¿La promesa ingresó al plan de desarrollo?",
            "verificadas": n_verif,
            "pendientes": reg_con_meta - n_verif,
            "pct_verificado": round(n_verif / n_reg * 100, 1) if n_reg else 0,
            "cautela": "El registro asigna meta a casi todas, pero solo una parte está contrastada "
                       "contra el documento electoral. La diferencia se declara: asignada no es verificada.",
        },
        "calidad": {         # 🔵 ÍNDICE del motor (pondera CÓMO ingresó)
            "pct": round(ife * 100, 1) if ife <= 1 else round(ife, 1),
            "pct_recalculado": round(suma_score / n_reg * 100, 2) if n_reg else 0,
            "suma_score": suma_score, "base": n_reg,
            "clasificacion": clasif,
            "interpretacion": interp if _fw(interp) else "",
            "escala": escala,
            "pregunta": "¿Con qué nivel de congruencia ingresó?",
        },
        # Auditoría del canon (2026-07-15 · CURADA por Javo el mismo día):
        # El parámetro rotulado "Promesas_Con_Meta_PDOT = 48" NO era un conteo sino la suma de
        # scores. Se curó: A8 pasó a "Suma_Score_Vinculación" y el conteo real (64) vive en su
        # propia celda, sin alimentar el índice. Ninguna fórmula fue tocada; el centinela del
        # motor volvió a ✅ y el ICPI quedó idéntico. Este bloque queda como acta de trazabilidad.
        "auditoria_canon": {
            "parametro_conteo": con_meta, "conteo_real_con_meta": reg_con_meta,
            "suma_score_real": suma_score,
            "coherente": con_meta == reg_con_meta,
            "nota": "Canon curado: el rótulo ya distingue la suma de puntajes del conteo de promesas.",
        },
        "por_eje": ejes,
        "promesas": promesas,
        "autoridades": {"detalle": autoridades, "sin_verificar": sin_dato, "total": len(autoridades)},
        "publicado": True,
    }


def main() -> None:
    block = build_block()
    snap = {}
    if os.path.exists(SNAP):
        with open(SNAP, encoding="utf-8") as f:
            snap = json.load(f)
    snap["mandato_dom"] = block
    with open(SNAP, "w", encoding="utf-8") as f:
        json.dump(snap, f, ensure_ascii=False, indent=2)
    inc, cal, aud = block["incorporacion"], block["calidad"], block["auditoria_canon"]
    print("OK - bloque 'mandato_dom' escrito en gm_snapshot.json")
    print(f"   INCORPORACIÓN (hecho) : {inc['con_meta']}/{inc['total']} = {inc['pct']}% con meta")
    print(f"     · verificadas       : {inc['verificadas']} ({inc['pct_verificado']}%) "
          f"· pendientes de contraste: {inc['pendientes']} · sin meta: {inc['sin_meta']}")
    print(f"   CALIDAD (índice motor): {cal['pct']}% (suma score {cal['suma_score']}/{cal['base']})")
    print(f"   AUDITORÍA canon       : conteo declarado={aud['parametro_conteo']} · real="
          f"{aud['conteo_real_con_meta']} · coherente={aud['coherente']}")
    print(f"   autoridades sin verificar: {block['autoridades']['sin_verificar']}/{block['autoridades']['total']}")


if __name__ == "__main__":
    main()
