"""
QUIRA OS — Enriquecedor del snapshot · bloque `planificacion` (QINV-001)
═══════════════════════════════════════════════════════════════════════════════
Puente Excel→snapshot (Regla 1: la fuente alimenta al deploy, nunca al revés).
Lee el Gold Master LOCAL (solo lectura · openpyxl data_only · NO corrompe) y
escribe el bloque `planificacion` en data/gm_snapshot.json para que el cajón
QINV-001 renderice NATIVO (sin cosechar páginas viejas) con dato verificado.

Secciones: A·PDOT (H04 metas+competencia) · B·POA (H05 direcciones) ·
C·PAC (H05b contratación) · D·Coherencia (H21b SAT-0). Corte Q1-2026.

Reproducible: correr tras refrescar el Gold Master. NO toca H12!B33 ni el canon.
Uso:  python scripts/enrich_planificacion.py
Dylus Lab © 2026
"""
from __future__ import annotations

import json
import os
import re
from collections import Counter, defaultdict

import openpyxl

EXCEL = r"C:\Users\DELL\Desktop\Javo\Dylus Lab\ProyecT\SIAP-ICPI_GOLD_MASTER_v5.5_TGI.xlsx"
SNAP = os.path.join(os.path.dirname(__file__), "..", "data", "gm_snapshot.json")


def _temp(estado: str) -> str:
    """Traduce el estado del motor (con emoji) a temperatura del firewall."""
    e = estado or ""
    if "🔴" in e:
        return "critico"
    if "⚠️" in e or "🟠" in e:
        return "alerta"
    if "✅" in e or "🟢" in e or "🔵" in e:
        return "verde"
    return "dim"


def _clean(s: str) -> str:
    """Quita emojis/marcadores de estado para almacenar texto limpio (firewall)."""
    for mk in ("🔴", "⚠️", "✅", "🟢", "🔵", "🟠", "⏳", "❌", "★"):
        s = s.replace(mk, "")
    return s.strip(" ·—-")


def build_block() -> dict:
    wb = openpyxl.load_workbook(EXCEL, read_only=True, data_only=True)

    def sh(prefix: str):
        return wb[next(s for s in wb.sheetnames if s.startswith(prefix))]

    # ── A · PDOT — metas por competencia COOTAD (H04 col D) ──────────────────
    ws = sh("H04_S2_PLAN")
    comp = Counter()
    metas = {}
    for r in range(15, 45):
        mid = ws.cell(r, 1).value
        v = ws.cell(r, 4).value
        if v:
            comp[str(v).strip().replace("_", " ")] += 1
        if mid and v:                      # solo metas reales del PDOT (con competencia COOTAD)
            metas[str(mid).strip()] = {
                "id": str(mid).strip(),
                "sistema": str(ws.cell(r, 2).value or "").strip(),
                "meta": str(ws.cell(r, 3).value or "").strip()[:95],
                "competencia": str(v or "").strip().replace("_", " "),
                "direccion": "",
            }
    metas_total = sum(comp.values())
    # orden canónico de severidad de competencia
    orden = ["Exclusiva Crítica", "Concurrente Crítica", "Exclusiva Importante", "Concurrente"]
    competencia = [{"label": k, "n": comp[k]} for k in orden if comp.get(k)]
    for k, n in comp.items():            # cualquier otra categoría no prevista
        if k not in orden:
            competencia.append({"label": k, "n": n})

    # ── B · POA — metas por dirección responsable (H05 col C) ────────────────
    ws = sh("H05_S3_OPER")
    dirs = Counter()
    for r in range(14, 45):
        mid = ws.cell(r, 1).value
        v = ws.cell(r, 3).value
        if v:
            name = str(v).split("+")[0].strip()       # dirección primaria (antes de "+")
            if name.startswith("Dir."):
                name = name[4:].strip()
            dirs[name] += 1
            if mid and str(mid).strip() in metas:
                metas[str(mid).strip()]["direccion"] = name
    direcciones = [{"dir": k, "n": n} for k, n in dirs.most_common()]
    n_direcciones = len(direcciones)

    # ── C · PAC — contratación (H05b) ────────────────────────────────────────
    ws = sh("H05b_S3b")
    total_pac = ws["B11"].value or 0
    tipos = Counter()
    n_proc = 0
    gap_proc = None
    for r in range(15, 60):
        pid = ws.cell(r, 1).value
        if not pid:
            continue
        n_proc += 1
        tipo = ws.cell(r, 3).value
        if tipo:
            tipos[str(tipo).strip()] += 1
        desc = str(ws.cell(r, 2).value or "")
        if "grupos prio" in desc.lower() or "atenci" in desc.lower():
            gap_proc = desc[:60]
    pac = {
        "total_usd": total_pac,
        "n_procesos": n_proc,
        "tipos": [{"tipo": t, "n": n} for t, n in tipos.most_common()],
        "gap_proceso": gap_proc,
    }

    # ── D · Coherencia — SAT-0 (H21b) ────────────────────────────────────────
    ws = sh("H21b_SAT")
    def cell(addr):
        return str(ws[addr].value or "")
    sat0 = {
        "componentes": [
            {"label": "Brecha POA-PAC", "estado": _clean(cell("B14")), "temp": _temp(cell("B14"))},
            {"label": "Downcoding contractual", "estado": _clean(cell("B15")), "temp": _temp(cell("B15"))},
            {"label": "Monto mínimo", "estado": _clean(cell("B16")), "temp": _temp(cell("B16"))},
            {"label": "Reloj de evidencia", "estado": _clean(cell("B17")), "temp": _temp(cell("B17"))},
        ],
        "global": _clean(cell("B19")),
        "global_temp": _temp(cell("B19")),
        "diagnostico": _clean(cell("B21")),
    }

    # ── PRESUPUESTO — eSIGEF inversión codificado (H07) ──────────────────────
    ws = sh("H07_S5")
    presupuesto = {
        "codificado_inversion": ws["B18"].value or 0,    # ~$30.3M
        "bienes": ws["B14"].value or 0,
        "obras": ws["B16"].value or 0,
        "devengado": ws["B19"].value or 0,
        "ti_pct": round((ws["B20"].value or 0) * 100, 1),
        "corte": str(ws["B10"].value or ""),
    }

    # cédula oficial ingerida (Zona Cruda H07 A46:E) — partidas + TOTAL municipal real
    _cat, _parts, _tc, _td = {}, [], 0.0, 0.0
    for r in range(46, 200):
        cuenta = ws.cell(r, 1).value
        cod = ws.cell(r, 4).value
        if not cuenta or not isinstance(cod, (int, float)):
            continue
        dev = ws.cell(r, 5).value if isinstance(ws.cell(r, 5).value, (int, float)) else 0
        grupo = str(ws.cell(r, 3).value or "")
        _tc += cod
        _td += dev
        _cat[grupo] = _cat.get(grupo, 0) + cod
        _parts.append({"cuenta": str(cuenta), "desc": str(ws.cell(r, 2).value or "")[:50],
                       "grupo": grupo, "cod": cod, "dev": dev})
    presupuesto["total_municipal"] = round(_tc, 2)
    presupuesto["devengado_municipal"] = round(_td, 2)
    presupuesto["por_categoria"] = sorted(
        [{"cat": k, "cod": round(v, 2)} for k, v in _cat.items()], key=lambda x: -x["cod"])
    presupuesto["partidas"] = _parts

    # ── PAC detalle — procesos vinculados a meta + coherencia (H05b A-H) ─────
    ws = sh("H05b_S3b")
    pac_detalle = []
    for r in range(15, 60):
        pid = ws.cell(r, 1).value
        if not pid:
            continue
        pac_detalle.append({
            "id": str(pid).strip(),
            "desc": str(ws.cell(r, 2).value or "").strip()[:55],
            "tipo": str(ws.cell(r, 3).value or "").strip(),
            "monto": ws.cell(r, 4).value or 0,
            "meta": str(ws.cell(r, 5).value or "").strip(),
            "alerta": _clean(str(ws.cell(r, 8).value or "")),
            "alerta_temp": _temp(str(ws.cell(r, 8).value or "")),
        })

    # ── POA detalle — monto anual por meta + cronograma mensual (H05 E-R) ────
    ws = sh("H05_S3_OPER")
    poa_detalle, poa_total = [], 0
    for r in range(14, 45):
        mid = ws.cell(r, 1).value
        anual = ws.cell(r, 5).value
        if not mid or not isinstance(anual, (int, float)):
            continue
        if str(mid).strip() not in metas:            # solo metas reales del PDOT (sin filas espurias)
            continue
        crono = [ws.cell(r, c).value for c in range(6, 18)]   # ENE..DIC %
        poa_detalle.append({
            "id": str(mid).strip(),
            "anual_usd": anual,
            "crono": [round(x, 3) if isinstance(x, (int, float)) else 0 for x in crono],
        })
        poa_total += anual

    # POA proyectos — detalle oficial ingerido (H05 sección filas 47+ · 257 proyectos)
    ws = sh("H05_S3_OPER")
    poa_proyectos = []
    for r in range(47, 320):
        d = ws.cell(r, 1).value
        monto = ws.cell(r, 6).value
        if not d or str(d).startswith("TOTAL") or not isinstance(monto, (int, float)):
            continue
        poa_proyectos.append({
            "dir": str(d).strip(),
            "meta": str(ws.cell(r, 2).value or "").strip(),
            "proyecto": str(ws.cell(r, 3).value or "").strip(),
            "desc": str(ws.cell(r, 4).value or "").strip(),
            "partida": str(ws.cell(r, 5).value or "").strip(),
            "anual": monto,
        })

    # ── E · PUBLICADO SERCOP — estado vivo verificado (H06 sección VERIFICADO) ─
    # Normalizador→silo: el conector escribió aquí el estado real publicado en SERCOP.
    # El cajón lo lee del Canon, NO de la API (Regla 1). Cruce plan-PAC ↔ publicado.
    ws = sh("H06_S4")
    pub_proc, pub_total, pub_etapas = [], 0.0, Counter()
    pub_corte, start_pub = "", None
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if v and str(v).startswith("▌ PUBLICADO SERCOP 2026"):
            start_pub = r
            m = re.search(r"corte (\d{4}-\d{2}-\d{2})", str(v))
            pub_corte = m.group(1) if m else ""
            break
    if start_pub:
        for r in range(start_pub + 2, ws.max_row + 1):
            ent, cod = ws.cell(r, 1).value, ws.cell(r, 2).value
            if not cod or not str(cod).startswith("ocds"):
                if str(cod) == "TOTAL VERIFICADO":
                    continue
                if not ent and not cod:
                    break
                continue
            monto = ws.cell(r, 5).value
            monto = monto if isinstance(monto, (int, float)) else 0
            etapa = str(ws.cell(r, 7).value or "")
            pub_total += monto
            pub_etapas[etapa] += 1
            pub_proc.append({
                "cod": str(cod),
                "desc": str(ws.cell(r, 3).value or "").strip()[:75],
                "partida": str(ws.cell(r, 4).value or "").strip(),
                "monto": monto,
                "monto_tipo": str(ws.cell(r, 6).value or "").strip(),
                "etapa": etapa,
                "fecha": str(ws.cell(r, 8).value or "")[:10],
            })
    plan_pac = total_pac or 0
    publicado = {
        "fuente": "SERCOP Contrataciones Abiertas (OCDS) · GAD · verificado en vivo",
        "corte": pub_corte,
        "n_procesos": len(pub_proc),
        "total_usd": round(pub_total, 2),
        "por_etapa": [{"etapa": k, "n": v} for k, v in pub_etapas.most_common()],
        "procesos": pub_proc,
        "cruce": {
            "plan_pac_usd": plan_pac,
            "publicado_usd": round(pub_total, 2),
            "cobertura_pct": round(pub_total / plan_pac * 100, 1) if plan_pac else 0,
        },
    }

    # ── Cobertura de metas con dotación POA (H16b!B12 · cirugía 2026-07-01) ────
    # 24/25 metas del PDOT tienen presupuesto POA del GAD; la restante (AH-I-X-03,
    # grupos prioritarios · Art. 249 COOTAD) la ejecuta el Patronato adscrito (d12).
    try:
        cobertura_metas_poa = round((sh("H16b")["B12"].value or 0) * 100, 1)
    except Exception:
        cobertura_metas_poa = None

    # ── IPE-ejecutado (H16b · % del gasto de inversión vinculado a metas · cirugía 2026-07-01) ──
    # Headline = valor de record del canon (Excel=Estado · B15). Desglose por objetivo = presentación
    # (match jerárquico partida eSIGEF → proyecto POA → objetivo · "el Excel no cruza solo").
    ipe_ejecutado = None
    ipe_por_objetivo: list[dict] = []
    try:
        _w16 = sh("H16b")
        ipe_ejecutado = {
            "pct": round((_w16["B15"].value or 0) * 100, 1),
            "vinculado": round(_w16["B16"].value or 0, 2),
            "total": round(_w16["B7"].value or 0, 2),
            "no_pdot": round(_w16["B17"].value or 0, 2),
        }
        import re as _re
        def _nrm(t): return _re.sub(r"\s+", " ", (t or "").strip().upper())
        _poa = defaultdict(set)
        for _p in poa_proyectos:
            _pt = (_p.get("partida") or "").strip()
            _ob = _nrm(_p.get("meta"))
            if _pt and _ob:
                _poa[_pt].add(_ob)
        _codes = sorted(_poa, key=len, reverse=True)
        _objdev: dict = defaultdict(float)
        for _x in presupuesto.get("partidas", []):
            _cta = str(_x.get("cuenta") or "").strip()
            if _cta[:1] not in ("7", "8"):
                continue
            _pc = next((c for c in _codes if _cta == c or _cta.startswith(c)), None)
            if _pc:
                _dev = _x.get("dev") or 0
                for _o in _poa[_pc]:
                    _objdev[_o] += _dev / len(_poa[_pc])
        ipe_por_objetivo = [{"objetivo": _o.title(), "dev": round(_v, 2)}
                            for _o, _v in sorted(_objdev.items(), key=lambda kv: -kv[1]) if _v > 0]
    except Exception:
        pass

    # ── SERIE MULTI-AÑO — trayectoria de ejecución + proyección (5º motor · H07b/H12c) ──
    # Regla 1/4: se LEE la serie que el Gold Master ya calculó (histórico REAL + proyección
    # del motor); QUIRA no recalcula, solo la narra. Firewall: etiqueta pública, sin códigos.
    # Métrica = ejecución de la inversión (nativa del cajón · misma fuente eSIGEF, consistente).
    serie_multianio = None
    try:
        _w7 = sh("H07b_Ti")
        _ejec = []
        for r in range(7, 11):                         # filas 2023..2026
            _anio = _w7.cell(r, 1).value
            _ti = _w7.cell(r, 4).value                 # Ti_Inversión (fracción 0-1)
            _tipo = str(_w7.cell(r, 5).value or "").upper()
            if isinstance(_anio, (int, float)) and isinstance(_ti, (int, float)):
                _ejec.append({
                    "anio": int(_anio),
                    "pct": round(_ti * 100, 1),
                    "cerrado": "VIVO" not in _tipo,     # 2026 = en curso (parcial, fuera de comparación)
                })
        _w12 = sh("H12c")
        def _n12(addr):
            v = _w12[addr].value
            return round(v, 1) if isinstance(v, (int, float)) else None
        serie_multianio = {
            "metrica": "ejecución de la inversión planificada",
            "ejecucion": _ejec,
            "proyeccion": {"proyeccion": _n12("B17"), "promedio": _n12("B13")},
            "fuente": "Cédulas presupuestarias eSIGEF 2023-2025 (dato real) · proyección del canon",
        }
    except Exception:
        pass

    return {
        "_fuente": "PDOT (planificación) · POA (operación) · PAC (contratación) · coherencia · corte Q1-2026",
        "metas_total": metas_total,
        "serie_multianio": serie_multianio,
        "cobertura_metas_poa": cobertura_metas_poa,
        "ipe_ejecutado": ipe_ejecutado,
        "ipe_por_objetivo": ipe_por_objetivo,
        "metas_detalle": list(metas.values()),
        "competencia": competencia,
        "direcciones": direcciones,
        "n_direcciones": n_direcciones,
        "pac": pac,
        "pac_detalle": pac_detalle,
        "presupuesto": presupuesto,
        "poa_detalle": poa_detalle,
        "poa_total": poa_total,
        "poa_proyectos": poa_proyectos,
        "sat0": sat0,
        "publicado": publicado,
    }


def main() -> None:
    block = build_block()
    with open(SNAP, encoding="utf-8") as f:
        snap = json.load(f)
    # Preservar la base normativa (la refresca scripts/normativa_planificacion.py) — que un
    # re-run del enricher NO la borre. Pipeline: enricher → normativa_planificacion.
    _prev_bn = (snap.get("planificacion") or {}).get("base_normativa")
    if _prev_bn:
        block["base_normativa"] = _prev_bn
    snap["planificacion"] = block
    with open(SNAP, "w", encoding="utf-8") as f:
        json.dump(snap, f, ensure_ascii=False, indent=2)
    pub = block["publicado"]
    print("OK - bloque 'planificacion' escrito en gm_snapshot.json")
    print(f"   metas={block['metas_total']} · direcciones={block['n_direcciones']} "
          f"· PAC=${block['pac']['total_usd']:,} ({block['pac']['n_procesos']} proc) "
          f"· SAT-0={block['sat0']['global']}")
    print(f"   PUBLICADO SERCOP: {pub['n_procesos']} proc · ${pub['total_usd']:,.2f} "
          f"· cobertura {pub['cruce']['cobertura_pct']}% del plan-PAC (corte {pub['corte']})")


if __name__ == "__main__":
    main()
