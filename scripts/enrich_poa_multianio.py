"""
QUIRA OS — Extractor POA MULTI-AÑO desde la FUENTE (Javo · 2026-07-11)
═══════════════════════════════════════════════════════════════════════════════
Re-vinculación meta↔actividad↔partida DESDE LA FUENTE (no inferida · Principio Rector:
la ausencia de evidencia es un RESULTADO, jamás una autorización a inferir). Los POA
oficiales del GAD por año (Excel) SÍ traen el vínculo — el 2025 tiene META·ACTIVIDAD·
PARTIDA explícitos. De ahí sale el mapa meta↔partida que ancla los años sin columna meta.

Fuente: Holding_Municipal_Montecristi\\POA 2023-2026\\GAD Montecristi\\*.xlsx
Salida: data/poa_multianio.json (artefacto curado · QUIRA lo lee · promoción candidata al canon).
Regla 1: la app lee el snapshot, no el Excel. Firewall: sin códigos internos en la salida pública.

Uso:  python scripts/enrich_poa_multianio.py
Dylus Lab © 2026
"""
from __future__ import annotations

import glob
import json
import os
import re
from collections import defaultdict
from difflib import SequenceMatcher

import openpyxl

POA_DIR = r"C:\Users\DELL\Desktop\Javo\Dylus Lab\ProyecT\Holding_Municipal_Montecristi\POA 2023-2026\GAD Montecristi"
OUT = os.path.join(os.path.dirname(__file__), "..", "data", "poa_multianio.json")

# esquema por año (col idx · descubierto por inspección de la fuente):
#   2025: META=9 · ACTIVIDAD=15 · PARTIDA=16 · MONTO=Σ(36..39) · header f4, data f5+
#   2024: ACTIVIDAD=1 · PARTIDA=2 · MONTO=3 (sin meta) · header f1, data f2+
#   2023: DESC=0 · PARTIDA=1 · MONTO=3 (sin meta · partida estructural) · data f6+
_ARCH = {
    2025: {"file": "GAD Monteristi POA 2025.xlsx", "data0": 5, "meta": 9, "act": 15, "part": 16, "monto": (36, 37, 38, 39)},
    2024: {"file": "GAD Montecristi POA 2024.xlsx", "data0": 2, "meta": None, "act": 1, "part": 2, "monto": (3,)},
    2023: {"file": "GAD Montecristi POA 2023.xlsx", "data0": 6, "meta": None, "act": 0, "part": 1, "monto": (3,)},
}


def _num(s) -> float:
    if isinstance(s, (int, float)):
        return float(s)
    try:
        return float(str(s).replace(".", "").replace(",", "."))
    except Exception:
        return 0.0


def _cell(row, i):
    return str(row[i]).strip() if (i is not None and i < len(row) and row[i] is not None) else ""


def _partida_econ(raw: str) -> str:
    """Normaliza la partida al CÓDIGO ECONÓMICO 6-díg (5-8XXXXX). Puro en 2025; embebido en
    el código estructural en 2023/24 (tras el año: '.2024.730813.'). Así el puente cruza años."""
    raw = str(raw or "").strip()
    if re.fullmatch(r"[5-8]\d{5}", raw):                 # 2025: ya es económico puro
        return raw
    m = re.search(r"\.20\d\d\.([5-8]\d{5})", raw)        # 2023/24: económico tras el año
    if m:
        return m.group(1)
    m = re.search(r"\b([5-8]\d{5})\b", raw)              # fallback: primer económico 6-díg
    return m.group(1) if m else ""


def _norm(s: str) -> str:
    s = (s or "").upper()
    for a, b in (("Á", "A"), ("É", "E"), ("Í", "I"), ("Ó", "O"), ("Ú", "U"), ("Ü", "U")):
        s = s.replace(a, b)                              # transliterar acentos (validación colega: subían falsos negativos)
    return re.sub(r"[^A-ZÑ ]", " ", s)


def _contratos_2025_por_meta(poa_2025: list) -> dict:
    """INNOVACIÓN (Javo · 2026-07-11): reconciliar PAC↔POA por DESCRIPCIÓN salta el muro de la partida
    compartida. El PAC y el POA describen el MISMO trabajo real (dos sistemas) → matching de descripción
    (restringido a la misma partida) atribuye el proceso a la actividad → su meta. Cadena de integridad
    INTERSISTÉMICA (no inferencia: reconciliación de dos fuentes del mismo hecho). Devuelve {meta: {...}}."""
    hits = glob.glob(os.path.join(POA_DIR, "..", "..", "PAC 2023-2026", "**", "GAD_Montecristi_PAC_2025.docx"),
                     recursive=True)
    if not hits:
        return {}
    try:
        import docx
    except Exception:
        return {}
    by_part: dict = defaultdict(list)
    for a in poa_2025:
        if a.get("meta") and a.get("partida"):
            by_part[a["partida"]].append(a)
    porm: dict = defaultdict(lambda: {"n": 0, "monto": 0.0, "nombres": [], "sim": []})
    for t in docx.Document(hits[0]).tables:
        for row in t.rows:
            c = [x.text.strip() for x in row.cells]
            if len(c) < 9 or not re.fullmatch(r"\d+", c[0].split("\n")[0]):
                continue
            pt = _partida_econ(c[1])
            if not pt or pt not in by_part:
                continue
            desc = c[4]
            monto = _num(c[8].split("\n")[0])
            best = max(by_part[pt], key=lambda a: SequenceMatcher(None, _norm(desc), _norm(a["actividad"])).ratio())
            s = SequenceMatcher(None, _norm(desc), _norm(best["actividad"])).ratio()
            if s < 0.55:                                  # reconciliación de baja confianza → no se atribuye
                continue
            d = porm[best["meta"][:110]]
            d["n"] += 1
            d["monto"] += monto
            d["sim"].append(round(s, 2))
            if len(d["nombres"]) < 4:
                d["nombres"].append(re.sub(r"\s+", " ", desc)[:52])
    return {m: {"n": v["n"], "monto": round(v["monto"], 2), "nombres": v["nombres"],
                "conf": round(sum(v["sim"]) / len(v["sim"]), 2) if v["sim"] else 0} for m, v in porm.items()}


def _leer_anio(anio: int) -> list[dict]:
    a = _ARCH[anio]
    p = os.path.join(POA_DIR, a["file"])
    if not os.path.exists(p):
        return []
    ws = openpyxl.load_workbook(p, read_only=True, data_only=True)[
        openpyxl.load_workbook(p, read_only=True).sheetnames[0]]
    acts = []
    for row in list(ws.iter_rows(values_only=True))[a["data0"]:]:
        act = _cell(row, a["act"])
        part = _cell(row, a["part"])
        if not act and not part:
            continue
        monto = sum(_num(row[i]) for i in a["monto"] if i < len(row))
        acts.append({
            "meta": _cell(row, a["meta"]) if a["meta"] is not None else "",
            "actividad": act[:120],
            "partida": _partida_econ(part),              # normalizada al código económico (cruza años)
            "monto": round(monto, 2),
        })
    return acts


def main() -> None:
    por_anio = {y: _leer_anio(y) for y in _ARCH}
    # mapa meta↔partida DESDE LA FUENTE (2025, único con META explícita): partida → metas
    p2m: dict[str, set] = defaultdict(set)
    for a in por_anio.get(2025, []):
        if a["meta"] and a["partida"]:
            p2m[a["partida"]].add(a["meta"])
    # solo las partidas DETERMINISTAS (una sola meta) sirven de ancla verificable
    ancla = {pt: next(iter(ms)) for pt, ms in p2m.items() if len(ms) == 1}

    # anclar 2023/24 a meta por partida DETERMINISTA (de la fuente 2025) — ANTES de cruzar años
    for y in (2023, 2024):
        for a in por_anio.get(y, []):
            if not a["meta"] and a["partida"] in ancla:
                a["meta_ancla"] = ancla[a["partida"]]

    # ── ejecución 2025 (cédula de cierre LOTAIP diciembre) por partida económica ──
    CED_2025 = (r"C:\Users\DELL\Desktop\Javo\Dylus Lab\ProyecT\Holding_Municipal_Montecristi"
                r"\Cedulas Presupuestarias 2023-2026\Presupuestos 2025\GAD Montecristi"
                r"\2025-Diciembre-Numeral 6-6. Conjunto de datos_nov.csv.xlsx")
    ejec: dict = {}
    try:
        wsx = openpyxl.load_workbook(CED_2025, read_only=True, data_only=True)["Sheet1"]
        for row in list(wsx.iter_rows(values_only=True))[1:]:
            cta = _partida_econ(_cell(row, 0))
            if cta:
                ejec[cta] = {"cod": _num(row[5]), "dev": _num(row[8])}
    except Exception:
        pass
    # ── biografías 2025 por meta: plan (POA fuente) + ejecución (cédula) — año cerrado ──
    bio = defaultdict(lambda: {"act": 0, "part": set(), "plan": 0.0})
    for a in por_anio.get(2025, []):
        if a["meta"]:
            b = bio[a["meta"]]
            b["act"] += 1
            b["plan"] += a["monto"]
            if a["partida"]:
                b["part"].add(a["partida"])
    # CONTRATOS por meta — reconciliación intersistémica PAC↔POA (salta el muro de la partida · innovación)
    contratos = _contratos_2025_por_meta(por_anio.get(2025, []))
    biografias = []
    for mk, b in bio.items():
        inv = sum(1 for pt in b["part"] if pt[:1] in ("7", "8"))
        con = contratos.get(mk[:110], {})
        biografias.append({"meta": mk[:110], "actividades": b["act"], "partidas": len(b["part"]),
                           "inversion": inv, "plan": round(b["plan"], 2),
                           "contratos": con.get("n", 0), "contratos_monto": con.get("monto", 0),
                           "contratos_nombres": con.get("nombres", []), "contratos_conf": con.get("conf", 0)})
    biografias.sort(key=lambda x: -x["plan"])

    # ── biografía MULTI-AÑO: la misma meta a través de los años (CONTINUIDAD del compromiso) ──
    # 2025 = nativo (vínculo de la fuente); 2023/24 = anclado por partida determinista (piso verificable).
    # El valor es la PERSISTENCIA, no el monto (el histórico solo captura el subconjunto anclado).
    ma: dict = defaultdict(lambda: defaultdict(lambda: {"act": 0, "plan": 0.0, "nativo": False}))
    for y in (2023, 2024, 2025):
        for a in por_anio.get(y, []):
            meta = a["meta"] or a.get("meta_ancla")
            if meta:
                m = ma[meta[:110]][y]
                m["act"] += 1
                m["plan"] += a["monto"]
                if a["meta"]:
                    m["nativo"] = True
    bio_ma = []
    for meta, years in ma.items():
        if len(years) >= 2 and not any(k in meta.lower() for k in ("roles", "institucional")):
            bio_ma.append({"meta": meta,
                           "anios": {str(y): {"act": v["act"], "plan": round(v["plan"], 2), "nativo": v["nativo"]}
                                     for y, v in sorted(years.items())}})
    bio_ma.sort(key=lambda x: (-len(x["anios"]), -sum(a["plan"] for a in x["anios"].values())))

    salida = {
        "biografia_multianio": bio_ma,
        "_fuente": "POA oficial GAD Montecristi por año (Excel) — vínculo meta↔actividad↔partida de la fuente",
        "biografias_2025": biografias,
        "_nota_metodologica": ("La 'meta' del POA es operativa (indicador). El mapa meta↔partida se toma del "
                               "2025 (único año con META explícita en la fuente); solo partidas DETERMINISTAS "
                               "(una sola meta) anclan otros años — la ausencia de vínculo NO se infiere."),
        "anios": {},
        "mapa_meta_partida_deterministas": len(ancla),
    }
    for y, acts in por_anio.items():
        con_meta = sum(1 for a in acts if a["meta"])
        # ancla los años sin meta por partida DETERMINISTA (de la fuente 2025)
        anclados = 0
        if con_meta == 0:
            for a in acts:
                if a["partida"] in ancla:
                    a["meta_ancla"] = ancla[a["partida"]]
                    anclados += 1
        salida["anios"][str(y)] = {
            "n_actividades": len(acts),
            "monto_total": round(sum(a["monto"] for a in acts), 2),
            "con_meta_fuente": con_meta,
            "anclados_por_partida": anclados,
            "actividades": acts,
        }
    with open(OUT, "w", encoding="utf-8") as f:
        json.dump(salida, f, ensure_ascii=False, indent=2)
    print("OK — data/poa_multianio.json")
    print(f"   mapa meta↔partida deterministas (de la fuente 2025): {len(ancla)} partidas")
    for y in sorted(salida["anios"]):
        d = salida["anios"][y]
        print(f"   {y}: {d['n_actividades']:>3} act · ${d['monto_total']/1e6:5.1f}M · "
              f"meta-fuente={d['con_meta_fuente']} · anclados={d['anclados_por_partida']}")


if __name__ == "__main__":
    main()
