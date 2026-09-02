"""
QUIRA OS — Enriquecedor del snapshot · bloque `presupuesto_dom` (DOM d02)
═══════════════════════════════════════════════════════════════════════════════
Puente Excel→snapshot (Regla 1). Lee el Gold Master LOCAL (solo lectura · openpyxl
data_only · NO corrompe) y escribe el bloque `presupuesto_dom` en data/gm_snapshot.json.

VISIÓN d02 (corrección de Javo · 2026-07-14): la salud financiera del municipio COMO
BASE para captar financiamiento internacional (reembolsable y no reembolsable).
  ① La base financiera:  ejecución presupuestaria (H07 eSIGEF) · salud presup. (ISP · H19)
  ② La capacidad de captación:  eficiencia/fondos externos (IEF · H20c) + llaves de
     elegibilidad (alineación PND H11b · Agenda 2030) — consumidas, no propias.
  ③ Reporte de cooperación (H32) — a futuro.
"Qué fondo específico aplica" = QUIRA Cooperación (producto · ADR-024), NO este cajón.

Uso:  python scripts/enrich_presupuesto.py
Dylus Lab © 2026
"""
from __future__ import annotations

import json
import os
import re

import openpyxl

try:
    from config import DATOS_DIR as _DATOS
except Exception:                                        # noqa: BLE001
    import os as _os
    from pathlib import Path as _P
    _DATOS = _P(_os.environ.get("QUIRA_DATOS", "."))
def _gold_master_vigente() -> str:
    """El Gold Master VIGENTE, resuelto por `config` (sufijo `_TGI`,
    versión más alta). Antes aquí había un literal `v5.5` y por eso este
    módulo leía una versión que el canon ya había superado — D-002."""
    try:
        from config import SIAP_PATH
        return str(SIAP_PATH)
    except Exception:                                    # noqa: BLE001
        return str(_DATOS / "SIAP-ICPI_GOLD_MASTER_v5.5_TGI.xlsx")


EXCEL = _gold_master_vigente()
SNAP = os.path.join(os.path.dirname(__file__), "..", "data", "gm_snapshot.json")

_HCODE = re.compile(r"\bH\d{1,2}[a-z]?\b")


def _fw(s) -> bool:
    """Seguro para público: sin nomenclatura canónica (H01-H99) ni fila-nota (Firewall)."""
    s = str(s or "")
    return not _HCODE.search(s) and not s.strip().upper().startswith("NOTA")


def _num(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return None



def _umbral_de_la_regla() -> dict:
    """El umbral COOTAD, **leído del puente BRN verificable** — no copiado.

    D-005 era: `65` escrito aquí mientras `RO-IV-001` declara 65 hasta
    2026-12-31 y **70 desde 2027-01-01**. Hoy coincidían; el 1 de enero de 2027
    dejarían de coincidir y nada avisaría.

    Detectar la copia caduca era el remedio; **hacerla imposible** es la cura:
    d02 deja de tener el valor y lo pide al lector, que sólo lo entrega si la
    pieza está vigente, el catálogo al día y el sello acredita (las tres).

    ⚠️ SI NO ES CONSUMIBLE, NO SE INVENTA UN NÚMERO. Se devuelve `None` con el
    motivo, y quien lea el bloque sabrá que el umbral no está acreditado en vez
    de recibir un valor que nadie respalda. Volver aquí al literal como respaldo
    reintroduciría exactamente la deuda que este cambio cierra.

    Y el lector NO es un segundo motor: dice qué regla hay, con qué estado y qué
    procedencia. La lógica de dominio sigue siendo de d02 (ADR-047)."""
    try:
        import sys as _s
        from pathlib import Path as _P
        _r = str(_P(__file__).resolve().parents[1])
        if _r not in _s.path:
            _s.path.insert(0, _r)
        from app.agents import brn_lector as _L
        regla = _L.regla("RO-IV-001")
    except Exception as e:                               # noqa: BLE001
        return {"valor": None, "estado": "no_determinable",
                "por_que": f"no se alcanzó el puente BRN: {type(e).__name__}"}

    if regla is None:
        return {"valor": None, "estado": "no_consta",
                "por_que": "RO-IV-001 no está en el catálogo compilado"}
    if not regla.es_consumible_como_vigente:
        return {"valor": None, "estado": "no_consumible",
                "por_que": f"pieza «{regla.estado_pieza}» · catálogo al día: "
                           f"{regla.catalogo_al_dia} · sello acredita: "
                           f"{regla.sello.acredita if regla.sello else False}"}
    return {"valor": regla.umbral_vigente,
            "estado": "consumible",
            "regla": regla.id,
            "variable": regla.variable,
            "vigencia_operativa": regla.vigencia_operativa,
            "deriva_de": regla.deriva_de,
            "procedencia": "puente BRN verificable · sello "
                           f"{regla.sello.validado_por} {regla.sello.fecha}"}


def build_block() -> dict:
    wb = openpyxl.load_workbook(EXCEL, read_only=True, data_only=True)

    def sh(prefix: str):
        return wb[next(s for s in wb.sheetnames if s.startswith(prefix))]

    def find(ws, etiqueta: str, col_val: int = 2, maxr: int = 60):
        """Busca una fila cuyo col 1 contenga `etiqueta` y devuelve el valor de col_val."""
        for r in ws.iter_rows(min_row=1, max_row=maxr, values_only=True):
            if r and r[0] and etiqueta.lower() in str(r[0]).lower():
                return r[col_val - 1] if len(r) >= col_val else None
        return None

    # ── ① SALUD PRESUPUESTARIA (ISP · H19) ────────────────────────────────────
    ws = sh("H19_ICS")
    isp_ref = str(find(ws, "ISP_Global", col_val=3) or "")     # "58.40% — Transición Crítica" (col 2 = Ti, no ISP)
    _m = re.search(r"([\d.]+)\s*%", isp_ref)
    isp_pct = round(float(_m.group(1)), 1) if _m else 0
    isp_clasif = str(find(ws, "Clasificación_ISP") or isp_ref).strip()
    isp = {
        "global_pct": isp_pct,
        "clasificacion": re.sub(r"[🔴🟡🟢🟠⚠️✅]", "", isp_clasif).strip(" —·"),
        # Piso TRANSITORIO 2026 (65%) de la regla de asignación mínima prioritaria — COOTAD-2026
        # Art. 198.1 fija la regla plena en 70%; la Disposición Transitoria Primera fija el piso del
        # 65% con seguimiento desde el 1-dic-2026. NO es el Art. 192 (ese es "monto a transferir 21%").
        "umbral_cootad": _umbral_de_la_regla(),
    }

    # ── ① EJECUCIÓN PRESUPUESTARIA (H07 eSIGEF) ───────────────────────────────
    ws = sh("H07_S5")
    cod = _num(find(ws, "Codificado_Total_Inversión")) or 0
    dev = _num(find(ws, "Devengado_Total_Inversión")) or 0
    ti = _num(find(ws, "Ti_Global_2026")) or 0
    corte = str(find(ws, "Fecha_Corte") or "").strip()
    ejecucion = {
        "codificado": round(cod), "devengado": round(dev),
        "ti_pct": round(ti * 100, 1) if ti <= 1 else round(ti, 1),
        "corte": corte,
    }
    # serie multi-año del Ti (2023-2026)
    serie = []
    for r in ws.iter_rows(min_row=24, max_row=32, values_only=True):
        if r and str(r[0]).strip().isdigit() and _num(r[1]) is not None:
            v = _num(r[1])
            serie.append({"anio": int(r[0]), "ti_pct": round(v * 100, 1) if v <= 1 else round(v, 1)})

    # ── ② CAPTACIÓN — FONDOS EXTERNOS (IEF · H20c) ────────────────────────────
    ws = sh("H20c")
    fondos, total_ext = [], 0.0
    for r in ws.iter_rows(min_row=14, max_row=60, values_only=True):
        if not r or not r[0]:
            continue
        mid = str(r[0]).strip()
        nombre = str(r[1] or "").strip()
        tipo = str(r[2] or "").strip()
        monto = _num(r[3]) or 0
        if not _fw(mid) or not _fw(nombre) or monto <= 0:
            continue
        # reembolsable vs no reembolsable (convenios/bonos = no reembolsable; crédito = reembolsable)
        reemb = "reembolsable" if any(k in (nombre + tipo).lower() for k in ("crédit", "credit", "préstamo", "prestamo", "banco")) else "no reembolsable"
        fondos.append({"meta": mid, "nombre": nombre[:70], "tipo": tipo.replace("_", " ").title(),
                       "monto": round(monto), "modalidad": reemb})
        total_ext += monto

    ief_umbral = {"alto": _num(find(ws, "Umbral_Alto")), "bueno": _num(find(ws, "Umbral_Bueno"))}
    captacion = {
        "total_externo": round(total_ext), "n_convenios": len(fondos),
        "detalle": fondos, "umbrales": ief_umbral,
    }

    # ── ② ELEGIBILIDAD — vinculación ODS (H11 · Agenda 2030) + PND consumido ───
    ws = sh("H11_S9")
    icods = _num(find(ws, "ICODS_Preliminar"))
    ods_cub = _num(find(ws, "ODS_Cubiertos_PDOT"))
    ods_por_meta = {}
    for r in ws.iter_rows(min_row=14, max_row=45, values_only=True):
        if r and r[0] and _fw(str(r[0])) and "ODS" in str(r[2] or ""):
            ods_por_meta[str(r[0]).strip()] = {"principal": str(r[2] or "").strip(),
                                               "meta_ods": str(r[4] or "").strip()}
    ods = {
        "icods_pct": round(icods * 100, 1) if icods and icods <= 1 else None,
        "ods_cubiertos": int(ods_cub) if ods_cub else None, "total_ods": 17,
        "n_metas_vinculadas": len(ods_por_meta),
    }
    # Alineación PND (H11b) + eje PND por meta — objeto compartido que NACE en d01 (se consume).
    try:
        _snap = json.load(open(SNAP, encoding="utf-8"))
        _al = (_snap.get("planificacion", {}) or {}).get("alineacion_pnd", {}) or {}
        alineacion_pnd = round((_al.get("vinculacion_media") or 0) * 100) or None
        pnd_metas = {m.get("id"): m.get("eje", "") for m in (_al.get("metas") or [])}
    except Exception:
        alineacion_pnd, pnd_metas = None, {}
    # Trazabilidad del financiamiento: cada convenio con su ODS + eje PND (la cadena que se financia).
    for f in fondos:
        o = ods_por_meta.get(f["meta"], {})
        f["ods"] = o.get("principal", "")
        f["pnd_eje"] = pnd_metas.get(f["meta"], "")

    # ── ③ SEÑALES PREVENTIVAS · SAT presupuestario (H22/H23/H24 — LEÍDOS del motor, no recalculados) ──
    # Decisión Javo · 2026-07-15: se elimina el DOM Alertas → cada SAT vive en su dominio (12 DOM).
    # d02 toma las 3 señales FINANCIERAS. Descriptivo → PREDICTIVO (Datos→Umbral→Alerta→Diagnóstico).
    def _estado(ws) -> str:
        raw = str(find(ws, "_Estado") or "")
        activa = ("sin señal" not in raw.lower()) and any(x in raw for x in ("⚠", "🔴", "🟠"))
        return "activa" if activa else "sin_senal"

    ws2, ws3, ws4 = sh("H22"), sh("H23"), sh("H24_SAT")
    e2, e3, e4 = _estado(ws2), _estado(ws3), _estado(ws4)
    u2 = _num(find(ws2, "Umbral_Reforma")) or 0.05
    v2 = _num(find(ws2, "Pct_Reforma")) or 0
    u3 = _num(find(ws3, "Umbral_Deveng")) or 0.10
    np3 = _num(find(ws3, "Metas_Con_Paralisis")) or 0
    u4 = _num(find(ws4, "Pct_Inversion_Minimo")) or 0.65
    # Cada señal expone su CADENA normativa (aporte del colega · 2026-07-15): la norma engendra una REGLA
    # (condición lógica), la regla se mide con un indicador, y el indicador dispara —o no— la señal.
    # QUIRA no inventa la regla: la deriva del umbral que el motor ya calcula sobre la norma verificada.
    # AUDITORÍA Javo 2026-07-18: las citas de artículo estaban ERRADAS (COOTAD 192 = "monto a
    # transferir 21%", no la regla del 65%; COPLAFIP 113/115 tampoco corresponden). Se QUITA la
    # cita de artículo (norma="") — la cadena verificada la proveerá la BRN (ADR-038). Se conserva
    # la REGLA operativa y el umbral, que el motor sí calcula. Regla 3: sin norma verificada, sin cita.
    senales = [
        {"nombre": "Reforma presupuestaria tardía", "estado": e2, "norma": "",
         "regla": f"las reformas no deben superar el {u2 * 100:.0f}% del presupuesto anual",
         "indicador": f"{v2 * 100:.0f}% del presupuesto reformado", "umbral": f"máx {u2 * 100:.0f}%",
         "vigila": "reformas significativas fuera del cronograma anual — señal de programación inestable"},
        {"nombre": "Parálisis presupuestaria", "estado": e3, "norma": "",
         "regla": f"ninguna meta debe ejecutar menos del {u3 * 100:.0f}% de lo asignado",
         "indicador": f"{int(np3)} meta(s) en parálisis", "umbral": f"ejecución mínima {u3 * 100:.0f}%",
         "vigila": "metas cuya ejecución financiera se estanca bajo el mínimo — riesgo de subejecución"},
        {"nombre": "Alerta fiscal · estructura COOTAD", "estado": e4, "norma": "",
         "regla": f"la inversión no debe caer bajo el {u4 * 100:.0f}% del presupuesto",
         "indicador": "estructura inversión/corriente conforme" if e4 == "sin_senal" else "estructura fiscal en alerta",
         "umbral": f"inversión ≥ {u4 * 100:.0f}% del presupuesto",
         "vigila": "que la inversión no caiga bajo el mínimo legal frente al gasto corriente"},
    ]
    senales = [s for s in senales if _fw(s["nombre"]) and _fw(s["indicador"])]
    sat = {"senales": senales, "n_senales": len(senales),
           "n_activas": sum(1 for s in senales if s["estado"] == "activa")}

    return {
        "_fuente": "Presupuesto (cédula eSIGEF) · Salud presupuestaria (ISP) · Eficiencia financiera / fondos "
                   "externos (IEF) · alineaciones consumidas · corte Abril 2026",
        "vision": "La salud financiera del municipio como base para captar financiamiento internacional "
                  "(reembolsable y no reembolsable).",
        "isp": isp,
        "ejecucion": ejecucion,
        "serie": serie,
        "captacion": captacion,
        "ods": ods,
        "elegibilidad": {
            "alineacion_pnd_pct": alineacion_pnd,     # consumido de H11b (d01)
            "nota": "Qué fondo específico aplica lo resuelve QUIRA Cooperación (producto), no este dominio.",
        },
        "sat_presupuestario": sat,   # señales preventivas leídas del motor (H22/H23/H24)
        "publicado": True,
    }


def main() -> None:
    block = build_block()
    snap = {}
    if os.path.exists(SNAP):
        with open(SNAP, encoding="utf-8") as f:
            snap = json.load(f)
    snap["presupuesto_dom"] = block
    with open(SNAP, "w", encoding="utf-8") as f:
        json.dump(snap, f, ensure_ascii=False, indent=2)
    print("OK - bloque 'presupuesto_dom' escrito en gm_snapshot.json")
    print(f"   ISP={block['isp']['global_pct']}% · ejecución Ti={block['ejecucion']['ti_pct']}% "
          f"· fondos externos ${block['captacion']['total_externo']:,} ({block['captacion']['n_convenios']} convenios)")


if __name__ == "__main__":
    main()
