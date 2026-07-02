# -*- coding: utf-8 -*-
"""
scripts/enrich_rdc.py — Cable Excel del MCD Rendición de Cuentas (QINV-009)
═══════════════════════════════════════════════════════════════════════════════
Réplica del patrón de Planificación (ADR-031). LEE del Gold Master (canon) y escribe
el bloque `rendicion` en data/gm_snapshot.json. El cajón lo lee de ahí (Regla 1).

Tres pilares (data verificada real):
  · FIDELIDAD NARRATIVA (H34b) — discurso oficial ↔ evidencia (el diferenciador · IF_n)
  · PRESUPUESTO PARTICIPATIVO (H10b) — serie 2024-2026 + componentes (COOTAD 238)
  · REPORTE CPCCS (H31) — compromisos + brecha (LOPC Art.88)

Firewall: la UI usa lenguaje de gobernanza. No se exponen ICPI/Ti/Vi/H-series.
Dylus Lab © 2026
"""
from __future__ import annotations

import json
import os

import openpyxl

EXCEL = r"C:\Users\DELL\Desktop\Javo\Dylus Lab\ProyecT\SIAP-ICPI_GOLD_MASTER_v5.5_TGI.xlsx"
SNAP = os.path.join(os.path.dirname(__file__), "..", "data", "gm_snapshot.json")


def _clean(s) -> str:
    s = str(s or "")
    for mk in ("🔴", "⚠️", "✅", "🟢", "🔵", "🟠", "⬜", "🔄", "★", "▌"):
        s = s.replace(mk, "")
    return s.strip(" ·—-\"")


def _temp_if(v: float) -> str:
    if v >= 0.85:
        return "verde"
    if v >= 0.60:
        return "alerta"
    return "critico"


def build_block() -> dict:
    wb = openpyxl.load_workbook(EXCEL, read_only=True, data_only=True)

    def sh(prefix: str):
        return wb[next(s for s in wb.sheetnames if s.startswith(prefix))]

    # ── 1 · FIDELIDAD NARRATIVA (H34b) — discurso ↔ evidencia · el diferenciador ──
    ws = sh("H34b")
    claims = []
    for r in range(11, 20):                       # MFN-001 … MFN-009
        cid = ws.cell(r, 1).value
        try:                                      # L = IF_n (guardado como texto '1.00')
            ifn = float(str(ws.cell(r, 12).value).strip())
        except (TypeError, ValueError):
            continue
        if not cid:
            continue
        claims.append({
            "id": str(cid).strip(),
            "entidad": _clean(ws.cell(r, 3).value),
            "meta": _clean(ws.cell(r, 5).value),
            "categoria": _clean(ws.cell(r, 6).value),
            "timestamp": _clean(ws.cell(r, 7).value),
            "discurso": _clean(ws.cell(r, 8).value),
            "evidencia": _clean(ws.cell(r, 9).value),
            "if_n": round(ifn, 2),
            "clasificacion": _clean(ws.cell(r, 13).value),
            "temp": _temp_if(ifn),
        })
    try:                                          # B21 = AVERAGE(J11:J19)
        fid_global = float(str(ws.cell(21, 2).value).strip())
    except (TypeError, ValueError):
        fid_global = None
    fidelidad = {
        "global_pct": round(float(fid_global) * 100, 1) if isinstance(fid_global, (int, float)) else None,
        "n_afirmaciones": len(claims),
        "n_alta": sum(1 for c in claims if c["temp"] == "verde"),
        "n_baja": sum(1 for c in claims if c["temp"] == "critico"),
        "claims": claims,
        "fuente": "Matriz de Fidelidad Narrativa — Rendición de Cuentas 2024 (video oficial ↔ evidencia verificada)",
    }

    # ── 2 · PRESUPUESTO PARTICIPATIVO (H10b) — serie 2024-2026 verificada ──────────
    ws = sh("H10b")
    serie = []
    for r in range(22, 25):                       # PP 2024 / 2025 / 2026
        anio = ws.cell(r, 1).value
        if not anio:
            continue
        serie.append({
            "anio": _clean(anio),
            "metodo": _clean(ws.cell(r, 2).value),
            "talleres": _clean(ws.cell(r, 3).value),
            "fichas": _clean(ws.cell(r, 4).value),
            "parroquias": _clean(ws.cell(r, 5).value),
            "presupuesto": ws.cell(r, 6).value if isinstance(ws.cell(r, 6).value, (int, float)) else _clean(ws.cell(r, 6).value),
            "acta": _clean(ws.cell(r, 7).value)[:70],
        })
    comp_2025 = []
    for r in range(31, 36):                        # componentes PP 2025 (verificado)
        c = ws.cell(r, 1).value
        m = ws.cell(r, 2).value
        if c and isinstance(m, (int, float)):
            comp_2025.append({"componente": _clean(c), "monto": m,
                              "pct": _clean(ws.cell(r, 3).value), "area": _clean(ws.cell(r, 4).value)})
    prioridades = []
    for r in range(13, 18):                        # prioridades PP 2026 (fichas verificadas)
        pid = ws.cell(r, 1).value
        if not pid:
            continue
        prioridades.append({"id": _clean(pid), "desc": _clean(ws.cell(r, 2).value),
                            "id_meta": _clean(ws.cell(r, 3).value), "fichas": _clean(ws.cell(r, 7).value)})
    participativo = {
        "marco_legal": _clean(ws.cell(8, 2).value) or "COOTAD Art. 238 / LOPC Art. 93",
        "fichas_2026": ws.cell(10, 2).value,
        "serie": serie,
        "componentes_2025": comp_2025,
        "prioridades": prioridades,
    }

    # ── 3 · REPORTE CPCCS (H31) — marco + compromisos (lenguaje gobernanza) ─────────
    ws = sh("H31")
    cpccs = {
        "marco_legal": _clean(ws.cell(11, 2).value) or "LOPC Art. 88 · Constitución Art. 204",
        "brecha_compromisos": _clean(ws.cell(65, 2).value),
        "fecha_rdc": _clean(ws.cell(61, 2).value),
    }

    wb.close()
    return {
        "_fuente": "RDC: Fidelidad Narrativa (H34b) · Participativo (H10b) · CPCCS (H31) · corte 2024-2026",
        "fidelidad": fidelidad,
        "participativo": participativo,
        "cpccs": cpccs,
    }


def main() -> None:
    block = build_block()
    with open(SNAP, encoding="utf-8") as f:
        snap = json.load(f)
    snap["rendicion"] = block
    with open(SNAP, "w", encoding="utf-8") as f:
        json.dump(snap, f, ensure_ascii=False, indent=2)
    fi = block["fidelidad"]
    print("OK - bloque 'rendicion' escrito en gm_snapshot.json")
    print(f"   Fidelidad narrativa: {fi['n_afirmaciones']} afirmaciones · global {fi['global_pct']}% "
          f"· {fi['n_alta']} alta · {fi['n_baja']} baja")
    print(f"   Participativo: {len(block['participativo']['serie'])} años · "
          f"{len(block['participativo']['componentes_2025'])} componentes")
    print(f"   CPCCS: {block['cpccs']['marco_legal']}")


if __name__ == "__main__":
    main()
