# -*- coding: utf-8 -*-
"""
scripts/gm_omega/cobertura_icpi.py — GM-Ω-ICPI-008 · COBERTURA

    ¿Qué universo mide realmente el ICPI, y el 27,4582 % representa el universo
    declarado o sólo una muestra?

    ⚠️ LA REGLA EN ROJO DE 008: **no se corrige el 25/66.** Primero hay que
    saber cuál de los tres es:

        25/66 = error
        25/66 = muestra estratégica metodológicamente prevista
        25/66 = muestra prevista pero ejecutada incorrectamente

    Son tres diagnósticos distintos y sólo uno es cierto.

    Y LO PRIMERO QUE ENCONTRÓ ESTE FRENTE FUE QUE YA ESTABA RESPONDIDO:
    `ADR-036` ratificó el 2026-07-15 que las 25 son el **universo operacional**
    del modelo v1, que las 25 existen todas en el PDOT y que ampliar a 66 sería
    **evolución, no corrección**. Buscar antes de declarar — la lección que esta
    auditoría ya tuvo que aprender dos veces.

    Lo que 008 aporta sobre el ADR son los tres huecos que éste no cierra:

        1. el CRITERIO de selección no está declarado en ninguna parte
        2. nadie ha medido si la muestra está SESGADA
        3. el ADR §1 exige declarar el alcance en toda publicación de d01/d03
           — y eso no se ha verificado nunca

    LECTURA PURA. No toca el Gold Master ni la cifra madre.

Uso:  python scripts/gm_omega/cobertura_icpi.py
Dylus Lab © 2026
"""
from __future__ import annotations

import json
import re
import sys
from collections import Counter
from pathlib import Path

_RAIZ = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(_RAIZ))

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

_SALIDA = _RAIZ / "docs" / "architecture" / "GM-OMEGA_ICPI_COBERTURA_008.md"
_PDOT = _RAIZ / "data" / "pdot" / "metas_plurianual_extraccion.json"
_FUERA = _RAIZ / "data" / "pdot" / "metas_fuera_del_motor.json"

_SUPERFICIES = ("quira_pages", "components", "views")

# Los cuatro veredictos posibles. No se fuerza ninguno.
_VEREDICTOS = ("COBERTURA ADECUADA",
               "COBERTURA LIMITADA PERO METODOLÓGICAMENTE JUSTIFICADA",
               "COBERTURA INSUFICIENTE",
               "NO DETERMINABLE")


def _norm(t: str) -> str:
    """Normaliza el texto de una meta para poder cruzarla entre catálogos."""
    return re.sub(r"\s+", " ", str(t or "")).strip().lower()[:60]


def leer_universo() -> dict:
    """El PDOT completo y el catálogo de exclusiones, con su procedencia."""
    if not (_PDOT.exists() and _FUERA.exists()):
        return {}
    pdot = json.loads(_PDOT.read_text(encoding="utf-8"))
    fuera = json.loads(_FUERA.read_text(encoding="utf-8"))
    return {"pdot": pdot, "fuera": fuera}


def leer_motor() -> dict:
    """Las metas del motor y el universo que el Gold Master DECLARA tener."""
    import openpyxl

    import config
    if not getattr(config, "GOLD_MASTER_RESUELTO", False):
        return {}
    wv = openpyxl.load_workbook(config.SIAP_PATH, data_only=True, read_only=True)
    h12, h04 = wv["H12_MOTOR_ICPI_CANÓNICO"], wv["H04_S2_PLANIFICACIÓN_PDOT"]

    metas = []
    for r in range(6, 31):
        mid = h12.cell(row=r, column=1).value
        if mid:
            metas.append(str(mid))

    # H04!B7 · el parámetro que el motor usa como universo. Su NOMBRE importa.
    declarado, etiqueta = None, None
    for r in range(1, 20):
        k = h04.cell(row=r, column=1).value
        if isinstance(k, str) and "Total_Metas" in k:
            etiqueta, declarado = k.strip(), h04.cell(row=r, column=2).value

    # Descripciones de las 25, para poder cruzarlas contra el PDOT.
    desc = {}
    for r in range(15, 40):
        mid = h04.cell(row=r, column=1).value
        if mid:
            desc[str(mid)] = {"sistema": h04.cell(row=r, column=2).value,
                              "descripcion": h04.cell(row=r, column=3).value,
                              "competencia": h04.cell(row=r, column=4).value}
    wv.close()
    return {"metas": metas, "declarado": declarado, "etiqueta": etiqueta,
            "desc": desc}


def declaracion_de_alcance() -> dict:
    """`ADR-036` §1: «Toda publicación de d01/d03 debe declararlo: *se mide
    contra las 25 metas estratégicas del modelo*».

    ⚠️ Nunca se verificó si eso se cumple. Un ADR ratificado cuya consecuencia
    práctica nadie comprueba es una decisión que existe en el papel."""
    patron = re.compile(r"(25 metas estratégicas|universo operacional|"
                        r"25 metas del PDOT|metas estratégicas del modelo)", re.I)
    en_codigo, en_texto_visible = [], []
    for carpeta in _SUPERFICIES:
        base = _RAIZ / carpeta
        if not base.exists():
            continue
        for py in sorted(base.rglob("*.py")):
            if "_deprecated" in py.parts:
                continue
            txt = py.read_text(encoding="utf-8", errors="replace")
            for m in patron.finditer(txt):
                linea = txt[:m.start()].count("\n") + 1
                # ¿está en un comentario, o en una cadena que la UI pinta?
                inicio = txt.rfind("\n", 0, m.start()) + 1
                es_comentario = txt[inicio:m.start()].lstrip().startswith("#")
                ref = f"{py.relative_to(_RAIZ).as_posix()}:{linea}"
                (en_codigo if es_comentario else en_texto_visible).append(ref)
    return {"comentarios": en_codigo, "visible": en_texto_visible}


def criterio_declarado() -> list[str]:
    """¿Existe en alguna parte el CRITERIO por el que esas 25 fueron elegidas?

    El `ADR-036` verifica que las 25 **existen** en el PDOT —ninguna inventada—
    pero eso responde «son legítimas», no «por qué éstas». Es la misma forma del
    problema de `E_i`: valores sin regla generadora.

    ⚠️ El patrón exige que la mención esté REFERIDA A LAS METAS. La primera
    versión buscaba «criterio de selección» a secas y devolvió `GATE-007`, que
    declara cómo elegir un segundo MUNICIPIO para validación externa —nada que
    ver—. Un detector que no acota el sujeto encuentra cualquier cosa."""
    patron = re.compile(r"(criterio[^.\n]{0,40}selecci[^.\n]{0,60}meta|"
                        r"por qué (?:esas|estas) 25|selección de las 25 metas|"
                        r"cómo se eligieron las 25|25 metas[^.\n]{0,40}criterio)",
                        re.I)
    hallados = []
    for sub in ("docs", "governance", "identity"):
        base = _RAIZ / sub
        if not base.exists():
            continue
        for md in base.rglob("*.md"):
            # ⚠️ Los documentos de GM-Ω se excluyen porque HABLAN del criterio
            # sin declararlo — y sin esto el detector se encuentra a sí mismo:
            # la primera pasada devolvió el propio informe de 008 como si fuera
            # una fuente que responde la pregunta que el informe hace.
            if "historico" in md.parts or md.name.startswith("GM-OMEGA_"):
                continue
            if patron.search(md.read_text(encoding="utf-8", errors="replace")):
                hallados.append(md.relative_to(_RAIZ).as_posix())
    return hallados


def main() -> int:
    uni, motor = leer_universo(), leer_motor()
    if not uni:
        print("[no determinable] faltan los catálogos del PDOT.")
        return 2
    if not motor:
        print("[no determinable] Gold Master no resuelto.")
        return 2

    pdot, fuera = uni["pdot"], uni["fuera"]
    total_pdot = pdot.get("total_metas")
    metas_pdot = pdot.get("metas", [])
    n_motor = len(motor["metas"])

    # ── LA ARITMÉTICA, que es lo primero que hay que cuadrar ────────────────
    textos_fuera = {_norm(m.get("meta")) for m in fuera}
    textos_pdot = [_norm(m.get("meta")) for m in metas_pdot]
    dentro_inferido = [t for t in textos_pdot if t not in textos_fuera]
    huecos = {
        "pdot_total": total_pdot,
        "pdot_filas": len(metas_pdot),
        "motor": n_motor,
        "fuera_catalogo": len(fuera),
        "dentro_inferido": len(dentro_inferido),
        "resta_simple": (total_pdot - n_motor) if total_pdot else None,
    }
    cuadra = huecos["resta_simple"] == huecos["fuera_catalogo"]

    # ── SESGO: ¿cómo se reparten los sistemas del PDOT? ─────────────────────
    sis_pdot = Counter(m.get("sistema", "?") for m in metas_pdot)
    sis_fuera = Counter(m.get("sistema", "?") for m in fuera)
    sis_dentro = {k: sis_pdot[k] - sis_fuera.get(k, 0) for k in sis_pdot}

    alcance = declaracion_de_alcance()
    criterio = criterio_declarado()

    print(f"PDOT {total_pdot} · motor {n_motor} · fuera {len(fuera)} · "
          f"cuadra: {'sí' if cuadra else 'NO'}")
    print(f"declaración de alcance: {len(alcance['visible'])} visibles · "
          f"{len(alcance['comentarios'])} en comentarios")
    print(f"criterio de selección documentado en: {len(criterio)} archivos")

    _escribir(motor, pdot, fuera, huecos, cuadra, sis_pdot, sis_fuera,
              sis_dentro, alcance, criterio)
    print(f"→ {_SALIDA.relative_to(_RAIZ).as_posix()}")
    return 0


def _escribir(motor, pdot, fuera, h, cuadra, sis_pdot, sis_fuera, sis_dentro,
              alcance, criterio) -> None:
    o: list[str] = []
    A = o.append
    proc = pdot.get("_procedencia", {})

    A("# GM-Ω · ICPI — COBERTURA  `008`")
    A("")
    A("**DERIVADO — no editar a mano.** Lo regenera "
      "`scripts/gm_omega/cobertura_icpi.py`.")
    A("")
    A("> ### La regla en rojo de 008")
    A("> **No se corrige el 25/66.** Primero hay que saber cuál de los tres es: "
      "un error, una muestra estratégica prevista, o una muestra prevista pero "
      "mal ejecutada. Y la cifra madre **27,4582 % sigue congelada** durante "
      "todo el diagnóstico.")
    A("")

    A("## ⚠️ Lo primero que encontró 008 fue que ya estaba respondido")
    A("")
    A("`ADR-036 · Universo Operacional del Modelo`, **RATIFICADO el 2026-07-15**:")
    A("")
    A("> «El Gold Master utiliza el **subconjunto estratégico de 25 metas del "
      "PDOT** como **universo operacional** para el cálculo del ICPI. No mide "
      "el PDOT completo (66 metas).» […] «las 25 existen todas en el PDOT — "
      "ninguna inventada. Es un subconjunto legítimo, no un error de carga.» "
      "[…] «Ampliar el rango de 25 a 66 **no sería una corrección**.»")
    A("")
    A("**El veredicto de las tres hipótesis, entonces:**")
    A("")
    A("| Hipótesis | |")
    A("|---|---|")
    A("| `25/66 = error` | ❌ descartada — las 25 existen todas en el PDOT |")
    A("| `25/66 = muestra estratégica prevista` | ✅ **ratificada por ADR** |")
    A("| `25/66 = muestra mal ejecutada` | ⚠️ no descartable: falta el criterio |")
    A("")
    A("Y una constancia metodológica: **este frente empezó a investigar algo "
      "que el canon ya había decidido.** Es la tercera vez en esta auditoría "
      "—`E_i`, el mapa índice→dominio, y ahora la cobertura—. La regla ya está "
      "escrita y hay que aplicarla antes, no después: **buscar donde debía "
      "estar, antes de declarar nada.**")
    A("")

    A("## 1 · La aritmética no cuadra")
    A("")
    A("| | |")
    A("|---|---:|")
    A(f"| Metas del PDOT (declaradas en el catálogo) | {h['pdot_total']} |")
    A(f"| Metas del PDOT (filas reales del catálogo) | {h['pdot_filas']} |")
    A(f"| Metas en el motor (`H12` filas 6-30) | {h['motor']} |")
    A(f"| Metas en el catálogo «fuera del motor» | {h['fuera_catalogo']} |")
    A(f"| Diferencia simple `PDOT − motor` | {h['resta_simple']} |")
    A(f"| Metas del PDOT que NO están en «fuera» (cruce por texto) | "
      f"{h['dentro_inferido']} |")
    A("")
    if not cuadra:
        A(f"⚠️ **`{h['pdot_total']} − {h['motor']} = {h['resta_simple']}`, pero "
          f"el catálogo de exclusiones tiene {h['fuera_catalogo']}.** Sobran "
          f"**{h['fuera_catalogo'] - h['resta_simple']}**. Los tres artefactos "
          "—el PDOT extraído, el motor y el catálogo de exclusiones— no "
          "describen el mismo universo.")
        A("")
        A("Y el cruce por texto lo confirma desde el otro lado: de las "
          f"{h['pdot_filas']} metas del PDOT, **{h['dentro_inferido']} no "
          f"aparecen en el catálogo de exclusiones** — pero el motor opera con "
          f"{h['motor']}. Faltan **{h['motor'] - h['dentro_inferido']}** por "
          "explicar.")
        A("")
        A("Eso **no invalida el ADR-036**: la decisión de operar sobre 25 sigue "
          "en pie, y el ADR verificó que las 25 existen en el PDOT. Lo que dice "
          "es que **el complemento nunca se cuadró**: los catálogos que "
          "describen lo que queda fuera no reconstruyen el universo. Hoy no se "
          "puede afirmar con precisión qué metas quedan excluidas, y por tanto "
          "**tampoco publicar un porcentaje de cobertura**.")
        A("")
        A("Las causas posibles son ordinarias —el catálogo de exclusiones puede "
          "incluir metas de otra versión del PDOT, o el cruce por texto puede "
          "fallar por redacciones que difieren— y **ninguna se elige aquí**: "
          "distinguirlas exige comparar meta a meta contra el documento, que es "
          "trabajo de curación, no de conteo.")
        A("")
    else:
        A("✅ Los tres artefactos cuadran.")
        A("")

    A("## 2 · El denominador y su estado — con una corrección de fondo")
    A("")
    A("Antes de dividir 25 entre 66, de dónde sale el 66:")
    A("")
    A(f"- **Fuente**: `{proc.get('fuente', '—')}`")
    A(f"- **SHA256**: `{str(proc.get('sha256', '—'))[:32]}…`")
    A(f"- **Carácter**: {proc.get('caracter', '—')}")
    A(f"- **Verificabilidad**: {proc.get('verificabilidad', '—')}")
    A(f"- **Corroborar contra**: `{proc.get('corroborar_contra', '—')}`")
    A("")
    A("### ⚠️ «No remitido formalmente» NO es «no oficial»")
    A("")
    A("Javo lo corrigió y va al fondo de la doctrina: **el documento se obtuvo "
      "del portal del GAD, y eso lo hace oficial.** La `LOTAIP Art. 7` obliga a "
      "publicarlo, y el canon de QUIRA dice literalmente que **el portal es la "
      "materialización de una obligación**. Degradar lo publicado a «no "
      "oficial» porque no llegó por oficio contradice el modelo entero: si sólo "
      "valiera lo remitido a solicitud, **toda la transparencia activa valdría "
      "cero** — y con ella `V_LOTAIP`, que puntúa 1,0 justamente por «documento "
      "en URL pública del GAD, accesible y verificable» (`H13!C10`).")
    A("")
    A("Son dos cosas distintas que la etiqueta mezclaba:")
    A("")
    A("| | |")
    A("|---|---|")
    A("| **no remitido formalmente** | no hubo entrega institucional por "
      "solicitud — es un hecho sobre el CANAL |")
    A("| **no oficial** | el documento no tiene carácter oficial — es un juicio "
      "sobre el DOCUMENTO, y aquí es **falso** |")
    A("")
    A("### Pero la reserva se mantiene, por otra razón")
    A("")
    A("La `verificabilidad: parcial` **sigue siendo correcta**, y no por la "
      "oficialidad: lo leído fue el **Plan Plurianual `.xlsx`**, y la propia "
      "procedencia pide corroborarlo contra el **PDOT Bicentenario `.docx`, "
      "tablas #341-352**. Son **dos documentos distintos**; que ambos sean "
      "oficiales no los hace el mismo.")
    A("")
    A("Es el **escalón 7** de la escalera prueba↔verificador: *lo leído ≠ la "
      "fuente*. El conteo de 66 procede del instrumento de programación "
      "plurianual, no del PDOT aprobado por ordenanza. Mientras no se corrobore, "
      "el denominador es **oficial y provisional a la vez** — dos atributos que "
      "no se contradicen.")
    A("")

    A("## 3 · Composición y sesgo por sistema del PDOT")
    A("")
    A("| Sistema | En el PDOT | Fuera del motor | Dentro (inferido) | % dentro |")
    A("|---|---:|---:|---:|---:|")
    for s in sorted(sis_pdot):
        tot, fu = sis_pdot[s], sis_fuera.get(s, 0)
        den = sis_dentro.get(s, 0)
        A(f"| {s} | {tot} | {fu} | {den} | {den / tot * 100:.0f} % |")
    A("")
    A("⚠️ **Esta tabla mide composición, NO sesgo.** Un reparto desigual puede "
      "ser exactamente lo que una muestra **estratégica** debe producir: si el "
      "criterio era «las metas de mayor peso presupuestario», concentrar en "
      "unos sistemas es el resultado correcto, no una distorsión.")
    A("")
    A("**Para hablar de sesgo hace falta el criterio, y el criterio no está "
      "declarado** (§4). Sin él, cualquier lectura de esta tabla sería inferir "
      "la regla desde el patrón de sus resultados — `DOC-009`.")
    A("")

    A("## 4 · ★ EL CRITERIO, DECLARADO POR SU FUENTE LEGÍTIMA")
    A("")
    A("> **«Las 25 fueron tomadas por contener el monto económico más amplio en "
      "relación al total de metas, las 66. Eso fue para fines sólo de tesis.»**")
    A("> — Javo, 2026-09-03")
    A("")
    A("**El criterio deja de ser `NOT_DETERMINABLE`.** Y la forma en que se "
      "resolvió importa tanto como el contenido: **no se dedujo mirando las 25** "
      "—eso habría sido `DOC-009`— sino que **lo declaró quien lo aplicó**. Es "
      "exactamente lo que a `E_i` le sigue faltando: una fuente con autoridad "
      "sobre la regla, no una explicación que encaje con los datos.")
    A("")
    A("### Y el criterio tiene una consecuencia medible")
    A("")
    A("Si la selección fue **por monto**, entonces la muestra es representativa "
      "del **gasto**, no del PDOT como instrumento de planificación. Contrastado "
      "con la composición por sistema (§3), el efecto es sistemático:")
    A("")
    sis_ord = sorted(sis_pdot, key=lambda s: sis_dentro.get(s, 0) / sis_pdot[s])
    for s in sis_ord[:2]:
        A(f"- **{s}** queda al "
          f"**{sis_dentro.get(s, 0) / sis_pdot[s] * 100:.0f} %** de cobertura")
    A("")
    A("El sistema **institucional** es el de menor cobertura — y es precisamente "
      "donde viven la gobernanza, la transparencia y la participación: metas de "
      "**bajo costo y alta relevancia** para un observatorio de integridad. Un "
      "criterio de monto las excluye por construcción.")
    A("")
    A("⚠️ **Esto no es un defecto de la tesis.** Para validar un modelo con "
      "recursos limitados, tomar las metas de mayor peso económico es una "
      "decisión metodológica razonable y transparente: concentra la validación "
      "donde está el dinero. **Lo que dice es qué puede afirmar el ICPI v1** — "
      "desempeño sobre el gasto estratégico— **y qué no**: desempeño sobre el "
      "PDOT como mandato completo.")
    A("")
    A("### La regla que queda")
    A("")
    A("> **La justificación del universo operacional no implica la "
      "justificación de su mecanismo de selección.**")
    A("")
    A("`ADR-036` justificó **usar 25 como universo operacional v1**. Eso no era "
      "lo mismo que justificar **por qué esas 25 son representativas** — y hasta "
      "hoy sólo teníamos lo primero. Son dos afirmaciones distintas y confundirlas "
      "es la misma trampa que `E_i`: conocer el valor no es conocer la regla que "
      "lo produjo. → `DOC-018`")
    A("")

    A("### Nota histórica · lo que este apartado decía antes")
    A("")
    if criterio:
        A("Documentos que mencionan un criterio de selección:")
        A("")
        for c in criterio:
            A(f"- `{c}`")
        A("")
    else:
        A("**Búsqueda en `docs/`, `governance/` e `identity/`: ningún documento "
          "declara por qué esas 25 y no otras 25.**")
        A("")
    A("`ADR-036` verifica que las 25 **existen** en el PDOT —ninguna "
      "inventada—, y eso responde «son legítimas». **No responde «por qué "
      "éstas».** Es la misma forma del problema de `E_i`: valores conocidos, "
      "regla generadora no reconstruible.")
    A("")
    A("Y aquí importa más que en `E_i`, porque de ese criterio depende si la "
      "muestra es **representativa**:")
    A("")
    A("| Si el criterio fue… | Entonces la muestra… |")
    A("|---|---|")
    A("| mayor peso presupuestario | es materialmente representativa aunque sea "
      "el 38 % de las metas |")
    A("| competencia exclusiva crítica | representa el mandato, no el gasto |")
    A("| disponibilidad de evidencia | **está sesgada hacia lo verificable**, y "
      "el ICPI mediría lo que es fácil de medir |")
    A("| conveniencia o disponibilidad | no es una muestra estratégica |")
    A("")
    A("Las cuatro producen el mismo conjunto de 25 y **significados "
      "completamente distintos del 27,4582 %**.")
    A("")

    A("## 5 · El ADR-036 §1 exige declarar el alcance · ¿se cumple?")
    A("")
    A("> «Toda publicación de d01/d03 debe declararlo: *se mide contra las 25 "
      "metas estratégicas del modelo*.»")
    A("")
    A(f"- En **texto que la interfaz puede pintar**: "
      f"**{len(alcance['visible'])}**"
      + (" — " + ", ".join(f"`{x}`" for x in alcance["visible"][:4])
         if alcance["visible"] else ""))
    A(f"- Sólo en **comentarios de código**: {len(alcance['comentarios'])}"
      + (" — " + ", ".join(f"`{x}`" for x in alcance["comentarios"][:4])
         if alcance["comentarios"] else ""))
    A("")
    if not alcance["visible"]:
        A("⚠️ **La consecuencia práctica del ADR no se ejecutó.** El alcance se "
          "declaró en el ADR y en un comentario de código; **el producto no lo "
          "dice**. Un usuario que lee el ICPI en cualquier superficie recibe un "
          "índice que se presenta como global sobre un universo del que nadie "
          "le informa.")
        A("")
        A("Es el patrón del «48,33 %» invertido: allí una cifra retirada seguía "
          "publicándose; aquí **una declaración obligatoria nunca llegó a "
          "publicarse**. Un ADR ratificado cuya consecuencia práctica nadie "
          "comprueba es una decisión que existe sólo en el papel.")
        A("")

    A("## Veredicto de 008")
    A("")
    A("> ### COBERTURA LIMITADA, METODOLÓGICAMENTE JUSTIFICADA EN SU ALCANCE v1,")
    A("> ### con criterio de selección DECLARADO y correspondencia "
      "exclusión/universo AÚN NO RECONCILIADA.")
    A("")
    A("La formulación es deliberadamente estrecha. Decir «metodológicamente "
      "justificada» a secas sonaría a que está demostrada la "
      "**representatividad** de las 25, y lo que `ADR-036` justifica es algo más "
      "específico: **la decisión de usar 25 como universo operacional v1**.")
    A("")
    A("| | Estado |")
    A("|---|---|")
    A("| `25/66` como cobertura documental | **37,88 %** — relación válida |")
    A("| 25 = universo operacional v1 | **RATIFICADO** (`ADR-036`) |")
    A("| Criterio de selección | **DECLARADO**: mayor monto económico (Javo) |")
    A("| Representatividad respecto del PDOT | **del gasto sí · del mandato no** |")
    A("| Identidad de las metas excluidas | ⚠️ **pendiente de reconciliación** |")
    A("| Sesgo | no es sesgo: es el criterio operando como fue definido |")
    A("| Ampliar 25→66 ahora | **NO** — es metodología nueva (`ADR-036 §4`) |")
    A("| ICPI 27,4582 % | **CONGELADO** · no se recalcula |")
    A("")
    A("**La afirmación que el ICPI v1 sostiene**, y ninguna más amplia:")
    A("")
    A("> El ICPI v1 opera sobre un subconjunto de 25 metas —las de mayor monto "
      "económico— de un PDOT que contiene 66. Su resultado **no representa el "
      "desempeño del PDOT completo**, sino el desempeño respecto de su universo "
      "operacional v1.")
    A("")
    A("**Reserva única que queda abierta · la correspondencia meta a meta.** No "
      "existe todavía un catálogo canónico que enlace las 66 con las 25. La "
      "resta `66−25=41` es aritméticamente correcta, pero **la identidad de esas "
      "41 no está demostrada documentalmente** — y los 50 del catálogo de "
      "exclusiones no pueden asumirse equivalentes. Ése es el único pendiente "
      "técnico real de 008.")
    A("")
    A("**Y la obligación del `ADR-036 §1` sigue incumplida**: el alcance no se "
      "declara en ninguna superficie visible (`DOC-017`).")
    A("")

    A("## ★ DECISIÓN v2 · el universo completo")
    A("")
    A("> **«Ahora, como ecosistema de Ecuador para LATAM, debemos trabajar con "
      "todo el universo del PDOT.»** — Javo, 2026-09-03")
    A("")
    A("Es exactamente la evolución que `ADR-036 §3/§4` anticipó, y la decisión "
      "es correcta: un observatorio que aspira a 222 GAD no puede medir sobre "
      "una muestra tomada para validar una tesis. **El criterio de monto sirvió "
      "para demostrar que el modelo funciona; no sirve para observar un "
      "mandato.**")
    A("")
    A("Pero el `ADR-036 §4` fija cómo: **versión nueva del motor · recalibración "
      "· nueva validación empírica · ADR específico**. Y hay una razón de "
      "secuencia que conviene respetar:")
    A("")
    A("> `011` todavía no ha dictaminado si la fórmula es válida. **Cargar 66 "
      "metas en un álgebra que puede cambiar sería hacer el trabajo dos veces** "
      "— justo lo que este proyecto decidió evitar al construir el mapa de "
      "frentes.")
    A("")
    A("Por eso la decisión se **registra ahora** y su ejecución va **después de "
      "`011`**. Con una excepción importante, y es la parte más cara:")
    A("")
    A("**La reconciliación meta a meta (`66 ↔ 25`) puede y debe empezar ya.** No "
      "depende de `011` —hay que hacerla sea cual sea el dictamen—, es "
      "prerequisito de v2, y además cierra la única reserva que 008 deja "
      "abierta. Es el trabajo que desbloquea todo lo demás.")
    A("")
    A("### Lo que 008 NO hace")
    A("")
    A("- **No amplía de 25 a 66.** `ADR-036 §2/§4`: sería una versión "
      "metodológica nueva, con recalibración y ADR propio. No entra por la "
      "puerta de una cura.")
    A("- **No toca la cifra madre.** 27,4582 % sigue congelada.")
    A("- **No declara sesgo.** Falta el criterio, y sin él sería `DOC-009`.")
    A("- **No cuadra el complemento.** Señala que 41 ≠ 50 y deja el trabajo "
      "identificado, porque cuadrarlo exige cruzar meta a meta contra el PDOT.")
    A("")
    A("---")
    A(f"*GM-Ω-ICPI-008 · universo operacional {h['motor']} · PDOT "
      f"{h['pdot_total']} (verificabilidad parcial) · el Gold Master no se "
      "modificó · Dylus Lab © 2026*")

    _SALIDA.write_text("\n".join(o) + "\n", encoding="utf-8")


if __name__ == "__main__":
    raise SystemExit(main())
