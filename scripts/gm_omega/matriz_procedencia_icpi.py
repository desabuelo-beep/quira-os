# -*- coding: utf-8 -*-
"""
scripts/gm_omega/matriz_procedencia_icpi.py — GM-Ω-ICPI-004/005

Genera la matriz de procedencia 25 × 6 del ICPI: para cada meta y cada variable,
de qué celda sale, con qué fórmula, de qué período y en qué estado de
trazabilidad queda.

POR QUÉ SE DERIVA Y NO SE ESCRIBE. El colega exigió la trazabilidad celda a
celda —«poder contestar, para cualquier meta, por qué tiene exactamente ese
número»— y la primera pasada comprimió las 150 celdas en 6 patrones. Comprimir
era legítimo mientras el patrón fuera idéntico, pero deja sin respuesta la
pregunta concreta.

Y escribirla a mano habría reproducido el defecto que esta misma auditoría
persigue: una tabla copiada del Excel se queda atrás el día que el Excel cambia
—el patrón del «48,33 %»—. Este script la vuelve a derivar cada vez.

    LECTURA PURA. No escribe en el Gold Master. Regla de Oro 1 y 4:
    el Excel es el estado, aquí sólo se le pregunta.

Uso:  python scripts/gm_omega/matriz_procedencia_icpi.py
Dylus Lab © 2026
"""
from __future__ import annotations

import re
import sys
from pathlib import Path

_RAIZ = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(_RAIZ))

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

_SALIDA = _RAIZ / "docs" / "architecture" / "GM-OMEGA_ICPI_MATRIZ_004.md"

# Columnas del motor, con el nombre que la TESIS les da (no el del Excel: el
# nombre canónico es el del constructo, y GM-Ω-ICPI-001 lo reconstruyó).
_VARIABLES = {
    2: ("P_i", "Coeficiente de Peso Presupuestario"),
    3: ("R_i", "Coeficiente de Relevancia Normativa"),
    4: ("V_i", "Inmutabilidad Documental"),
    5: ("E_i", "Coeficiente de Fricción de Autonomía"),
    6: ("T_i", "Materialización Temporal"),
    9: ("C_i", "Trazabilidad Orgánica"),
}

# ── ESTADOS DE PROCEDENCIA (taxonomía del colega, 2026-09-03) ────────────────
# NO son veredictos: son estados de la auditoría mientras GM-Ω-ICPI-011 no
# dictamine. Y la distinción entre los dos últimos es la que el colega exigió:
#
#   VERIFICADO              la cadena completa es reconstruible
#   PARCIALMENTE_VERIFICADO fórmula y fuente inmediata conocidas; falta la
#                           fuente documental o la justificación metodológica
#   NO_DETERMINABLE         no se puede reconstruir la cadena
#   UNTRACEABLE             hay valor y NO existe fuente identificable
#                           **después de agotar las fuentes razonables**
#   TEMPORAL_SEMANTIC_GAP   la fuente existe, pero su período o función no está
#                           correctamente declarado
#
# ⚠️ `E_i` estuvo clasificado aquí como UNTRACEABLE, y era afirmar más de lo
# medido. Se agotó la búsqueda: no deriva de `Competencia_GAD` (Exclusiva_Crítica
# toma 0,75 y 1), no deriva de la entidad ejecutora (EP Aseo toma los tres
# valores) — pero **la tesis SÍ define su regla**, con COOTAD Art. 54 y NCI
# 200-04 CGE: directa 1,00 · convenio 0,90 · delegación a adscrita 0,75. Luego
# la regla existe y lo que no consta en el libro es la MODALIDAD de cada meta.
_ESTADO = {
    "P_i": ("PARCIALMENTE_VERIFICADO", "referencia directa a H14!G; falta la "
                                       "cédula presupuestaria por meta"),
    "R_i": ("VERIFICADO", "fórmula + artículo del COOTAD citado meta a meta"),
    "V_i": ("TEMPORAL_SEMANTIC_GAP", "la columna leída se llama `Vi_2025`"),
    # ⚠️ ACTUALIZADO 2026-09-04 · los documentos de abril cerraron el hueco.
    # `Metodologia_SIAP_ICPI.docx` (TERRA/QUADRUM) define «Eᵢ — Autonomía
    # Orgánica: el CONTROL DEL DIRECTOR sobre la ejecución. 1.0 autónomo · 0.9
    # compartido · 0.75 difuso», y `H12!A4` la cita TEXTUALMENTE. La regla está
    # VERIFICADA; lo que falta es el insumo para auditar su aplicación por meta.
    #
    # Esta auditoría había contrastado los valores contra OTRA definición —la de
    # la tesis, fricción por delegación con COOTAD 54— y por eso «no cuadraban».
    "E_i": ("REGLA_VERIFICADA · aplicación pendiente",
            "«Autonomía Orgánica» definida en Metodologia_SIAP_ICPI (abril) y "
            "citada en H12!A4; el CONTROL DEL DIRECTOR por meta no consta"),
    "T_i": ("VERIFICADO · sensibilidad pendiente", "ratio por ENTIDAD ejecutora; "
                                                   "el tope MIN(1,…) se juzga en 007"),
    "C_i": ("PARCIALMENTE_VERIFICADO", "VLOOKUP a TBL_CALIBRACION_Ci; falta la "
                                       "regla que calibra la tabla"),
}

# La regla de la tesis para E_i, para poder contrastarla — no para aplicarla.
_ADSCRITAS = {"ENTE-02 Patronato", "ENTE-03 Bomberos", "ENTE-04 EP Aseo"}
_E_ESPERADA_ADSCRITA = 0.75

# Qué entidad ejecutora representa cada columna de `Ti_norm_2026` (H07b fila 20).
_ENTIDAD_TI = {"B": "ENTE-01 GAD central", "C": "ENTE-02 Patronato",
               "D": "ENTE-03 Bomberos", "E": "ENTE-04 EP Aseo"}


def _entidad_de(formula: str) -> str:
    """De `=H07b_Ti_INVERSIÓN_eSIGEF!B20` a la entidad que ese ratio mide."""
    m = re.search(r"!([A-Z]+)\d+", str(formula))
    return _ENTIDAD_TI.get(m.group(1), "—") if m else "—"


def construir() -> list[dict]:
    import openpyxl

    import config
    if not getattr(config, "GOLD_MASTER_RESUELTO", False):
        return []

    # Dos lecturas del mismo libro: fórmulas y valores. Sin ellas no se puede
    # responder a la vez «de dónde sale» y «cuánto vale».
    wf = openpyxl.load_workbook(config.SIAP_PATH, data_only=False, read_only=True)
    wv = openpyxl.load_workbook(config.SIAP_PATH, data_only=True, read_only=True)
    hf, hv = wf["H12_MOTOR_ICPI_CANÓNICO"], wv["H12_MOTOR_ICPI_CANÓNICO"]

    # Entidad ejecutora de cada meta, según de qué columna de `Ti_norm_2026`
    # toma su ratio. Es el mejor proxy disponible en el libro, y se declara como
    # proxy: la MODALIDAD de ejecución (directa/convenio/delegada) no consta.
    entidad_meta = {
        str(hv.cell(row=r, column=1).value):
            _entidad_de(hf.cell(row=r, column=6).value) for r in range(6, 31)
    }

    filas = []
    for r in range(6, 31):                       # las 25 metas del motor
        meta = hv.cell(row=r, column=1).value
        for col, (var, nombre) in _VARIABLES.items():
            formula = hf.cell(row=r, column=col).value
            valor = hv.cell(row=r, column=col).value
            es_literal = not (isinstance(formula, str) and formula.startswith("="))
            estado, por_que = _ESTADO[var]

            # COMPROBACIÓN CRUZADA que pidió el colega: ¿el valor de `E_i`
            # concuerda con la regla de la tesis, dada la entidad que ejecuta?
            # No demuestra un defecto —la modalidad real no consta— pero señala
            # dónde la regla documentada y el valor asignado no concuerdan.
            # ⚠️ REFORMULADA 2026-09-04. Antes decía «la regla de la tesis pide
            # 0,75» y se leía como una incoherencia del motor. No lo es: son DOS
            # DEFINICIONES distintas de `E_i` conviviendo en la genealogía —
            #
            #   A · Metodologia_SIAP_ICPI (abril) → CONTROL DEL DIRECTOR
            #       1,0 autónomo · 0,9 compartido · 0,75 difuso  ← la que cita H12!A4
            #   B · tesis → FRICCIÓN POR DELEGACIÓN (COOTAD 54 · NCI 200-04)
            #       1,0 directa · 0,90 convenio · 0,75 adscrita
            #
            # El motor implementa A. Estos casos sólo divergen bajo B, y esa
            # divergencia es EVIDENCIA GENEALÓGICA, no un defecto.
            alerta = ""
            if var == "E_i":
                ent = entidad_meta.get(str(meta), "—")
                if ent in _ADSCRITAS and valor != _E_ESPERADA_ADSCRITA:
                    alerta = (f"↔ lo ejecuta {ent} (adscrita) y la regla de la "
                              f"tesis pide {_E_ESPERADA_ADSCRITA} — divergencia "
                              f"entre definiciones A y B, no defecto")

            filas.append({
                "meta": str(meta), "var": var, "nombre": nombre,
                "celda": f"H12!{hf.cell(row=r, column=col).coordinate}",
                "valor": valor,
                "origen": "LITERAL (sin origen declarado)" if es_literal else str(formula),
                "entidad": entidad_meta.get(str(meta), "—") if var in ("T_i", "E_i") else "—",
                "estado": estado, "por_que": por_que, "alerta": alerta,
            })
    wf.close(); wv.close()
    return filas


def main() -> int:
    filas = construir()
    if not filas:
        print("[no determinable] Gold Master no resuelto — no se genera la matriz.")
        return 2                                  # ni verde ni rojo: no pude mirar

    metas = sorted({f["meta"] for f in filas})
    print(f"matriz de procedencia: {len(filas)} celdas · {len(metas)} metas × "
          f"{len(_VARIABLES)} variables")

    out = ["# GM-Ω · ICPI — MATRIZ DE PROCEDENCIA  `004/005`", "",
           "**DERIVADO — no editar a mano.** Lo regenera "
           "`scripts/gm_omega/matriz_procedencia_icpi.py` leyendo el Gold Master "
           "vigente. Escribirlo a mano lo dejaría atrás el día que el Excel "
           "cambie, que es el patrón que esta auditoría persigue.", "",
           f"`{len(filas)}` celdas · {len(metas)} metas × {len(_VARIABLES)} variables · "
           "baseline congelado **27,4582 %** (regla GM-Ω-ICPI-000)", "",
           "## Estado de trazabilidad por variable", "",
           "| Var | Constructo (tesis) | Estado provisional | Por qué |",
           "|---|---|---|---|"]
    for var, nombre in [(v, n) for v, n in _VARIABLES.values()]:
        estado, por_que = _ESTADO[var]
        out.append(f"| `{var}` | {nombre} | **{estado}** | {por_que} |")

    out += ["", "⚠️ Estos NO son veredictos: son estados de la auditoría mientras "
            "`GM-Ω-ICPI-011` no dictamine.", "", "## Las 150 celdas", ""]
    for meta in metas:
        out += [f"### `{meta}`", "",
                "| Var | Celda | Valor | Origen | Entidad |", "|---|---|---|---|---|"]
        for f in [x for x in filas if x["meta"] == meta]:
            v = f["valor"]
            v = f"{v:.6f}" if isinstance(v, float) else str(v)
            ent = f"{f['entidad']} {f['alerta']}".strip()
            out.append(f"| `{f['var']}` | `{f['celda']}` | {v} | "
                       f"`{f['origen'][:72]}` | {ent} |")
        out.append("")

    _SALIDA.write_text("\n".join(out) + "\n", encoding="utf-8")
    print(f"→ {_SALIDA.relative_to(_RAIZ).as_posix()}")

    literales = [f for f in filas if f["origen"].startswith("LITERAL")]
    print(f"celdas sin origen declarado: {len(literales)} "
          f"({', '.join(sorted({f['var'] for f in literales}))})")
    alertas = [f for f in filas if f["alerta"]]
    if alertas:
        print(f"E_i · divergencia entre definiciones A y B: {len(alertas)} "
              f"→ {', '.join(f['meta'] for f in alertas)}")
        print("  ↔ El motor implementa la definición A (control del director, "
              "H12!A4). Bajo la B (fricción por delegación, tesis) estos casos")
        print("    darían otro valor. Es evidencia genealógica, NO un defecto.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
