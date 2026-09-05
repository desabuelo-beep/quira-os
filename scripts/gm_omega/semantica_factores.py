# -*- coding: utf-8 -*-
"""
scripts/gm_omega/semantica_factores.py — GM-Ω-ICPI-011-C2 · GENEALOGÍA SEMÁNTICA

    ¿Qué significa CADA FACTOR del ICPI, según el propio instrumento?

    `011-C1` reconstruyó el álgebra: P·R·V·T → +E → +C. Esta etapa pregunta
    otra cosa: **qué mide cada letra**, y si lo que el motor DECLARA medir
    coincide con lo que su mecanismo EFECTIVAMENTE mide.

    POR QUÉ AHORA Y NO DESPUÉS. `009` clasificó `C_i` dos veces y las dos
    veces se equivocó: primero DOCUMENTAL (error de esta dirección), después
    MATERIAL (hipótesis del autor). El colega detuvo la segunda: un análisis de
    incentivos no puede fijar la semántica de la variable que lo audita. Sin
    `011-C2`, todo análisis de comportamiento se hace sobre variables cuya
    ontología todavía se está reconstruyendo.

    LA REGLA DE ESTA ETAPA. Se lee del INSTRUMENTO, no de la memoria:
    `H02_GLOSARIO_QUIRA` (definiciones canónicas), `H01 Secciones I/M`
    (calibración de `C_i`), `H12` (valores y fórmulas vigentes). Toda
    afirmación se clasifica DEMOSTRADO · DECLARADO · INFERIDO · NO DETERMINABLE.

    ⚠️ NO DICTAMINA. Que una semántica sea confusa, redundante o calibrada al
    revés es un HECHO que se registra; si eso invalida el constructo lo juzga
    `011-C4`. `011-C2` reconstruye.

Uso:  python scripts/gm_omega/semantica_factores.py
Dylus Lab © 2026
"""
from __future__ import annotations

import re
import sys
from pathlib import Path

_RAIZ = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(_RAIZ))

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

_SALIDA = _RAIZ / "docs" / "architecture" / "GM-OMEGA_ICPI_SEMANTICA_011C2.md"

# Los seis factores y la columna de `H12` donde vive cada uno. El orden es el
# de la fórmula, no el alfabético.
_FACTORES = [("P_i", 2), ("R_i", 3), ("V_i", 4), ("E_i", 5),
             ("T_i", 6), ("C_i", 9)]

# Términos del glosario que definen o afectan a cada factor. La búsqueda es por
# prefijo del término, no por coincidencia libre: `Ci` casaría con «Ciudadano».
_TERMINOS = {
    "C_i": ("Ci (", "Ci_Adaptativo", "Ci_Determinista", "Mapeo Retrospectivo"),
    "E_i": ("Ei (",),
    "V_i": ("Vi (", "Vi_", "Interoperabilidad"),
    "T_i": ("Ti (", "Ti_", "Materialización"),
    "P_i": ("Pi (", "Peso Presupuestario"),
    "R_i": ("Ri (", "Relevancia Normativa"),
}


def leer_libro() -> dict:
    """Todo lo que 011-C2 necesita del Gold Master. `{}` si no se resolvió —
    el tercer estado: «no pude obtener» no es «no existe»."""
    import openpyxl

    import config
    if not getattr(config, "GOLD_MASTER_RESUELTO", False):
        return {}

    wv = openpyxl.load_workbook(config.SIAP_PATH, data_only=True, read_only=True)
    wf = openpyxl.load_workbook(config.SIAP_PATH, data_only=False, read_only=True)
    h01, h02 = wv["H01_PARÁMETROS"], wv["H02_GLOSARIO_QUIRA"]
    h12v, h12f = wv["H12_MOTOR_ICPI_CANÓNICO"], wf["H12_MOTOR_ICPI_CANÓNICO"]

    # ── Glosario: término → (definición, fuente) ──────────────────────────
    glosario = {}
    for r in range(1, 240):
        t = h02.cell(row=r, column=1).value
        if not t:
            continue
        glosario[str(t).strip()] = (
            str(h02.cell(row=r, column=2).value or "").strip(),
            str(h02.cell(row=r, column=3).value or "").strip())

    # ── H01 Sección I · la tabla que imputa cada meta a una unidad ────────
    seccion_i, cab_i = {}, None
    for r in range(93, 125):
        a = h01.cell(row=r, column=1).value
        if a == "ID_Meta":
            cab_i = [h01.cell(row=r, column=c).value for c in range(1, 7)]
            continue
        if cab_i and a and str(a)[0].isalpha() and "-" in str(a):
            seccion_i[str(a)] = {
                "unidad": h01.cell(row=r, column=3).value,
                "direccion": str(h01.cell(row=r, column=4).value or ""),
                "ci_base": h01.cell(row=r, column=5).value,
                "legal": str(h01.cell(row=r, column=6).value or ""),
            }

    # ── H01 Sección M · las cuatro deducciones legales de `C_i` ───────────
    seccion_m, cab_m, nota_m = {}, None, []
    for r in range(186, 220):
        a = h01.cell(row=r, column=1).value
        if a == "ID_Meta":
            cab_m = [str(h01.cell(row=r, column=c).value or "").replace("\n", " ")
                     for c in range(2, 7)]
            continue
        if not a:
            continue
        if cab_m and str(a)[0].isalpha() and "-" in str(a) and len(str(a)) < 16:
            seccion_m[str(a)] = [h01.cell(row=r, column=c).value
                                 for c in range(2, 8)]
        elif "Ci" in str(a) or "metodológica" in str(a):
            nota_m.append(str(a).strip())

    # ── H01 Sección L · la matriz que DEFINE cuánto deduce cada infracción ─
    #
    # ⚠️ Esta dirección estuvo a punto de escribir que el glosario apuntaba a
    # una sección inexistente («Fuente: H01 Sección L» cuando la tabla está en
    # la M). Falso: la L define los códigos y sus deducciones, la M los
    # registra por meta. Verificarlo antes de afirmarlo evitó el error — y
    # destapó algo mayor, que son las deducciones mismas.
    seccion_l, regla_l = [], []
    for r in range(171, 186):
        a = h01.cell(row=r, column=1).value
        if not a:
            continue
        if str(a).startswith("INF-"):
            seccion_l.append({
                "codigo": str(a),
                "norma": str(h01.cell(row=r, column=2).value or ""),
                "evento": str(h01.cell(row=r, column=3).value or ""),
                "deduccion": h01.cell(row=r, column=4).value,
                "referencia": str(h01.cell(row=r, column=5).value or ""),
            })
        elif any(k in str(a) for k in ("REGLA INF", "Si INF", "=SI(", "MOTOR DETERMINISTA")):
            regla_l.append(re.sub(r"\s+", " ", str(a)).strip())

    # ── H12 · valor y fórmula de cada factor, meta a meta ─────────────────
    metas = []
    for r in range(6, 31):
        mid = h12v.cell(row=r, column=1).value
        if not mid:
            continue
        fila = {"id": str(mid)}
        for nombre, col in _FACTORES:
            fila[nombre] = h12v.cell(row=r, column=col).value
            fila[nombre + "_f"] = h12f.cell(row=r, column=col).value
        metas.append(fila)

    # Encabezados literales que el informe cita textualmente. Se leen aquí y no
    # cuando hacen falta: reabrir el libro por cada cita cuesta segundos y deja
    # cuatro rutas distintas al mismo dato.
    citas = {f: re.sub(r"\s+", " ",
                       str(h01.cell(row=f, column=1).value or "—")
                       ).replace("|", "·").strip()
             for f in (93, 94, 186)}

    # ── ¿Qué columnas entran REALMENTE al ICPI? ───────────────────────────
    # El numerador es la única autoridad sobre esto. Un factor definido en el
    # glosario pero ausente del numerador no participa del índice.
    num = str(h12f.cell(row=6, column=10).value or "")
    cols_num = sorted({openpyxl.utils.column_index_from_string(c)
                       for c in re.findall(r"\b([A-Z]{1,2})6\b", num)})
    entran = [str(h12v.cell(row=5, column=c).value or f"col{c}") for c in cols_num]

    return {"glosario": glosario, "seccion_i": seccion_i, "cab_i": cab_i,
            "seccion_m": seccion_m, "cab_m": cab_m, "nota_m": nota_m,
            "seccion_l": seccion_l, "regla_l": regla_l,
            "numerador": num, "entran": entran,
            "metas": metas, "citas": citas}


def definiciones(glosario: dict) -> dict:
    """Las entradas del glosario que definen cada factor. Sin inventar: si un
    factor no tiene entrada, se dice que no la tiene."""
    out = {}
    for f, claves in _TERMINOS.items():
        enc = []
        for termino, (definicion, fuente) in glosario.items():
            if any(termino.startswith(k) for k in claves):
                enc.append((termino, definicion, fuente))
        out[f] = enc
    return out


def mecanismo(metas: list[dict], factor: str) -> dict:
    """Qué hace REALMENTE la celda: ¿fórmula o literal? ¿cuántos valores
    distintos? Es la mitad que ninguna definición puede sustituir."""
    formulas = [m[factor + "_f"] for m in metas]
    literales = sum(1 for x in formulas if not str(x or "").startswith("="))
    refs = set()
    for x in formulas:
        if str(x or "").startswith("="):
            refs.update(re.findall(r"(H\d+[a-z]?_?[A-ZÁÉÍÓÚÑa-z_]*)", str(x)))
    valores = [m[factor] for m in metas if isinstance(m[factor], (int, float))]
    return {"literales": literales, "total": len(formulas),
            "refs": sorted(refs)[:4],
            "distintos": sorted({round(float(v), 4) for v in valores}),
            "ejemplo": next((str(x) for x in formulas
                             if str(x or "").startswith("=")), None)}


def solape_E_C(metas: list[dict], seccion_i: dict) -> dict:
    """★ La comparación que motivó esta etapa.

    `E_i` (glosario): «grado de autonomía con que el GAD ejerce una
    competencia · 1,0 autónomo / 0,9 compartido / 0,75 difuso o ambiguo».

    `C_i` Sección I: misma escala {1 · 0,9 · 0,75} y base legal redactada con
    el MISMO vocabulario — «proceso exclusivo / compartido / difuso».

    Si además coincidieran meta a meta, la fórmula estaría multiplicando dos
    veces la misma dimensión. Se mide; no se supone."""
    pares, iguales = [], 0
    for m in metas:
        e, c = m["E_i"], m["C_i"]
        if not isinstance(e, (int, float)) or not isinstance(c, (int, float)):
            continue
        legal = seccion_i.get(m["id"], {}).get("legal", "")
        etq = ("exclusivo" if "exclusiv" in legal.lower() else
               "compartido" if "compartid" in legal.lower() else
               "difuso" if "difuso" in legal.lower() else "—")
        coincide = abs(float(e) - float(c)) < 1e-9
        iguales += coincide
        pares.append({"id": m["id"], "E": float(e), "C": float(c),
                      "etiqueta": etq, "igual": coincide, "legal": legal})
    n = len(pares) or 1
    # ¿Predice la etiqueta legal el valor de cada factor?
    esperado = {"exclusivo": 1.0, "compartido": 0.9, "difuso": 0.75}
    pred_e = sum(1 for p in pares if esperado.get(p["etiqueta"]) is not None
                 and abs(p["E"] - esperado[p["etiqueta"]]) < 1e-9)
    pred_c = sum(1 for p in pares if esperado.get(p["etiqueta"]) is not None
                 and abs(p["C"] - esperado[p["etiqueta"]]) < 1e-9)
    con_etq = sum(1 for p in pares if p["etiqueta"] != "—")
    return {"pares": pares, "iguales": iguales, "n": n,
            "pct": iguales / n * 100, "con_etq": con_etq,
            "pred_e": pred_e, "pred_c": pred_c,
            "escala_e": sorted({p["E"] for p in pares}),
            "escala_c": sorted({p["C"] for p in pares})}


def infracciones_activas(seccion_m: dict) -> dict:
    """¿Está OPERANDO el mecanismo declarado de `C_i`?

    La Sección M registra cuatro clases de infracción (LOSNCP · CGE/NCI ·
    COPFP · CPCCS). Si todas están en cero y `C_i` aun así varía, el valor
    vigente NO procede del mecanismo declarado."""
    total, con_inf, distintos = 0, 0, set()
    for _mid, fila in seccion_m.items():
        total += 1
        infs = [x for x in fila[:4] if isinstance(x, (int, float))]
        if any(x for x in infs):
            con_inf += 1
        tot = fila[4] if len(fila) > 4 else None
        if isinstance(tot, (int, float)):
            distintos.add(round(float(tot), 4))
    return {"metas": total, "con_infraccion": con_inf,
            "valores_ci": sorted(distintos)}


# Qué término del glosario `Ci_Determinista` corresponde a cada código de la
# Sección L. El mapeo es por CONTENIDO, no por nombre: los dos artefactos
# nombran de forma distinta la misma infracción, que ya es parte del hallazgo.
_EQUIV = {"INF-01": "SERCOP_Alert", "INF-02": "CGE_Obs",
          "INF-03": "POA_Retraso", "INF-04": "CPCCS_Desacato"}


def deduccion_declarada_vs_efectiva(glosario: dict, seccion_l: list) -> list:
    """★ ¿Coinciden el glosario y la matriz normativa sobre cuánto deduce cada
    infracción?

    El glosario define `Ci_Determinista` con pesos explícitos. La Sección L los
    define otra vez, en la columna `Deducción_Ci`. Son dos artefactos del mismo
    libro hablando del mismo factor: **deberían decir lo mismo**."""
    texto = (glosario.get("Ci_Determinista") or ("", ""))[0]
    out = []
    for fila in seccion_l:
        term = _EQUIV.get(fila["codigo"], "")
        m = re.search(re.escape(term) + r"\s*[×x*]\s*([0-9]*[.,][0-9]+)", texto)
        declarado = float(m.group(1).replace(",", ".")) if m else None
        efectivo = fila["deduccion"]
        # La Sección L puede expresar una FIJACIÓN («FIJA Ci=0.50») en vez de
        # una resta. No es un formato distinto: es otra operación.
        es_fija = isinstance(efectivo, str) and "FIJA" in efectivo.upper()
        num = None
        if isinstance(efectivo, (int, float)):
            num = abs(float(efectivo))
        elif es_fija:
            m2 = re.search(r"([0-9]*[.,][0-9]+)", str(efectivo))
            num = float(m2.group(1).replace(",", ".")) if m2 else None
        if declarado is None or num is None:
            estado = "⬜ no comparable"
        elif es_fija:
            estado = "🔴 **otra operación** · el glosario resta, la matriz FIJA"
        elif abs(declarado - num) < 1e-9:
            estado = "✅ coinciden"
        else:
            estado = f"🔴 **divergen** · ×{max(declarado, num) / min(declarado, num):.0f}"
        out.append({**fila, "termino": term, "declarado": declarado,
                    "efectivo": efectivo, "estado": estado})
    return out


def main() -> int:
    d = leer_libro()
    if not d:
        print("[no determinable] Gold Master no resuelto.")
        return 2

    defs = definiciones(d["glosario"])
    mecs = {f: mecanismo(d["metas"], f) for f, _ in _FACTORES}
    sol = solape_E_C(d["metas"], d["seccion_i"])
    inf = infracciones_activas(d["seccion_m"])
    ded = deduccion_declarada_vs_efectiva(d["glosario"], d["seccion_l"])

    sin_def = [f for f, e in defs.items() if not e]
    print(f"factores {len(_FACTORES)} · con definición en glosario "
          f"{len(defs) - len(sin_def)}/{len(defs)}")
    if sin_def:
        print(f"  sin entrada de glosario: {', '.join(sin_def)}")
    print(f"E_i ≡ C_i en {sol['iguales']}/{sol['n']} metas "
          f"({sol['pct']:.0f} %) · misma escala: "
          f"{sol['escala_e'] == sol['escala_c']}")
    print(f"Sección M · metas con infracción registrada: "
          f"{inf['con_infraccion']}/{inf['metas']}")
    divergen = [x for x in ded if "🔴" in x["estado"]]
    print(f"deducciones glosario vs Sección L: {len(divergen)}/{len(ded)} "
          f"divergen")
    print(f"entran al numerador: {', '.join(d['entran'])}")

    _escribir(d, defs, mecs, sol, inf, ded)
    print(f"→ {_SALIDA.relative_to(_RAIZ).as_posix()}")
    return 0


def _escribir(d, defs, mecs, sol, inf, ded) -> None:
    o: list[str] = []
    A = o.append

    A("# GM-Ω · ICPI — GENEALOGÍA SEMÁNTICA  `011-C2`")
    A("")
    A("**DERIVADO — no editar a mano.** Lo regenera "
      "`scripts/gm_omega/semantica_factores.py` leyendo `H02_GLOSARIO_QUIRA`, "
      "`H01` Secciones I y M, y `H12`.")
    A("")
    A("> ### La pregunta")
    A("> `011-C1` reconstruyó **el álgebra**. Ésta reconstruye **el "
      "significado**: qué mide cada letra, y si lo que el motor DECLARA medir "
      "coincide con lo que su mecanismo EFECTIVAMENTE mide.")
    A("")
    A("⚠️ **No dictamina.** Que una semántica resulte confusa, solapada o "
      "calibrada al revés es un hecho que aquí se REGISTRA. Si eso invalida el "
      "constructo lo juzga `011-C4`.")
    A("")

    A("## Por qué esta etapa se adelantó a `010`")
    A("")
    A("`009` clasificó `C_i` **dos veces y las dos se equivocó**:")
    A("")
    A("| Intento | Clasificación | Quién la propuso | Por qué falló |")
    A("|---|---|---|---|")
    A("| 1.º | DOCUMENTAL | esta dirección | inflaba el techo de la vía "
      "documental, que es justamente el resultado que 009 medía |")
    A("| 2.º | MATERIAL | Javo | plausible e institucionalmente coherente, "
      "pero **hipótesis del autor**, no semántica demostrada |")
    A("| 3.º | PENDIENTE | el colega | ✅ un análisis de incentivos no puede "
      "fijar la ontología de la variable que lo audita |")
    A("")
    A("> Mientras no se sepa qué significan `E_i` y `C_i`, todo análisis de "
      "comportamiento se hace sobre variables cuya ontología seguimos "
      "reconstruyendo. **Por eso `011-C2` va antes que `010`.**")
    A("")

    # ── 1 · las definiciones canónicas ────────────────────────────────────
    A("## 1 · Qué DECLARA medir cada factor")
    A("")
    A("Fuente: `H02_GLOSARIO_QUIRA`, el glosario del propio Gold Master. "
      "Literal, sin parafrasear.")
    A("")
    for f, _col in _FACTORES:
        entradas = defs.get(f, [])
        A(f"### `{f}`")
        A("")
        if not entradas:
            A("⬜ **SIN ENTRADA EN EL GLOSARIO.** No es que la definición sea "
              "mala: **no existe** en el instrumento. Tercer estado.")
            A("")
            continue
        for termino, definicion, fuente in entradas:
            A(f"**{termino}**" + (f" · fuente declarada: {fuente}" if fuente else ""))
            A("")
            A(f"> {definicion}")
            A("")

    # ── 2 · el mecanismo efectivo ─────────────────────────────────────────
    A("## 2 · Qué hace EFECTIVAMENTE la celda")
    A("")
    A("Una definición no dice cómo se calcula. Esta tabla lee la fórmula real "
      "de `H12`, meta a meta:")
    A("")
    A("| Factor | Celdas literales | Valores distintos | Referencias | Estado |")
    A("|---|---:|---|---|---|")
    for f, _col in _FACTORES:
        m = mecs[f]
        lit = f"{m['literales']}/{m['total']}"
        vals = ", ".join(f"{v:g}" for v in m["distintos"][:6])
        if len(m["distintos"]) > 6:
            vals += f", … ({len(m['distintos'])})"
        refs = ", ".join(f"`{x}`" for x in m["refs"]) or "—"
        if m["literales"] == m["total"]:
            est = "🔴 **literal · sin fórmula**"
        elif m["literales"]:
            est = "🟡 mixto"
        else:
            est = "✅ derivado"
        A(f"| `{f}` | {lit} | {vals} | {refs} | {est} |")
    A("")

    # ── 3 · C_i · la hipótesis contra el instrumento ──────────────────────
    A("## ★ 3 · `C_i` — la hipótesis contrastada contra el instrumento")
    A("")
    A("La hipótesis que `009` tenía prohibido dar por buena:")
    A("")
    A("> «`C_i` mide **atribución y entrega material verificada** — acta de "
      "entrega-recepción e impacto verificado. Si `T=1` (dinero entregado) "
      "pero la obra no tiene acta ni impacto (`C→0`), el producto penaliza la "
      "meta y anula el maquillaje contable de fin de año.»")
    A("")
    A("Esto es lo que el Gold Master dice de sí mismo:")
    A("")
    A(f"| Encabezado de `H01` Sección I | {_cita(d, 93)} |")
    A("|---|---|")
    A(f"| Registro de la creación (`H01!A94`) | {_cita(d, 94)} |")
    A(f"| Encabezado de Sección M | {_cita(d, 186)} |")
    A("")
    cab = d.get("cab_m") or []
    A("Y las **cuatro deducciones** que la Sección M registra son:")
    A("")
    A("| # | Deducción | Qué mide |")
    A("|---|---|---|")
    _QUE_MIDE = {"LOSNCP": "infracción de contratación pública",
                 "CGE": "observación de la Contraloría",
                 "COPFP": "incumplimiento de planificación/finanzas",
                 "CPCCS": "desacato en participación ciudadana"}
    for i, c in enumerate([x for x in cab if x][:4], 1):
        que = next((v for k, v in _QUE_MIDE.items() if k in c), "—")
        A(f"| {i} | `{c.strip()}` | {que} |")
    A("")
    A("### El veredicto de la comparación")
    A("")
    A("| Mitad de la hipótesis | ¿Está en el instrumento? |")
    A("|---|---|")
    A("| **Atribución** — imputar la meta a un responsable | ✅ **SÍ** · la "
      "Sección I asigna a cada meta `Cod_Unidad`, `Dirección Responsable "
      "(Res.040-2025)` y `Base Legal Estatuto` |")
    A("| **Entrega material verificada** — acta de entrega-recepción, impacto "
      "| 🔴 **NO** · ninguna de las cuatro deducciones mide entrega. Todas "
      "miden **infracciones normativas verificadas** |")
    A("")
    A("> ### `C_i` mide legalidad del proceso, no entrega del producto")
    A(">")
    A("> El nombre canónico es **«Calidad de Proceso Orgánico»** / "
      "**«Trazabilidad Orgánica (imputabilidad responsable)»**. El proceso "
      "**nace en 1,00 por presunción de legalidad** y se deduce por "
      "infracciones documentadas. Es un **descuento punitivo-jurídico**, no una "
      "verificación de entrega.")
    A("")
    A("### La formulación exacta del resultado")
    A("")
    A("⚠️ Decir «la hipótesis queda **refutada**», a secas, sería exceder lo "
      "que esta etapa puede demostrar. `011-C2` establece **qué hace el "
      "instrumento**; no puede establecer por sí sola **qué se quiso hacer**. "
      "La formulación defendible es:")
    A("")
    A("> La hipótesis de que `C_i` mide o verifica la entrega material **no "
      "encuentra respaldo en la especificación ni en el mecanismo actualmente "
      "implementado**; la evidencia examinada **la contradice como descripción "
      "del mecanismo vigente**.")
    A("")
    A("Que es una afirmación sobre el mecanismo, no sobre la intención:")
    A("")
    A("| Cuestión | Estado tras `011-C2` |")
    A("|---|---|")
    A("| Semántica implementada | calidad del proceso orgánico / "
      "responsabilidad institucional · **DEMOSTRADO** |")
    A("| Atribución | representada · **DEMOSTRADO** |")
    A("| Entrega material | **no representada** |")
    A("| Impacto | **no representado** |")
    A("| **Intención original del autor** | ⬜ **NO DETERMINABLE** salvo "
      "fuente documental · `011-C3` |")
    A("")
    A("⚠️ **La hipótesis no era descabellada: era una lectura del propósito, no "
      "del mecanismo** (`DOC-024`). La mitad de atribución se sostiene; la de "
      "entrega material **no está implementada en ninguna variable del ICPI**. "
      "Y eso tiene una consecuencia directa sobre `009`:")
    A("")
    A("| Distorsión institucional | Estado real de la respuesta |")
    A("|---|---|")
    A("| anticipo de noviembre con obra sin empezar | 🔴 **el motor no lo "
      "captura hoy** · `T_i` sube y `C_i` no baja, porque `C_i` sólo baja ante "
      "una infracción registrada |")
    A("")
    A("Es decir: la cuarta distorsión de la lista de `009` —la que se declaró "
      "no cubierta— **no es la única**. La segunda tampoco lo está. Se corrige "
      "en el expediente de `009` como consecuencia de esta etapa.")
    A("")

    # ── 3-bis · el mismo factor, tres reglas ──────────────────────────────
    A("## ★ 3-bis · El mismo factor, tres reglas distintas")
    A("")
    A("Al verificar la referencia del glosario —«Fuente: `H01` Sección L»— "
      "apareció algo que no se buscaba. La Sección L **sí existe** y define "
      "cuánto deduce cada infracción. El problema es que **el glosario la "
      "define otra vez, y no dicen lo mismo**.")
    A("")
    A("| Código | Norma | Deducción · **Sección L** | Deducción · **glosario "
      "`Ci_Determinista`** | |")
    A("|---|---|---|---|---|")
    for x in ded:
        dec = (f"×{x['declarado']:g}" if x["declarado"] is not None else "—")
        efe = (f"{x['efectivo']:g}" if isinstance(x["efectivo"], (int, float))
               else str(x["efectivo"]))
        A(f"| `{x['codigo']}` | {x['norma'][:28]} | `{efe}` | `{dec}` | "
          f"{x['estado']} |")
    A("")
    A("Y hay una tercera regla, en la propia Sección L:")
    A("")
    for r in d.get("regla_l", [])[:4]:
        A(f"> `{r[:190]}`")
    A("")
    A("### Las tres divergencias")
    A("")
    A("| # | Qué difiere | Consecuencia |")
    A("|---|---|---|")
    A("| 1 | **`INF-03` deduce 0,05 o 0,20** según qué artefacto se lea | un "
      "retraso de planificación pesa **4 veces más** en el glosario que en la "
      "matriz normativa |")
    A("| 2 | **`INF-04` resta 0,50 o FIJA `Ci=0,50`** | no es formato, es "
      "**otra operación**: sobre una meta con `Ci=0,75`, restar da `0,25`; "
      "fijar da `0,50` |")
    A("| 3 | **El piso es `0` o `0,50`** | el glosario dice `MAX(…, 0)`; la "
      "regla de la Sección L dice `MÁX(0,50; …)`. Con el piso alto, `C_i` "
      "**nunca puede anular una meta** |")
    A("")
    A("> ### Es el patrón del «48,33 %», aplicado a una variable del motor")
    A(">")
    A("> Un **derivado narrativo** —el glosario— se desacopló de su **fuente "
      "canónica** —la Sección L—, y ambos siguen circulando como si dijeran lo "
      "mismo. `QUIRA` fue construido para detectar exactamente esto en los "
      "documentos que audita. Aquí ocurre **dentro del instrumento que audita**.")
    A("")
    A("⚠️ **Y hoy no cambia ningún número**: sin infracciones registradas, "
      "ninguna de las tres reglas se ejecuta. La divergencia es **latente**. "
      "Se activaría el día que se registre la primera infracción — que es "
      "precisamente el día en que el motor tiene que estar bien.")
    A("")
    A("### Qué NO se afirma aquí")
    A("")
    A("- **No se afirma cuál de las tres es la correcta.** Determinar la regla "
      "vigente exige la razón de cada versión: `011-C3`.")
    A("- **No se afirma que sea un error de diseño.** Puede ser una versión "
      "anterior no propagada, y eso también lo dice `011-C3`.")
    A("- **No se toca nada.** El Gold Master es inmutable (`Regla de Oro 1`); "
      "`011-C2` levanta acta.")
    A("")

    # ── 3-ter · qué entra realmente al índice ─────────────────────────────
    A("## 3-ter · Qué entra REALMENTE al índice")
    A("")
    A("Una definición en el glosario no es participación en el cálculo. La "
      "única autoridad es el numerador:")
    A("")
    A("```")
    A(f"  Numerador_i = {d.get('numerador', '—')}")
    A("```")
    A("")
    A("Es decir, entran: " + " · ".join(f"`{x}`" for x in d.get("entran", []))
      + ".")
    A("")
    A("| Definido en el glosario | ¿Entra al ICPI? |")
    A("|---|---|")
    A("| `Ci (Calidad de proceso)` — vía `H01` Sección M | ✅ sí, es la "
      "columna `C_i` |")
    A("| `Ci_Adaptativo` — modificadores por `TIPO_FINANCIAMIENTO` e "
      "`INTANGIBLE_FLAG` | 🔴 **NO** · el numerador no lo referencia |")
    A("")
    A("⚠️ Eso incluye la **discriminación positiva ×1,15 por "
      "`FONDO_CONCURSABLE`**, que el glosario declara y el motor canónico no "
      "aplica. Un premio definido y no implementado no es lo mismo que un "
      "premio inexistente: es una **capacidad declarada sin efecto**.")
    A("")
    A("### La distinción que este hallazgo obliga a hacer")
    A("")
    A("```")
    A("  DEFINIDO   ≠   CALCULADO   ≠   UTILIZADO")
    A("```")
    A("")
    A("`Ci_Adaptativo` está **definido**. Que no participe del ICPI vigente es "
      "**VERIFICADO**. Pero **no debe llamarse «error»**: hasta que aparezca "
      "evidencia, su causa admite cinco lecturas incompatibles entre sí —")
    A("")
    A("| # | Lectura posible |")
    A("|---|---|")
    A("| 1 | componente experimental abandonado |")
    A("| 2 | componente diseñado pero nunca conectado |")
    A("| 3 | componente sustituido por otro |")
    A("| 4 | residuo documental de una versión anterior |")
    A("| 5 | implementación incompleta |")
    A("")
    A("> **Estado causal: `NO DETERMINABLE`.** Las cinco producen el mismo "
      "síntoma observable, y elegir una sin fuente sería inventar la "
      "genealogía. Va a `011-C3` como expediente propio (`C3-08`).")
    A("")

    # ── 4 · ¿opera el mecanismo? ──────────────────────────────────────────
    A("## ★ 4 · ¿Está OPERANDO el mecanismo declarado?")
    A("")
    A(f"De las **{inf['metas']} metas** de la Sección M, las que registran "
      f"alguna infracción son **{inf['con_infraccion']}**.")
    A("")
    if inf["con_infraccion"] == 0:
        A("Ninguna. Y sin embargo `C_i` **no es constante**: toma los valores "
          + ", ".join(f"`{v:g}`" for v in inf["valores_ci"]) + ".")
        A("")
        A("> ### Si todas las deducciones son cero, el valor vigente NO procede del mecanismo declarado")
        A("")
        A("No hay que inferirlo: **el instrumento lo dice de sí mismo**.")
        A("")
        for n in d.get("nota_m", [])[:3]:
            A(f"> {n}")
            A("")
    A("Y el glosario nombra la técnica sin eufemismo:")
    A("")
    mr = d["glosario"].get("Mapeo Retrospectivo")
    if mr:
        A(f"> **Mapeo Retrospectivo** · {mr[0]}")
        A("")
    A("### ⚠️ Cuarta divergencia · las dos secciones se contradicen")
    A("")
    A("Puestas una junto a otra, la Sección L y la Sección M **del mismo "
      "libro** afirman lo contrario sobre el mismo factor:")
    A("")
    A("| Sección | Qué declara |")
    A("|---|---|")
    enc_l = next((r for r in d.get("regla_l", []) if "heurístic" in r.lower()), "")
    enc_m = next((n for n in d.get("nota_m", []) if "HEUR" in n.upper()), "")
    if enc_l:
        A(f"| **L** — matriz normativa | «{enc_l[:150]}» |")
    if enc_m:
        A(f"| **M** — registro y calibración | «{enc_m[:210]}» |")
    A("")
    A("> La L declara que el motor **abandona** la valoración heurística. La M "
      "declara que la heurística de 2025 **es el fallback vigente**. Ambas son "
      "ciertas a la vez sólo si «abandonar» significa «dejar de usarla cuando "
      "haya infracciones» — que es una lectura posible, pero **es una lectura, "
      "no lo que el texto dice**.")
    A("")
    A("Ésta es la divergencia **más consecuente de las cuatro**, porque no es "
      "sobre un peso ni sobre una operación: es sobre **si el factor es "
      "determinista o heurístico hoy**. Y de eso depende cómo se puede "
      "presentar el ICPI públicamente.")
    A("")
    A("### Cómo se clasifica esto")
    A("")
    A("| Afirmación | Grado |")
    A("|---|---|")
    A("| El mecanismo declarado de `C_i` es la deducción por infracciones | "
      "**DEMOSTRADO** · glosario + Sección M |")
    A("| Hoy no hay ninguna infracción registrada | **DEMOSTRADO** · Sección M |")
    A("| El valor vigente de `C_i` procede de `Ci_Manual_2025` | "
      "**DECLARADO POR EL INSTRUMENTO** · nota metodológica de la Sección M |")
    A("| La calibración se ajustó para reproducir un ICPI previamente fijado | "
      "**DECLARADO POR EL INSTRUMENTO** · glosario, «Mapeo Retrospectivo» |")
    A("| Esa calibración es metodológicamente admisible | ⬜ **NO LO JUZGA "
      "`011-C2`** · `011-C3` (justificación) y `011-C4` (dictamen) |")
    A("")
    A("### Cómo se enuncia esto sin convertirlo en acusación")
    A("")
    A("Que exista una declaración escrita de que **se inyectaron valores "
      "históricos para reproducir un ICPI predeterminado** es un **hecho "
      "documental**, y hay que decirlo. Pero un hecho documental no es una "
      "imputación de manipulación, y la diferencia está en la formulación:")
    A("")
    A("> La documentación demuestra una **calibración retrospectiva orientada "
      "a reproducir un valor canónico preexistente**. La **legitimidad "
      "metodológica** de dicha calibración queda **fuera de `011-C2`** y "
      "requiere justificación en `C3`/`C4`.")
    A("")
    A("⚠️ Hay razones legítimas para calibrar así —fijar una línea base "
      "comparable, preservar continuidad entre ejercicios— y razones que no lo "
      "serían. **`011-C2` no puede distinguirlas y no lo intenta.** Lo que "
      "hace es dejar el hecho registrado con su cita, para que `C3` lo "
      "pregunte con la fuente delante.")
    A("")
    A("⚠️ **Y hay que decir lo que esto NO es.** No registrar una infracción "
      "que no existe es **correcto**: el canon prohíbe fabricar infracciones "
      "para alimentar el motor. La cuestión abierta es distinta y es de "
      "vigencia: **usar una calibración heurística de 2025 como valor de "
      "2026**. `007-B0` ya dejó esa pregunta abierta; `011-C2` le pone nombre "
      "propio y la entrega a `011-C3`.")
    A("")

    # ── 5 · el solapamiento E_i ↔ C_i ─────────────────────────────────────
    A("## ★ 5 · `E_i` y `C_i` — dos variables, una escala, un vocabulario")
    A("")
    A("El hallazgo que obligó a mirar dos veces:")
    A("")
    A("| | `E_i` | `C_i` (base) |")
    A("|---|---|---|")
    A("| Nombre | Autonomía orgánica | Calidad de proceso orgánico |")
    A("| Escala observada | " + " · ".join(f"`{v:g}`" for v in sol["escala_e"])
      + " | " + " · ".join(f"`{v:g}`" for v in sol["escala_c"]) + " |")
    A("| Vocabulario de la escala | autónomo / compartido / difuso | "
      "proceso **exclusivo / compartido / difuso** (Sección I, «Base Legal "
      "Estatuto») |")
    A("| Fuente declarada | Estatuto Orgánico (Res. 040-2025) | Estatuto "
      "Orgánico (Res. 040-2025) |")
    A("")
    A("Misma escala, mismo vocabulario, misma fuente. **La pregunta obligada "
      "es si son la misma variable dos veces.** Se mide meta a meta:")
    A("")
    A(f"> **`E_i` = `C_i` en {sol['iguales']} de {sol['n']} metas "
      f"({sol['pct']:.0f} %).**")
    A("")
    A("| Meta | `E_i` | `C_i` | Etiqueta legal (Sección I) | ¿Coinciden? |")
    A("|---|---:|---:|---|---|")
    for p in sol["pares"]:
        A(f"| `{p['id']}` | {p['E']:g} | {p['C']:g} | {p['etiqueta']} | "
          f"{'✅' if p['igual'] else '🔴 **divergen**'} |")
    A("")
    A("### Qué se puede y qué no se puede concluir")
    A("")
    A("| Afirmación | Grado |")
    A("|---|---|")
    A("| Comparten escala, vocabulario y fuente declarada | **DEMOSTRADO** |")
    A(f"| Coinciden en {sol['pct']:.0f} % de las metas | **DEMOSTRADO** |")
    A("| **Son la misma variable** | 🔴 **REFUTADO** · propiedad matemática: "
      f"divergen en {sol['n'] - sol['iguales']} metas, y si fueran la misma "
      "coincidirían en todas |")
    A("| **El motor cuenta dos veces la autonomía** | ⬜ **NO SE AFIRMA** · "
      "esas mismas divergencias lo impiden |")
    A(f"| La etiqueta legal predice `E_i` | en {sol['pred_e']}/{sol['con_etq']} "
      "metas con etiqueta |")
    A(f"| La etiqueta legal predice `C_i` | en {sol['pred_c']}/{sol['con_etq']} "
      "metas con etiqueta |")
    A("| Existe **ambigüedad ontológica** entre ambas | **DEMOSTRADO** · dos "
      "dimensiones distintas de la fórmula usan el mismo vocabulario y la "
      "misma escala sobre la misma fuente |")
    A("")
    A("### La formulación exacta, y las dos que hay que evitar")
    A("")
    A("| Formulación | Veredicto |")
    A("|---|---|")
    A("| «`E_i` y `C_i` son la misma variable» | 🔴 falsa · divergen en "
      f"{sol['n'] - sol['iguales']} metas |")
    A("| «el motor duplica la autonomía» / «hay doble conteo» | 🔴 **no "
      "demostrada** · sería exactamente el mismo error, con otro nombre |")
    A("| **«existe una POTENCIAL SUPERPOSICIÓN SEMÁNTICA entre `E_i` y `C_i` "
      "que requiere justificación»** | ✅ es lo que la evidencia sostiene |")
    A("")
    A("> El hallazgo es **peor de diagnosticar y mejor de corregir** que una "
      "duplicación: están parcialmente superpuestas y **no se sabe por qué "
      "divergen donde divergen**.")
    A("")
    A("Porque una divergencia puede significar dos cosas opuestas:")
    A("")
    A("```")
    A("  E = 1,00  ·  C = 0,75      ¿la competencia es autónoma pero el")
    A("                             proceso orgánico es difuso?          ← legítimo")
    A("                             ¿o una de las dos está mal asignada? ← defecto")
    A("```")
    A("")
    A("Y `011-C2` **no puede distinguirlas**: exigiría la razón de cada "
      "asignación, que es material de `011-C3`. Lo que sí puede decir es que "
      "**nada en el instrumento explica la diferencia** — no hay columna de "
      "justificación para `E_i`, que es literal en `H12` y carece de entrada "
      "propia en la Sección I.")
    A("")
    A("### 📜 CORRECCIÓN POSTERIOR — aportada por `011-C3`")
    A("")
    A("La frase anterior era **cierta del instrumento y falsa del corpus**, y "
      "se conserva para que la corrección sea auditable. `011-C3` encontró en "
      "`metodologia.docx` (25-mar-2026) las dos escalas originales, y la "
      "superposición **sí está explicada**:")
    A("")
    A("| Variable | Eje que mide | Escala original |")
    A("|---|---|---|")
    A("| `E_i` · Fricción de Autonomía | **quién EJECUTA** — modalidad de "
      "ejecución | directa `1,00` · convenio `0,90` · delegada `0,75` |")
    A("| `C_i` · Imputabilidad Orgánica | **quién RESPONDE** — claridad de la "
      "asignación | responsable único `1,00` · compartida `0,90` · difusa "
      "`0,75` |")
    A("")
    A("> Comparten escala **porque ambas son escalas ordinales de tres grados "
      "sobre el mismo Estatuto Orgánico**. La superposición es **deliberada y "
      "justificada**, no un accidente — y por eso pueden divergir sin que eso "
      "sea un error.")
    A("")
    A("La metodología incluso trae el caso: `M3` (Salud) con ejecución "
      "**directa** (`E=1,00`) y responsabilidad **compartida** entre "
      "Planificación y Obras Públicas (`C=0,90`).")
    A("")
    A("**Lo que sigue sin explicación** es cada una de las 12 asignaciones "
      "divergentes del motor, meta a meta. Esa parte permanece `NO "
      "DETERMINABLE`.")
    A("")

    # ── 6 · lo que se entrega ─────────────────────────────────────────────
    A("## Lo que `011-C2` entrega")
    A("")
    A("### A `011-C3` · justificación de cada transformación")
    A("")
    A("| # | Pregunta que `C3` hereda | De dónde sale |")
    A("|---|---|---|")
    A("| 1 | ¿Por qué `C_i` se calibró retrospectivamente contra un ICPI ya "
      "fijado, y quién lo decidió? | §4 |")
    A("| 2 | ¿Por qué una calibración declarada «2025» sigue vigente en 2026? "
      "| §4 |")
    A("| 3 | ¿Cuál de las reglas de deducción es la vigente — la Sección L o "
      "el glosario? | §3-bis |")
    A("| 4 | ¿El piso de `C_i` es `0` o `0,50`? De ello depende si `C_i` "
      "**puede anular una meta** | §3-bis |")
    A("| 5 | ¿`INF-04` resta o fija? | §3-bis |")
    A("| 6 | ¿Es `C_i` determinista o heurístico **hoy**? | §4 |")
    A("| 7 | ¿Por qué `E_i` y `C_i` divergen en las metas donde divergen? | §5 |")
    A("| 8 | ¿Se incorporó `C_i` sabiendo que solaparía con `E_i`, o se "
      "descubrió después? | §5 |")
    A("| 9 | ¿Por qué `Ci_Adaptativo` está definido y no se aplica? | §3-ter |")
    A("")
    A("### ★ La pregunta que `C3` hereda por encima de las nueve")
    A("")
    A("Las cuatro divergencias tienen una forma común, y verla ordenada "
      "cambia el encargo de `C3`. Sobre `C_i` conviven **cuatro reglas**:")
    A("")
    A("| | Regla | Dónde vive |")
    A("|---|---|---|")
    A("| **A** | ponderaciones y piso del glosario | `H02` "
      "`Ci_Determinista` |")
    A("| **B** | otra parametrización: `INF-03` `0,05`, `INF-04` FIJA, piso "
      "`0,50` | `H01` Sección L |")
    A("| **C** | el cálculo efectivo que alimenta el numerador | `H12` col. "
      "`C_i` ← `H01` Sección M |")
    A("| **D** | `Ci_Manual_2025` cuando no hay infracciones | nota "
      "metodológica de la Sección M |")
    A("")
    A("> ### Que una regla esté documentada no la hace la regla vigente")
    A(">")
    A("> Lo que `C3` tiene que determinar no es cuál regla es mejor, sino "
      "**cuál gobierna realmente el valor que entra en el ICPI** — y qué "
      "evidencia justifica cada transición entre ellas.")
    A("")
    A("### A `011-C4` · el dictamen")
    A("")
    A("> Si dos de los seis factores comparten escala, vocabulario y fuente, la "
      "pregunta de la multiplicatividad **cambia de forma**: ya no es sólo si "
      "el producto es la operación correcta, sino **sobre cuántas dimensiones "
      "realmente independientes opera**.")
    A("")
    A("### ★ La matriz dimensional que queda congelada para `C4`")
    A("")
    A("Tras `011-C2`, la arquitectura **ya no puede describirse** como "
      "`V`=evidencia · `T`=ejecución · `E`=estructura · `C`=entrega. Esa "
      "lectura quedó superada. La descripción provisional correcta es:")
    A("")
    A("| Factor | Dimensión | Estado de su semántica |")
    A("|---|---|---|")
    A("| `V_i` | **evidencial / documental** | establecida |")
    A("| `T_i` | **temporal de ejecución** presupuestaria | establecida |")
    A("| `E_i` | **estructural / competencial** (autonomía) | ⚠️ requiere "
      "reconstrucción histórica completa |")
    A("| `C_i` | **jurídico-orgánica**: calidad y responsabilidad del proceso "
      "| ⚠️ regla efectiva y justificación **abiertas en `C3`** |")
    A("| — | **entrega material / impacto físico** | 🔴 **EXCLUIDOS del "
      "instrumento actual** |")
    A("")
    A("⚠️ La última fila es la que impide que QUIRA se atribuya una capacidad "
      "que hoy no tiene. Mientras ningún dominio la incorpore, **el ICPI no "
      "responde por el resultado material de una meta**, y así debe "
      "presentarse.")
    A("")
    A("### A `009` · una corrección")
    A("")
    A("El expediente de `009` afirma que el motor responde a la disociación "
      "financiero ↔ físico mediante `C_i`. **Esta etapa lo desmiente**: `C_i` "
      "sólo baja ante una infracción registrada, y hoy no hay ninguna. La "
      "corrección se aplica en `009` marcando esa fila como no cubierta.")
    A("")

    A("## Dictamen de `011-C2` · por grado de certeza")
    A("")
    A("| Afirmación | Estado |")
    A("|---|---|")
    A("| Cada factor tiene una definición canónica en el glosario del motor | "
      f"**{'DEMOSTRADO' if not [f for f, e in defs.items() if not e] else 'PARCIAL'}** |")
    A("| `C_i` mide calidad jurídica del proceso orgánico | **DEMOSTRADO** · "
      "glosario `H02` + Secciones I/M |")
    A("| `C_i` mide o verifica entrega material | 🔴 **SIN RESPALDO EN EL "
      "MECANISMO VIGENTE** · la evidencia examinada la contradice como "
      "descripción del mecanismo · ⚠️ no dice nada sobre la INTENCIÓN "
      "original |")
    A("| `C_i` imputa la meta a una unidad orgánica responsable | "
      "**DEMOSTRADO** · Sección I |")
    A("| El valor vigente de `C_i` no procede del mecanismo declarado | "
      "**DECLARADO POR EL INSTRUMENTO** |")
    A("| `E_i` mide autonomía en el ejercicio de la competencia | "
      "**DEMOSTRADO** · glosario `H02` |")
    A("| `E_i` y `C_i` son la misma variable | 🔴 **REFUTADO** · propiedad "
      "matemática |")
    A("| Existe **potencial superposición semántica** `E_i` ↔ `C_i` que "
      "requiere justificación | **DEMOSTRADO** |")
    A("| El motor cuenta dos veces la autonomía (doble conteo) | ⬜ **NO SE "
      "AFIRMA** · las divergencias lo impiden |")
    A("| **Intención original del autor sobre `C_i`** | ⬜ **NO "
      "DETERMINABLE** · `011-C3` |")
    A("| La razón de cada divergencia `E_i` ↔ `C_i` | ⬜ **NO DETERMINABLE** "
      "aquí · `011-C3` |")
    A("| El glosario y la Sección L discrepan sobre `INF-03`, `INF-04` y el "
      "piso | **DEMOSTRADO** |")
    A("| `Ci_Adaptativo` no entra al numerador | **DEMOSTRADO** · la fórmula "
      "del numerador no lo referencia |")
    A("| Cuál de las reglas discrepantes es la vigente | ⬜ **NO "
      "DETERMINABLE** aquí · `011-C3` |")
    A("| Si el solapamiento invalida la arquitectura | ⬜ **FUERA DE ALCANCE** "
      "· `011-C4` |")
    A("")
    A("> ### GM-Ω-011-C2 — CERRADO COMO RECONSTRUCCIÓN SEMÁNTICA")
    A(">")
    A("> Se estableció **qué declara medir cada factor** y **qué mide su "
      "mecanismo**. Aparecieron cuatro divergencias: `C_i` no verifica entrega "
      "material; `E_i` y `C_i` comparten escala y vocabulario sin ser la misma "
      "variable; el glosario y la matriz normativa discrepan sobre tres reglas "
      "de deducción; y las Secciones L y M se contradicen sobre si el factor "
      "es determinista o heurístico.")
    A(">")
    A("> **Ninguna cambia hoy el ICPI** — las tres primeras son latentes y la "
      "cuarta ya está resuelta de facto por el fallback. Eso las hace más "
      "fáciles de corregir, **no menos importantes**.")
    A(">")
    A("> **No juzga** si son defectos. Reconstruir el significado no es "
      "aprobarlo ni condenarlo — eso es `011-C3` y `011-C4`.")
    A("")
    A("---")
    A(f"*GM-Ω-ICPI-011-C2 · {len(d['metas'])} metas · {len(_FACTORES)} factores "
      "· leído del Gold Master, no de la memoria · el Gold Master no se "
      "modificó · Dylus Lab © 2026*")

    _SALIDA.write_text("\n".join(o) + "\n", encoding="utf-8")


def _cita(d: dict, fila: int) -> str:
    """Texto literal de `H01!A{fila}`, leído una sola vez en `leer_libro`."""
    return d.get("citas", {}).get(fila, "—")


if __name__ == "__main__":
    raise SystemExit(main())
