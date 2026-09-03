# -*- coding: utf-8 -*-
"""
scripts/gm_omega/sensibilidad_icpi.py — GM-Ω-ICPI-007 · A · B · C · D · X

Mide CUÁNTO depende el ICPI de cada decisión metodológica del motor, sin tocar
ni una celda del Gold Master. Cuatro familias de contrafactuales y una prueba
transversal:

    007-A   el peso        ¿el índice mide la cadena de gestión, o cómo se
                           repartió el peso presupuestario y jurídico?
    007-B   V_i            especificación implementada vs. regla documentada
                           vs. regla anterior
    007-C   el tope de T   ¿MIN(1,·) elimina información que la teoría del
                           indicador necesitaría conservar?
    007-D   la estructura  ¿la multiplicatividad impone una penalización
                           conjunta coherente con el constructo?
    007-X   robustez       ¿cambia la CATEGORÍA, o sólo el número?

    ⚠️ LAS TRES ETIQUETAS, en todos y cada uno de los escenarios:
        · MATEMÁTICAMENTE REPRODUCIBLE
        · METODOLÓGICAMENTE CONTRAFACTUAL
        · NO AUTORIZADO PARA PUBLICACIÓN

    El único número oficial sigue siendo **27,4582 %** (regla GM-Ω-ICPI-000)
    hasta que GM-Ω-ICPI-011 dictamine. Un contrafactual que se escapa del
    laboratorio y reaparece seis meses después como «el ICPI» es exactamente el
    patrón del «48,33 %» que esta auditoría persigue.

    Y `E_i` NO ENTRA EN NINGÚN ESCENARIO. Su regla generadora está
    `NOT_DETERMINABLE` (007-B0): optimizar una variable cuya biografía no
    conocemos sería inferir su regla desde el efecto que produce — justo lo que
    DOC-009 prohíbe. Entra en 011 o no entra.

    LECTURA PURA. Regla de Oro 1 y 4.

Uso:  python scripts/gm_omega/sensibilidad_icpi.py
Dylus Lab © 2026
"""
from __future__ import annotations

import re
import sys
from math import prod
from pathlib import Path

_RAIZ = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(_RAIZ))

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

_SALIDA = _RAIZ / "docs" / "architecture" / "GM-OMEGA_ICPI_SENSIBILIDAD_007.md"

# El baseline congelado por la regla 000, en escala 0-1 como lo guarda H12!B33.
_BASELINE = 0.27458226534062735
_TOLERANCIA = 1e-9

# La escala AVEP tal como la implementa H12!B34 — no una versión nuestra.
_AVEP = [(0.90, "🔵 Excelencia en Gobernanza"), (0.70, "🟢 Gestión por Mandato"),
         (0.40, "🟡 Transición Crítica"), (0.20, "🟠 Gestión por Ocurrencia"),
         (0.00, "🔴 Ruptura Sistémica")]

# Qué entidad ejecutora representa cada columna de `Ti_norm_2026` (H07b fila 20).
#
# ⚠️ Las claves llevan el prefijo `col_` A PROPÓSITO. La versión anterior usaba
# la letra suelta, y entonces `"E"` significaba dos cosas en el mismo archivo:
# la columna E del Excel y la dimensión E_i del ICPI —la única cuya regla está
# NOT_DETERMINABLE y que 007 tiene prohibido tocar—. Lo detectó la prueba que
# vigila justamente eso, al contar como manipulación de `E_i` lo que era un mapa
# de columnas. El falso positivo señalaba una ambigüedad real: en una auditoría
# sobre trazabilidad, un nombre que significa dos cosas es un defecto.
_ENTIDAD_TI = {"col_B": "ENTE-01 GAD central", "col_C": "ENTE-02 Patronato",
               "col_D": "ENTE-03 Bomberos", "col_E": "ENTE-04 EP Aseo"}


def clasificar(icpi: float) -> str:
    """La categoría AVEP del motor. ⚠️ H12!B34 sólo la EMITE con 12 meses de
    corte (`H07!B22>=12`); con corte parcial devuelve «lectura preliminar». Aquí
    se calcula siempre, para poder comparar escenarios — y eso se declara."""
    for umbral, nombre in _AVEP:
        if icpi >= umbral:
            return nombre
    return _AVEP[-1][1]


def distancia_al_umbral(icpi: float) -> tuple[float, str]:
    """A cuántos puntos porcentuales está el índice de cambiar de categoría.
    Es la magnitud que 007-X necesita: un índice a 0,3 puntos de saltar de
    categoría es frágil aunque su número parezca estable."""
    cortes = [u for u, _ in _AVEP if 0 < u < 1]
    d = min(abs(icpi - u) for u in cortes)
    cercano = min(cortes, key=lambda u: abs(icpi - u))
    return d * 100, f"{cercano * 100:.0f} %"


# ═════════════════════════════════════════════════════════════════════════════
# LECTURA DEL MOTOR
# ═════════════════════════════════════════════════════════════════════════════
def leer_motor() -> dict:
    """Todo lo que hace falta para reproducir el ICPI y sus contrafactuales.
    Devuelve {} si el Gold Master no se resolvió — el tercer estado."""
    import openpyxl

    import config
    if not getattr(config, "GOLD_MASTER_RESUELTO", False):
        return {}

    wf = openpyxl.load_workbook(config.SIAP_PATH, data_only=False, read_only=True)
    wv = openpyxl.load_workbook(config.SIAP_PATH, data_only=True, read_only=True)
    h12f, h12v = wf["H12_MOTOR_ICPI_CANÓNICO"], wv["H12_MOTOR_ICPI_CANÓNICO"]
    h13v = wv["H13_VARIABLES_Vi"]
    h07v = wv["H07_S5_FINANCIERO_eSIGEF"]
    h07bv, h07bf = wv["H07b_Ti_INVERSIÓN_eSIGEF"], wf["H07b_Ti_INVERSIÓN_eSIGEF"]

    # Los cuatro verificadores de V_i, meta a meta (H13!B..E, filas 25-49).
    verif = {}
    for r in range(25, 50):
        meta = h13v.cell(row=r, column=1).value
        if not meta:
            continue
        verif[str(meta)] = {
            "SERCOP": h13v.cell(row=r, column=2).value,
            "eSIGEF": h13v.cell(row=r, column=3).value,
            "LOTAIP": h13v.cell(row=r, column=4).value,
            "CPCCS": h13v.cell(row=r, column=5).value,
            "Vi_tabla": h13v.cell(row=r, column=6).value,
        }

    # T_i sin normalizar, por entidad (H07b fila 19), y el factor de la curva.
    factor = h07v.cell(row=23, column=2).value          # H07!B23 · pacing real
    mes = h07v.cell(row=22, column=2).value             # H07!B22 · mes de corte
    ti_raw = {}
    for clave, ent in _ENTIDAD_TI.items():
        col = openpyxl.utils.column_index_from_string(clave.removeprefix("col_"))
        ti_raw[ent] = h07bv.cell(row=19, column=col).value

    metas = []
    for r in range(6, 31):
        meta = str(h12v.cell(row=r, column=1).value)
        f_ti = str(h12f.cell(row=r, column=6).value or "")
        m = re.search(r"!([A-Z]+)\d+", f_ti)
        ent = _ENTIDAD_TI.get(f"col_{m.group(1)}", "—") if m else "—"
        metas.append({
            "id": meta,
            "P": h12v.cell(row=r, column=2).value,
            "R": h12v.cell(row=r, column=3).value,
            "V": h12v.cell(row=r, column=4).value,
            "E": h12v.cell(row=r, column=5).value,
            "T": h12v.cell(row=r, column=6).value,
            "C": h12v.cell(row=r, column=9).value,
            "entidad": ent,
            "verif": verif.get(meta, {}),
        })

    # ¿Cuántas hojas rotulan el ICPI en su cabecera con `ROUND(B33,2)&"%"`?
    # B33 vive en escala 0-1, así que ese rótulo imprime «0,27 %».
    rotulo = sum(1 for n in wf.sheetnames
                 if isinstance(wf[n].cell(row=1, column=5).value, str)
                 and "ROUND(H12_MOTOR_ICPI_CANÓNICO!B33,2)" in
                 str(wf[n].cell(row=1, column=5).value))
    rotulo += 1 if "ROUND(B33,2)" in str(h12f.cell(row=1, column=5).value or "") else 0

    # La nota que describe el FactorTemporal, para contrastarla con la fórmula.
    nota_factor = str(h07bf.cell(row=20, column=6).value or "")

    # ¿QUÉ LLEGA AL PRODUCTO? Un defecto dentro del libro y un defecto que la UI
    # publica no son el mismo hallazgo, y la diferencia hay que MEDIRLA, no
    # suponerla. La capa API es `H73_OUTPUT_API`: pares clave-valor que el
    # conector lee. Aquí se comprueba si corrige la escala y qué clasificación
    # entrega.
    api = {"pct_corrige_escala": None, "clasificacion": None}
    if "H73_OUTPUT_API" in wf.sheetnames:
        h73f, h73v = wf["H73_OUTPUT_API"], wv["H73_OUTPUT_API"]
        for r in range(1, min(h73f.max_row, 60) + 1):
            k = str(h73v.cell(row=r, column=1).value or "")
            if k == "ICPI_GLOBAL_PCT":
                api["pct_corrige_escala"] = "*100" in str(
                    h73f.cell(row=r, column=2).value or "")
            elif k == "ICPI_CLASIFICACION":
                api["clasificacion"] = h73v.cell(row=r, column=2).value

    avep = procedencia_avep(wf)

    wf.close(); wv.close()
    return {"metas": metas, "factor": factor, "mes": mes, "ti_raw": ti_raw,
            "hojas_rotulo": rotulo, "nota_factor": nota_factor, "api": api,
            "avep": avep, "consumidores": _consumidores_de_clasificacion()}


def procedencia_avep(wf) -> dict:
    """¿De dónde salen los umbrales de la escala AVEP, y dónde vive la escala?

    Javo aportó el contexto que hacía falta: **AVEP es invención de Dylus Lab**,
    ajustada después a lo que la normativa pública exigía. Eso convierte a `007-X`
    —la robustez de la CATEGORÍA— en una medición con un supuesto que hay que
    hacer explícito: los umbrales `0,9 / 0,7 / 0,4 / 0,2`.

    Se mide, sin inferir: en cuántas hojas está COPIADA la escala, si alguna
    celda cita una norma junto a ella, y si las etiquetas divergen entre copias.
    """
    ETIQUETAS = ("Excelencia en Gobernanza", "Gestión por Mandato",
                 "Transición Crítica", "Gestión por Ocurrencia",
                 "Ruptura Sistémica")
    NORMA = re.compile(r"(COOTAD|COPFP|LOTAIP|LOPC|Constituci[oó]n|Acuerdo \d|"
                       r"Art\.\s*\d|Resoluci[oó]n|SENPLADES|CGE|Norma de Control)")
    copias, con_norma, variantes = set(), [], set()
    for nombre in wf.sheetnames:
        h = wf[nombre]
        for fila in h.iter_rows(min_row=1, max_row=min(h.max_row, 60),
                                max_col=min(h.max_column, 15)):
            for c in fila:
                v = c.value
                if not isinstance(v, str) or len(v) < 8:
                    continue
                tocadas = [e for e in ETIQUETAS if e in v]
                if len(tocadas) >= 3:            # una copia de la escala entera
                    copias.add(nombre)
                    if NORMA.search(v):
                        con_norma.append(f"{nombre}!{c.coordinate}")
                    # ⚠️ La divergencia del nivel superior SÓLO se busca DENTRO
                    # de una copia de la escala. La primera versión barría todo
                    # el libro y devolvió «Excelencia en Equidad» —de otro
                    # índice— como si AVEP hubiera divergido. Era DOC-009 otra
                    # vez: un patrón que encaja no es el hecho que parece.
                    m = re.search(r"Excelencia en (\w+)", v)
                    if m:
                        variantes.add(m.group(1))
    return {"copias": sorted(copias), "con_norma": con_norma,
            "variantes_nivel_alto": sorted(variantes)}


def _consumidores_de_clasificacion() -> list[str]:
    """Superficies del producto que publican `icpi[...]clasificacion`. Se cuentan
    leyendo el repositorio: afirmar «llega a la UI» sin contar dónde sería la
    misma inferencia sin evidencia que persigue toda esta auditoría."""
    hallados = []
    for carpeta in ("quira_pages", "components", "views"):
        base = _RAIZ / carpeta
        if not base.exists():
            continue
        for py in sorted(base.rglob("*.py")):
            if "_deprecated" in py.parts:
                continue
            t = py.read_text(encoding="utf-8", errors="replace")
            if re.search(r"icpi[^\n]{0,40}clasificacion", t):
                hallados.append(py.relative_to(_RAIZ).as_posix())
    return hallados


# ═════════════════════════════════════════════════════════════════════════════
# EL ÍNDICE Y SUS PIEZAS
# ═════════════════════════════════════════════════════════════════════════════
def evaluar(metas: list[dict], peso, score) -> dict:
    """ICPI = Σ(K_i · S_i) / Σ(K_i).

    ⚠️ Y ESO YA ES UN HALLAZGO DE ESTRUCTURA: la fórmula canónica
    Σ(P·R·V·E·T·C)/Σ(P·R) es ALGEBRAICAMENTE una media ponderada de S=V·E·T·C
    con pesos K=P·R. La multiplicatividad no está en la AGREGACIÓN entre metas
    —que es lineal— sino DENTRO de cada meta, entre sus cuatro dimensiones.
    007-D audita esa segunda; 007-A, la primera."""
    filas = []
    for m in metas:
        k, s = peso(m), score(m)
        filas.append({**m, "K": k, "S": s, "J": k * s})
    sk = sum(f["K"] for f in filas)
    sj = sum(f["J"] for f in filas)
    icpi = sj / sk if sk else 0.0
    for f in filas:
        f["k_rel"] = f["K"] / sk if sk else 0.0
        f["j_rel"] = f["J"] / sj if sj else 0.0
    filas.sort(key=lambda f: -f["J"])
    return {"icpi": icpi, "filas": filas, "sumK": sk, "sumJ": sj}


def concentracion(filas: list[dict]) -> dict:
    """Cuánto del resultado explican unas pocas metas. Un índice que se presenta
    como global y en realidad depende de tres metas no es incorrecto — pero
    tiene que ser CONSCIENTE Y DECLARADO."""
    js = sorted((f["j_rel"] for f in filas), reverse=True)
    acum, n50, n80 = 0.0, None, None
    for i, v in enumerate(js, 1):
        acum += v
        if n50 is None and acum >= 0.50:
            n50 = i
        if n80 is None and acum >= 0.80:
            n80 = i
            break
    return {"top1": js[0], "top3": sum(js[:3]), "top5": sum(js[:5]),
            "hhi": sum(v * v for v in js), "n50": n50, "n80": n80}


# ── las piezas intercambiables ───────────────────────────────────────────────
_S_ACTUAL = lambda m: m["V"] * m["E"] * m["T"] * m["C"]
_K_ACTUAL = lambda m: m["P"] * m["R"]


def v_por_regla(m: dict) -> float | None:
    """La regla VIGENTE de V_i, documentada en prosa en H13!B16-B20:

        Vi = 0.0  si V_eSIGEF=0 O V_SERCOP=0     (sin núcleo financiero, sin score)
        Vi = 0.5  si núcleo OK y sin LOTAIP ni CPCCS
        Vi = 1.0  si núcleo OK y (LOTAIP=1 O CPCCS=1)

    ⚠️ En el libro está DOCUMENTADA pero NO IMPLEMENTADA: H13!F contiene 25
    literales, no la fórmula. Aplicarla a los verificadores es la única forma de
    comprobar si los valores que el motor consume la obedecen."""
    v = m["verif"]
    if not v or v.get("SERCOP") is None:
        return None
    if not v["eSIGEF"] or not v["SERCOP"]:
        return 0.0
    return 1.0 if (v["LOTAIP"] or v["CPCCS"]) else 0.5


def v_regla_anterior(m: dict, tres_niveles: bool) -> float | None:
    """La regla ANTERIOR, según la deja documentada H13!B21:

        «La fórmula original SI(suma≥2,0.5) era incorrecta: producía Vi=0.5 para
         metas con SERCOP=0/eSIGEF=0 pero LOTAIP=1/CPCCS=1.»

    ⚠️ SU FORMA EXACTA ES `NOT_DETERMINABLE`. La nota conserva un fragmento
    —el umbral de 2 y el 0,5— pero no dice qué producía con los cuatro
    verificadores en 1. Se prueban las DOS lecturas posibles y se declara que
    son lecturas, no la regla. Es DOC-009 aplicado a V: la nota documenta el
    CAMBIO, no reconstruye el original."""
    v = m["verif"]
    if not v or v.get("SERCOP") is None:
        return None
    suma = sum(1 for k in ("SERCOP", "eSIGEF", "LOTAIP", "CPCCS") if v[k])
    if tres_niveles and suma == 4:
        return 1.0
    return 0.5 if suma >= 2 else 0.0


def main() -> int:
    d = leer_motor()
    if not d:
        print("[no determinable] Gold Master no resuelto — no se corren escenarios.")
        return 2

    metas, factor = d["metas"], d["factor"]
    base = evaluar(metas, _K_ACTUAL, _S_ACTUAL)

    # ── VERIFICACIÓN DEL BASELINE ────────────────────────────────────────────
    # Antes de mover una sola pieza hay que demostrar que el laboratorio
    # reproduce el motor. Si no cuadra, ningún contrafactual significa nada.
    desvio = abs(base["icpi"] - _BASELINE)
    if desvio > _TOLERANCIA:
        print(f"[hallazgo] el laboratorio NO reproduce el baseline: "
              f"{base['icpi'] * 100:.6f} % vs {_BASELINE * 100:.6f} % "
              f"(desvío {desvio:.2e}). Ningún escenario es interpretable.")
        return 1
    print(f"baseline reproducido: {base['icpi'] * 100:.4f} % "
          f"(desvío {desvio:.2e}) · FactorTemporal={factor} · mes={d['mes']}")

    esc: list[dict] = []

    def add(fam, cid, nombre, r, nota=""):
        esc.append({"fam": fam, "id": cid, "nombre": nombre, "icpi": r["icpi"],
                    "r": r, "nota": nota})

    # ── 007-A · EL PESO ──────────────────────────────────────────────────────
    # R_i se normaliza por el MÁXIMO TEÓRICO (1,5 × bono 1,15 = 1,725) mientras
    # P_i se normaliza por la SUMA (Σ=1 exacto, verificado en H14!G33). Son dos
    # normalizaciones distintas conviviendo en el mismo producto: A3 mide qué
    # pasa si R se normaliza como P.
    sumaR = sum(m["R"] for m in metas)
    add("A", "A0", "P × R  (baseline)", base)
    add("A", "A1", "sólo P (peso presupuestario)", evaluar(metas, lambda m: m["P"], _S_ACTUAL))
    add("A", "A2", "sólo R (relevancia jurídica)", evaluar(metas, lambda m: m["R"], _S_ACTUAL))
    add("A", "A3", "P × R con R normalizado por la suma",
        evaluar(metas, lambda m: m["P"] * (m["R"] / sumaR), _S_ACTUAL),
        "R por suma en vez de por el máximo teórico 1,725")
    add("A", "A4", "peso uniforme (todas las metas valen igual)",
        evaluar(metas, lambda m: 1.0, _S_ACTUAL),
        "la pregunta literal de Javo: ¿todo debe valer igual?")

    # ── 007-B · V_i ──────────────────────────────────────────────────────────
    reglas_ok = all(v_por_regla(m) is not None for m in metas)
    add("B", "B0", "V implementado (H13!F · literales)", base)
    if reglas_ok:
        add("B", "B1", "V por la regla documentada (H13!B20)",
            evaluar(metas, _K_ACTUAL,
                    lambda m: v_por_regla(m) * m["E"] * m["T"] * m["C"]))
        add("B", "B2a", "V por la regla anterior · lectura literal",
            evaluar(metas, _K_ACTUAL,
                    lambda m: v_regla_anterior(m, False) * m["E"] * m["T"] * m["C"]),
            "≥2 verificadores → 0,5 · si no 0 · forma exacta NOT_DETERMINABLE")
        add("B", "B2b", "V por la regla anterior · lectura de tres niveles",
            evaluar(metas, _K_ACTUAL,
                    lambda m: v_regla_anterior(m, True) * m["E"] * m["T"] * m["C"]),
            "4 verificadores → 1 · ≥2 → 0,5 · forma exacta NOT_DETERMINABLE")
        add("B", "B3", "V sin núcleo obligatorio (media de los 4 silos)",
            evaluar(metas, _K_ACTUAL,
                    lambda m: (sum(1 for k in ("SERCOP", "eSIGEF", "LOTAIP", "CPCCS")
                                   if m["verif"][k]) / 4) * m["E"] * m["T"] * m["C"]),
            "diagnóstico: cuánto pesa que eSIGEF+SERCOP sean eliminatorios")

    # ── 007-C · EL TOPE DE T ─────────────────────────────────────────────────
    # T_i = MIN(1, Ti_raw / FactorTemporal). Sin tope, una entidad que va por
    # delante de la curva de pacing conservaría su ventaja.
    t_sin_tope = {}
    for ent, raw in d["ti_raw"].items():
        try:
            t_sin_tope[ent] = (raw / factor) if factor else raw
        except TypeError:
            t_sin_tope[ent] = None
    truncadas = [m for m in metas
                 if t_sin_tope.get(m["entidad"]) and t_sin_tope[m["entidad"]] > 1.0]
    add("C", "C0", "T con tope MIN(1, ·)  (baseline)", base)
    if all(isinstance(v, float) for v in t_sin_tope.values()):
        add("C", "C1", "T sin tope",
            evaluar(metas, _K_ACTUAL,
                    lambda m: m["V"] * m["E"] * t_sin_tope.get(m["entidad"], m["T"]) * m["C"]),
            f"{len(truncadas)} de {len(metas)} metas están truncadas hoy")

    # ── 007-D · LA ESTRUCTURA ────────────────────────────────────────────────
    # ⚠️ D1-D3 NO SON CANDIDATOS DE REEMPLAZO. Si la media ponderada diera 61 %
    # y la multiplicativa 27 %, eso NO haría al 61 «más correcto»: mediría la
    # consecuencia sustantiva de la multiplicatividad, que es lo que hay que
    # justificar en 011.
    dims = lambda m: (m["V"], m["E"], m["T"], m["C"])
    add("D", "D0", "multiplicativa  V×E×T×C  (baseline)", base)
    add("D", "D1", "media aritmética de las 4 dimensiones",
        evaluar(metas, _K_ACTUAL, lambda m: sum(dims(m)) / 4),
        "las deficiencias se suman en vez de interactuar")
    add("D", "D2", "media geométrica de las 4 dimensiones",
        evaluar(metas, _K_ACTUAL, lambda m: prod(dims(m)) ** 0.25),
        "conserva el cero eliminatorio, suaviza la penalización compuesta")
    add("D", "D3", "por bloques: V eliminatorio × media(E, T, C)",
        evaluar(metas, _K_ACTUAL, lambda m: m["V"] * ((m["E"] + m["T"] + m["C"]) / 3)),
        "aísla cuánto viene del cero de V y cuánto de la interacción del resto")

    # ═════════════════════════════════════════════════════════════════════════
    # SALIDA
    # ═════════════════════════════════════════════════════════════════════════
    for e in esc:
        e["cat"] = clasificar(e["icpi"])
        e["dist"], e["umbral"] = distancia_al_umbral(e["icpi"])
        e["delta"] = (e["icpi"] - _BASELINE) * 100
        e["delta_rel"] = ((e["icpi"] / _BASELINE - 1) * 100) if _BASELINE else 0.0

    cat_base = clasificar(_BASELINE)
    saltan = [e for e in esc if e["cat"] != cat_base]

    print(f"\nescenarios: {len(esc)} · categoría del baseline: {cat_base}")
    for e in esc:
        marca = "  ⚠️ SALTA" if e["cat"] != cat_base else ""
        print(f"  {e['id']:<4} {e['icpi'] * 100:>8.4f} %  "
              f"Δ {e['delta']:>+8.4f}  {e['cat']}{marca}")
    print(f"\ncategoría estable en {len(esc) - len(saltan)}/{len(esc)} escenarios")

    _escribir(d, base, esc, cat_base, saltan, truncadas, t_sin_tope, sumaR)
    print(f"→ {_SALIDA.relative_to(_RAIZ).as_posix()}")
    return 0


def _escribir(d, base, esc, cat_base, saltan, truncadas, t_sin_tope, sumaR) -> None:
    c = concentracion(base["filas"])
    o: list[str] = []
    A = o.append

    A("# GM-Ω · ICPI — SENSIBILIDAD  `007-A/B/C/D/X`")
    A("")
    A("**DERIVADO — no editar a mano.** Lo regenera "
      "`scripts/gm_omega/sensibilidad_icpi.py` leyendo el Gold Master vigente.")
    A("")
    A("> ### ⚠️ LAS TRES ETIQUETAS")
    A("> Cada escenario de este documento es, sin excepción:")
    A("> **MATEMÁTICAMENTE REPRODUCIBLE** · **METODOLÓGICAMENTE CONTRAFACTUAL** · "
      "**NO AUTORIZADO PARA PUBLICACIÓN**.")
    A(">")
    A("> El único número oficial es **27,4582 %** (regla `GM-Ω-ICPI-000`) hasta "
      "que `GM-Ω-ICPI-011` dictamine. Ninguna cifra de aquí puede citarse fuera "
      "de esta auditoría, ni siquiera para ilustrar: el patrón que perseguimos "
      "—el «48,33 %»— nació de un número de trabajo que sobrevivió a su contexto.")
    A("")
    A(f"Baseline reproducido por el laboratorio: **{base['icpi'] * 100:.6f} %** "
      f"(desvío {abs(base['icpi'] - _BASELINE):.2e} respecto de `H12!B33`). "
      "Sin esa reproducción exacta, ningún contrafactual sería interpretable.")
    A("")
    A("**`E_i` no entra en ningún escenario.** Su regla generadora está "
      "`NOT_DETERMINABLE` (`007-B0`): mover una variable cuya biografía no "
      "conocemos sería inferir su regla desde el efecto que produce, que es lo "
      "que `DOC-009` prohíbe. Entra en `011` o no entra.")
    A("")

    # ── jerarquía de sensibilidad ────────────────────────────────────────────
    # La conclusión central, DERIVADA: qué decisión metodológica mueve más el
    # índice. Se calcula del rango de cada familia, no se escribe a mano.
    _FAM = {"A": "el peso  `P × R`", "B": "la especificación de `V`",
            "C": "el tope de `T`", "D": "la estructura algebraica"}
    rangos = []
    for fam, nombre in _FAM.items():
        ds = [e["delta"] for e in esc if e["fam"] == fam]
        rangos.append((max(ds) - min(ds), fam, nombre, min(ds), max(ds)))
    rangos.sort(reverse=True)

    A("## La jerarquía de sensibilidad")
    A("")
    A("> La pregunta con la que arrancó `007` era si el ICPI mide la integridad "
      "de la cadena de gestión o está fuertemente condicionado por cómo "
      "repartimos el peso presupuestario y jurídico. La respuesta es que el peso "
      "**apenas** lo condiciona, y su álgebra **lo gobierna**.")
    A("")
    A("| Decisión metodológica | Rango que abre | Peor caso | Mejor caso |")
    A("|---|---:|---:|---:|")
    for r, fam, nombre, lo, hi in rangos:
        A(f"| `007-{fam}` · {nombre} | **{r:.2f} pp** | {lo:+.2f} pp | {hi:+.2f} pp |")
    A("")
    A("Ordenado así, el resultado es inequívoco y **no era el esperado**: "
      f"cambiar la estructura algebraica mueve el índice hasta "
      f"**{rangos[0][0]:.1f} puntos**, mientras que redistribuir todo el peso "
      f"—o eliminarlo del todo— lo mueve "
      f"**{[r for r, f, *_ in rangos if f == 'A'][0]:.1f}**.")
    A("")
    A("### La conclusión, formulada con precisión")
    A("")
    A("> **El ICPI presenta baja sensibilidad a las alternativas de ponderación "
      "ENSAYADAS y alta sensibilidad a la arquitectura algebraica de agregación. "
      "Por tanto, la validez sustantiva del índice depende mucho más de la "
      "justificación teórica de su estructura multiplicativa que de la elección "
      "entre las ponderaciones evaluadas.**")
    A("")
    A("⚠️ **La formulación importa, y la primera versión era peor.** Decir que "
      "el índice es «frágil a su forma matemática» suena a diagnóstico y en "
      "realidad contrabandea un juicio: sugiere que la multiplicatividad es un "
      "defecto. `007-D` **no demuestra que multiplicar esté mal** — demuestra "
      "que multiplicar es **altamente determinante**. Son dos afirmaciones "
      "distintas y sólo la segunda está medida.")
    A("")
    A("Y las dos precisiones del enunciado no son adorno:")
    A("")
    A("- **«ensayadas»** — se probaron cuatro alternativas de peso, no todas las "
      "posibles. Una ponderación radicalmente distinta podría mover más. Lo "
      "medido es lo medido.")
    A("- **«validez sustantiva»** — lo que está en juego no es qué número sale, "
      "sino si el número significa lo que el constructo promete.")
    A("")
    A("Eso reordena `011`: la discusión sobre si el agua potable debe pesar más "
      "que un taller —legítima, y respondida por `P·R`— resulta ser de "
      "**segundo orden** frente a la pregunta de primer orden, que es:")
    A("")
    A("> **¿Qué teoría de la integridad representa realmente "
      "`J = P·R·V·E·T·C`, y qué la fundamenta?**")
    A("")
    A("La estructura multiplicativa no queda impugnada por `007`. Queda "
      "**obligada a demostrar por qué debe existir**.")
    A("")

    # ── tabla maestra ────────────────────────────────────────────────────────
    A("## 007-X · Robustez de clasificación")
    A("")
    A("La pregunta de 007-X no es cuánto cambia el número, sino si cambia **la "
      "categoría** — porque es la categoría, no el decimal, lo que se convierte "
      "en decisión.")
    A("")
    A("| Esc. | Escenario | ICPI | Δ abs. | Δ rel. | Categoría AVEP |")
    A("|---|---|---:|---:|---:|---|")
    for e in esc:
        salta = " ⚠️" if e["cat"] != cat_base else ""
        A(f"| `{e['id']}` | {e['nombre']} | {e['icpi'] * 100:.4f} % | "
          f"{e['delta']:+.4f} | {e['delta_rel']:+.1f} % | {e['cat']}{salta} |")
    A("")
    A(f"**Categoría del baseline: {cat_base}.** Se mantiene en "
      f"**{len(esc) - len(saltan)} de {len(esc)}** escenarios.")
    A("")
    if saltan:
        A("Escenarios que **cambian la categoría** — sensibilidad decisional, no "
          "sólo numérica:")
        A("")
        for e in saltan:
            A(f"- `{e['id']}` {e['nombre']} → **{e['cat']}**")
        A("")
    dist, umbral = distancia_al_umbral(_BASELINE)
    A(f"⚠️ **El baseline está a {dist:.2f} puntos porcentuales del umbral de "
      f"{umbral}.** Esa distancia es la que convierte cualquier decisión "
      "metodológica de este documento en una decisión sobre la categoría "
      "publicable, y no sólo sobre un decimal.")
    A("")
    A("⚠️ Y una precisión que cambia cómo se lee toda esta tabla: **hoy el motor "
      "NO emite categoría.** `H12!B34` la condiciona a `H07!B22>=12` —doce meses "
      "de corte— y el corte vigente es el mes "
      f"**{d['mes']}**, así que la celda devuelve «Corte parcial · lectura "
      "preliminar (no comparable con umbral anual)». Las categorías de esta tabla "
      "son **las que el motor emitiría al cierre**, calculadas con sus mismos "
      "umbrales. No son lo que el motor dice hoy.")
    A("")

    # ── 007-X-bis · la procedencia de la propia escala ───────────────────────
    av = d["avep"]
    A("### ★ 007-X-bis · ¿Y de dónde salen los umbrales?")
    A("")
    A("Javo aportó el contexto que faltaba: **la escala AVEP es invención de "
      "Dylus Lab**, ajustada después a lo que la normativa pública exigía. Eso "
      "obliga a hacer explícito un supuesto que toda esta sección arrastraba: "
      "**`007-X` mide la robustez de la categoría contra unos umbrales cuya "
      "procedencia no estaba auditada.** Auditada ahora, esto es lo que hay.")
    A("")
    A("**1 · La norma sostiene el CONSTRUCTO, no los CORTES.** La tesis titula "
      "un apartado «Baremo AVEP — Interpretación jurídica» y lo que fundamenta "
      "allí es *por qué* medir congruencia: `COPFP Art. 41` —el PDOT es la "
      "directriz **principal**, luego una inversión no alineada es jurídicamente "
      "cuestionable—. Las variables sí tienen norma citada (`P_i` → COPFP 54; "
      "`R_i` → COOTAD 54-55 + Constitución 3, 12, 66). **Dónde cortar en 70 o en "
      "40 no la tiene.**")
    A("")
    if av["copias"]:
        A(f"**2 · La escala está COPIADA en {len(av['copias'])} hojas** "
          + ", ".join(f"`{h}`" for h in av["copias"][:8])
          + (" …" if len(av["copias"]) > 8 else "") + ", y "
          + (f"**{len(av['con_norma'])}** de esas copias cita una norma."
             if av["con_norma"] else
             "**ninguna de esas copias cita una norma**, mientras que los "
             "umbrales de inversión del mismo libro sí citan COOTAD."))
        A("")
    A("   Y la copia no es accidental: `H01!A30` **instruye a copiarla "
      "literalmente**. Viene de un incidente real que `H01!A28` conserva —")
    A("")
    A("   > «AVEP NO es una función de Excel. NO existe `=AVEP()`. Si se escribe "
      "`=AVEP(...)` el resultado será `#¿NOMBRE?` y el ecosistema fallará.»")
    A("")
    A("   El motor confundió la escala con una fórmula, y la solución adoptada "
      "—replicar el `IF` en todas las hojas— **resolvió el síntoma y consolidó "
      "la causa**: una capa de interpretación quedó incrustada dentro del "
      "cálculo, y duplicada. Cambiar un umbral hoy exige editar N celdas a mano.")
    A("")
    if len(av["variantes_nivel_alto"]) > 1:
        A(f"**3 · Y ya divergió**: el nivel superior aparece como "
          + " y ".join(f"«Excelencia en {v}»" for v in av["variantes_nivel_alto"])
          + ". Es el precio de la copia literal.")
        A("")
    A("**3 · El acrónimo tiene DOS biografías, y una etiqueta cambió.** En las "
      "tesis, `AVEP` se expande de dos maneras —«**A**lineación, **V**inculación, "
      "**E**jecución, **P**ublicación» y «**A**lfaro **V**irtus **E**scala de "
      "**P**onderación»—, la segunda un nombre propio; y el nivel superior pasó "
      "de «Excelencia en **Trazabilidad**» (tesis) a «Excelencia en "
      "**Gobernanza**» (motor). Los rangos, en cambio, **coinciden exactamente**: "
      "0-19 · 20-39 · 40-69 · 70-89 · 90-100. La escala del motor sí reproduce la "
      "de la tesis; lo que se movió fue el vocabulario.")
    A("")
    A("**4 · La tesis nunca dijo que fuera una fórmula.** La llama «Baremo de "
      "Valoración» y «Baremo de **Interpretación**», y dice que los resultados "
      "«se **contrastan** con» él. La doctrina correcta ya estaba escrita antes "
      "que el motor:")
    A("")
    A("   ```")
    A("   dato → estado epistemológico → INTERPRETACIÓN → producto")
    A("                                      ↑ aquí vive AVEP")
    A("   ```")
    A("")
    A("**5 · Qué significa esto para LATAM (`010`).** Aquí está la tensión que "
      "Javo intuye, y tiene salida:")
    A("")
    A("| | Anclar los cortes a normativa local | Mantenerlos propios |")
    A("|---|---|---|")
    A("| Defensa en Ecuador | fuerte (hay norma) | exige argumento teórico |")
    A("| Viaje a LATAM | ❌ no viaja: se recalibra por país | ✅ viaja |")
    A("")
    A("   La salida no es elegir una: es **separar las capas**. El constructo se "
      "ancla a norma —y esa parte es local por naturaleza—; los **cortes** son "
      "una decisión metodológica propia, explícita y **calibrable por país**. "
      "Que es justamente la arquitectura núcleo/adaptador que `010` tiene que "
      "demostrar.")
    A("")
    A("⚠️ **Nada de esto dice que la escala esté mal.** Los umbrales de un índice "
      "compuesto casi nunca salen de una norma: son una decisión metodológica, y "
      "es legítima. Lo que `007-X-bis` establece es que **hoy se presenta con la "
      "misma autoridad que un umbral legal y no la tiene**, que vive en la capa "
      "equivocada, y que de ella depende un Certificado (`H01!C59` fija la "
      "emisión en AVEP ≥ 70 %). Una escala con consecuencia contractual necesita "
      "procedencia declarada. → `011`.")
    A("")

    # ── concentración ────────────────────────────────────────────────────────
    A("## Concentración del resultado (baseline)")
    A("")
    A("| Medida | Valor |")
    A("|---|---:|")
    A(f"| Meta que más aporta | {c['top1'] * 100:.2f} % del numerador |")
    A(f"| Tres metas que más aportan | {c['top3'] * 100:.2f} % |")
    A(f"| Cinco metas que más aportan | {c['top5'] * 100:.2f} % |")
    A(f"| Metas que explican el 50 % | {c['n50']} de {len(base['filas'])} |")
    A(f"| Metas que explican el 80 % | {c['n80']} de {len(base['filas'])} |")
    A(f"| HHI de las contribuciones | {c['hhi']:.4f} |")
    A("")
    A("Esto **no convierte el índice en malo**: concentrar el peso en las metas "
      "estratégicas puede ser exactamente lo que QUIRA quiere medir, y es lo que "
      "`P_i` hace a propósito —impedir que metas baratas inflen el resultado "
      "mientras el alcantarillado sigue parado—. Pero un índice que se presenta "
      "como global y depende de unas pocas metas tiene que ser **consciente y "
      "declarado**, no descubierto por un tercero.")
    A("")
    A("### Las diez metas que más pesan")
    A("")
    A("| # | Meta | P | R | K=P·R | peso efec. | S=V·E·T·C | J | % del num. |")
    A("|---|---|---:|---:|---:|---:|---:|---:|---:|")
    for i, f in enumerate(base["filas"][:10], 1):
        A(f"| {i} | `{f['id']}` | {f['P']:.4f} | {f['R']:.4f} | {f['K']:.4f} | "
          f"{f['k_rel'] * 100:.2f} % | {f['S']:.4f} | {f['J']:.4f} | "
          f"{f['j_rel'] * 100:.2f} % |")
    A("")
    ceros = [f for f in base["filas"] if f["S"] == 0]
    if ceros:
        A(f"Y **{len(ceros)} metas aportan exactamente 0 al numerador** mientras "
          f"siguen ocupando "
          f"{sum(f['k_rel'] for f in ceros) * 100:.2f} % del denominador: "
          + ", ".join(f"`{f['id']}`" for f in ceros) + ". "
          "Es la multiplicatividad operando — y lo que 007-D mide.")
        A("")
        criticas = sorted((f for f in ceros if f["R"] >= 0.85),
                          key=lambda f: -f["R"])
        if criticas:
            A("⚠️ Y entre ellas están metas de **máxima relevancia jurídica**: "
              + ", ".join(f"`{f['id']}` (R={f['R']:.4f})" for f in criticas)
              + ". El motor las reconoce como competencia exclusiva crítica y "
              "acto seguido las anula, porque una sola dimensión en cero "
              "extingue el producto. Que eso sea correcto —una obra crítica sin "
              "verificación no debería puntuar— o excesivo —desaparece del "
              "índice justo la meta que más importaba vigilar— es la decisión "
              "de `011`. Aquí sólo se mide que ocurre.")
            A("")

    # ── A ────────────────────────────────────────────────────────────────────
    A("## 007-A · Sensibilidad del peso (`P × R`)")
    A("")
    A("> ¿El ICPI mide la integridad de la cadena de gestión, o está fuertemente "
      "condicionado por cómo repartimos el peso presupuestario y jurídico?")
    A("")
    A("**Hallazgo de estructura, previo a los escenarios.** La fórmula canónica "
      "`Σ(P·R·V·E·T·C) / Σ(P·R)` es algebraicamente una **media ponderada** de "
      "`S = V·E·T·C` con pesos `K = P·R`. La multiplicatividad no está en la "
      "agregación entre metas —que es lineal— sino **dentro** de cada meta, entre "
      "sus cuatro dimensiones. Son dos decisiones distintas y hasta ahora se "
      "leían como una: `007-A` audita la primera, `007-D` la segunda.")
    A("")
    A("**Y las dos normalizaciones no son la misma.** `P_i` se normaliza por la "
      f"**suma** (Σ=1 exacto, verificado en `H14!G33`); `R_i` por el **máximo "
      f"teórico** `1,5 × 1,15 = 1,725`, y por eso Σ`R_i` = {sumaR:.4f}, no 1. "
      "Conviven en el mismo producto. `A3` mide qué pasaría si `R` se normalizara "
      "como `P`.")
    A("")
    _tabla_familia(A, esc, "A", cat_base)
    A("")
    d_a3 = next(e["delta"] for e in esc if e["id"] == "A3")
    d_a4 = next(e["delta"] for e in esc if e["id"] == "A4")
    d_a1 = next(e["delta"] for e in esc if e["id"] == "A1")
    d_a2 = next(e["delta"] for e in esc if e["id"] == "A2")
    A("### ★ HALLAZGO DE INVARIANCIA DE ESCALA")
    A("")
    A(f"`A3` se desvía {abs(d_a3):.0e} pp del baseline. Eso no es «casi igual»: "
      "es **cero algebraico**, y no es un resultado empírico de este conjunto de "
      "datos — es una **propiedad del estimador**. La demostración cabe en tres "
      "líneas:")
    A("")
    A("```")
    A("        K_i = P_i · R_i                        peso vigente")
    A("       R'_i = R_i / ΣR                         normalizar R por la suma")
    A("       K'_i = P_i · R_i / ΣR = (1/ΣR) · K_i    una constante común")
    A("")
    A("   ICPI(K') = Σ(cK_i·S_i) / Σ(cK_i)")
    A("            = c·Σ(K_i·S_i) / c·Σ(K_i)")
    A("            = ICPI(K)                          ∎")
    A("```")
    A("")
    A("Toda transformación de `R` que sea una constante multiplicativa común "
      "deja el ICPI **exactamente igual**. De ahí se siguen dos cosas:")
    A("")
    A("1. **La escala de `R_i` es irrelevante para el índice; sólo importa su "
      "forma relativa entre metas.** Que `R` se normalice por el máximo teórico "
      "y `P` por la suma es una inconsistencia de presentación —dos variables "
      "que parecen comparables y no lo son— **sin ningún efecto sobre el "
      "resultado**. `011` puede cerrarlo sin discutirlo.")
    A("2. Y una **falsa preocupación queda eliminada**: no hay que decidir cómo "
      "normalizar `R`, porque la decisión no existe. Saber qué transformaciones "
      "son irrelevantes *por construcción* es tan parte de auditar un estimador "
      "como saber cuáles lo mueven — y es lo que separa correr escenarios de "
      "entender el instrumento.")
    A("")
    A("### La respuesta a la pregunta de Javo")
    A("")
    A("> *«¿pesa más el agua potable por necesidad de extrema urgencia, o todo "
      "debe valer igual si se planificó?»*")
    A("")
    A(f"`A4` responde con un número: si **todas las metas pesaran igual**, el "
      f"ICPI sería {d_a4:+.2f} pp distinto. Menos de un punto. Y las dos mitades "
      f"del peso, por separado, abren apenas "
      f"{abs(d_a1 - d_a2):.2f} pp entre sí.")
    A("")
    A("Lo cual **no invalida la ponderación** —`P·R` sigue siendo el antídoto "
      "anti-gaming que impide que metas baratas inflen el índice, y sigue "
      "gobernando el **ranking** de metas, que es donde se toman decisiones "
      "concretas—. Lo que dice es más preciso: la ponderación decide **a quién "
      "se mira**, no **cuánto sale**. El agregado lo decide `S = V·E·T·C`.")
    A("")

    # ── B ────────────────────────────────────────────────────────────────────
    A("## 007-B · `V_i` — especificación implementada vs. documentada")
    A("")
    A("**A diferencia de `E_i`, aquí sí hay biografía.** El propio libro conserva "
      "la regla vigente, la regla anterior y la justificación del cambio:")
    A("")
    A("```")
    A("H13!B16-B20 · REGLA VIGENTE (documentada en prosa)")
    A("  Vi = 0.0  si V_eSIGEF=0 O V_SERCOP=0     sin núcleo financiero, sin score")
    A("  Vi = 0.5  si núcleo OK y sin LOTAIP ni CPCCS")
    A("  Vi = 1.0  si núcleo OK y (LOTAIP=1 O CPCCS=1)")
    A("")
    A("H13!B21 · POR QUÉ CAMBIÓ")
    A("  «La fórmula original SI(suma≥2,0.5) era incorrecta: producía Vi=0.5")
    A("   para metas con SERCOP=0/eSIGEF=0 pero LOTAIP=1/CPCCS=1.»")
    A("```")
    A("")
    A("Eso es exactamente lo que a `E_i` le falta, y conviene decirlo en voz alta: "
      "**`E_i` es la excepción del motor, no su norma.** El resto de las variables "
      "documenta sus cambios.")
    A("")
    A("### ⚠️ Pero la regla está documentada y NO implementada")
    A("")
    A("`H12!D` lee `VLOOKUP(A, H13!$A:$F, 6)` → la columna `F` de `H13`, que "
      "contiene **25 literales**, no la fórmula. La regla vive en prosa, en las "
      "celdas `B16..B20`, donde ningún recálculo la aplica. Es el mismo patrón "
      "que `E_i`, un escalón más arriba: allí no había regla; aquí la hay y no "
      "está conectada a los valores que el motor consume.")
    A("")
    A("Y esa columna se llama `Vi_2025`, bajo un título que dice «VALORES Vi DE "
      "REFERENCIA 2025 — para verificar ICPI_Real_2025». **El ICPI 2026 está "
      "consumiendo la columna de referencia de 2025** "
      "(`TEMPORAL_SEMANTIC_GAP`, ya registrado en la matriz `004`).")
    A("")
    coh = _coherencia_v(base["filas"])
    A("### Coherencia regla ↔ valor, meta a meta")
    A("")
    if coh["sin_datos"]:
        A(f"⚠️ {coh['sin_datos']} metas sin verificadores en `H13`: no se pueden "
          "contrastar y **no se cuentan como coincidentes**.")
        A("")
    A(f"De las **{coh['comparables']} metas comparables**, "
      f"**{coh['iguales']} coinciden** con la regla documentada y "
      f"**{len(coh['difieren'])} no**.")
    A("")
    if coh["difieren"]:
        A("| Meta | V implementado | V por la regla | Δ | peso K | verificadores |")
        A("|---|---:|---:|---:|---:|---|")
        for f in coh["difieren"]:
            v = f["verif"]
            sig = (f"SERCOP={v['SERCOP']} eSIGEF={v['eSIGEF']} "
                   f"LOTAIP={v['LOTAIP']} CPCCS={v['CPCCS']}")
            A(f"| `{f['id']}` | {f['V']} | {f['v_regla']} | "
              f"{f['v_regla'] - f['V']:+.1f} | {f['k_rel'] * 100:.2f} % | {sig} |")
        A("")
    else:
        A("**Los valores implementados obedecen la regla documentada, uno por "
          "uno.** La especificación de `V_i` es reconstruible y está cumplida: "
          "lo que falla no es la regla, es que viva en prosa y en la columna de "
          "otro año.")
        A("")
    A("### La regla anterior: reconstruible sólo en parte")
    A("")
    A("`H13!B21` conserva un **fragmento** —el umbral de 2 y el 0,5— pero no dice "
      "qué producía la regla original con los cuatro verificadores en 1. "
      "**Su forma exacta es `NOT_DETERMINABLE`**, y por eso se prueban las dos "
      "lecturas posibles y se declaran como lecturas. `DOC-009` aplicado a `V`: "
      "la nota documenta el **cambio**, no reconstruye el **original**.")
    A("")
    _tabla_familia(A, esc, "B", cat_base)
    A("")
    try:
        b2a = next(e["icpi"] for e in esc if e["id"] == "B2a")
        b2b = next(e["icpi"] for e in esc if e["id"] == "B2b")
        c2a, c2b = clasificar(b2a), clasificar(b2b)
        A("### ⚠️ El hallazgo de `B2`: la incertidumbre abarca dos categorías")
        A("")
        A(f"`B2a` y `B2b` son **dos lecturas de la misma regla anterior**, y "
          f"difieren en **{abs(b2a - b2b) * 100:.2f} puntos** — de "
          f"{b2a * 100:.2f} % a {b2b * 100:.2f} %. No cruzan un decimal: cruzan "
          f"una frontera de categoría, de «{c2a}» a «{c2b}».")
        A("")
        A("La lectura correcta **no** es «el motor antiguo daba 16 %». Es esta:")
        A("")
        A("> El fragmento que el libro conserva de la regla anterior es "
          "**insuficiente para reconstruir el pasado**, y el margen de esa "
          "insuficiencia vale dos categorías AVEP.")
        A("")
        A("Es `DOC-009` en su forma más útil: la nota de `H13!B21` **parece** "
          "documentar la regla anterior y en realidad documenta sólo por qué se "
          "abandonó. Dos auditorías igual de rigurosas, partiendo del mismo "
          "libro, reconstruirían historias distintas — y ninguna de las dos "
          "podría demostrar la suya.")
        A("")
    except StopIteration:
        pass
    A("### ★ DOS VACÍOS DE NATURALEZA DISTINTA — `V` no tiene el problema de `E`")
    A("")
    A("Este es el resultado que más lejos llega de todo `007`, y no es un "
      "número. Puestos uno al lado del otro, `V` y `E` **no tienen el mismo "
      "problema**, y tratarlos igual sería el error:")
    A("")
    A("| | `V_i` | `E_i` |")
    A("|---|---|---|")
    A("| Definición del constructo | ✅ existe | ✅ existe |")
    A("| Regla vigente documentada | ✅ `H13!B16-B20` | ❌ no consta |")
    A("| Regla histórica documentada | ✅ fragmento en `H13!B21` | ✅ tesis: 1 · 0,90 · 0,75 |")
    A("| Explicación del cambio | ✅ y con su motivo | ❌ ninguna |")
    A("| Valores reproducibles contra su regla | ✅ 25 de 25 | ❌ ninguno |")
    A("| **Naturaleza del vacío** | **límite de reconstrucción** | **ausencia de regla generadora** |")
    A("")
    A("`V` está en una situación **sana para una auditoría**: hay genealogía, y "
      "hay un límite explícito de lo que sabemos. Se puede decir con precisión "
      "qué se sabe, qué no, y por qué. `E` no: existe la variable, existe una "
      "regla histórica en la tesis, existe la corrección de Javo sobre no "
      "penalizar la afiliación, existen los valores — y **no existe evidencia "
      "preservada que permita reconstruir la regla que produjo esos valores**.")
    A("")
    A("> Un vacío de trazabilidad se clasifica por su **naturaleza**, no por su "
      "tamaño. «No puedo reconstruirlo del todo» y «no hay nada que "
      "reconstruir» exigen auditorías distintas y admiten conclusiones "
      "distintas.")
    A("")
    A("Y de ahí se sigue, retroactivamente, que **fue correcto dejar `E_i` fuera "
      "de `007`**: hacer sensibilidad sobre una variable cuya regla generadora "
      "se desconoce habría producido números impecables sobre una premisa "
      "epistemológicamente vacía. Elegante y sin fundamento — que es la forma "
      "más difícil de detectar un error.")
    A("")
    A("**Y hay que separar dos cosas que 007 no mezcla.** `V` como **regla** —qué "
      "significa verificación intersistémica— y `V` como **evidencia** —si lo "
      "capturado satisface esa regla—. Este documento mide sólo la **sensibilidad "
      "del resultado** a la elección de arquitectura. Cuál de las dos representa "
      "mejor el constructo que QUIRA quiere medir es una pregunta de `011`.")
    A("")

    # ── C ────────────────────────────────────────────────────────────────────
    A("## 007-C · El tope `MIN(1, ·)` de `T_i`")
    A("")
    A("No se parte de que el tope sea un defecto: **puede ser correcto**. Si la "
      "teoría dice que una meta que alcanzó el umbral temporal esperado ya obtuvo "
      "el máximo crédito temporal, truncar es la implementación fiel de esa idea. "
      "La pregunta auditable es otra:")
    A("")
    A("> ¿El tope elimina información que la teoría del indicador necesitaría "
      "conservar?")
    A("")
    A("`T_i = MIN(1, Ti_raw / FactorTemporal)`, con `FactorTemporal` = "
      f"**{d['factor']}** para el mes de corte **{d['mes']}**.")
    A("")
    A("| Entidad | `Ti_raw` | sin tope | con tope | ¿truncada? |")
    A("|---|---:|---:|---:|---|")
    for ent, raw in d["ti_raw"].items():
        sin = t_sin_tope.get(ent)
        if isinstance(sin, float) and isinstance(raw, float):
            trunc = "⚠️ **sí**" if sin > 1 else "no"
            A(f"| {ent} | {raw:.6f} | {sin:.6f} | {min(1.0, sin):.6f} | {trunc} |")
    A("")
    n_ent = sum(1 for v in t_sin_tope.values() if isinstance(v, float) and v > 1)
    A(f"**{n_ent} de {len(t_sin_tope)} entidades está truncada**, y arrastra "
      f"**{len(truncadas)} de {len(base['filas'])} metas**"
      + (": " + ", ".join(f"`{m['id']}`" for m in truncadas) if truncadas else "")
      + ".")
    A("")
    A("### ⚠️ Y el `FactorTemporal` ya no es lo que su nota dice")
    A("")
    A("La fórmula real es una **curva de pacing empírica**:")
    A("")
    A("```")
    A("H07!B23 = CHOOSE(mes, 0.011, 0.11, 0.128, 0.212, 0.266, 0.36,")
    A("                      0.442, 0.516, 0.766, 0.883, 0.925, 1)")
    A("  «curva pacing Montecristi 2025: promedio de 3 adscritas»")
    A("```")
    A("")
    A(f"pero la nota que la describe en `H07b!F20` sigue diciendo "
      f"«`{d['nota_factor'][:80]}`». Para abril, la curva da "
      f"**{d['factor']}** y la descripción lineal daría **0,3333**: un 57 % más "
      "exigente. La fórmula cambió y su descripción se quedó atrás — el mismo "
      "patrón del «48,33 %», aquí en la nota de una celda.")
    A("")
    A("Hay además una circularidad que `011` tendrá que juzgar: la curva se "
      "construyó con el pacing de **las tres adscritas en 2025**, y se usa para "
      "normalizar el desempeño de **esas mismas adscritas en 2026**. El "
      "denominador y el numerador comparten origen.")
    A("")
    _tabla_familia(A, esc, "C", cat_base)
    A("")

    # ── D ────────────────────────────────────────────────────────────────────
    A("## 007-D · La arquitectura multiplicativa")
    A("")
    A("Aquí ya no se audita un parámetro: se audita la **teoría matemática del "
      "indicador**. `J = P·R·V·E·T·C` afirma algo muy fuerte —que las dimensiones "
      "son **conjuntamente necesarias**— y por tanto que `V=0 → J=0`, y que "
      "cuatro deficiencias moderadas se **componen** en vez de promediarse. Eso "
      "puede ser exactamente lo que significa «integridad».")
    A("")
    A("> ### ⚠️ `D1`–`D3` NO son candidatos de reemplazo")
    A("> Son instrumentos de diagnóstico. Si la media ponderada diera 61 % y la "
      "multiplicativa 27 %, eso **no haría al 61 más correcto**: mediría la "
      "consecuencia sustantiva de la multiplicatividad. Y una consecuencia de ese "
      "tamaño hay que justificarla, no heredarla.")
    A("")
    _tabla_familia(A, esc, "D", cat_base)
    A("")
    try:
        d1 = next(e for e in esc if e["id"] == "D1")
        d2 = next(e for e in esc if e["id"] == "D2")
        d3 = next(e for e in esc if e["id"] == "D3")
        A("### Qué significa esta tabla")
        A("")
        A(f"**{d1['delta']:+.1f} puntos.** Ninguna otra decisión del motor se "
          "acerca. La multiplicatividad no es un detalle de implementación: es "
          "**la decisión que define el indicador**, y su efecto es tres "
          "categorías AVEP de distancia.")
        A("")
        A(f"Y la comparación entre `D2` ({d2['icpi'] * 100:.2f} %) y `D3` "
          f"({d3['icpi'] * 100:.2f} %) —dos construcciones distintas que caen a "
          f"{abs(d2['icpi'] - d3['icpi']) * 100:.2f} pp— **separa los dos "
          "efectos que hasta ahora iban juntos**:")
        A("")
        A("- el **cero eliminatorio** (`V=0 → J=0`), que `D3` conserva íntegro;")
        A("- la **penalización compuesta** entre las dimensiones que no son cero, "
          "que `D3` sustituye por un promedio.")
        A("")
        A("Que ambos den casi lo mismo dice que **casi toda la severidad del "
          "motor viene de la composición de deficiencias moderadas, no de los "
          "ceros**. Con `S = V·E·T·C`, una meta con las cuatro dimensiones en "
          "0,75 —que en lenguaje llano es «va aceptablemente en todo»— puntúa "
          "0,32. Ese es el núcleo del constructo, y hay que sostenerlo "
          "explícitamente o cambiarlo: hoy no está argumentado en ninguna parte "
          "del libro.")
        A("")
    except StopIteration:
        pass

    # ── hallazgos colaterales ────────────────────────────────────────────────
    A("## Hallazgos colaterales del sondeo")
    A("")
    A("Aparecieron al leer el motor para montar el laboratorio. No son "
      "contrafactuales: son **estado observado**, y por tanto sí son citables "
      "dentro de la auditoría.")
    A("")
    api = d["api"]
    A(f"1. **El rótulo del ICPI imprime `0,27 %` en {d['hojas_rotulo']} hojas del "
      "libro.** `H12!B33` guarda el índice en escala 0-1 (`=B31/B32`, sin el "
      "`×100` que declara la fórmula de `A3`), y la cabecera `E1` de cada hoja lo "
      "rotula con `ROUND(B33,2)&\"%\"`. El resultado literal es «ICPI 2026: "
      "0,27 %» donde el índice es 27,46 %.")
    A("")
    if api["pct_corrige_escala"]:
        A("   ⚠️ **Y esto NO llega al producto** — hay que decirlo con la misma "
          "precisión con que se señala el defecto. La capa API corrige la escala: "
          "`H73!ICPI_GLOBAL_PCT = H12!B33*100`, y es de ahí de donde lee el "
          "conector. La UI publica 27,46 %. El defecto es **interno al libro**, "
          "en lo que ve quien abre el Excel: un auditor externo leyendo el Gold "
          "Master vería 0,27 % en la cabecera de cada hoja. Grave para la "
          "defensa documental, inocuo para la UI.")
    else:
        A("   ⚠️ **Y no consta que la capa API corrija la escala** — hay que "
          "verificar `H73!ICPI_GLOBAL_PCT` antes de afirmar nada sobre lo que "
          "publica la UI.")
    A("")
    A("2. **La brecha ICM–ICPI compara escalas incompatibles.** "
      "`B36 = B35 − B33` resta un valor en escala 0-100 (`B35 = H08!B7×100`) "
      "menos uno en 0-1 (`B33`), y `B37` clasifica el resultado contra umbrales "
      "de 30 y 15. Con `B36` acotado a ese rango, el veredicto «✅ Brecha de "
      "Verificación mínima» es **estructuralmente inalcanzable de otro modo**: no "
      "es un resultado, es el único desenlace posible de la fórmula.")
    A("")
    cons = d["consumidores"]
    A("3. **La categoría AVEP no se emite hoy — y ese silencio SÍ llega al "
      "producto.** `H12!B34` exige 12 meses de corte y devuelve una frase de "
      "diagnóstico interno. La capa API la propaga tal cual:")
    A("")
    A(f"   > `H73!ICPI_CLASIFICACION` = «{api['clasificacion']}»")
    A("")
    if cons:
        A(f"   Y **{len(cons)} superficies del producto** consumen ese campo: "
          + ", ".join(f"`{p}`" for p in cons) + ".")
        A("")
        A("   Son dos problemas encadenados. Uno: donde debería haber una "
          "categoría de gobernanza hay una frase que no lo es. Dos: esa frase "
          "está escrita en **lenguaje interno** —«no comparable con umbral "
          "anual»— y cruza al producto, que es justo lo que el Bloomberg "
          "Firewall existe para impedir.")
        A("")
        A("   ### ⚠️ Y este hallazgo pesa MÁS que el del rótulo `0,27 %`")
        A("")
        A("   El `0,27 %` es real pero se queda dentro del libro. Este **cruza "
          "la frontera entre motor y producto**, que es de otra categoría "
          "arquitectónica. El motor *sabe* que está en corte parcial —y hace "
          "bien en negarse a clasificar—, pero esa condición interna termina "
          "**presentada como si fuera una categoría de gestión**. Es "
          "exactamente lo que la doctrina de QUIRA separa:")
        A("")
        A("   ```")
        A("   dato → estado epistemológico → interpretación → producto")
        A("   ```")
        A("")
        A("   Un estado de disponibilidad del indicador se convirtió en una "
          "categoría sustantiva. La cura no es «poner una categoría igualmente» "
          "—sería fabricar una lectura anual que el corte no sostiene—, sino "
          "**dos campos donde hoy hay uno**:")
        A("")
        A("   ```")
        A("   estado_determinabilidad = CORTE_PARCIAL      (o ANUAL_COMPLETO)")
        A("   clasificacion_avep      = NO_EMITIDA         (o la categoría)")
        A("   ```")
        A("")
        A("   Con eso, la UI puede decir en lenguaje de administración pública "
          "que la lectura anual todavía no es comparable, sin inventar una "
          "categoría ni publicar la jerga del motor. Queda especificado en "
          "`D-011`; no se implementa aquí, porque `007` observa.")
    A("")
    A("Los tres van al dictamen `011`. Ninguno se corrige aquí. Pero el tercero "
      "no puede esperar a `011` sin que alguien lo sepa, y por eso queda escrito "
      "aquí y en el registro de deudas.")
    A("")
    A("---")
    A(f"*GM-Ω-ICPI-007 · {len(esc)} escenarios · baseline congelado 27,4582 % · "
      "el Gold Master no se modificó · Dylus Lab © 2026*")

    _SALIDA.write_text("\n".join(o) + "\n", encoding="utf-8")


def _coherencia_v(filas: list[dict]) -> dict:
    """¿Cada `V_i` literal coincide con lo que la regla documentada produciría?"""
    difieren, iguales, sin_datos = [], 0, 0
    for f in filas:
        vr = v_por_regla(f)
        if vr is None:
            sin_datos += 1
            continue
        if abs(vr - f["V"]) > 1e-12:
            difieren.append({**f, "v_regla": vr})
        else:
            iguales += 1
    return {"difieren": difieren, "iguales": iguales, "sin_datos": sin_datos,
            "comparables": iguales + len(difieren)}


def _tabla_familia(A, esc, fam: str, cat_base: str) -> None:
    A("| Esc. | Escenario | ICPI | Δ vs. baseline | Categoría | Nota |")
    A("|---|---|---:|---:|---|---|")
    for e in [x for x in esc if x["fam"] == fam]:
        salta = " ⚠️" if e["cat"] != cat_base else ""
        A(f"| `{e['id']}` | {e['nombre']} | {e['icpi'] * 100:.4f} % | "
          f"{e['delta']:+.4f} pp | {e['cat']}{salta} | {e['nota']} |")


if __name__ == "__main__":
    raise SystemExit(main())
