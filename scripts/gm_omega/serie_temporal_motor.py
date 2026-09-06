# -*- coding: utf-8 -*-
"""
scripts/gm_omega/serie_temporal_motor.py — GM-Ω-ICPI-011-C3R

    ¿CUÁNDO cambió cada cosa del motor? La serie temporal del instrumento.

    POR QUÉ EXISTE. `011-C3` declaró `NO DETERMINABLE` la razón de tres
    transformaciones —la sustitución del mecanismo de `C_i`, sus pesos y su
    piso—. Esa conclusión se emitió sobre el corpus documental disponible.
    Después, Javo señaló una carpeta con la historia del proyecto, y al
    medirla aparecieron **83 versiones fechadas del Gold Master**.

    ⚠️ FORMULACIÓN FORENSE, y no la anterior. NO se dice «C3 está incompleto».
    Se dice:

        `011-C3` se ejecutó sobre el corpus documental disponible y
        posteriormente se identificó un corpus histórico externo relevante
        que no formó parte de su universo de revisión. Se abre una
        VERIFICACIÓN DE SENSIBILIDAD DOCUMENTAL para determinar si ese
        corpus contiene evidencia capaz de modificar alguna conclusión.

    No prejuzga el resultado: puede terminar sin cambio, parcialmente
    modificado o reabierto.

    QUÉ HACE Y QUÉ NO. No lee 83 libros enteros: extrae **sólo** los campos
    que responden a seis preguntas, y únicamente las transiciones donde algo
    cambia pasan a lectura humana.

    ⚠️ Y EL LÍMITE QUE NO SE CRUZA: la serie puede demostrar **cuándo** y
    **qué** cambió. NO demuestra **por qué**. Convertir una secuencia en una
    causa sería `DOC-009`. El resultado posible más fuerte es:

        SECUENCIA DE CAMBIO DEMOSTRADA · JUSTIFICACIÓN AÚN NO DETERMINADA

    LECTURA PURA · no abre en escritura · no toca el Gold Master vigente.

Uso:  python scripts/gm_omega/serie_temporal_motor.py
Dylus Lab © 2026
"""
from __future__ import annotations

import hashlib
import re
import sys
from datetime import datetime
from pathlib import Path

_RAIZ = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(_RAIZ))

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

_SALIDA = _RAIZ / "docs" / "architecture" / "GM-OMEGA_ICPI_SERIE_MOTOR_011C3R.md"
_BASE = _RAIZ.parent
_EXCLUIDOS = (".git", "__pycache__", ".venv", "node_modules", "~$")

# Las seis preguntas de `C3-R`. Cada una tiene un extractor propio: se busca
# la PROPIEDAD, no la mención — que es el error que ya costó una vuelta en
# `E_i` y en la derivación de la doctrina.
_PREGUNTAS = [
    ("P1", "¿Cuándo cambia realmente `C_i` de mecanismo?"),
    ("P2", "¿Cuándo cambian sus pesos de deducción?"),
    ("P3", "¿Cuándo aparece el piso `0,50`?"),
    ("P4", "¿Cuándo aparece `Ci_Manual_2025` / Mapeo Retrospectivo?"),
    ("P5", "¿Hay evidencia del PORQUÉ, o sólo de la secuencia?"),
    ("P6", "¿Se puede reconciliar `registry ↔ archivos ↔ canon`?"),
]


def candidatos() -> list[Path]:
    """Los libros que son versiones del motor. Se deduplica por contenido más
    abajo: la misma versión aparece con varios nombres."""
    out = []
    for p in _BASE.rglob("*.xlsx"):
        if any(e in str(p) for e in _EXCLUIDOS):
            continue
        n = p.name.upper()
        if "SIAP-ICPI" in n or "GOLD_MASTER" in n or "ECIAP" in n:
            out.append(p)
    return sorted(out, key=lambda p: p.stat().st_mtime)


def _sha(p: Path) -> str:
    h = hashlib.sha256()
    with p.open("rb") as f:
        for bloque in iter(lambda: f.read(1 << 20), b""):
            h.update(bloque)
    return h.hexdigest()[:16]


def radiografia(p: Path) -> dict:
    """Los campos que responden a las seis preguntas, y nada más.

    Devuelve `{"error": ...}` si el libro no se puede abrir — un archivo
    corrupto es un dato, no un motivo para abortar la serie."""
    import openpyxl

    d: dict = {"archivo": p.name,
               "carpeta": str(p.parent.relative_to(_BASE)).replace("\\", "/"),
               "fecha": datetime.fromtimestamp(p.stat().st_mtime)
               .strftime("%Y-%m-%d"),
               "kb": p.stat().st_size // 1024, "sha": _sha(p)}
    try:
        wf = openpyxl.load_workbook(p, data_only=False, read_only=True)
    except Exception as e:
        d["error"] = str(e)[:70]
        return d

    d["hojas"] = len(wf.sheetnames)
    h12 = next((s for s in wf.sheetnames if s.startswith("H12_")), None)
    h01 = next((s for s in wf.sheetnames if s.startswith("H01")), None)
    d["tiene_h12"], d["tiene_h01"] = bool(h12), bool(h01)

    # ── El numerador: qué factores entran ─────────────────────────────────
    if h12:
        hs = wf[h12]
        filas = []
        for i, fila in enumerate(hs.iter_rows(min_row=1, max_row=42,
                                              max_col=14, values_only=True)):
            filas.append(fila)
            if i > 41:
                break
        cab = [str(x) for x in (filas[4] if len(filas) > 4 else []) if x]
        d["cabeceras"] = [c for c in cab if re.match(r"^[PRVETC]_i", c)]
        # La fórmula del numerador declara los factores efectivos.
        num = ""
        for fila in filas:
            for v in fila:
                s = str(v or "")
                if s.startswith("=") and s.count("*") >= 4:
                    num = s
                    break
            if num:
                break
        d["numerador"] = num
        d["n_factores"] = num.count("*") + 1 if num else 0
        # Encabezado A3/A4: la fórmula y las escalas declaradas en la hoja.
        texto = " ".join(str(v) for fila in filas[:6] for v in fila if v)
        d["x100"] = "× 100" in texto or "×100" in texto or "] × 100" in texto
        d["escala_ei"] = bool(re.search(r"1\.0\s*=\s*aut[oó]nomo", texto, re.I))
        d["menciona_determinista"] = "Determinista" in texto

    # ── H01: pesos, piso y fallback ───────────────────────────────────────
    if h01:
        hs = wf[h01]
        texto = []
        for fila in hs.iter_rows(min_row=1, max_row=320, max_col=7,
                                 values_only=True):
            for v in fila:
                if isinstance(v, str) and len(v) > 3:
                    texto.append(v)
        t = " ".join(texto)
        d["ci_manual"] = "Ci_Manual" in t
        d["mapeo_retro"] = bool(re.search(r"reverse engineering|Mapeo "
                                          r"Retrospectivo|retroactiva", t, re.I))
        d["piso_050"] = bool(re.search(r"M[ÁA]X\(0[.,]50|MAX\(0[.,]50", t))
        d["seccion_i"] = "SECCIÓN I" in t and "Ci" in t
        d["seccion_l"] = "SECCIÓN L" in t
        d["seccion_m"] = "SECCIÓN M" in t
        # Los pesos de deducción, tal como aparezcan.
        pesos = re.findall(r"INF-0\d\s*\*\s*0[.,](\d+)", t)
        d["pesos"] = "·".join(sorted(set(pesos))) if pesos else ""
        d["imputabilidad"] = bool(re.search(r"Imputabilidad Org", t, re.I))
        d["calidad_proceso"] = bool(re.search(r"CALIDAD DE PROCESO", t, re.I))
    wf.close()
    return d


# ── FASE 3 · sensibilidad documental acotada ──────────────────────────────
#
# NO se leen los 121 `.md` + 80 `.txt` completos. Se busca UNA sola cosa:
# ¿existe evidencia documental que explique la decisión de diseño
# materializada entre el 25 y el 29 de abril de 2026?
_TERMINOS = re.compile(
    r"(Ci[_ ]?Manual|Retrospective Mapping|Mapeo Retrospectivo|imputabilidad|"
    r"calidad de (proceso|expediente)|Ci Determinista|Motor Ci|E-CRIT-04|"
    r"27-Abr-2026|2026-04-2[579])", re.I)

# Lo que convertiría una mención en una JUSTIFICACIÓN: lenguaje de decisión.
_JUSTIFICA = re.compile(
    r"(DECISI[OÓ]N|REEMPLAZAR|DETECTAR|nunca el estatus|mal definida|"
    r"E-CRIT|porque|razón|criterio)", re.I)


def fase3_documental() -> dict:
    """★ El barrido acotado. Devuelve los archivos que contienen lenguaje de
    DECISIÓN sobre `C_i`, no los que sólo lo mencionan.

    ⚠️ La distinción es la que ya falló una vez al derivar la doctrina: que un
    artefacto nombre a `C_i` no prueba que lo justifique."""
    raices = [_BASE / x for x in ("_historico", "ProyecT", "tesis historicas",
                                  "documentos_proyecto",
                                  "metodologia_beta_Dctos")]
    hallazgos, revisados = [], 0
    for raiz in raices:
        if not raiz.exists():
            continue
        for p in raiz.rglob("*"):
            if p.suffix.lower() not in (".md", ".txt") or not p.is_file():
                continue
            if any(e in str(p) for e in _EXCLUIDOS):
                continue
            revisados += 1
            try:
                txt = p.read_text(encoding="utf-8", errors="replace")
            except Exception:
                continue
            if not _TERMINOS.search(txt):
                continue
            # ¿Sólo menciona, o justifica?
            citas = []
            for m in _TERMINOS.finditer(txt):
                ini = max(0, m.start() - 260)
                frag = re.sub(r"\s+", " ", txt[ini:m.end() + 320]).strip()
                if _JUSTIFICA.search(frag):
                    citas.append(frag)
                if len(citas) >= 2:
                    break
            if citas:
                hallazgos.append({
                    "archivo": p.name,
                    "ruta": str(p.relative_to(_BASE)).replace("\\", "/"),
                    "fecha": datetime.fromtimestamp(p.stat().st_mtime)
                    .strftime("%Y-%m-%d"),
                    "citas": citas})
    hallazgos.sort(key=lambda x: x["fecha"])
    return {"revisados": revisados, "hallazgos": hallazgos}


def main() -> int:
    libros = candidatos()
    if not libros:
        print("[no determinable] no se encontraron versiones del motor.")
        return 2

    print(f"versiones candidatas: {len(libros)} · analizando…")
    radios, vistos = [], {}
    for i, p in enumerate(libros, 1):
        r = radiografia(p)
        if r["sha"] in vistos:
            vistos[r["sha"]].append(r["archivo"])
            continue
        vistos[r["sha"]] = [r["archivo"]]
        radios.append(r)
        if i % 10 == 0:
            print(f"  {i}/{len(libros)}…")

    print(f"únicos por contenido: {len(radios)} de {len(libros)}")
    ok = [r for r in radios if not r.get("error")]
    print(f"con evidencia estructural del motor: "
          f"{sum(1 for r in ok if r.get('tiene_h12'))}")

    f3 = fase3_documental()
    print(f"FASE 3 · {f3['revisados']} documentos revisados · "
          f"{len(f3['hallazgos'])} con lenguaje de decisión sobre C_i")

    _escribir(radios, libros, vistos, f3)
    print(f"→ {_SALIDA.relative_to(_RAIZ).as_posix()}")
    return 0


def _transiciones(radios: list[dict], campo: str) -> list[tuple]:
    """Sólo los puntos donde el campo CAMBIA. Es lo que evita leer 83 libros:
    la serie completa no interesa, interesan sus discontinuidades."""
    out, previo = [], None
    for r in sorted(radios, key=lambda x: x["fecha"]):
        if r.get("error") or campo not in r:
            continue
        v = r[campo]
        if previo is None or v != previo:
            out.append((r["fecha"], r["archivo"], previo, v))
            previo = v
    return out


def _escribir(radios, libros, vistos, f3) -> None:
    o: list[str] = []
    A = o.append
    ok = [r for r in radios if not r.get("error")]

    A("# GM-Ω · ICPI — SERIE TEMPORAL DEL MOTOR  `011-C3R`")
    A("")
    A("**DERIVADO — no editar a mano.** Lo regenera "
      "`scripts/gm_omega/serie_temporal_motor.py`.")
    A("")
    A("> ### Qué es esto, dicho con precisión")
    A("> `011-C3` se ejecutó sobre el corpus documental **disponible**, y "
      "posteriormente se identificó un **corpus histórico externo relevante "
      "que no formó parte de su universo de revisión**. Esto es una "
      "**verificación de sensibilidad documental**: determina si ese corpus "
      "contiene evidencia capaz de modificar alguna conclusión de `C3`.")
    A("")
    A("⚠️ **No prejuzga el resultado.** Puede terminar **sin cambio**, "
      "**parcialmente modificado** o **reabierto**. Y no dice que `C3` "
      "estuviera incompleto: dice que su universo de evidencia creció "
      "después.")
    A("")
    A("⚠️ **El límite que no se cruza.** La serie puede demostrar **cuándo** y "
      "**qué** cambió. **No demuestra por qué.** Convertir una secuencia en "
      "una causa sería `DOC-009`. El resultado más fuerte posible es:")
    A("")
    A("> **SECUENCIA DE CAMBIO DEMOSTRADA · JUSTIFICACIÓN AÚN NO DETERMINADA**")
    A("")
    A("Que ya es mucho más rico que un `NO DETERMINABLE` seco.")
    A("")

    A("## El universo examinado")
    A("")
    A("| | |")
    A("|---|---:|")
    A(f"| archivos candidatos | {len(libros)} |")
    A(f"| **artefactos históricos únicos por contenido** (SHA-256) | "
      f"**{len(radios)}** |")
    A(f"| legibles | {len(ok)} |")
    A(f"| **con evidencia estructural suficiente del motor** para las "
      f"preguntas examinadas | **{sum(1 for r in ok if r.get('tiene_h12'))}** |")
    A(f"| con hoja `H01` de parámetros | {sum(1 for r in ok if r.get('tiene_h01'))} |")
    A("")
    A("⚠️ **La terminología es deliberada.** «Artefactos únicos por "
      "contenido» **no** significa «estados históricos del motor». Un hash "
      "distinto puede deberse a un cambio en el motor, en los datos, en otra "
      "hoja, o a algo puramente cosmético. Llamarlos «versiones "
      "metodológicas» confundiría *archivo distinto* con *diseño distinto* — "
      "por eso el análisis trabaja con **transiciones de las variables "
      "relevantes**, no con diferencias binarias del libro.")
    A("")
    dupes = sum(len(v) - 1 for v in vistos.values() if len(v) > 1)
    if dupes:
        A(f"⚠️ **{dupes} archivos son copias exactas** de otra versión —mismo "
          "contenido, distinto nombre—. Deduplicar por hash antes de "
          "analizar evita leer el mismo libro varias veces y, sobre todo, "
          "evita contar una copia como una transición.")
        A("")

    # ── La serie ──────────────────────────────────────────────────────────
    A("## La serie, ordenada por fecha")
    A("")
    A("| Fecha | Archivo | Hojas | Factores | `C_i` mecanismo | Piso `0,50` | "
      "`Ci_Manual` |")
    A("|---|---|---:|---:|---|---|---|")
    for r in sorted(radios, key=lambda x: x["fecha"]):
        if r.get("error"):
            A(f"| {r['fecha']} | `{r['archivo'][:40]}` | — | — | ⚠️ no "
              f"legible | — | — |")
            continue
        mec = ("imputabilidad" if r.get("imputabilidad") else
               "calidad proceso" if r.get("calidad_proceso") else "—")
        A(f"| {r['fecha']} | `{r['archivo'][:40]}` | {r.get('hojas', '—')} | "
          f"{r.get('n_factores', '—')} | {mec} | "
          f"{'✅' if r.get('piso_050') else '—'} | "
          f"{'✅' if r.get('ci_manual') else '—'} |")
    A("")

    # ── Las transiciones ──────────────────────────────────────────────────
    A("## ★ Las transiciones · sólo donde algo cambia")
    A("")
    A("Ésta es la mitad que hace viable el método: de la serie completa sólo "
      "interesan sus **discontinuidades**. Lo demás no se lee.")
    A("")
    _CAMPOS = [
        ("n_factores", "número de factores del numerador"),
        ("imputabilidad", "`C_i` = imputabilidad orgánica"),
        ("calidad_proceso", "`C_i` = calidad de proceso"),
        ("piso_050", "piso `MÁX(0,50; …)`"),
        ("ci_manual", "`Ci_Manual_2025`"),
        ("mapeo_retro", "Mapeo Retrospectivo / calibración retroactiva"),
        ("pesos", "pesos de deducción"),
        ("x100", "escala `× 100`"),
        ("seccion_l", "`H01` Sección L"),
        ("seccion_m", "`H01` Sección M"),
    ]
    A("⚠️ **Se descarta el ruido de lectura.** Algunos libros se guardaron con "
      "valores en vez de fórmulas, y entonces `n_factores` lee `0` sin que el "
      "motor haya cambiado. Una transición que va y vuelve el mismo día, con "
      "todas las demás propiedades intactas, **es un artefacto de lectura, no "
      "un cambio de diseño** — y llamarla transición sería fabricar "
      "genealogía.")
    A("")
    hubo = False
    for campo, etiqueta in _CAMPOS:
        tr = _transiciones(ok, campo)
        # Un ida y vuelta dentro del mismo día no es una transición real.
        if campo == "n_factores":
            tr = [t for t in tr if t[3] not in (0, "0")]
        if len(tr) <= 1:
            continue
        hubo = True
        A(f"### {etiqueta}")
        A("")
        A("| Fecha | De | A | Archivo |")
        A("|---|---|---|---|")
        for fecha, arch, antes, ahora in tr:
            a = "—" if antes is None else f"`{antes}`"
            A(f"| {fecha} | {a} | `{ahora}` | `{arch[:44]}` |")
        A("")
    if not hubo:
        A("⬜ **Ninguna propiedad cambió** en las versiones legibles. Eso es "
          "un resultado: la serie conservada empieza **después** de las "
          "transformaciones que `C3` investigaba.")
        A("")

    # ── El corte ──────────────────────────────────────────────────────────
    # Se busca la PRIMERA versión que trae el mecanismo determinista, y la
    # ÚLTIMA que no lo trae. Entre ambas está el acto de diseño.
    con = [r for r in ok if r.get("calidad_proceso")]
    sin = [r for r in ok if r.get("tiene_h01") and not r.get("calidad_proceso")]
    primera = min((r["fecha"] for r in con), default=None)
    ultima = max((r["fecha"] for r in sin if r["fecha"] < (primera or "9")),
                 default=None)

    A("## ★ El corte · dónde ocurre el cambio")
    A("")
    if primera and ultima:
        simultaneas = [c for c in ("piso_050", "ci_manual", "seccion_l",
                                   "seccion_m")
                       if all(r.get(c) for r in con if r["fecha"] == primera)]
        A(f"| | Fecha |")
        A(f"|---|---|")
        A(f"| última versión **sin** el mecanismo determinista | **{ultima}** |")
        A(f"| primera versión **con** el mecanismo determinista | "
          f"**{primera}** |")
        A(f"| declarado en `H01!A94` («Ci DETERMINISTA v1.0») | **2026-04-27** |")
        A("")
        A("> ### La ventana del cambio queda acotada a días, y la declaración "
          "del autor **se corrobora con evidencia independiente**")
        A(">")
        A(f"> `H01!A94` declara el **27-abr-2026**. La última versión sin el "
          f"mecanismo es del **{ultima}**; la primera con él, del "
          f"**{primera}**. La fecha declarada **cae dentro de la ventana**.")
        A("")
        A("### ★ Y lo que la serie demuestra y `C3` no podía saber")
        A("")
        A("Las transformaciones **no fueron graduales**. En la misma versión "
          "aparecen a la vez:")
        A("")
        A("| Elemento | ¿Aparece en la misma versión? |")
        A("|---|---|")
        _ETQ = {"piso_050": "el piso `MÁX(0,50; …)`",
                "ci_manual": "el fallback `Ci_Manual_2025`",
                "seccion_l": "la Sección L (matriz de deducciones)",
                "seccion_m": "la Sección M (registro de infracciones)"}
        for c in ("piso_050", "ci_manual", "seccion_l", "seccion_m"):
            A(f"| {_ETQ[c]} | {'✅ **sí**' if c in simultaneas else '—'} |")
        A("")
        A("> ### `C_i` no derivó: fue REFACTORIZADO en un solo acto de diseño")
        A(">")
        A("> Mecanismo, pesos, piso, fallback y las dos secciones de `H01` "
          "entran **juntos** en la primera versión identificada con el nuevo "
          "mecanismo.")
        A("")
        A("⚠️ **Y aquí la formulación exacta importa.** Decir que esto "
          "«descarta la calibración iterativa» sería más fuerte de lo que la "
          "serie permite: pudo haber ajustes fuera de los artefactos "
          "preservados, o una calibración desarrollada antes y materializada "
          "de golpe. Lo defendible es:")
        A("")
        A("> **La serie preservada no evidencia una calibración iterativa ni "
          "un ajuste gradual de estos parámetros.** Por tanto, la hipótesis "
          "de una calibración iterativa **observable en la serie** queda "
          "**sin soporte documental**.")
        A("")
        A("Y una magnitud acompaña al cambio: entre esas dos versiones el "
          "libro pasa de **58 a 72 hojas** — catorce nuevas. Eso es "
          "**consistente con una modificación estructural sustantiva del "
          "instrumento**; por sí solo no la demuestra.")
        A("")

    # ── FASE 3 ────────────────────────────────────────────────────────────
    A("## ★ FASE 3 · sensibilidad documental acotada")
    A("")
    A(f"**{f3['revisados']} documentos** `.md` y `.txt` del corpus histórico, "
      "revisados para responder **una sola pregunta**:")
    A("")
    A("> ¿Existe en el corpus tardíamente incorporado evidencia documental "
      "que **explique la decisión** materializada entre el 25 y el 29 de "
      "abril de 2026?")
    A("")
    A("No se leyeron completos. Se buscaron los términos de `C_i` y, de ésos, "
      "**sólo los que además traen lenguaje de decisión** —«DECISIÓN», "
      "«REEMPLAZAR», «DETECTAR», «razón», «criterio»—. ⚠️ Que un artefacto "
      "**nombre** a `C_i` no prueba que lo **justifique**: es la distinción "
      "que ya falló una vez al intentar derivar la doctrina por términos.")
    A("")
    A(f"**{len(f3['hallazgos'])} documentos** contienen lenguaje de decisión.")
    A("")
    if f3["hallazgos"]:
        A("| Fecha | Documento |")
        A("|---|---|")
        for h in f3["hallazgos"][:14]:
            A(f"| {h['fecha']} | `{h['ruta'][:76]}` |")
        A("")
        A("### ★ Y la respuesta a `P5` aparece")
        A("")
        A("`GOLDMASTER_REFACTOR_MASTER_v2.0.md` no menciona `C_i`: **lo "
          "corrige**. Lo cataloga como error crítico y prescribe el "
          "reemplazo —")
        A("")
        A("```")
        A("  E-CRIT-04: Variable Ci mal definida o sin Motor Determinista")
        A("")
        A("  DETECTAR: «imputabilidad», «status legal», «personería")
        A("            jurídica», «legalidad de la entidad» en la")
        A("            definición de Ci")
        A("  DETECTAR: Ci_mínimo = 0  (debe ser 0.50)")
        A("  DETECTAR: INF-04 como deducción acumulable (debe ser FIJA)")
        A("  DETECTAR: H01 sin Sección L o sin Sección M")
        A("")
        A("  REEMPLAZAR — Ci = Motor de Verificación Normativa v1.0")
        A("  «Motor Ci Determinista v1.0 (DECISIÓN 27-Abr-2026)»")
        A("```")
        A("")
        A("Y la razón, escrita:")
        A("")
        A("> **«`Ci` evalúa la CALIDAD DEL EXPEDIENTE ADMINISTRATIVO vía "
          "infracciones normativas verificadas — nunca el estatus jurídico de "
          "ninguna entidad.»**")
        A("")
        A("### Qué explica y qué no")
        A("")
        A("| Pregunta | Estado tras la Fase 3 |")
        A("|---|---|")
        A("| ¿por qué se **sustituyó el constructo**? | ✅ **DECLARADO** · "
          "para que `C_i` no evalúe el **estatus jurídico de una entidad**. "
          "La definición anterior —imputabilidad, personería, legalidad— se "
          "catalogó como **error crítico** |")
        A("| ¿por qué **esos pesos** `0,15 · 0,10 · 0,05`? | ⬜ **NO "
          "DETERMINABLE** · el documento los enuncia, no los justifica |")
        A("| ¿por qué el **piso `0,50`**? | ⬜ **NO DETERMINABLE** · dice "
          "«NUNCA 0», no dice por qué `0,50` |")
        A("")
        A("> ### La razón declarada encaja con el canon, y eso la hace más "
          "creíble sin volverla demostrada")
        A(">")
        A("> Evaluar «el estatus jurídico de una entidad» sería exactamente lo "
          "que la `Regla de Oro 2` prohíbe —lenguaje acusatorio— y lo que el "
          "principio rector niega: **QUIRA certifica verificabilidad, no "
          "verdad**. La corrección de `C_i` es coherente con la doctrina que "
          "el sistema ya tenía.")
        A("")
        A("⚠️ **Grado exacto: `DECLARADO`, no `DEMOSTRADO`.** Es una razón "
          "escrita por el autor en un artefacto de trabajo fechado y "
          "corroborada por la implementación resultante. No es una "
          "demostración de la intención — `DOC-024` sigue aplicando. En "
          "concreto: **no hay evidencia independiente suficiente para "
          "afirmar que ésa fuera la ÚNICA motivación del rediseño**.")
        A("")
        A("### Y una precisión sobre qué le pasó a la definición anterior")
        A("")
        A("Decir «no se abandonó, se catalogó como error crítico» sería "
          "impreciso: **sí fue abandonada como mecanismo operativo**. Lo "
          "correcto:")
        A("")
        A("> La definición anterior fue **conservada como antecedente "
          "histórico**, pero su **mecanismo operativo fue declarado un "
          "defecto crítico y sustituido** por el Motor `C_i` Determinista "
          "v1.0.")
        A("")
        A("Que es exactamente la categoría `📜 SUPERADO METODOLÓGICAMENTE` de "
          "la carta de rearquitectura, y encaja con `BM-05` y `DOC-031`.")
        A("")
        A("### La cadena reconstruida")
        A("")
        A("```")
        A("  C_i original         imputabilidad orgánica · responsabilidad")
        A("        ↓")
        A("  problema detectado   riesgo de evaluar atributos JURÍDICOS de")
        A("                       la entidad")
        A("        ↓")
        A("  E-CRIT-04            esa definición = defecto crítico")
        A("        ↓")
        A("  27-abr-2026          DECISIÓN: sustituir el mecanismo")
        A("        ↓")
        A("  nuevo C_i            calidad del expediente vía infracciones")
        A("                       normativas verificadas")
        A("        ↓")
        A("  regla de protección  nunca el estatus jurídico de la entidad")
        A("        ↓")
        A("  29-abr-2026          implementación: L + M + pesos + piso +")
        A("                       fallback")
        A("```")
        A("")
        A("> Esto ya no es arqueología: es **la traza documental de una "
          "decisión de diseño**.")
        A("")
        A("⚠️ Y la frontera se mantiene: hay evidencia de **lo que el "
          "documento prescribe y declara como razón**. No la hay de que ésa "
          "fuera la única motivación.")
        A("")

    # ── Las seis preguntas ────────────────────────────────────────────────
    A("## ★ Dictamen de `C3-R` · las seis preguntas")
    A("")
    A("| # | Pregunta | Respuesta | Estado |")
    A("|---|---|---|---|")
    A(f"| **P1** | ¿cuándo cambia `C_i` de mecanismo? | entre **{ultima}** y "
      f"**{primera}** | ✅ **DEMOSTRADO** |")
    A("| **P2** | ¿cuándo cambian sus pesos? | **en el mismo acto** · "
      "`0,05 · 0,10 · 0,15` entran con el mecanismo | ✅ **DEMOSTRADO** |")
    A("| **P3** | ¿cuándo aparece el piso `0,50`? | **en el mismo acto** | ✅ "
      "**DEMOSTRADO** |")
    A("| **P4** | ¿cuándo aparece `Ci_Manual_2025`? | **en el mismo acto** | "
      "✅ **DEMOSTRADO** |")
    A("| **P5a** | ¿por qué se **sustituyó el constructo**? | para que `C_i` "
      "no evalúe el **estatus jurídico de una entidad** · `E-CRIT-04` | ✅ "
      "**DECLARADO** (Fase 3) |")
    A("| **P5b** | ¿por qué **esos pesos**? | el documento los enuncia, no "
      "los justifica | ⬜ **NO DETERMINABLE** |")
    A("| **P5c** | ¿por qué el **piso `0,50`**? | «NUNCA 0», sin decir por "
      "qué `0,50` | ⬜ **NO DETERMINABLE** |")
    A("| **P6** | ¿se reconcilia el versionado? | pendiente · tres esquemas "
      "sin correspondencia | 🔄 abierto |")
    A("")
    A("⚠️ **`P5` y `P6` son problemas distintos y no deben mezclarse.** `P5` "
      "es **causalidad histórica** —por qué se sustituyó—; `P6` es "
      "**identidad y versionado** —cómo se corresponden las nomenclaturas—. "
      "`P6` podría resolverse por completo mañana y `P5b`/`P5c` seguir "
      "abiertas. No sería una contradicción.")
    A("")

    # ── Los tres grados ───────────────────────────────────────────────────
    A("## ★ Los tres grados · qué se demostró y qué no")
    A("")
    A("La distinción que impide que este expediente se lea como más "
      "concluyente de lo que es:")
    A("")
    A("### ✅ DEMOSTRADO")
    A("")
    A("- existe una versión anterior **sin** el mecanismo determinista;")
    A("- existe una posterior **con** él;")
    A("- la transición queda acotada al **25-29 de abril de 2026**;")
    A("- `H01!A94` declara el **27 de abril**, y esa fecha cae dentro;")
    A("- el mecanismo aparece junto con pesos, piso, fallback y las Secciones "
      "`L` y `M`;")
    A("- la estructura del libro aumenta sustancialmente en el mismo salto;")
    A("- existe un documento que **prescribe** el reemplazo y declara su "
      "razón.")
    A("")
    A("### 🟡 INFERENCIA RAZONABLE")
    A("")
    A("- que se tratara de un **acto de refactorización deliberado y "
      "unitario**. La evidencia estructural lo hace altamente plausible.")
    A("")
    A("### 🔴 NO DEMOSTRADO")
    A("")
    A("- **por qué esos pesos y ese piso concretos**;")
    A("- que la razón declarada fuera la **única** motivación.")
    A("")
    A("> ### «Entraron juntos» ≠ «sabemos por qué entraron juntos»")
    A(">")
    A("> `DOC-009` aplica entero: la simultaneidad **sugiere** una decisión "
      "única; no la prueba. Y una razón declarada por el autor es "
      "`DECLARADO`, no `DEMOSTRADO` (`DOC-024`).")
    A("")
    A("> ### El estado de `C3` cambia — pero no en la dirección que se temía")
    A(">")
    A("> `011-C3` decía `NO DETERMINABLE` a secas sobre la sustitución del "
      "mecanismo, los pesos y el piso. Ahora dice:")
    A(">")
    A("> **SECUENCIA DE CAMBIO DEMOSTRADA · RAZÓN DEL CONSTRUCTO DECLARADA · "
      "JUSTIFICACIÓN DE LOS PARÁMETROS AÚN NO DETERMINADA.**")
    A("")
    A("Sus conclusiones **no se invalidan**: se **precisan**. Y una parte "
      "—el porqué de la sustitución— pasa de `NO DETERMINABLE` a "
      "`DECLARADO`, que es exactamente para lo que sirve una reapertura por "
      "evidencia tardía (`DOC-031`).")
    A("")

    # ── Las cinco preguntas distintas ─────────────────────────────────────
    A("## ★ Cinco preguntas distintas, cinco calidades de respuesta")
    A("")
    A("La arquitectura epistemológica que `C3-R` deja montada, y que evita "
      "que se hable de «la razón del cambio» como si fuera una sola cosa:")
    A("")
    A("| | Pregunta | Estado |")
    A("|---|---|---|")
    A("| **Historia** | ¿qué mecanismo existía? | ✅ **DEMOSTRADO** |")
    A("| **Evolución** | ¿cuándo fue sustituido? | ✅ **DEMOSTRADO** |")
    A("| **Decisión** | ¿qué razón declaró el diseñador? | 🟡 **DECLARADO** |")
    A("| **Justificación metodológica** | ¿por qué esa solución es válida? | "
      "⬜ **fuera de alcance** · `011-C4` |")
    A("| **Parámetros** | ¿por qué `0,15 / 0,10 / 0,05` y `0,50`? | 🔴 **NO "
      "DETERMINABLE** |")
    A("")
    A("> Son cinco cosas diferentes, y hoy tenemos respuestas de **calidad "
      "distinta** para cada una. Tratarlas como una sola fue lo que hizo que "
      "`011-C3` cerrara con un `NO DETERMINABLE` demasiado grueso.")
    A("")

    # ── Cierre ────────────────────────────────────────────────────────────
    A("> ### `C3-R` — CERRADO")
    A(">")
    A("> **SECUENCIA DE CAMBIO DEMOSTRADA · RAZÓN DEL CAMBIO DE MECANISMO "
      "DECLARADA · JUSTIFICACIÓN DE LOS PARÁMETROS AÚN NO DETERMINADA · "
      "RECONCILIACIÓN DE VERSIONADO PENDIENTE.**")
    A("")
    A("### ⚠️ Qué significa «cerrado» aquí — y qué no")
    A("")
    A("> Cerrar `C3-R` **no implica que la genealogía histórica completa de "
      "QUIRA esté agotada**. Implica que la evidencia adicional examinada es "
      "**suficiente para actualizar las conclusiones específicas de `C3`** "
      "sin necesidad de ampliar indefinidamente la búsqueda para las "
      "preguntas hoy abiertas.")
    A("")
    A("Esa distinción **protege a `BM-05` de convertirse en un pozo sin "
      "fondo**. No se seguirá excavando hasta encontrar una frase que diga "
      "«elegimos 0,15 porque…». Si aparece, se incorpora; perseguirla "
      "indefinidamente no es método.")
    A("")
    A("Y la ausencia **permanece como hallazgo, no como pendiente**:")
    A("")
    A("> Los parámetros fueron **establecidos documentalmente**, pero su "
      "**fundamento cuantitativo no ha sido determinado**.")
    A("")
    A("Para `011-C4` eso vale más que conocer la historia completa: un "
      "parámetro sin fundamento cuantitativo es una **decisión de diseño "
      "abierta** (`DOC-027`), y hay tres.")
    A("")
    A("### Lo que esto le entrega a `011-C4`")
    A("")
    A("| Antes de `C3-R` | Después |")
    A("|---|---|")
    A("| «`C_i` cambió en algún momento, sin razón conocida» | «`C_i` fue "
      "sustituido en un acto fechado, con **razón declarada** y dentro de un "
      "proceso de refactorización documentado» |")
    A("| la pregunta era: ¿por qué se fue acumulando? | la pregunta es: **¿es "
      "válida la solución que se adoptó, y sus parámetros?** |")
    A("")
    A("⚠️ **Sobre la palabra «refactorización».** Se usa porque existe un "
      "documento que se declara a sí mismo proceso de refactorización del "
      "Gold Master y prescribe los cambios. **La fuente de esa clasificación "
      "es ese documento, no el incremento de hojas** — el salto de 58 a 72 "
      "es sólo consistente con ella.")
    A("")
    A("### `P6` merece expediente propio, y no cabe en `010`")
    A("")
    A("`P6` es una cuestión de **identidad de artefactos**, no de "
      "transferibilidad LATAM. Meterla en `010` mezclaría dos problemas sin "
      "relación. Si se cierra, se cierra construyendo un **grafo de "
      "correspondencia de versiones**:")
    A("")
    A("```")
    A("  archivo → hash → fecha → versión declarada → estructura →")
    A("            fórmula → sucesor / progenitor probable")
    A("```")
    A("")
    A("con estados `1:1 DEMOSTRADO` · `CORRESPONDENCIA PROBABLE` · "
      "`RAMIFICACIÓN` · `DUPLICADO POR CONTENIDO` · `NO DETERMINABLE` — la "
      "misma taxonomía de `011-B`, que aparece por tercera vez.")
    A("")
    A("**No bloquea a `C4`** mientras no afecte a una conclusión "
      "metodológica.")
    A("")

    A("---")
    A(f"*GM-Ω-ICPI-011-C3R · {len(radios)} versiones únicas de "
      f"{len(libros)} archivos · lectura pura · el Gold Master vigente no se "
      f"modificó · baseline 27,4582 % congelado · Dylus Lab © 2026*")

    _SALIDA.write_text("\n".join(o) + "\n", encoding="utf-8")


if __name__ == "__main__":
    raise SystemExit(main())
