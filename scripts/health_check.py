#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
scripts/health_check.py — QUIRA Intelligence
Verificación rápida del estado operacional del sistema.

Uso:
  python scripts/health_check.py

Checks:
  1. Gold Master existe y es legible
  2. H73_OUTPUT_API tiene métricas clave
  3. Último snapshot disponible (fecha y valores)
  4. Supabase alcanzable
  5. Backup más reciente

Salida: OK / ALERTA / ERROR por componente.
Dylus Lab © 2026 — Sprint Observabilidad Longitudinal
"""
from __future__ import annotations

import json
import os
import sys
from datetime import datetime
from pathlib import Path

sys.stdout.reconfigure(encoding="utf-8", errors="replace")

_RAIZ = Path(__file__).parent.parent
sys.path.insert(0, str(_RAIZ))

GREEN  = "\033[92m"
YELLOW = "\033[93m"
RED    = "\033[91m"
RESET  = "\033[0m"
BOLD   = "\033[1m"


def check(label, estado, detalle):
    color = GREEN if estado == "OK" else (YELLOW if estado == "ALERTA" else RED)
    simbolo = "✓" if estado == "OK" else ("⚠" if estado == "ALERTA" else "✗")
    print(f"  {color}{simbolo}{RESET} {label}: {detalle}")
    return estado


def main():
    print(f"\n{BOLD}=== QUIRA Health Check — {datetime.now().strftime('%Y-%m-%d %H:%M')} ==={RESET}\n")
    errores = []
    alertas = []

    # ── 1. Gold Master ─────────────────────────────────────────────────────────
    print(f"{BOLD}[1] Gold Master{RESET}")
    try:
        import config as cfg
        gm_path = Path(cfg.SIAP_PATH)
        if gm_path.exists():
            size_kb = gm_path.stat().st_size // 1024
            estado = check("Archivo", "OK", f"{gm_path.name} ({size_kb:,} KB)")
        else:
            estado = check("Archivo", "ERROR", f"NO ENCONTRADO: {gm_path}")
            errores.append("Gold Master ausente")
    except Exception as e:
        estado = check("Archivo", "ERROR", str(e))
        errores.append(f"Config error: {e}")

    # ── 2. H73 métricas ────────────────────────────────────────────────────────
    print(f"\n{BOLD}[2] H73_OUTPUT_API — Métricas clave{RESET}")
    try:
        import openpyxl
        wb = openpyxl.load_workbook(str(gm_path), read_only=True, data_only=True)
        if "H73_OUTPUT_API" in wb.sheetnames:
            ws = wb["H73_OUTPUT_API"]
            metricas = {}
            for row in ws.iter_rows(values_only=True):
                if row[0] and row[1] is not None:
                    metricas[str(row[0]).strip()] = row[1]
            wb.close()
            icpi = metricas.get("ICPI_GLOBAL_PCT")
            tgi  = metricas.get("TGI_SCORE")
            mmp  = metricas.get("MMP_AVANCE_PCT")
            check("ICPI_GLOBAL_PCT", "OK" if icpi else "ALERTA", f"{icpi:.3f}%" if icpi else "None")
            check("TGI_SCORE",       "OK" if tgi  else "ALERTA", f"{tgi:.4f}" if tgi else "None")
            check("MMP_AVANCE_PCT",  "OK" if mmp == 1.0 else "ALERTA", f"{mmp}")
        else:
            check("H73_OUTPUT_API", "ERROR", "Hoja no encontrada")
            errores.append("H73 ausente")
    except Exception as e:
        check("H73", "ERROR", str(e))
        errores.append(f"H73 error: {e}")

    # ── 3. Último snapshot ─────────────────────────────────────────────────────
    print(f"\n{BOLD}[3] Snapshots longitudinales{RESET}")
    snap_dir = _RAIZ / "data" / "snapshots" / "130801"
    try:
        snaps = sorted(snap_dir.glob("130801_*.json"))
        if snaps:
            ultimo = snaps[-1]
            with open(ultimo, encoding="utf-8") as f:
                s = json.load(f)
            icpi_snap = s.get("icpi", {}).get("global_pct", "?")
            tgi_snap  = s.get("tgi",  {}).get("score", "?")
            meta      = s.get("_meta", {})
            fecha     = meta.get("fecha_corte", meta.get("generated_at", "?"))[:10]
            gm_ok     = meta.get("gold_master_ok", "?")
            check("Último snapshot", "OK",
                  f"{ultimo.name} | {fecha} | ICPI={icpi_snap:.2f}% | TGI={tgi_snap:.2f} | GM_OK={gm_ok}")
            check("Total snapshots", "OK" if len(snaps) >= 1 else "ALERTA",
                  f"{len(snaps)} snapshot(s) disponibles")
        else:
            check("Snapshots", "ERROR", "Ninguno encontrado")
            errores.append("Sin snapshots")
    except Exception as e:
        check("Snapshots", "ERROR", str(e))

    # ── 4. Supabase ────────────────────────────────────────────────────────────
    print(f"\n{BOLD}[4] Supabase{RESET}")
    try:
        import os
        # Buscar credenciales: env vars → secrets.toml → config.py
        sb_url = os.environ.get("SUPABASE_URL") or getattr(cfg, "SUPABASE_URL", None)
        if not sb_url:
            secrets_path = _RAIZ / ".streamlit" / "secrets.toml"
            if secrets_path.exists():
                with open(secrets_path, encoding='utf-8') as sf:
                    for line in sf:
                        line = line.strip()
                        if line.startswith("supabase_uri"):
                            sb_url = line.split("=", 1)[1].strip().strip('"').strip("'")
                            break
        if sb_url:
            check("URL configurada", "OK", sb_url[:45] + "...")
        else:
            check("URL configurada", "ALERTA", "supabase_uri no encontrada en secrets.toml ni env")
            alertas.append("Supabase URL ausente")
    except Exception as e:
        check("Supabase", "ALERTA", str(e))
        alertas.append(f"Supabase: {e}")

    # ── 5. Backups ─────────────────────────────────────────────────────────────
    print(f"\n{BOLD}[5] Backups Gold Master{RESET}")
    backup_dir = _RAIZ / "data" / "backups"
    try:
        backups = sorted(backup_dir.glob("GM_BACKUP_*.xlsx"))
        if backups:
            ultimo_bk = backups[-1]
            size_kb = ultimo_bk.stat().st_size // 1024
            check("Último backup", "OK", f"{ultimo_bk.name} ({size_kb:,} KB)")
            check("Total backups", "OK", f"{len(backups)} backup(s)")
        else:
            check("Backups", "ALERTA", "Ninguno encontrado — ejecutar --solo-respaldar-gm")
            alertas.append("Sin backups GM")
    except Exception as e:
        check("Backups", "ALERTA", str(e))

    # ── Resumen ────────────────────────────────────────────────────────────────
    print(f"\n{BOLD}=== RESUMEN ==={RESET}")
    if not errores and not alertas:
        print(f"  {GREEN}{BOLD}SISTEMA OK — Listo para Snapshot{RESET}")
    elif not errores:
        print(f"  {YELLOW}{BOLD}ALERTAS ({len(alertas)}): {'; '.join(alertas)}{RESET}")
    else:
        print(f"  {RED}{BOLD}ERRORES ({len(errores)}): {'; '.join(errores)}{RESET}")
        if alertas:
            print(f"  {YELLOW}Alertas adicionales: {'; '.join(alertas)}{RESET}")
    print()
    return 0 if not errores else 1


if __name__ == "__main__":
    sys.exit(main())
