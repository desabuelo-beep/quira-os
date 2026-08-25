# -*- coding: utf-8 -*-
"""
QUIRA OS — Ingesta PAC GAD 2026 (SERCOP) → silo H06 (sección DETALLE PROCESOS)
═══════════════════════════════════════════════════════════════════════════════
Pipeline Canon. Parsea el PAC oficial (PDF SERCOP · por POSICIÓN de columna, ya que
el texto envuelve — la columna V.Total está en x1~732, 2 decimales). Escribe los
91 procesos crudos como sección de detalle en H06, DEBAJO de las 25 metas curadas.

Validado: suma V.Total itemizada = $19,735,931.53 (los 91 procesos). El total
oficial $29.85M incluye régimen especial no itemizado. Sobre COPIA WORK.
REQUIERE recálculo COM posterior. Dylus Lab © 2026
"""
import re
import shutil

import openpyxl
import pdfplumber

try:
    from config import DATOS_DIR as _DATOS
except Exception:                                        # noqa: BLE001
    import os as _os
    from pathlib import Path as _P
    _DATOS = _P(_os.environ.get("QUIRA_DATOS", "."))

BASE = str(_DATOS)
PAC = (BASE + r"\Holding_Municipal_Montecristi\PAC 2023-2026\GAD Montecristi"
       r"\GAD_Montecristi_PAC_2026.pdf")
TGI = BASE + r"\SIAP-ICPI_GOLD_MASTER_v5.5_TGI.xlsx"
WORK = BASE + r"\SIAP-ICPI_GOLD_MASTER_v5.5_WORK_20260623_CANON-SPRINT.xlsx"

_PROC = re.compile(r"(Subasta Inversa|Menor Cuant\w+|[ÍI]nfima Cuant\w+|Cat[aá]logo|"
                   r"Licitaci[oó]n|Cotizaci[oó]n|Consultor\w+|R[eé]gimen Especial)", re.I)


def parse_pac() -> list:
    procs = []
    with pdfplumber.open(PAC) as pdf:
        for pg in pdf.pages:
            words = pg.extract_words()
            vts = [w for w in words if re.match(r"^[\d,]+\.\d{2}$", w["text"]) and 724 <= w["x1"] <= 744]
            for tok in vts:
                top = tok["top"]
                rw = sorted([w for w in words if abs(w["top"] - top) < 3.5], key=lambda z: z["x0"])
                text = " ".join(w["text"] for w in rw)
                monto = float(tok["text"].replace(",", ""))
                cm = re.search(r"\b(\d{9})\b", text)
                cpc = cm.group(1) if cm else ""
                tm = re.search(r"\b(Bien|Obra|Servicio|Consultor[ií]a)\b", text)
                tipo = tm.group(1) if tm else ""
                pm = _PROC.search(text)
                proc = pm.group(1).title() if pm else ""
                dm = re.search(r"[A-ZÁÉÍÓÚÑ][A-ZÁÉÍÓÚÑ0-9\s,\.\-]{8,}", text)
                desc = re.split(r"\s\d+[.,]\d{2}", dm.group(0))[0].strip()[:55] if dm else ""
                procs.append({"cpc": cpc, "tipo": tipo, "proc": proc, "desc": desc, "monto": monto})
    return procs


def main() -> None:
    shutil.copy(TGI, WORK)
    procs = parse_pac()
    wb = openpyxl.load_workbook(WORK)
    h06 = next(s for s in wb.sheetnames if s.startswith("H06"))
    ws = wb[h06]
    start = 62                                   # debajo de las metas curadas (33-57)
    ws.cell(start, 1).value = ("▌ DETALLE PROCESOS PAC 2026 — Fuente: PAC oficial SERCOP GAD "
                               "Montecristi (" + str(len(procs)) + " procesos itemizados)")
    for j, h in enumerate(["CPC", "Tipo", "Procedimiento", "Descripción", "Monto V.Total"], 1):
        ws.cell(start + 1, j).value = h
    for i, p in enumerate(procs):
        r = start + 2 + i
        ws.cell(r, 1).value = p["cpc"]
        ws.cell(r, 2).value = p["tipo"]
        ws.cell(r, 3).value = p["proc"]
        ws.cell(r, 4).value = p["desc"]
        ws.cell(r, 5).value = p["monto"]
    tr = start + 2 + len(procs)
    ws.cell(tr, 4).value = "TOTAL ITEMIZADO"
    ws.cell(tr, 5).value = sum(p["monto"] for p in procs)
    wb.save(WORK)
    print("OK - " + str(len(procs)) + " procesos en " + h06
          + " filas " + str(start + 2) + "-" + str(start + 1 + len(procs)))
    print("total itemizado: $" + format(sum(p["monto"] for p in procs), ",.2f"))
    print("muestra:")
    for p in procs[:4]:
        print("  $" + format(p["monto"], ",.0f").rjust(11) + "  " + p["tipo"] + " · "
              + p["proc"] + " · " + p["desc"][:34])


if __name__ == "__main__":
    main()
