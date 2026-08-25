# -*- coding: utf-8 -*-
"""
QUIRA OS — Ingesta POA GAD 2026 → silo H05 (sección DETALLE PROYECTOS)
═══════════════════════════════════════════════════════════════════════════════
Pipeline Canon. Parsea el POA oficial (PDF matriz de planificación · extract_tables
por los 123 rects de celda) y escribe los 257 proyectos como sección de detalle en
H05, DEBAJO de las 25 metas (fórmulas E14:E38 intactas). Sobre COPIA WORK.

Columnas POA (mapeadas): 04 meta · 05 dirección · 09 proyecto · 11 descripción ·
15 partida · 20 monto anual · 21-24 trimestral Q1-Q4. Validado: suma = $39,310,032
= H05 col E (exacto, al dólar).

REQUIERE recálculo COM posterior (openpyxl borra cache). NO toca fórmulas.
Dylus Lab © 2026
"""
import shutil
import sys

import openpyxl
import pdfplumber

try:
    from config import DATOS_DIR as _DATOS
except Exception:                                        # noqa: BLE001
    import os as _os
    from pathlib import Path as _P
    _DATOS = _P(_os.environ.get("QUIRA_DATOS", "."))

BASE = str(_DATOS)
POA = BASE + r"\Holding_Municipal_Montecristi\POA 2023-2026\GAD Montecristi\GAD Montecristi POA 2026.pdf"
TGI = BASE + r"\SIAP-ICPI_GOLD_MASTER_v5.5_TGI.xlsx"
WORK = BASE + r"\SIAP-ICPI_GOLD_MASTER_v5.5_WORK_20260623_CANON-SPRINT.xlsx"


def money(s: str) -> float:
    s = (s or "").replace("$", "").replace(",", "").replace("—", "").strip()
    if s in ("", "-"):
        return 0.0
    try:
        return float(s)
    except ValueError:
        return 0.0


def parse_poa() -> list:
    rows = []
    with pdfplumber.open(POA) as pdf:
        for pg in pdf.pages:
            for t in pg.extract_tables():
                rows += t
    projs = []
    for row in rows:
        if len(row) < 30:
            continue
        anual = money(row[20])
        desc = (row[11] or "").replace("\n", " ").strip()
        if anual > 0 and desc and "DESCRIP" not in desc.upper():
            projs.append({
                "dir": (row[5] or "").replace("\n", " ").strip()[:50],
                "meta": (row[4] or "").replace("\n", " ").strip()[:60],
                "proy": (row[9] or "").replace("\n", " ").strip()[:50],
                "desc": desc[:80],
                "part": (row[15] or "").strip()[:18],
                "anual": anual,
                "q": [money(row[21]), money(row[22]), money(row[23]), money(row[24])],
            })
    return projs


def main() -> None:
    shutil.copy(TGI, WORK)                       # fresh WORK desde TGI (con PAC ya promovido)
    projs = parse_poa()
    wb = openpyxl.load_workbook(WORK)
    h05 = next(s for s in wb.sheetnames if s.startswith("H05_"))
    ws = wb[h05]
    start = 45
    ws.cell(start, 1).value = ("▌ DETALLE PROYECTOS POA 2026 — Fuente: POA oficial GAD "
                               "Montecristi (" + str(len(projs)) + " proyectos)")
    for j, h in enumerate(["Dirección", "Meta", "Proyecto", "Descripción", "Partida",
                           "Monto Anual", "Q1", "Q2", "Q3", "Q4"], 1):
        ws.cell(start + 1, j).value = h
    for i, p in enumerate(projs):
        r = start + 2 + i
        ws.cell(r, 1).value = p["dir"]
        ws.cell(r, 2).value = p["meta"]
        ws.cell(r, 3).value = p["proy"]
        ws.cell(r, 4).value = p["desc"]
        ws.cell(r, 5).value = p["part"]
        ws.cell(r, 6).value = p["anual"]
        for k in range(4):
            ws.cell(r, 7 + k).value = p["q"][k]
    tr = start + 2 + len(projs)
    ws.cell(tr, 5).value = "TOTAL POA"
    ws.cell(tr, 6).value = sum(p["anual"] for p in projs)
    wb.save(WORK)
    total = sum(p["anual"] for p in projs)
    print("OK - " + str(len(projs)) + " proyectos escritos en " + h05
          + " filas " + str(start + 2) + "-" + str(start + 1 + len(projs)))
    print("total escrito: $" + format(total, ",.2f"))


if __name__ == "__main__":
    main()
