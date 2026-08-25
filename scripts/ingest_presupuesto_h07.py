# -*- coding: utf-8 -*-
"""
QUIRA OS — Ingesta cédula presupuestaria GAD 2026 → silo H07 (Zona Cruda)
═══════════════════════════════════════════════════════════════════════════════
Pipeline Canon. Parsea la cédula oficial (XLSX · cédula de gastos eSIGEF, formato
europeo 1.234,56) y escribe las 135 partidas en la ZONA CRUDA de H07 (A46:E,
cols Codigo·Descripción·Grupo·Codificado·Devengado). NO toca la Zona Inteligente
(col N+, fórmulas que leen la zona cruda). Sobre COPIA WORK.

Presupuesto TOTAL municipal 2026: $45,977,893.81 codificado · $7,752,517.84 devengado.
REQUIERE recálculo COM posterior. Dylus Lab © 2026
"""
import shutil

import openpyxl

try:
    from config import DATOS_DIR as _DATOS
except Exception:                                        # noqa: BLE001
    import os as _os
    from pathlib import Path as _P
    _DATOS = _P(_os.environ.get("QUIRA_DATOS", "."))

BASE = str(_DATOS)
CEDULA = (BASE + r"\Holding_Municipal_Montecristi\Cedulas Presupuestarias 2023-2026"
          r"\Presupuestos 2026\GAD Montecristi 2026\GAD Montecristi Presupuesto abril 2026.xlsx")
TGI = BASE + r"\SIAP-ICPI_GOLD_MASTER_v5.5_TGI.xlsx"
WORK = BASE + r"\SIAP-ICPI_GOLD_MASTER_v5.5_WORK_20260623_CANON-SPRINT.xlsx"


def eur(v) -> float:
    """Convierte formato europeo (1.234.567,89) a float."""
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v or "").strip().replace(".", "").replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return 0.0


def parse_cedula() -> list:
    wb = openpyxl.load_workbook(CEDULA, read_only=True, data_only=True)
    ws = wb["Sheet1"]
    parts = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r[0]:
            continue
        parts.append({
            "cuenta": str(r[0]).strip(),
            "cat": str(r[1] or "").strip(),
            "desc": str(r[2] or "").strip(),
            "cod": eur(r[5]),     # Codificado
            "dev": eur(r[8]),     # Devengado
        })
    return parts


def main() -> None:
    shutil.copy(TGI, WORK)
    parts = parse_cedula()
    wb = openpyxl.load_workbook(WORK)
    h07 = next(s for s in wb.sheetnames if s.startswith("H07_"))
    ws = wb[h07]
    start = 46                                   # Zona Cruda: datos tras headers fila 45
    for i, p in enumerate(parts):
        r = start + i
        ws.cell(r, 1).value = p["cuenta"]
        ws.cell(r, 2).value = p["desc"][:60]
        ws.cell(r, 3).value = p["cat"][:30]
        ws.cell(r, 4).value = p["cod"]
        ws.cell(r, 5).value = p["dev"]
    wb.save(WORK)
    cod = sum(p["cod"] for p in parts)
    dev = sum(p["dev"] for p in parts)
    print("OK - " + str(len(parts)) + " partidas en " + h07
          + " filas " + str(start) + "-" + str(start + len(parts) - 1))
    print("codificado total: $" + format(cod, ",.2f"))
    print("devengado total:  $" + format(dev, ",.2f"))


if __name__ == "__main__":
    main()
