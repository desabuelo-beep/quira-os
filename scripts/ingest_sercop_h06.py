# -*- coding: utf-8 -*-
"""
QUIRA OS — Ingesta SERCOP (publicado) → silo H06 · ARQUITECTURA DEL NORMALIZADOR
═══════════════════════════════════════════════════════════════════════════════
REGLA CANÓNICA (consenso 2026-06-24): el Canon JAMÁS sabe de dónde viene el dato.

    Conector (fetch_sercop · OCDS)  →  NORMALIZADOR  →  silo H06  →  Canon

El conector habla con la API; el NORMALIZADOR traduce al esquema del Canon; el
Canon recibe filas limpias, source-agnostic. Mañana SERCOP→SECOP→ChileCompra solo
toca conector+normalizador, NUNCA el Canon.

H06 ya tiene: 25 metas curadas (filas 33-57) + INVENTARIO PASIVO histórico
(filas 60-837 · 772 procesos · montos por_verificar · NO incluye GAD-2026).
Este ingestor AÑADE — en espacio vacío al final, sin tocar lo existente — el bloque
"PUBLICADO 2026 VERIFICADO" (estado vivo del GAD vía OCDS, montos reales) para el
cruce plan-PAC ↔ publicado-SERCOP. Sobre COPIA WORK · requiere recálculo COM posterior.
Dylus Lab © 2026
"""
import shutil

import openpyxl

from scripts.fetch_sercop import build_contratacion_block

try:
    from config import DATOS_DIR as _DATOS
except Exception:                                        # noqa: BLE001
    import os as _os
    from pathlib import Path as _P
    _DATOS = _P(_os.environ.get("QUIRA_DATOS", "."))

BASE = str(_DATOS)
TGI = BASE + r"\SIAP-ICPI_GOLD_MASTER_v5.5_TGI.xlsx"
WORK = BASE + r"\SIAP-ICPI_GOLD_MASTER_v5.5_WORK_20260624_SERCOP.xlsx"

# Etapa OCDS → lenguaje de gestión pública (Firewall: sin jerga OCDS cruda en el Canon)
_ETAPA = {
    "planning": "Planificado", "active": "En proceso", "tender": "En licitación",
    "award": "Adjudicado", "contract": "Contratado", "complete": "Ejecutado",
    "cancelled": "Cancelado", "unsuccessful": "Desierto",
}
COLS = ["Entidad", "Cod_Proceso", "Descripción", "Partida", "Monto", "Monto_Tipo", "Etapa", "Fecha"]


def normalizar(block: dict, entidad: str = "GAD") -> list[dict]:
    """NORMALIZADOR — OCDS (del conector) → esquema del Canon (source-agnostic).
    El Canon recibe filas limpias sin saber que el origen es SERCOP/OCDS."""
    filas = []
    for p in block.get("procesos", []):
        etapa = p.get("estado") or (p.get("tag") or ["?"])[-1] or "?"
        filas.append({
            "entidad": entidad,
            "cod": str(p.get("ocid", "")),
            "descripcion": (p.get("descripcion") or "").strip()[:80],
            "partida": str(p.get("partida") or ""),
            "monto": p.get("monto") or 0,
            "monto_tipo": p.get("monto_tipo") or "?",
            "etapa": _ETAPA.get(str(etapa).lower(), str(etapa)),
            "fecha": str(p.get("fecha") or "")[:10],
        })
    return filas


def _last_data_row(ws, cols=6) -> int:
    """Última fila REAL con datos (escanea de abajo hacia arriba · ignora formato vacío)."""
    for r in range(ws.max_row, 1, -1):
        if any(ws.cell(r, c).value not in (None, "") for c in range(1, cols + 1)):
            return r
    return 1


def main() -> None:
    block = build_contratacion_block(2026, "montecristi", "GOBIERNO AUTONOMO")
    filas = normalizar(block, "GAD")
    total = sum(f["monto"] for f in filas)

    shutil.copy(TGI, WORK)                          # copia fresca (descarta WORK previo)
    wb = openpyxl.load_workbook(WORK)
    h06 = next(s for s in wb.sheetnames if s.startswith("H06"))
    ws = wb[h06]
    start = _last_data_row(ws) + 4                  # DESPUÉS de todo el inventario, con aire

    ws.cell(start, 1).value = (
        f"▌ PUBLICADO SERCOP 2026 — GAD · VERIFICADO (live OCDS) · corte {block.get('fecha_corte','')} · "
        f"{len(filas)} procesos · ${total:,.2f}")
    for j, h in enumerate(COLS, 1):
        ws.cell(start + 1, j).value = h
    for i, f in enumerate(filas):
        r = start + 2 + i
        ws.cell(r, 1).value = f["entidad"]
        ws.cell(r, 2).value = f["cod"]
        ws.cell(r, 3).value = f["descripcion"]
        ws.cell(r, 4).value = f["partida"]
        ws.cell(r, 5).value = f["monto"]
        ws.cell(r, 6).value = f["monto_tipo"]
        ws.cell(r, 7).value = f["etapa"]
        ws.cell(r, 8).value = f["fecha"]
    tr = start + 2 + len(filas)
    ws.cell(tr, 2).value = "TOTAL VERIFICADO"
    ws.cell(tr, 5).value = round(total, 2)
    wb.save(WORK)

    print(f"OK - {len(filas)} procesos PUBLICADO 2026 VERIFICADO en {h06}")
    print(f"     seccion fila {start} · datos {start+2}-{start+1+len(filas)} · total fila {tr}")
    print(f"     total verificado: ${total:,.2f} | etapas: {block.get('conteos_por_etapa')}")
    for f in filas[:8]:
        print(f"  [{f['etapa']:<11}|{f['monto_tipo']:<11}] ${f['monto']:>11,.0f}  {f['descripcion'][:40]}")


if __name__ == "__main__":
    main()
