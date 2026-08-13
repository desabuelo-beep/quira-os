# -*- coding: utf-8 -*-
"""
scripts/normativa/extraer_pai.py — el Plan Anual de Inversiones del GAD
═══════════════════════════════════════════════════════════════════════
POR QUÉ EXISTE (2026-08-12). El POA 2026 no declara meta ni indicador (OBS-027),
y quedaba la pregunta de si otro instrumento del mismo año sí lo hacía. El PAI es
el candidato natural: ordena la inversión y trae la partida.

RESPUESTA, verificada sobre los originales: **el PAI 2026 tampoco los trae.**

    2025 · POA meta ✓   PAI meta ✓
    2026 · POA meta ✗   PAI meta ✗

Eso convierte a OBS-027 en un hallazgo de instrumento y no de documento: en 2026
la meta desaparece de TODA la cadena operativa, no de una hoja suelta.

LO QUE EL PAI 2026 SÍ APORTA, y el POA no tenía:
  · `CÓDIGO DE LA ACTIVIDAD` (`APAA-01`, `OOPP-25`) — identificador estable por
    dirección. No llega a la meta, pero es el primer identificador propio que el
    GAD asigna a un objeto de gestión (contrastar con OBS-026).
  · `BENEFICIARIOS DE LA ACTIVIDAD` (`CANTONAL` / `INSTITUCIONAL`) — alcance
    territorial mínimo, pertinente a OBS-020.

POR QUÉ SE MAPEA POR CABECERA Y NO POR POSICIÓN. El PAI 2026 viene en cuatro
hojas y no todas alinean sus columnas: `Table 3` trae 23 en vez de 21 por celdas
combinadas. Anclar por posición leería el objetivo estratégico como unidad
administrativa sin dar un solo error.

Uso:  python scripts/normativa/extraer_pai.py [--anios 2025,2026] [--json salida]
Dylus Lab © 2026
"""
from __future__ import annotations

import argparse
import json
import re
import sys
import unicodedata
from pathlib import Path

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

BASE = Path(r"C:\Users\DELL\Desktop\Javo\Dylus Lab\ProyecT\Holding_Municipal_Montecristi"
            r"\POA 2023-2026\GAD Montecristi\PAI GAD 2023-2026")

ARCHIVOS = {
    2023: "Plan Anual de Inversion (PAI) 2023.xlsx",
    2025: "Plan Anual de Inversion (PAI) 2025.xlsx",
    2026: "Plan Anual de inversion (PAI) 2026.xlsx",
}

# Cabecera normalizada (sin tildes, sin saltos, minúsculas) → campo canónico.
# Se compara por prefijo porque los títulos vienen recortados de distinta forma
# en cada hoja («OBJETIVO DE DESARROLLO DEL PDOT» / «OBJETIVO DE\nDESARROLLO…»).
COLUMNAS = [
    ("objetivo de desarrollo sostenible", "ods"),
    ("objetivo plan nacional", "objetivo_pnd"),
    ("sistema", "sistema_pdot"),
    ("componente", "sistema_pdot"),
    ("objetivo de desarrollo del", "objetivo_desarrollo"),
    ("objetivo de desarrollo", "objetivo_desarrollo"),
    ("objetivo estrategico", "objetivo_estrategico"),
    ("unidad administrativa", "unidad_administrativa"),
    ("objetivo de gestion", "objetivo_gestion"),
    ("indicador", "indicador_pdot"),
    ("meta", "meta_pdot"),
    ("programas y proyecto", "programa"),
    ("programa", "programa"),
    ("subprograma", "subprograma"),
    ("proyecto emblematico", "proyecto_emblematico"),
    ("proyecto", "proyecto"),
    ("no. de actividad", "n_actividad"),
    ("codigo de la actividad", "codigo_actividad"),
    ("actividad", "actividad"),
    ("beneficiarios", "beneficiarios"),
    ("grupo de gasto", "grupo_gasto"),
    ("partida completa", "partida_completa"),
    ("partida", "partida"),
    ("descripcion", "descripcion"),
    ("area presupuestaria", "area"),
    ("fuente de financiamiento", "financiamiento"),
    ("monto presupuesto", "monto"),
    ("presupuesto", "monto"),
    ("codigo", "partida_completa"),
    ("responsable", "responsable"),
]


def _norm(s) -> str:
    if s is None:
        return ""
    s = "".join(c for c in unicodedata.normalize("NFD", str(s))
                if unicodedata.category(c) != "Mn").lower()
    return " ".join(s.replace("\n", " ").replace("\xa0", " ").split())


def _campo(titulo: str) -> str | None:
    t = _norm(titulo)
    if not t:
        return None
    for pref, campo in COLUMNAS:          # orden: el más específico primero
        if t.startswith(pref):
            return campo
    return None


def _txt(v) -> str:
    if v is None:
        return ""
    return " ".join(str(v).replace("\n", " ").replace("\xa0", " ").split())


def _partida6(v) -> str:
    """El ítem de 6 dígitos, aunque venga espaciado (`73 0505`) o dentro del
    código estructurado (`B300.330.2026.73 0505`)."""
    s = _txt(v).replace(" ", "")
    if re.fullmatch(r"\d{6}", s):
        return s
    for parte in re.split(r"[.\-]", s):
        if re.fullmatch(r"\d{6}", parte):
            return parte
    m = re.search(r"(\d{6})(?!\d)", s)
    return m.group(1) if m else ""


def _cabecera(filas: list[tuple]) -> tuple[int, dict[int, str]] | None:
    """Busca la fila de títulos. No se asume que sea la primera: hay hojas con
    encabezado institucional arriba y otras que arrancan en datos."""
    mejor = None
    for i, f in enumerate(filas[:8]):
        mapa = {j: c for j, celda in enumerate(f) if (c := _campo(celda))}
        if len(mapa) >= 6 and (mejor is None or len(mapa) > len(mejor[1])):
            mejor = (i, mapa)
    return mejor


def _realinear(filas: list[tuple], mapa: dict[int, str]) -> dict[int, str]:
    """Corrige el desplazamiento de columnas de una hoja sin títulos propios.

    `Table 3` del PAI 2026 trae 23 columnas en vez de 21 —celdas combinadas que
    la exportación abrió— y todo su contenido corrido un lugar. Heredar el mapa
    tal cual leía el objetivo de desarrollo como objetivo estratégico y el grupo
    de gasto como partida, **sin producir un solo error**: las filas salían
    completas y equivocadas.

    El anclaje es la partida: un valor de exactamente 6 dígitos es reconocible
    por su forma, así que se busca en qué columna cae de verdad y se corre el
    mapa entero esa distancia. Si no aparece, se devuelve el mapa intacto —no se
    inventa una corrección sobre una hoja que no se entendió."""
    col_partida = next((j for j, c in mapa.items() if c == "partida"), None)
    if col_partida is None:
        return mapa
    from collections import Counter
    desvios: Counter = Counter()
    for f in filas[:25]:
        for j, v in enumerate(f):
            if re.fullmatch(r"\d{6}", _txt(v).replace(" ", "")):
                desvios[j - col_partida] += 1
                break
    if not desvios:
        return mapa
    d, n = desvios.most_common(1)[0]
    if d == 0 or n < 3:                      # sin desvío, o evidencia insuficiente
        return mapa
    return {j + d: c for j, c in mapa.items()}


def extraer(anio: int) -> list[dict]:
    import openpyxl
    wb = openpyxl.load_workbook(BASE / ARCHIVOS[anio], data_only=True, read_only=True)
    out: list[dict] = []
    mapa_previo: dict[int, str] | None = None

    for hoja in wb.sheetnames:
        filas = [tuple(f) for f in wb[hoja].iter_rows(values_only=True)]
        cab = _cabecera(filas)
        if cab:
            inicio, mapa = cab[0] + 1, cab[1]
            mapa_previo = mapa
        elif mapa_previo:
            # Sólo `Table 1` del PAI 2026 trae títulos; las demás arrancan en
            # datos y heredan el mapa en vez de descartarse en silencio.
            inicio, mapa = 0, _realinear(filas, mapa_previo)
        else:
            continue

        for n, f in enumerate(filas[inicio:], inicio + 1):
            reg = {}
            for j, campo in mapa.items():
                if j < len(f) and (v := _txt(f[j])):
                    reg.setdefault(campo, v)
            if not reg.get("actividad") and not reg.get("programa"):
                continue
            p = _partida6(reg.get("partida") or reg.get("partida_completa"))
            if p:
                reg["partida"] = p
            out.append({"anio": anio, "hoja": hoja, "fila": n, "campos": reg})
    wb.close()
    return out


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--anios", default="2023,2025,2026")
    ap.add_argument("--json", help="ruta donde volcar el resultado")
    args = ap.parse_args()

    print("PLAN ANUAL DE INVERSIONES · GAD Montecristi\n")
    todo = {}
    for a in [int(x) for x in args.anios.split(",")]:
        regs = extraer(a)
        todo[str(a)] = regs
        cm = sum(1 for r in regs if r["campos"].get("meta_pdot"))
        ci = sum(1 for r in regs if r["campos"].get("indicador_pdot"))
        co = sum(1 for r in regs if r["campos"].get("objetivo_estrategico"))
        cp = sum(1 for r in regs if r["campos"].get("partida"))
        cc = sum(1 for r in regs if r["campos"].get("codigo_actividad"))
        print(f"  {a}: {len(regs):4} filas · objetivo {co:3} · META {cm:3} · "
              f"indicador {ci:3} · partida {cp:3} · cód. actividad {cc:3}")
        print(f"        partidas distintas: {len({r['campos'].get('partida') for r in regs if r['campos'].get('partida')})}")

    if args.json:
        Path(args.json).write_text(json.dumps(todo, ensure_ascii=False, indent=1),
                                   encoding="utf-8")
        print(f"\n  → {args.json}")


if __name__ == "__main__":
    main()
