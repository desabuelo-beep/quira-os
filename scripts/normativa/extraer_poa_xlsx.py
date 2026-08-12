# -*- coding: utf-8 -*-
"""
scripts/normativa/extraer_poa_xlsx.py — POA del GAD desde el ORIGINAL, no desde la copia
═══════════════════════════════════════════════════════════════════════════════════════
POR QUÉ EXISTE (2026-08-12). El corpus tenía el POA del GAD con **53 % de texto
roto**: columnas de tabla leídas carácter a carácter —«ild,t o ifvo o sr: ot aM
cliee acj»—. De ahí se concluyó que «el POA no ancla al PDOT», y era falso: los
originales del GAD son `.xlsx` con columnas limpias. **La carencia era del
capturador, no de la fuente** (ADR-042 §6 aplicado a la ingesta documental).

Este extractor lee los `.xlsx` originales y produce un fragmento POR FILA con sus
campos rotulados. Así la meta, el indicador y la partida quedan recuperables sin
depender de que un chunker de prosa acierte a cortar una tabla.

QUÉ NO HACE:
  · No infiere la meta cuando el POA no la declara. El POA 2026 no tiene columna
    de meta (OBS-027) y aquí se refleja tal cual: ausente, no vacía ni supuesta.
  · No corrige ni normaliza cifras. Copia lo que el documento dice.

Uso:  python scripts/normativa/extraer_poa_xlsx.py [--dry-run]
Dylus Lab © 2026
"""
from __future__ import annotations

import argparse
import sys
from pathlib import Path

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

BASE = Path(r"C:\Users\DELL\Desktop\Javo\Dylus Lab\ProyecT\Holding_Municipal_Montecristi"
            r"\POA 2023-2026\GAD Montecristi")

# ══════════════════════════════════════════════════════════════════════════════
# MAPEO POR AÑO
#
# El formato del POA cambió cada año, así que no hay una convención común que
# aplicar — se declara la de cada uno, verificada a mano sobre el archivo.
#
# 2025 no tiene fila de cabecera utilizable: la columna 7 repite la etiqueta
# «INDICADOR ESTRATÉGICO PDOT» en cada fila. Se ancla por posición.
# 2026 sí tiene cabecera en la fila 6, y NO incluye meta ni indicador (OBS-027).
# ══════════════════════════════════════════════════════════════════════════════
MAPEO: dict[str, dict] = {
    "GAD Monteristi POA 2025.xlsx": {
        "anio": 2025, "fila_datos": 7, "por_posicion": True,
        "campos": {3: "sistema_pdot", 4: "objetivo_desarrollo", 5: "objetivo_estrategico",
                   6: "objetivo_gestion", 8: "indicador_pdot", 9: "meta_pdot",
                   11: "programa", 12: "subprograma", 13: "proyecto", 15: "actividad",
                   16: "partida", 17: "partida_completa", 19: "area", 20: "financiamiento",
                   21: "responsable", 22: "monto"},
    },
    "GAD Montecristi POA 2026.xlsx": {
        "anio": 2026, "fila_datos": 7, "por_posicion": True,
        "campos": {0: "ods", 1: "objetivo_pnd", 2: "sistema_pdot",
                   3: "objetivo_desarrollo", 4: "objetivo_estrategico",
                   5: "unidad_administrativa", 6: "objetivo_gestion", 7: "programa",
                   8: "subprograma", 9: "proyecto", 11: "actividad",
                   12: "codigo_actividad", 15: "partida", 16: "partida_completa",
                   18: "area", 19: "financiamiento", 20: "monto"},
        # sin `meta_pdot` ni `indicador_pdot`: el instrumento no los trae (OBS-027)
    },
}


def _txt(v) -> str:
    if v is None:
        return ""
    s = str(v).replace("\n", " ").strip()
    return " ".join(s.split())


def extraer(archivo: str) -> list[dict]:
    import openpyxl
    cfg = MAPEO[archivo]
    wb = openpyxl.load_workbook(BASE / archivo, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    filas = list(ws.iter_rows(values_only=True))
    wb.close()

    out: list[dict] = []
    for n, f in enumerate(filas[cfg["fila_datos"] - 1:], cfg["fila_datos"]):
        campos = {nom: _txt(f[i]) for i, nom in cfg["campos"].items() if i < len(f)}
        # una fila cuenta si tiene actividad Y partida: lo demás puede venir
        # arrastrado por celdas combinadas
        if not campos.get("actividad") or not campos.get("partida"):
            continue
        campos = {k: v for k, v in campos.items() if v}
        out.append({"anio": cfg["anio"], "fila": n, "campos": campos})
    return out


def como_texto(reg: dict) -> str:
    """Fragmento rotulado: cada campo con su nombre, una línea cada uno.

    Se prefiere esto a una frase en prosa porque el valor de este documento está
    en la CORRESPONDENCIA entre columnas, y una prosa la disuelve."""
    cab = f"POA GAD MONTECRISTI {reg['anio']} · fila {reg['fila']}"
    cuerpo = "\n".join(f"{k.upper().replace('_',' ')}: {v}" for k, v in reg["campos"].items())
    return f"{cab}\n{cuerpo}"


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--dry-run", action="store_true", help="no escribe; solo informa")
    args = ap.parse_args()

    print("EXTRACCIÓN DEL POA DESDE LOS ORIGINALES .xlsx\n")
    total = 0
    for archivo in MAPEO:
        regs = extraer(archivo)
        total += len(regs)
        con_meta = sum(1 for r in regs if r["campos"].get("meta_pdot"))
        con_sis = sum(1 for r in regs if r["campos"].get("sistema_pdot"))
        print(f"  {archivo}")
        print(f"     {len(regs)} filas · con SISTEMA PDOT: {con_sis} · con META: {con_meta}")
        if regs:
            print(f"     muestra:\n        " + como_texto(regs[0])[:300].replace("\n", "\n        "))
        print()
    print(f"  TOTAL: {total} fragmentos estructurados"
          f"{'  (dry-run: no se escribió nada)' if args.dry_run else ''}")


if __name__ == "__main__":
    main()
