"""
QUIRA OS — quira_extract.py  v1.1
Lee H73_OUTPUT_API + H99_ENGINE_CORE del Gold Master Excel
y genera los valores actualizados para demo_data.py
Uso: python scripts/quira_extract.py
Dylus Lab © 2026
"""
import sys
import io
from pathlib import Path

# Force UTF-8 output
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

import openpyxl

# ── RUTAS ──────────────────────────────────────────────────────────────────────
_BASE = Path(__file__).resolve().parent.parent.parent / "ProyecT"

EXCEL_CANDIDATES = [
    _BASE / "SIAP-ICPI_GOLD_MASTER_v4.0_QUIRA_20260514.xlsx",
    _BASE / "SIAP-ICPI_GOLD_MASTER_v3.0_QUIRA_20260514.xlsx",
    _BASE / "Dylus Lab - Sistema de Integridad Algorítmica Predictivo (SIAP-ICPI v1.0)222.xlsx",
]


def find_excel() -> Path:
    for p in EXCEL_CANDIDATES:
        if p.exists():
            return p
    raise FileNotFoundError(
        f"No se encontró el Gold Master en {_BASE}\n"
        f"Archivos buscados:\n" + "\n".join(f"  {c}" for c in EXCEL_CANDIDATES)
    )


def load_output_api(wb: openpyxl.Workbook) -> dict:
    """Lee H73_OUTPUT_API → {KEY: VALUE}."""
    sheet = None
    for name in wb.sheetnames:
        if "OUTPUT_API" in name:
            sheet = wb[name]
            break
    if not sheet:
        raise ValueError("Hoja H73_OUTPUT_API no encontrada")

    data = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        key = row[0]
        val = row[1]
        if key and val is not None and str(key).strip() != "---":
            data[str(key).strip()] = val
    return data


def load_h99_parroquias(wb: openpyxl.Workbook) -> list[dict]:
    """Lee H99_ENGINE_CORE → lista de dicts por parroquia."""
    sheet = None
    for name in wb.sheetnames:
        if "H99" in name or "ENGINE_CORE" in name:
            sheet = wb[name]
            break
    if not sheet:
        return []

    parroquias = []
    # Header fijo: A=ID, B=Nombre, C=Tipo, D=Población, E=NBI_Pct,
    #              F=Cobertura_Agua, G=Inv_PerCapita, H=Inv_Total,
    #              I=Composite_Need, J=IET_Local_Pct, K=IRS_Contrib,
    #              L=Fondos_Bloqueados, M=Lat, N=Lon
    for row in sheet.iter_rows(min_row=3, values_only=True):
        pid = row[0]
        if not pid or not str(pid).startswith("P-"):
            continue
        nombre = row[1] or ""
        tipo   = row[2] or ""
        try:
            parroquias.append({
                "nombre":       str(nombre).strip(),
                "tipo":         str(tipo).strip(),
                "poblacion":    int(row[3] or 0),
                "nbi_pct":      float(row[4] or 0),
                "agua_pct":     float(row[5] or 0),
                "inv_percapita":int(row[6] or 0),
                "inv_total":    int(row[7] or 0) if row[7] else 0,
                "composite_need": float(row[8] or 0) if row[8] else 0,
                "iet_local":    float(row[9] or 0) if row[9] else 0,
                "irs_contrib":  float(row[10] or 0) if row[10] else 0,
                "fondos_bloq":  int(row[11] or 0) if row[11] else 0,
                "lat":          float(row[12] or 0) if row[12] else None,
                "lon":          float(row[13] or 0) if row[13] else None,
            })
        except (TypeError, ValueError):
            continue
    return parroquias


def fmt(v) -> str:
    """Formatea valor para código Python."""
    if isinstance(v, str):
        return f'"{v}"'
    if isinstance(v, float):
        if v == int(v):
            return str(int(v))
        return f"{v:.4f}"
    return str(v)


# ── MAPEO H73_KEY → demo_data var_name ────────────────────────────────────────
# Columnas: (var_name_salida, api_key_exacto_H73, comentario)
# api_key DEBE coincidir exactamente con col A de H73_OUTPUT_API
MAPPING = [
    # var_name              api_key_H73              comentario
    ("ICPI_GLOBAL",         "ICPI_GLOBAL_PCT",        "# ICGI-T / ICPI global (%) → ICGIT_Q1_2026['score']"),
    ("ISP",                 "ISP_SALUD_PRESUP",       "# Salud Presupuestaria → INDICES['ISP']['valor']"),
    ("PSG",                 "PSG_EJECUCION",          "# Presupuesto Sensible Género → INDICES['PSG']['valor']"),
    ("IED",                 "IED_GLOBAL",             "# Eficiencia Direcciones → INDICES['IED']['valor']"),
    ("IFE_PDOT",            "IFE_PROMESAS_PDOT",      "# Promesas vinculadas PDOT → INDICES['IFE-A']['valor']"),
    ("IFE_TOTAL",           "IFE_PROMESAS_TOTAL",     "# Total promesas CNE"),
    ("IGP",                 "IGP_REF_2025",           "# Gobernanza Participativa → INDICES['IGP']['valor']"),
    ("IOC",                 "IOC_OPACIDAD",           "# Opacidad Crítica → INDICES['IOC']['valor'] × 100"),
    ("IET_PERCAPITA",       "IET_PERCAPITA_PEOR",     "# IET peor parroquia → INDICES['IET']['valor']"),
    ("IRS_GLOBAL",          "IRS_GLOBAL",             "# Índice Regresividad Social (H99)"),
    ("SAT_ACTIVAS",         "SAT_ACTIVAS_COUNT",      "# SAT alertas activas"),
    ("TRUST_SCORE",         "TRUST_SCORE",            "# Trust Score sistema"),
    ("PRESUPUESTO_TOTAL",   "PRESUPUESTO_TOTAL_4E",   "# Presupuesto total 4 entidades → ICGIT_Q1_2026['presupuesto_total']"),
    ("ICPI_2023",           "ICPI_2023",              "# Histórico 2023 → ICGIT_Q1_2026['historico']['2023']"),
    ("ICPI_2024",           "ICPI_2024",              "# Histórico 2024 → ICGIT_Q1_2026['historico']['2024']"),
    ("ICPI_2025",           "ICPI_2025",              "# Histórico 2025 → ICGIT_Q1_2026['historico']['2025']"),
    ("ICPI_ACUMULADO_Q1",   "ICPI_ACUMULADO_Q1",      "# ICPI acumulado Q1-2026"),
    ("BRECHA_FONDOS_BLOQ",  "BRECHA_FONDOS_BLOQ",     "# Fondos bloqueados ($USD)"),
    ("NBI_RURAL_PCT",       "NBI_RURAL_PCT",          "# NBI rural cantonal (%)"),
    ("GPS_PARROQUIAS_OK",   "GPS_PARROQUIAS_OK",      "# Parroquias con GPS válidas"),
]


def main():
    # ── Localizar Excel ────────────────────────────────────────────────────────
    try:
        excel_path = find_excel()
    except FileNotFoundError as e:
        print(f"ERROR: {e}", file=sys.stderr)
        sys.exit(1)

    print(f"# ═══════════════════════════════════════════════════════════")
    print(f"# QUIRA OS · quira_extract.py v1.1")
    print(f"# Fuente: {excel_path.name}")
    print(f"# ═══════════════════════════════════════════════════════════")
    print()

    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    data = load_output_api(wb)
    parroquias = load_h99_parroquias(wb)
    wb.close()

    print(f"# H73_OUTPUT_API: {len(data)} métricas extraídas")
    print(f"# H99_ENGINE_CORE: {len(parroquias)} parroquias extraídas")
    print(f"# Timestamp fuente: {data.get('EXTRACT_TIMESTAMP', 'N/A')}")
    print()

    # ── ICPI_DATA dict ────────────────────────────────────────────────────────
    print("# ── ICPI_DATA (para reference / demo_data.py) ───────────────────────────")
    not_found = []
    print("ICPI_DATA = {")
    for var_name, api_key, comment in MAPPING:
        val = data.get(api_key)
        if val is not None:
            print(f"    '{var_name}': {fmt(val)},  {comment}")
        else:
            not_found.append((var_name, api_key))
            print(f"    # '{var_name}': NOT FOUND  (H73 key buscada: '{api_key}')")
    print("}")
    print()

    if not_found:
        print(f"# ⚠ {len(not_found)} clave(s) no encontradas en H73_OUTPUT_API:")
        for vn, ak in not_found:
            print(f"#   var='{vn}' → key H73='{ak}'")
        print()

    # ── Bloque demo_data.py listo para pegar ─────────────────────────────────
    def _to_pct(v, name=""):
        """Convierte decimal→porcentaje si el valor viene como ratio (≤1)."""
        if v is None:
            return None
        return round(v * 100, 2) if abs(v) <= 1 else round(v, 2)

    icpi_pct     = data.get("ICPI_GLOBAL_PCT")           # ya es % (53.56)
    isp_pct      = _to_pct(data.get("ISP_SALUD_PRESUP")) # 0.1458 → 14.58
    psg_pct      = _to_pct(data.get("PSG_EJECUCION"))    # 0.1283 → 12.83
    ied_pct      = _to_pct(data.get("IED_GLOBAL"))       # 0.3399 → 33.99
    ife_count    = data.get("IFE_PROMESAS_PDOT")         # 48 (count)
    ife_total    = data.get("IFE_PROMESAS_TOTAL")        # 66 (count)
    ife_pct      = round(ife_count / ife_total * 100, 2) if ife_count and ife_total else None
    igp_pct      = _to_pct(data.get("IGP_REF_2025"))     # 0.2798 → 27.98
    ioc_raw      = data.get("IOC_OPACIDAD")              # 0.1771 (decimal)
    ioc_pct      = _to_pct(ioc_raw)                      # → 17.71
    iet_raw      = data.get("IET_PERCAPITA_PEOR")        # 0.4480 (decimal)
    iet_pct      = _to_pct(iet_raw)                      # → 44.80
    presup       = data.get("PRESUPUESTO_TOTAL_4E")      # 54,242,424 (4 entidades)
    h23_raw      = data.get("ICPI_2023")                 # 0.5736 decimal
    h24_raw      = data.get("ICPI_2024")                 # 0.6712 decimal
    h25_raw      = data.get("ICPI_2025")                 # 0.6993 decimal
    h23_pct      = _to_pct(h23_raw)                     # → 57.36
    h24_pct      = _to_pct(h24_raw)                     # → 67.12
    h25_pct      = _to_pct(h25_raw)                     # → 69.93
    irsv         = data.get("IRS_GLOBAL")               # 93.1 (ya es %)
    trust        = data.get("TRUST_SCORE")              # 89.6
    brecha       = data.get("BRECHA_FONDOS_BLOQ")       # 3_660_000
    nbi_rural    = data.get("NBI_RURAL_PCT")            # 67.9

    print("# ── FRAGMENTO ACTUALIZADO PARA demo_data.py ─────────────────────────────")
    print("# Valores convertidos a las unidades que usa demo_data.py (porcentajes)")
    print()
    print("# → ICGIT_Q1_2026")
    if icpi_pct is not None:
        print(f'ICGIT_Q1_2026["score"]               = {icpi_pct}')
    if presup is not None:
        pv = int(round(presup))
        print(f'# ICGIT_Q1_2026["presupuesto_total"] = {pv:,}  ← holding 4 entidades')
        print(f'#   (demo_data usa solo GAD = 26_689_147; actualizar si se muestra holding)')
    if h23_pct is not None:
        print(f'ICGIT_Q1_2026["historico"]["2023"]   = {h23_pct}')
    if h24_pct is not None:
        print(f'ICGIT_Q1_2026["historico"]["2024"]   = {h24_pct}')
    if h25_pct is not None:
        print(f'ICGIT_Q1_2026["historico"]["2025"]   = {h25_pct}')
    print()
    print("# → INDICES['valor'] — todos en porcentaje (0–100)")
    if isp_pct  is not None: print(f'INDICES["ISP"]["valor"]   = {isp_pct}')
    if psg_pct  is not None: print(f'INDICES["PSG"]["valor"]   = {psg_pct}')
    if ied_pct  is not None: print(f'INDICES["IED"]["valor"]   = {ied_pct}')
    if ife_pct  is not None:
        print(f'INDICES["IFE-A"]["valor"] = {ife_pct}  # {int(ife_count)}/{int(ife_total)} promesas × 100')
    if igp_pct  is not None: print(f'INDICES["IGP"]["valor"]   = {igp_pct}')
    if ioc_pct  is not None: print(f'INDICES["IOC"]["valor"]   = {ioc_pct}  # Opacidad % (raw decimal={ioc_raw})')
    if iet_pct  is not None: print(f'INDICES["IET"]["valor"]   = {iet_pct}  # IET per cápita peor parroquia (raw={iet_raw})')
    print()
    print("# → Nuevas métricas (agregar a demo_data.py si no existen)")
    if irsv  is not None: print(f'IRS_GLOBAL        = {irsv}  # Índice Regresividad Social (%)')
    if trust is not None: print(f'TRUST_SCORE       = {trust}  # Trust Score sistema ICPI')
    if brecha is not None: print(f'BRECHA_FONDOS_BLOQ = {int(brecha):,}  # USD fondos bloqueados por bajo ICPI')
    if nbi_rural is not None: print(f'NBI_RURAL_PCT     = {nbi_rural}  # % NBI parroquias rurales')
    print()

    # ── PARROQUIAS desde H99_ENGINE_CORE ─────────────────────────────────────
    if parroquias:
        print("# ── PARROQUIAS (desde H99_ENGINE_CORE) ───────────────────────────────────")
        print("# Reemplazar PARROQUIAS en demo_data.py con estos datos actualizados")
        print("PARROQUIAS_H99 = {")
        for p in parroquias:
            lat_str = f"{p['lat']}" if p['lat'] else "None"
            lon_str = f"{p['lon']}" if p['lon'] else "None"
            print(
                f"    '{p['nombre']}': {{"
                f"'poblacion': {p['poblacion']}, "
                f"'nbi_pct': {p['nbi_pct']}, "
                f"'agua_pct': {p['agua_pct']}, "
                f"'inv_percapita': {p['inv_percapita']}, "
                f"'inv_total': {p['inv_total']}, "
                f"'composite_need': {p['composite_need']:.4f}, "
                f"'lat': {lat_str}, "
                f"'lon': {lon_str}"
                f"}},"
            )
        print("}")
    else:
        print("# ⚠ H99_ENGINE_CORE no encontrada — usando datos hardcoded")
        print("PARROQUIAS_H99 = {")
        fallback = [
            ("Montecristi",        39800, 38.4, 89.3, 217, 8_629_600, -1.0450, -80.6578),
            ("Aníbal San Andrés",   5200, 52.1, 28.9,  58,   301_600, -1.0238, -80.5891),
            ("Colorado",            3800, 58.7, 34.7,  32,   121_600, -1.1253, -80.5612),
            ("Leónidas Proaño",     4100, 54.3, 41.2,  48,   196_800, -1.0978, -80.6021),
            ("Gral. Alfaro",        6300, 49.8, 55.8,  71,   447_300, -0.9876, -80.5234),
            ("Isabel Muentes",      5700, 61.2,  1.0,  40,   228_000, -0.9712, -80.6487),
            ("La Pila",             4600, 55.9, 51.2,  52,   239_200, -1.0689, -80.6923),
        ]
        for n, pob, nbi, agua, inv_h, inv_t, lat, lon in fallback:
            print(
                f"    '{n}': {{"
                f"'poblacion': {pob}, 'nbi_pct': {nbi}, 'agua_pct': {agua}, "
                f"'inv_percapita': {inv_h}, 'inv_total': {inv_t}, "
                f"'lat': {lat}, 'lon': {lon}}},"
            )
        print("}")

    print()
    print("# ═══════════════════════════════════════════════════════════")
    print(f"# QUIRA OS · Extracción completada · Gold Master v4.0")
    print("# Ejecutar: python scripts/quira_extract.py > extracted_values.txt")
    print("# Luego actualizar data/demo_data.py manualmente con los valores")
    print("# ═══════════════════════════════════════════════════════════")


if __name__ == "__main__":
    main()
