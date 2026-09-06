# -*- coding: utf-8 -*-
"""
scripts/rearq/matriz_ontologica.py — REARQ · Q-M0 · MATRIZ ONTOLÓGICO-ARQUITECTÓNICA

    El PRIMER artefacto de la Rearquitectura. **No es una fórmula nueva**: es
    una ficha por indicador con veinte campos, y su función es obligar a
    separar lo que un indicador CALCULA de lo que se AFIRMA con él.

    ⚠️ NO EMPIEZA POR EL ICPI, y es deliberado. Aunque sea el indicador más
    trabajado, entra **como uno más** —en orden alfabético— para no diseñar la
    arquitectura nueva alrededor de su forma histórica.

    ★ LA COLUMNA QUE MÁS OBLIGA es la 19:

        ¿qué afirmación NO permite hacer este indicador?

    Es el mecanismo que impide que la capa de presentación convierta una
    medición limitada en una afirmación mucho mayor. Y viene de lo aprendido
    en `D2`:

        dato → evidencia → inferencia → afirmación   NO son sinónimos

    LO QUE ESTA MATRIZ MIDE, y es su hallazgo: **cuánto no sabemos todavía**.
    Cada `POR DECLARAR` es una celda que nadie ha llenado, y contarlas es más
    útil que rellenarlas de memoria.

    ⚠️ La Rearquitectura está autorizada **exclusivamente para diseño
    conceptual y arquitectónico**. Esta matriz no decide destinos: los
    registra como evaluación inicial.

Uso:  python scripts/rearq/matriz_ontologica.py
Dylus Lab © 2026
"""
from __future__ import annotations

import re
import sys
from pathlib import Path

_RAIZ = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(_RAIZ))

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

_SALIDA = _RAIZ / "docs" / "architecture" / "REARQ_Q-M0_MATRIZ_ONTOLOGICA.md"
_CONTRATO = _RAIZ / "docs" / "architecture" / "GM-OMEGA_CONTRATO_INDICE_DOMINIO.md"

_PD = "⬜ POR DECLARAR"

# Los veinte campos, en el orden que impone el método: primero el fenómeno,
# al final la fórmula. Invertirlo sería empezar por la matemática, que es el
# hábito que `GM-Ω` vino a corregir.
_CAMPOS = [
    ("01", "Fenómeno que se quiere conocer"),
    ("02", "Pregunta de gestión que responde"),
    ("03", "Unidad que observa"),
    ("04", "Universo que cubre"),
    ("05", "Evidencia que necesita"),
    ("06", "Estado de esa evidencia"),
    ("07", "Inferencia que realiza"),
    ("08", "Relación normativa que lo sostiene"),
    ("09", "Relación teórica que lo sostiene"),
    ("10", "FONDO o FORMA"),
    ("11", "Subdominio o capacidad que representa"),
    ("12", "Residencia"),
    ("13", "Temporalidad"),
    ("14", "Comparabilidad territorial"),
    ("15", "Fórmula que utiliza hoy"),
    ("16", "Decisiones históricas de las que depende"),
    ("17", "Dependencia ecuatoriana"),
    ("18", "✅ Qué afirmación PERMITE hacer"),
    ("19", "🔴 Qué afirmación NO PERMITE hacer"),
    ("20", "Estado Rearquitectura"),
]

# ── LAS FICHAS ───────────────────────────────────────────────────────────
#
# Se declara ÚNICAMENTE lo que alguna etapa estableció con su fuente. Todo lo
# demás queda `POR DECLARAR` y se cuenta. Rellenar de memoria produciría una
# matriz completa y falsa — el defecto que esta matriz existe para evitar.
_FICHAS: dict[str, dict[str, str]] = {

    "IBSC": {
        "01": "Bienestar social del cantón, por sustitución de `V·E·T·C` por "
              "una variable `S_i` (glosario `H12b`)",
        "15": "misma fórmula canónica del ICPI con `S_i` en lugar de "
              "`V_i×E_i×T_i×C_i` · fuente `H04b`",
        "16": "hereda **entera** la arquitectura multiplicativa del ICPI "
              "(`ADR-054`/`D1`) sin declaración propia",
        "20": "⚠️ **revisar dependencia**: si `D1` se rediseña, `IBSC` cambia "
              "sin que nadie lo haya decidido para `IBSC`",
    },

    "ICODS": {
        "02": "alineación con la Agenda 2030 en el eje biofísico y de riesgo",
        "10": "FONDO — es sectorial",
        "12": "`d13` · Constitución §CAPA 0.5",
        "14": "🟢 alta — los ODS son marco internacional, no ecuatoriano",
        "17": "baja · el catálogo ODS no depende de Ecuador",
    },

    "ICPI": {
        "01": "**congruencia programática e intersistémica**: si la cadena "
              "`programa → norma → verificación → ejecución → tiempo → "
              "trazabilidad` se sostiene entera (`GM-Ω-001`)",
        "03": "meta del PDOT identificada por su ID canónico — ⚠️ **no "
              "declarada en el canon** (`011-A2`)",
        "04": "25 de 66 metas · universo operacional de `v1` (`ADR-036`) · "
              "correspondencia con las 66 sin reconciliar (`011-B`)",
        "05": "cuatro silos: `SERCOP` · `eSIGEF` · `LOTAIP` · `CPCCS`",
        "06": "🔴 `V_i = 0` en **6 de 25** metas — el 12,8 % del peso",
        "07": "que una cadena con un eslabón no acreditado **no acredita "
              "congruencia** (lectura `A`, sellada en `ADR-055`)",
        "08": "`COPFP 54` · `COOTAD 54-55` · `LOTAIP 7` · `COPFP 115-117` + "
              "Acuerdo 067 MEF",
        "09": "⬜ **no consta** · `011-C4` no halló razón teórica que funde "
              "la multiplicatividad (`D1`)",
        "10": "⚠️ **en disputa** — reside en `d06` (FONDO) pero mide una "
              "propiedad transversal (FORMA)",
        "12": "`d06` Salud Institucional — ⚠️ apoyada en «Cumplimiento "
              "Institucional», nombre que el canon **ya retiró**",
        "13": "corte parcial · abril 2026 · `T_i` sin cerrar el ejercicio",
        "14": "🔵 media — la estructura viaja; `SERCOP`/`eSIGEF`/`LOTAIP` son "
              "instancias sustituibles (`010`)",
        "15": "`Σ(P·R·V·E·T·C) / Σ(P·R)` · `H12!B33` · **inmutable**",
        "16": "`D1` multiplicatividad · `D2` `V_i` multiplicativo · `D3` "
              "pesos · `D4` piso · `D5` AVEP — las cinco selladas 2026-09-06",
        "17": "media · la función viaja, las instituciones no",
        "18": "«la **congruencia acreditada** de las 25 metas del universo "
              "operacional es 27,4582 % al corte de abril»",
        "19": "«el GAD cumple el 27 % de su PDOT» · «el desempeño anual es "
              "del 27 %» · «`n` metas **no se ejecutaron**» · «mide velocidad "
              "de ejecución» (`D-011`)",
        "20": "⚠️ **pendiente** — `C4` no lo declaró necesario ni incorrecto. "
              "Su lugar se decide en `Q-M1`, no aquí",
    },

    "IED": {
        "01": "eficiencia de la **dirección responsable**, no del sector",
        "02": "¿qué tan eficientemente funciona la unidad orgánica?",
        "03": "la dirección municipal del Estatuto Orgánico (`Res. 040-2025`)",
        "10": "**FORMA** — la pregunta aplica a Salud, Obras Públicas y "
              "Financiera por igual (`010`)",
        "12": f"{_PD} — ⚠️ y por eso importa: **no pertenece a ningún dominio "
              "sectorial**, y hoy no existe el eje que lo albergaría",
        "15": "ejecución de metas PDOT desglosada por dirección · `H17` · `H30`",
        "20": "🟢 **candidato a capacidad transversal de FORMA**",
    },

    "IEF": {
        "01": "capacidad del GAD de **captar fondos externos**",
        "15": "`Σ(fondos externos captados) / presupuesto codificado` · `H20c`",
        "10": "**FORMA** — es capacidad de gestión, no política sectorial",
    },

    # ⚠️ FICHA CORREGIDA. La primera versión declaró `IFE` = «eficiencia
    # financiera» y anunció una **duplicación con `IEF`** que no existe. Bastó
    # abrir la hoja: `IFE` es FIDELIDAD ELECTORAL y se alimenta de
    # `H03_S1_ELECTORAL_CNE`. El parecido de las siglas produjo un hallazgo
    # inventado, y sólo el instrumento lo desmintió.
    "IFE": {
        "02": "¿el gobierno cumple el plan con el que fue electo?",
        "10": "**FORMA** — fidelidad al mandato, no política sectorial",
        "12": "candidato natural a `d03` Gobernanza del Mandato Electoral",
        "17": "🔴 alta · el Plan CNE y su régimen electoral son ecuatorianos",
        "20": "⚠️ **residencia por declarar** · y su sigla se confunde con "
              "`IEF` (eficiencia financiera): son cosas distintas",
    },

    "IGP": {
        "02": "gobernanza participativa",
        "04": "🔴 mide **2 de 7** mecanismos de participación (`D-010`)",
        "10": "**FORMA** — capacidad institucional, no sector",
        "12": "`d08` · Constitución §CAPA 0.5",
        "19": "«el GAD tiene gobernanza participativa» — mide dos séptimas "
              "partes del fenómeno (`D-010`)",
        "20": "⚠️ **alcance en disputa** · `D-010` abierta",
    },

    "IPE": {
        "02": "¿qué proporción del gasto ejecutado está vinculada a metas del "
              "PDOT?",
        "10": "**FORMA** — es articulación plan↔presupuesto",
        "12": "`d01` · `PCD-D01` **cerrado**",
        "15": "fórmula nativa en `H16b` · convertida a fórmula Excel para "
              "trazabilidad celda a celda",
        "20": "🟢 **el más maduro**: dominio cerrado, pregunta declarada, "
              "fórmula trazable",
    },

    "ITAM": {
        "02": "transparencia activa municipal",
        "10": "**FORMA**",
        "12": "`d07` · asignación **por confirmar** · `d07` en curación",
        "17": "🔴 alta · `LOTAIP` y su matriz de cumplimiento son ecuatorianas",
    },

    "MMP": {
        "01": "monitoreo mensual, trimestral y anual del pulso institucional",
        "12": f"{_PD} · capa INTERNA · `H25`-`H27`",
        "13": "mensual · trimestral · anual",
        "20": "⚠️ ¿es un **indicador** o un **producto de seguimiento**? "
              "`T2` lo clasificó `INTERNO`",
    },

    "PSG": {
        "01": "presupuesto con enfoque de género",
        "10": "**FONDO** con lectura transversal — ⚠️ caso límite del eje",
        "12": "`d12` · Constitución §CAPA 0.5",
        "15": "`H16c`",
        "17": "media · la política de género es marco internacional; su "
              "instrumentación, nacional",
    },

    "TGI": {
        "01": "confianza metodológica del propio sistema — **mide a QUIRA, no "
              "al GAD**",
        "09": "🔴 `H95` `L-07`: los pesos `20/20/25/25/10` son **«criterio "
              "experto (Dylus Lab), no PCA ni regresión»**",
        "10": "**FORMA** · y de segundo orden: es metaevaluación",
        "12": f"{_PD} · «probablemente transversal» · capa INTERNA",
        "19": "🔴 «QUIRA está validado» — `H95` `L-09` declara que es **una "
              "evaluación interna, no verificada externamente**",
        "20": "⚠️ **caso especial**: un indicador que evalúa al evaluador "
              "necesita un estatuto propio en la arquitectura",
    },
}


def leer_contrato() -> dict[str, dict[str, str]]:
    """Dominio, rol y capa de cada índice, del contrato ya construido. No se
    reescribe lo que otro artefacto ya declara."""
    if not _CONTRATO.exists():
        return {}
    out = {}
    for linea in _CONTRATO.read_text(encoding="utf-8").splitlines():
        m = re.match(r"\|\s*`([A-Z]{3,5})`\s*\|([^|]*)\|([^|]*)\|([^|]*)\|"
                     r"([^|]*)\|", linea)
        if m:
            out[m.group(1)] = {"dominio": m.group(2).strip(),
                               "rol": m.group(3).strip(),
                               "pregunta": m.group(4).strip(),
                               "capa": m.group(5).strip()}
    return out


def leer_hojas_indice() -> dict[str, dict[str, str]]:
    """★ Lo que el Gold Master dice de CADA índice, en su propia hoja.

    ⚠️ ESTA FUNCIÓN EXISTE POR UN ERROR DE ESTA DIRECCIÓN. La primera versión
    de `Q-M0` declaró `POR DECLARAR` el fenómeno, la evidencia y la fórmula de
    casi todos los índices **sin haber abierto sus hojas**. Javo lo corrigió
    en una línea: *«pero todos están en el Excel canónico»*.

    Y tenía razón: cada hoja trae su **título completo**, una **descripción
    con su fuente**, y en varios casos la **fórmula escrita en texto**. Es
    exactamente el patrón que esta misma investigación cazó tres veces en
    otros —`E_i` declarado `UNTRACEABLE` cuando la tesis lo definía, «no
    existe artefacto índice→dominio» cuando la Constitución lo declaraba,
    `011-C3` cerrando `NO DETERMINABLE` lo que 83 versiones podían fechar— y
    esta vez lo cometí yo.

    El tercer estado obliga en las dos direcciones: «no pude obtener» no es
    «no existe», **y tampoco al revés**: no haber mirado no autoriza a
    declarar vacío."""
    try:
        import openpyxl

        import config
        if not getattr(config, "GOLD_MASTER_RESUELTO", False):
            return {}
        wv = openpyxl.load_workbook(config.SIAP_PATH, data_only=True,
                                    read_only=True)
    except Exception:
        return {}

    out: dict[str, dict[str, str]] = {}
    for sigla in _FICHAS:
        hojas = [h for h in wv.sheetnames
                 if re.search(rf"(^|_){sigla}(_|$)", h)]
        if not hojas:
            continue
        # La primera hoja es la canónica del índice; las demás son vistas.
        hs = wv[hojas[0]]
        filas = []
        for fila in hs.iter_rows(min_row=1, max_row=10, max_col=6,
                                 values_only=True):
            filas.append([str(x) for x in fila if x is not None])
        texto = " ".join(" ".join(f) for f in filas)

        d: dict[str, str] = {"hojas": " · ".join(f"`{h}`" for h in hojas)}
        # Título completo: la fila 2 suele traer «HXX — SIGLA — NOMBRE».
        for f in filas[1:4]:
            for celda in f:
                m = re.search(rf"{sigla}\s*[—–-]\s*(.+)", celda)
                if m and len(m.group(1)) > 8:
                    d["titulo"] = m.group(1).strip()[:120]
                    break
            if "titulo" in d:
                break
        # Descripción y fuente declarada.
        for f in filas[2:5]:
            for celda in f:
                if len(celda) > 40 and not celda.startswith("="):
                    d.setdefault("descripcion", re.sub(r"\s+", " ",
                                                       celda).strip()[:260])
        m = re.search(r"Fuente[s]?[:\s]+([^.|]{4,90})", texto)
        if m:
            d["fuente"] = m.group(1).strip()
        # Fórmula declarada en texto dentro de la hoja.
        m = re.search(r"F[ÓO]RMULA[:\s]+([^|]{6,120})", texto)
        if m:
            d["formula"] = m.group(1).strip()
        elif re.search(rf"{sigla}\s*=", texto):
            m2 = re.search(rf"({sigla}\s*=[^|]{{4,120}})", texto)
            if m2:
                d["formula"] = m2.group(1).strip()
        # ¿Usa la escala AVEP? Importa: `D5` la dejó sin objeto declarado.
        if re.search(r"(Gesti[oó]n por|Ruptura Sist|Excelencia en|"
                     r"Transici[oó]n Cr[ií]tica)", texto):
            d["avep"] = "sí"
        out[sigla] = d
    return out


def leer_glosario() -> dict[str, str]:
    """La definición canónica de cada índice, del glosario del Gold Master.
    `{}` si no se resuelve — el tercer estado."""
    try:
        import openpyxl

        import config
        if not getattr(config, "GOLD_MASTER_RESUELTO", False):
            return {}
        wv = openpyxl.load_workbook(config.SIAP_PATH, data_only=True,
                                    read_only=True)
        h02 = wv["H02_GLOSARIO_QUIRA"]
    except Exception:
        return {}

    out = {}
    for fila in h02.iter_rows(min_row=1, max_row=240, max_col=3,
                              values_only=True):
        termino = str(fila[0] or "").strip()
        for sigla in _FICHAS:
            if termino == sigla or termino.startswith(sigla + " "):
                if sigla not in out:
                    out[sigla] = re.sub(r"\s+", " ",
                                        str(fila[1] or "")).strip()
    return out


def fusionar(contrato: dict, hojas: dict) -> dict[str, dict[str, str]]:
    """Las fichas declaradas MÁS lo que el instrumento ya dice de sí mismo.

    Precedencia: lo declarado por `GM-Ω` gana sobre lo derivado, porque una
    etapa que estudió el caso sabe más que un encabezado. Pero **un campo
    vacío se llena con el Excel antes que con `POR DECLARAR`**."""
    out = {s: dict(f) for s, f in _FICHAS.items()}
    for sigla, f in out.items():
        h = hojas.get(sigla, {})
        c = contrato.get(sigla, {})
        # 01 fenómeno ← título + descripción de su PROPIA hoja.
        #
        # ⚠️ AQUÍ EL INSTRUMENTO TIENE PRECEDENCIA, y es la excepción a la
        # regla general de esta función. Para «qué fenómeno se quiere
        # conocer», la fuente primaria es lo que la hoja dice de sí misma —
        # no lo que esta dirección recuerde. El caso `IFE` lo demostró: se
        # declaró «eficiencia financiera» y la hoja dice **fidelidad
        # electoral**.
        if h.get("titulo") or h.get("descripcion"):
            partes = [x for x in (h.get("titulo"), h.get("descripcion")) if x]
            derivado = " · ".join(partes)[:280]
            f["01"] = (f"{derivado}\n\n— *lectura de `GM-Ω`:* {f['01']}"
                       if "01" in f and f["01"] not in derivado else derivado)
        elif "01" not in f:
            pass
        # 02 pregunta ← contrato, si la declara
        if "02" not in f and c.get("pregunta") and \
                "POR_DECLARAR" not in c["pregunta"]:
            f["02"] = c["pregunta"]
        # 05 evidencia ← la fuente que la hoja declara
        if "05" not in f and h.get("fuente"):
            f["05"] = f"declarada en la hoja: {h['fuente']}"
        # 12 residencia ← contrato
        if "12" not in f and c.get("dominio") and \
                "POR_DECLARAR" not in c["dominio"]:
            f["12"] = c["dominio"]
        # 15 fórmula ← la que la hoja escribe, o al menos dónde vive
        if "15" not in f:
            if h.get("formula"):
                f["15"] = f"`{h['formula']}`"
            elif h.get("hojas"):
                f["15"] = f"implementada en {h['hojas']}"
        # 16 · usar AVEP es depender de `D5`, que quedó sin objeto declarado
        if h.get("avep") and "16" not in f:
            f["16"] = ("usa la escala **AVEP** para clasificar su resultado — "
                       "depende de `D5` (`ADR-058`), cuyo objeto **no está "
                       "declarado**")
    return out


def main() -> int:
    contrato = leer_contrato()
    glosario = leer_glosario()
    hojas = leer_hojas_indice()
    fichas = fusionar(contrato, hojas)

    total = len(fichas) * len(_CAMPOS)
    llenas = sum(len(f) for f in fichas.values())
    solo_decl = sum(len(f) for f in _FICHAS.values())
    print(f"indicadores: {len(fichas)} · campos por ficha: {len(_CAMPOS)} · "
          f"celdas: {total}")
    print(f"declaradas por GM-Ω: {solo_decl} · derivadas del Excel: "
          f"{llenas - solo_decl} · total: {llenas} "
          f"({llenas / total * 100:.0f} %)")
    print(f"POR DECLARAR: {total - llenas} "
          f"({(total - llenas) / total * 100:.0f} %)")
    print(f"hojas del Gold Master leídas: {len(hojas)}/{len(fichas)} índices")

    _escribir(contrato, glosario, total, llenas, fichas, hojas, solo_decl)
    print(f"→ {_SALIDA.relative_to(_RAIZ).as_posix()}")
    return 0


def _escribir(contrato, glosario, total, llenas, fichas, hojas,
              solo_decl) -> None:
    o: list[str] = []
    A = o.append

    A("# REARQ · `Q-M0` — MATRIZ ONTOLÓGICO-ARQUITECTÓNICA")
    A("")
    A("**DERIVADO — no editar a mano.** Lo regenera "
      "`scripts/rearq/matriz_ontologica.py`. El **dominio, rol y capa** salen "
      "del contrato índice→dominio; las **definiciones**, del glosario del "
      "Gold Master; el resto **se declara**, porque es un juicio.")
    A("")
    A("> ### Qué es esto")
    A("> El **primer artefacto de la Rearquitectura**. No es una fórmula "
      "nueva: es una **ficha por indicador** con veinte campos, y su función "
      "es obligar a separar lo que un indicador **calcula** de lo que se "
      "**afirma** con él.")
    A("")
    A("⚠️ **No empieza por el ICPI, y es deliberado.** Aunque sea el "
      "indicador más trabajado, entra **como uno más** —orden alfabético— "
      "para no diseñar la arquitectura nueva alrededor de su forma "
      "histórica.")
    A("")
    A("⚠️ **La Rearquitectura está autorizada exclusivamente para diseño "
      "conceptual y arquitectónico.** Esta matriz **no decide destinos**: los "
      "registra como evaluación inicial. Gold Master intacto, baseline "
      "**27,4582 %** congelado.")
    A("")

    # ── El hallazgo ───────────────────────────────────────────────────────
    A("## ★ Lo que esta matriz mide primero: cuánto no sabemos")
    A("")
    A("| | |")
    A("|---|---:|")
    A(f"| indicadores | {len(fichas)} |")
    A(f"| campos por ficha | {len(_CAMPOS)} |")
    A(f"| celdas totales | {total} |")
    A(f"| declaradas por `GM-Ω` | {solo_decl} |")
    A(f"| **derivadas del Gold Master** | **{llenas - solo_decl}** |")
    A(f"| **establecidas en total** | **{llenas}** "
      f"({llenas / total * 100:.0f} %) |")
    A(f"| **`POR DECLARAR`** | **{total - llenas}** "
      f"({(total - llenas) / total * 100:.0f} %) |")
    A("")
    A("### 📜 CORRECCIÓN · la primera versión de esta matriz declaró vacío lo "
      "que el instrumento ya decía")
    A("")
    A("La `v1` publicó **71 % `POR DECLARAR`**. Javo lo corrigió en una línea:")
    A("")
    A("> *«Pero todos están en el Excel canónico.»*")
    A("")
    A("Y tenía razón. **Cada índice tiene su hoja**, y cada hoja trae su "
      "título completo, una descripción con su fuente y, en varios casos, la "
      "fórmula escrita en texto. Esta dirección declaró `POR DECLARAR` "
      "**sin haberlas abierto**.")
    A("")
    A("⚠️ **Es el mismo patrón que esta investigación cazó tres veces en "
      "otros** —`E_i` declarado `UNTRACEABLE` cuando la tesis lo definía, «no "
      "existe artefacto índice→dominio» cuando la Constitución lo declaraba, "
      "`011-C3` cerrando `NO DETERMINABLE` lo que 83 versiones podían "
      "fechar— y esta vez lo cometió quien lo venía señalando.")
    A("")
    A("> ### El tercer estado obliga en las dos direcciones")
    A(">")
    A("> «No pude obtener» no es «no existe». **Y tampoco al revés: no haber "
      "mirado no autoriza a declarar vacío.**")
    A("")
    A("> ### Y el porcentaje que queda **sí** es el hallazgo")
    A(">")
    A("> Lo que sigue `POR DECLARAR` después de leer el instrumento es lo que "
      "**nadie ha establecido en ninguna parte**. Rellenarlo de memoria "
      "produciría un documento completo y falso.")
    A("")
    sin19 = sum(1 for s, f in fichas.items() if "19" not in f)
    A(f"Y el dato más incómodo, que el Excel **no** puede llenar: "
      f"**{sin19} de {len(fichas)} indicadores no tienen declarado qué "
      f"afirmación NO permiten hacer.** Ninguna hoja lo dice — porque un "
      f"instrumento describe lo que calcula, no sus límites.")
    A("")

    # ── Resumen ───────────────────────────────────────────────────────────
    A("## Resumen · los doce, en orden alfabético")
    A("")
    A("| Índice | Eje | Residencia | Estado Rearquitectura |")
    A("|---|---|---|---|")
    for sigla in sorted(fichas):
        f = fichas[sigla]
        c = contrato.get(sigla, {})
        eje = f.get("10", _PD)
        res = f.get("12") or (c.get("dominio") or _PD)
        A(f"| **`{sigla}`** | {eje[:46]} | {res[:44]} | "
          f"{f.get('20', _PD)[:60]} |")
    A("")

    # ── Las fichas ────────────────────────────────────────────────────────
    A("## Las fichas")
    A("")
    A("⚠️ **El orden de los campos es el método**: primero el fenómeno, al "
      "final la fórmula. Invertirlo sería empezar por la matemática — el "
      "hábito que `GM-Ω` vino a corregir.")
    A("")
    for sigla in sorted(fichas):
        f = fichas[sigla]
        c = contrato.get(sigla, {})
        A(f"### `{sigla}`")
        A("")
        d = glosario.get(sigla)
        if d:
            A(f"> **Glosario del Gold Master:** {d[:300]}")
            A("")
        A("| # | Campo | |")
        A("|---|---|---|")
        for num, nombre in _CAMPOS:
            valor = f.get(num)
            if not valor:
                # Lo que el contrato ya declara no se cuenta como vacío.
                if num == "02" and c.get("pregunta") and \
                        "POR_DECLARAR" not in c["pregunta"]:
                    valor = c["pregunta"]
                elif num == "12" and c.get("dominio") and \
                        "POR_DECLARAR" not in c["dominio"]:
                    valor = c["dominio"]
                else:
                    valor = _PD
            A(f"| {num} | {nombre} | {valor} |")
        A("")

    # ── Lo que la matriz deja ver ─────────────────────────────────────────
    A("## ★ Lo que la matriz deja ver de un vistazo")
    A("")
    A("| Hallazgo | Consecuencia |")
    A("|---|---|")
    A("| **`IFE` es Fidelidad Electoral; `IEF` es Eficiencia Financiera** | "
      "🔴 **no son duplicados**: dos siglas casi idénticas para fenómenos "
      "**sin relación**. El riesgo no es la redundancia — es confundirlos |")
    A("| **`IED` no pertenece a ningún dominio sectorial** | es el caso que "
      "prueba el eje FORMA: hoy **no existe el sitio** donde debería vivir |")
    A("| **`IBSC` hereda entera la multiplicatividad del ICPI** | si `D1` se "
      "rediseña, `IBSC` cambia **sin que nadie lo haya decidido para "
      "`IBSC`** |")
    A("| **`TGI` evalúa al evaluador** | un indicador que mide a QUIRA "
      "necesita **estatuto propio**; y `H95` declara que no está verificado "
      "externamente |")
    A("| **`ICPI` reside en `d06`** apoyado en un nombre retirado | su "
      "residencia está abierta desde `010` |")
    A("| **`IPE` es el más maduro** | dominio cerrado, pregunta declarada, "
      "fórmula trazable — **el patrón a replicar** |")
    A("")
    A("> ### Y el patrón de fondo")
    A(">")
    A("> Los indicadores que **sí** tienen pregunta declarada son los que "
      "pasaron por curación de dominio (`PCD`). Los que están en blanco son "
      "los que nunca la tuvieron. **La matriz no descubre indicadores malos: "
      "descubre indicadores no curados.**")
    A("")
    A("### 📜 El caso `IFE` · cómo se fabricó un hallazgo falso")
    A("")
    A("La `v1` de esta matriz publicó:")
    A("")
    A("> ~~«`IFE` e `IEF` son dos siglas casi idénticas para materia "
      "contigua → verificar duplicación»~~")
    A("")
    A("**Falso.** Bastó abrir `H16_IFE`:")
    A("")
    A("| Sigla | Qué es | Se alimenta de |")
    A("|---|---|---|")
    A("| `IFE` | **Índice de Fidelidad Electoral** | `H03_S1_ELECTORAL_CNE` |")
    A("| `IEF` | **Índice de Eficiencia Financiera** | `H20c` |")
    A("")
    A("No son materia contigua: **no tienen relación**. El parecido de las "
      "siglas produjo un hallazgo inventado, y sólo el instrumento lo "
      "desmintió.")
    A("")
    A("> ### El riesgo real no era la duplicación: es la confusión")
    A(">")
    A("> Si esta dirección las confundió teniendo el Excel delante, cualquier "
      "lector puede hacerlo. **Eso sí es un hallazgo**, y va a `T6` "
      "(renombrar / deprecar) como caso de nombres que no distinguen.")
    A("")

    # ── La regla ──────────────────────────────────────────────────────────
    A("## La regla que hace útil el campo 19")
    A("")
    A("```")
    A("  dato  →  evidencia  →  inferencia  →  afirmación")
    A("```")
    A("")
    A("**No son sinónimos.** Es lo que `D2` dejó demostrado: entre «no hay "
      "evidencia» y «el fenómeno no ocurrió» hay dos saltos, y cada uno es "
      "una decisión.")
    A("")
    A("> El campo **19 — qué afirmación NO permite hacer** es el mecanismo "
      "que impide que la capa de presentación convierta una medición limitada "
      "en una afirmación mucho mayor.")
    A("")
    A("Y ya tiene un caso probado: el ICPI mide congruencia acreditada de 25 "
      "metas al corte de abril, y `data/gm_snapshot.json` lo publica como que "
      "«mide velocidad de ejecución» (`D-011`). **El campo 19 existe porque "
      "eso ya pasó.**")
    A("")

    A("## Lo que `Q-M0` NO hace")
    A("")
    A("- **No decide el destino de ningún indicador.** El campo 20 es una "
      "evaluación inicial, no un dictamen.")
    A("- **No rellena de memoria.** Lo no establecido queda `POR DECLARAR` y "
      "se cuenta.")
    A("- **No toca fórmulas** ni el Gold Master.")
    A("- **No resuelve FONDO/FORMA**: lo aplica donde `010` lo estableció y "
      "deja ver dónde falta.")
    A("")
    A("> ### La pregunta que abre `Q-M1`")
    A(">")
    A("> No «¿qué índices conservamos?», sino **«¿qué preguntas sobre la "
      "gestión pública necesita responder QUIRA, y qué evidencia permite "
      "responderlas legítimamente?»**. Los indicadores tendrán que ganarse su "
      "lugar dentro de esa respuesta — el ICPI incluido.")
    A("")
    A("---")
    A(f"*REARQ · `Q-M0` · {len(fichas)} indicadores × {len(_CAMPOS)} campos "
      f"· {total - llenas} celdas por declarar · el Gold Master no se "
      f"modificó · Dylus Lab © 2026*")

    _SALIDA.write_text("\n".join(o) + "\n", encoding="utf-8")


if __name__ == "__main__":
    raise SystemExit(main())
