#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
scripts/sercop_sprint0.py — QUIRA OS · Sprint 0
================================================================
Ingesta de contratación pública SERCOP para las 5 entidades del
Holding Municipal Montecristi. Fuente: OCDS API datos abiertos.

Almacena en:
  1. Supabase / SQLite  → tabla sercop_contratos
  2. Excel Gold Master  → H06 Zona Cruda (filas 66+)

Uso:
  python scripts/sercop_sprint0.py              # todos los años
  python scripts/sercop_sprint0.py --year 2025  # solo 2025
  python scripts/sercop_sprint0.py --dry-run     # sin escribir
  python scripts/sercop_sprint0.py --supabase-only  # solo BD

Dylus Lab © 2026
"""
from __future__ import annotations

import argparse
import json
import sys
import time
import urllib.request
import urllib.parse
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

sys.path.insert(0, str(Path(__file__).parent.parent))

# ── CONSTANTES ────────────────────────────────────────────────────────────────
API_BASE   = "https://datosabiertos.compraspublicas.gob.ec/PLATAFORMA/api"
GOLD_MASTER = Path(r"C:\Users\DELL\Desktop\Javo\Dylus Lab\ProyecT\SIAP-ICPI_GOLD_MASTER_v5.5_TGI.xlsx")
H06_SHEET   = "H06_S4_CONTRATACIÓN_SERCOP"
H06_ZONA_CRUDA_ROW = 66   # fila 1-indexed donde empieza la Zona Cruda

YEARS = [2023, 2024, 2025, 2026]

# ── ENTIDADES DEL HOLDING ─────────────────────────────────────────────────────
# buyer_keywords: términos de búsqueda (case insensitive, partial match)
HOLDING = {
    "GAD": {
        "codigo": "GAD",
        "nombre": "Gobierno Autónomo Descentralizado Montecristi",
        "ruc":    "1360001010001",
        "buyer_keywords": ["MUNICIPIO DE MONTECRISTI", "GADMCM", "GAD MUNICIPAL.*MONTECRISTI"],
        "sercop_id": 2896,
    },
    "EMAI": {
        "codigo": "EMAI",
        "nombre": "Empresa Municipal de Aseo Integral Montecristi EP",
        "ruc":    "1360086760001",
        "buyer_keywords": ["ASEO INTEGRAL MONTECRISTI", "EMAI", "EMPRESA MUNICIPAL DE ASEO.*MONTECRISTI"],
        "sercop_id": 827527,
    },
    "BOMBEROS": {
        "codigo": "BOMBEROS",
        "nombre": "Cuerpo de Bomberos de Montecristi",
        "ruc":    "1360051380001",
        "buyer_keywords": ["CUERPO DE BOMBEROS DE MONTECRISTI", "BOMBEROS.*MONTECRISTI"],
        "sercop_id": 132804,
    },
    "PATRONATO": {
        "codigo": "PATRONATO",
        "nombre": "Patronato Municipal de Amparo Social de Montecristi",
        "ruc":    "1360059440001",
        "buyer_keywords": ["PATRONATO.*MONTECRISTI", "PATRONATO MUNICIPAL.*MONTECRISTI"],
        "sercop_id": 362253,
    },
    "HABITAT": {
        "codigo": "HABITAT",
        "nombre": "Empresa Pública Municipal de Hábitat y Vivienda Montecristi (MONTEHOGAR-EP)",
        "ruc":    "",          # no provisto — se llena desde API
        "buyer_keywords": ["MONTEHOGAR", "HABITAT.*MONTECRISTI", "VIVIENDA.*MONTECRISTI"],
        "sercop_id": 922099,
    },
}


# ── HTTP ──────────────────────────────────────────────────────────────────────
def _get(url: str, timeout: int = 30) -> dict | None:
    """GET con reintentos."""
    for attempt in range(3):
        try:
            req = urllib.request.Request(
                url,
                headers={"User-Agent": "QUIRA-Scout/2.0", "Accept": "application/json"},
            )
            with urllib.request.urlopen(req, timeout=timeout) as r:
                return json.loads(r.read().decode("utf-8", errors="replace"))
        except Exception as e:
            if attempt < 2:
                time.sleep(2 ** attempt)
            else:
                print(f"  [WW] GET {url[:80]}... → {e}")
                return None


# ── CLASIFICACIÓN DE ENTIDAD ──────────────────────────────────────────────────
import re as _re

def _classify_buyer(buyer_name: str) -> str | None:
    """Retorna el codigo de entidad del Holding o None."""
    buyer_upper = buyer_name.upper()
    for codigo, info in HOLDING.items():
        for kw in info["buyer_keywords"]:
            if _re.search(kw, buyer_upper, _re.IGNORECASE):
                return codigo
    return None


# ── ESTADO del proceso ────────────────────────────────────────────────────────
def _estado_from_ocid(ocid: str) -> str:
    """
    Infiere estado a partir del OCID (sin llamar al record API para ahorrar requests).
    El record API se llama solo para los primeros 50 por año.
    """
    return "desconocido"   # se actualiza en enrich_estado()


def _enrich_estado(ocid: str) -> dict:
    """Llama al record API para obtener estado y URL del proceso."""
    url = f"{API_BASE}/record?ocid={urllib.parse.quote(ocid)}"
    data = _get(url)
    if not data:
        return {"estado": "desconocido", "url_proceso": ""}

    # El record OCDS tiene releases con tags que indican el estado
    releases = data.get("records", [{}])[0].get("compiledRelease", {})
    tags = releases.get("tag", [])
    tender = releases.get("tender", {})
    status = tender.get("status", "desconocido")

    # URL del proceso en el portal
    process_id = releases.get("tender", {}).get("id", "")
    url_proceso = (
        f"https://www.compraspublicas.gob.ec/ProcesoContratacion/compras/"
        f"PC/informacionProcesoContratacion2.cpe?idSoliCompra={process_id}"
        if process_id else ""
    )

    return {"estado": status, "url_proceso": url_proceso}


# ── FETCHING POR AÑO ──────────────────────────────────────────────────────────
def fetch_year(year: int, verbose: bool = True) -> list[dict]:
    """
    Descarga todos los procesos de entidades MONTECRISTI para un año dado.
    Clasifica cada proceso por entidad del Holding.
    """
    if verbose:
        print(f"\n[>>] Año {year} — consultando API SERCOP...")

    results: list[dict] = []
    page = 1

    while True:
        url = (
            f"{API_BASE}/search_ocds"
            f"?year={year}&buyer=MONTECRISTI&page={page}"
        )
        data = _get(url)
        if not data or not data.get("data"):
            break

        items = data["data"]
        total = data.get("total", 0)
        pages = data.get("pages", 1)

        if verbose and page == 1:
            print(f"  Total procesos MONTECRISTI {year}: {total} ({pages} páginas)")

        for item in items:
            buyer = item.get("buyer", "")
            codigo = _classify_buyer(buyer)
            if not codigo:
                continue   # no pertenece al Holding

            info = HOLDING[codigo]
            rec = {
                "sercop_id":           item.get("id"),
                "ocid":                item.get("ocid", ""),
                "entidad_codigo":      codigo,
                "entidad_nombre":      buyer,
                "entidad_ruc":         info["ruc"],
                "anio":                year,
                "mes":                 item.get("month"),
                "tipo_contratacion":   item.get("internal_type", ""),
                "tipo_metodo":         item.get("method", ""),
                "estado":              "por_verificar",
                "monto_presupuestado": float(item.get("budget", 0) or 0),
                "monto_adjudicado":    float(item.get("amount", 0) or 0),
                "titulo":              (item.get("title") or "")[:500],
                "descripcion":         (item.get("description") or "")[:1000],
                "proveedor":           item.get("suppliers", ""),
                "localidad":           item.get("locality", ""),
                "fecha_publicacion":   item.get("date", "")[:10] if item.get("date") else "",
                "url_proceso":         "",
                "fecha_ingesta":       datetime.now(timezone.utc).isoformat(),
                "fuente":              "SERCOP_OCDS_API",
            }
            results.append(rec)

        if page >= pages:
            break
        page += 1
        time.sleep(0.3)    # respetar rate limit

    if verbose:
        by_entity: dict[str, int] = {}
        for r in results:
            c = r["entidad_codigo"]
            by_entity[c] = by_entity.get(c, 0) + 1
        for ent, cnt in sorted(by_entity.items()):
            print(f"    {ent}: {cnt} procesos")

    return results


# ── SUPABASE / SQLite ─────────────────────────────────────────────────────────
def save_to_db(records: list[dict], dry_run: bool = False) -> int:
    """Inserta o actualiza registros en sercop_contratos. Retorna count."""
    if dry_run or not records:
        return 0

    from sentinel.db_config import get_connection, init_db
    init_db()   # crea tabla si no existe
    conn = get_connection()
    c    = conn.cursor()

    inserted = 0
    updated  = 0
    for rec in records:
        # Verificar si ya existe (por ocid)
        existing = c.execute(
            "SELECT id FROM sercop_contratos WHERE ocid = ?", (rec["ocid"],)
        ).fetchone()

        if existing:
            c.execute("""
                UPDATE sercop_contratos SET
                    entidad_nombre      = ?,
                    anio                = ?,
                    mes                 = ?,
                    tipo_contratacion   = ?,
                    tipo_metodo         = ?,
                    estado              = ?,
                    monto_presupuestado = ?,
                    monto_adjudicado    = ?,
                    titulo              = ?,
                    descripcion         = ?,
                    proveedor           = ?,
                    localidad           = ?,
                    fecha_publicacion   = ?,
                    url_proceso         = ?,
                    fecha_ingesta       = ?
                WHERE ocid = ?
            """, (
                rec["entidad_nombre"], rec["anio"], rec["mes"],
                rec["tipo_contratacion"], rec["tipo_metodo"], rec["estado"],
                rec["monto_presupuestado"], rec["monto_adjudicado"],
                rec["titulo"], rec["descripcion"], rec["proveedor"],
                rec["localidad"], rec["fecha_publicacion"], rec["url_proceso"],
                rec["fecha_ingesta"], rec["ocid"],
            ))
            updated += 1
        else:
            c.execute("""
                INSERT INTO sercop_contratos (
                    sercop_id, ocid, entidad_codigo, entidad_nombre, entidad_ruc,
                    anio, mes, tipo_contratacion, tipo_metodo, estado,
                    monto_presupuestado, monto_adjudicado, titulo, descripcion,
                    proveedor, localidad, fecha_publicacion, url_proceso,
                    fecha_ingesta, fuente
                ) VALUES (?,?,?,?,?, ?,?,?,?,?, ?,?,?,?, ?,?,?,?, ?,?)
            """, (
                rec["sercop_id"], rec["ocid"], rec["entidad_codigo"],
                rec["entidad_nombre"], rec["entidad_ruc"],
                rec["anio"], rec["mes"], rec["tipo_contratacion"],
                rec["tipo_metodo"], rec["estado"],
                rec["monto_presupuestado"], rec["monto_adjudicado"],
                rec["titulo"], rec["descripcion"],
                rec["proveedor"], rec["localidad"], rec["fecha_publicacion"],
                rec["url_proceso"], rec["fecha_ingesta"], rec["fuente"],
            ))
            inserted += 1

    conn.commit()
    conn.close()
    return inserted + updated


# ── EXCEL H06 ─────────────────────────────────────────────────────────────────
def save_to_h06(all_records: list[dict], dry_run: bool = False) -> None:
    """
    Escribe resumen SERCOP en H06 Zona Cruda del Gold Master.
    No toca las fórmulas existentes de la Zona Inteligente (cols N-T).
    Crea un bloque de datos tabulares en fila 66+ con los campos canónicos.
    """
    if dry_run:
        print("  [DRY] H06 Excel: no escribe")
        return
    if not GOLD_MASTER.exists():
        print(f"  [WW] Gold Master no encontrado: {GOLD_MASTER}")
        return

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        print("  [WW] openpyxl no instalado — saltando escritura Excel")
        return

    print(f"  Abriendo Gold Master ({GOLD_MASTER.name})...")
    wb = openpyxl.load_workbook(str(GOLD_MASTER))

    if H06_SHEET not in wb.sheetnames:
        print(f"  [WW] Hoja {H06_SHEET!r} no encontrada")
        wb.close()
        return

    ws = wb[H06_SHEET]

    # Cabecera en fila H06_ZONA_CRUDA_ROW - 1
    header_row = H06_ZONA_CRUDA_ROW - 1
    HEADERS = [
        "ENTIDAD", "OCID", "AÑO", "MES", "TIPO_CONTRATACION",
        "ESTADO", "MONTO_PRESUPUESTADO_USD", "MONTO_ADJUDICADO_USD",
        "PROVEEDOR", "TITULO", "FECHA_PUBLICACION", "FUENTE_INGESTA",
    ]
    for col_idx, h in enumerate(HEADERS, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=h)
        cell.font      = Font(bold=True, color="FFFFFF")
        cell.fill      = PatternFill("solid", fgColor="1E3A5F")
        cell.alignment = Alignment(horizontal="center")

    # Datos a partir de H06_ZONA_CRUDA_ROW
    for row_offset, rec in enumerate(all_records):
        row_idx = H06_ZONA_CRUDA_ROW + row_offset
        ws.cell(row=row_idx, column=1,  value=rec["entidad_codigo"])
        ws.cell(row=row_idx, column=2,  value=rec["ocid"])
        ws.cell(row=row_idx, column=3,  value=rec["anio"])
        ws.cell(row=row_idx, column=4,  value=rec["mes"])
        ws.cell(row=row_idx, column=5,  value=rec["tipo_contratacion"])
        ws.cell(row=row_idx, column=6,  value=rec["estado"])
        ws.cell(row=row_idx, column=7,  value=rec["monto_presupuestado"])
        ws.cell(row=row_idx, column=8,  value=rec["monto_adjudicado"])
        ws.cell(row=row_idx, column=9,  value=rec["proveedor"])
        ws.cell(row=row_idx, column=10, value=rec["titulo"])
        ws.cell(row=row_idx, column=11, value=rec["fecha_publicacion"])
        ws.cell(row=row_idx, column=12, value=rec["fuente"])

    wb.save(str(GOLD_MASTER))
    print(f"  [OK] H06 actualizado: {len(all_records)} filas en Zona Cruda ({GOLD_MASTER.name})")


# ── RESUMEN ───────────────────────────────────────────────────────────────────
def print_summary(all_records: list[dict]) -> None:
    """Imprime resumen ejecutivo por entidad × año."""
    print("\n" + "=" * 64)
    print("RESUMEN — SERCOP Sprint 0 · Holding Municipal Montecristi")
    print("=" * 64)

    # Agrupar por entidad
    by_entity: dict[str, dict] = {}
    for rec in all_records:
        cod = rec["entidad_codigo"]
        yr  = rec["anio"]
        key = f"{cod}_{yr}"
        if key not in by_entity:
            by_entity[key] = {
                "entidad": cod,
                "anio": yr,
                "count": 0,
                "monto_total": 0.0,
            }
        by_entity[key]["count"]        += 1
        by_entity[key]["monto_total"]  += rec["monto_adjudicado"]

    for key in sorted(by_entity.keys()):
        g = by_entity[key]
        print(
            f"  {g['entidad']:<10} {g['anio']}  "
            f"{g['count']:>4} procesos  "
            f"${g['monto_total']:>14,.2f} USD adjudicado"
        )

    total_monto = sum(r["monto_adjudicado"] for r in all_records)
    print(f"\n  TOTAL: {len(all_records)} procesos | ${total_monto:,.2f} USD")
    print("=" * 64)


# ── LEER DESDE BD → H06 ──────────────────────────────────────────────────────
def export_db_to_h06() -> None:
    """
    Lee TODOS los registros de sercop_contratos en Supabase/SQLite
    y los escribe en H06 Zona Cruda del Gold Master (completo, todos los años).
    Útil después de ingestas por año para consolidar H06.
    """
    from sentinel.db_config import get_connection, init_db
    init_db()
    conn = get_connection()
    c    = conn.cursor()

    rows = c.execute("""
        SELECT entidad_codigo, ocid, anio, mes, tipo_contratacion,
               estado, monto_presupuestado, monto_adjudicado,
               proveedor, titulo, fecha_publicacion, fuente
        FROM sercop_contratos
        ORDER BY entidad_codigo, anio, fecha_publicacion
    """).fetchall()
    conn.close()

    if not rows:
        print("[WW] sercop_contratos vacía — nada que exportar a H06")
        return

    # Convertir a lista de dicts compatible con save_to_h06
    # Supabase devuelve RealDictRow; SQLite devuelve sqlite3.Row — ambos soportan acceso por nombre
    records = [
        {
            "entidad_codigo":      r["entidad_codigo"],
            "ocid":                r["ocid"],
            "anio":                r["anio"],
            "mes":                 r["mes"],
            "tipo_contratacion":   r["tipo_contratacion"],
            "estado":              r["estado"],
            "monto_presupuestado": float(r["monto_presupuestado"] or 0),
            "monto_adjudicado":    float(r["monto_adjudicado"] or 0),
            "proveedor":           r["proveedor"],
            "titulo":              r["titulo"],
            "fecha_publicacion":   r["fecha_publicacion"],
            "fuente":              r["fuente"],
        }
        for r in rows
    ]

    print(f"[>>] Exportando {len(records)} registros de BD → H06 Zona Cruda...")
    save_to_h06(records, dry_run=False)


# ── MAIN ──────────────────────────────────────────────────────────────────────
def main() -> None:
    parser = argparse.ArgumentParser(
        description="Sprint 0 — Ingesta SERCOP Holding Municipal Montecristi | QUIRA OS"
    )
    parser.add_argument("--year",         type=int, default=None, help="Solo este año (default: 2023-2026)")
    parser.add_argument("--dry-run",      action="store_true",    help="No escribe en BD ni Excel")
    parser.add_argument("--supabase-only",action="store_true",    help="Solo BD, no Excel")
    parser.add_argument("--no-excel",     action="store_true",    help="No actualiza el Excel Gold Master")
    parser.add_argument("--from-db-to-h06", action="store_true", help="Lee BD y escribe H06 completo (sin fetch SERCOP)")
    parser.add_argument("--verbose",      action="store_true",    default=True)
    args = parser.parse_args()

    # Modo especial: sólo exportar BD → H06
    if args.from_db_to_h06:
        export_db_to_h06()
        return

    years_to_fetch = [args.year] if args.year else YEARS
    all_records: list[dict] = []

    print(f"[>>] Sprint 0 — Ingesta SERCOP")
    print(f"     Entidades: {list(HOLDING.keys())}")
    print(f"     Años: {years_to_fetch}")
    print(f"     Modo: {'DRY RUN (sin escritura)' if args.dry_run else 'PRODUCCIÓN'}")
    print()

    for year in years_to_fetch:
        records = fetch_year(year, verbose=args.verbose)
        all_records.extend(records)
        time.sleep(0.5)

    if not all_records:
        print("\n[WW] Sin registros encontrados. Verificar conectividad con SERCOP.")
        return

    # Guardar en BD
    if not args.dry_run:
        print(f"\n[>>] Guardando {len(all_records)} registros en BD...")
        saved = save_to_db(all_records, dry_run=args.dry_run)
        print(f"  [OK] BD actualizada: {saved} registros procesados")
    else:
        print(f"\n[DRY] BD: se guardarían {len(all_records)} registros")

    # Guardar en Excel H06
    if not args.supabase_only and not args.no_excel:
        print(f"\n[>>] Actualizando H06 en Gold Master...")
        save_to_h06(all_records, dry_run=args.dry_run)

    # Resumen
    print_summary(all_records)

    # Guardar JSON de respaldo
    out_path = Path(__file__).parent.parent / "data" / "scouting" / "sercop_sprint0_holding.json"
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(
            {
                "sprint": "0",
                "fecha": datetime.now(timezone.utc).isoformat(),
                "total_records": len(all_records),
                "records": all_records,
            },
            f, ensure_ascii=False, indent=2,
        )
    print(f"\n[OK] JSON respaldo: {out_path}")
    print("[OK] Sprint 0 completado.")


if __name__ == "__main__":
    main()
