"""
scripts/sync_protocol.py
QUIRA OS — Protocolo Canónico de Sincronización de 4 Fuentes de Verdad
Dylus Lab © 2026-05-20

REGLA CANÓNICA:
  Cada vez que se construye algo en una fuente, debe diseminarse
  automáticamente en las otras tres. "La parte que corresponda":

  quira-os (código)   → GitHub (commit+push) + Vault Obsidian (nota arch) + GM (CHANGELOG)
  Obsidian (doctrina) → quira-os (vault_registry rebuild) + GM (cross-ref) + GitHub
  Gold Master (datos) → quira-os (loader update) + Obsidian (data note) + GitHub
  GitHub (sync)       → automático post-commit

API:
  sync_delivery(component_name, component_type, description, version, files)
  → escribe nota vault + entrada GM changelog + muestra checklist GitHub

Tipos de componente: 'code' | 'data' | 'arch' | 'calibration'
"""
from __future__ import annotations

import io
import json
import os
import subprocess
import sys
from datetime import datetime
from pathlib import Path
from typing import Optional

# Fix Windows cp1252 console encoding for emoji/Unicode in print statements
if sys.stdout and hasattr(sys.stdout, "buffer"):
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

# ── RUTAS CANÓNICAS ───────────────────────────────────────────────────────────

_ROOT          = Path(__file__).parent.parent
_VAULT_ROOT    = Path(r"C:\Proyectos\QUIRA\knowledge_base\QUIRA_KB_Montecristi")
_GM_PATH       = Path(r"C:\Users\DELL\Desktop\Javo\Dylus Lab\ProyecT\SIAP-ICPI_GOLD_MASTER_v5.5_TGI.xlsx")
_VAULT_MODULE  = _VAULT_ROOT / "08_EJECUCION"   # D3_EJECUCION — motor analítico
_ARCH_MODULE   = _VAULT_ROOT / "00_CORE"          # meta-arquitectura

# ── MAPPING TIPO → VAULT MODULE ────────────────────────────────────────────────

_TYPE_TO_MODULE: dict[str, Path] = {
    "code":        _VAULT_MODULE,
    "calibration": _VAULT_MODULE,
    "data":        _VAULT_MODULE,
    "arch":        _ARCH_MODULE,
}


# ── GENERADOR DE NOTA OBSIDIAN ────────────────────────────────────────────────

def _write_vault_note(
    name:        str,
    comp_type:   str,
    description: str,
    version:     str,
    files:       list[str],
    sat_codes:   list[str],
    dimensions:  list[str],
    fecha:       str,
) -> Path:
    """Crea o actualiza una nota en el vault Obsidian para el componente."""
    module_dir = _TYPE_TO_MODULE.get(comp_type, _VAULT_MODULE)
    module_dir.mkdir(parents=True, exist_ok=True)

    slug      = name.replace(" ", "-").replace("/", "-").replace(".", "-")
    note_path = module_dir / f"{slug}.md"

    sats_yaml  = json.dumps(sat_codes)  if sat_codes  else "[]"
    dims_yaml  = json.dumps(dimensions) if dimensions else "[]"
    files_list = "\n".join(f"  - `{f}`" for f in files) if files else "  - (sin archivos)"

    content = f"""---
id: {slug}
nombre: {name}
tipo: {comp_type}
version: {version}
fecha: {fecha}
dimension_tgi: {dims_yaml}
sat_codes: {sats_yaml}
module: 08_EJECUCION
fuente: QUIRA-OS
keywords: [{name}, RC-7, longitudinal, calibracion, ejecucion presupuestaria]
---

# {name} — {version}

> Generado automáticamente por `scripts/sync_protocol.py` · {fecha}
> Fuente canónica: `quira-os/` → GitHub → este vault → Gold Master

## Descripción

{description}

## Archivos

{files_list}

## Dimensiones TGI activadas

{chr(10).join(f"- [[{d}]]" for d in dimensions) if dimensions else "- (todas las dimensiones)"}

## SAT Codes implicados

{chr(10).join(f"- [[{s}]]" for s in sat_codes) if sat_codes else "- (análisis sin SAT directo)"}

## Integración con el loop semántico

```
Gold Master (v5.5_TGI)
  ↓
{name}
  ↓
SAT Engine → Vault Enricher → Legal Router → Claude Haiku → PMV
```

## Referencias

- COPLAFIP Art.97-100 — programación y reforma presupuestaria
- COOTAD Art.432 — facultades Alcalde sobre presupuesto
- [[TRAYECTORIA_2023_2026]] — serie longitudinal GAD Montecristi
- [[08_EJECUCION/_Índice_Ejecucion]]
"""

    note_path.write_text(content, encoding="utf-8")
    return note_path


# ── ACTUALIZACIÓN GOLD MASTER CHANGELOG ───────────────────────────────────────

def _update_gm_changelog(
    name:        str,
    comp_type:   str,
    description: str,
    version:     str,
    files:       list[str],
    fecha:       str,
) -> bool:
    """
    Agrega una fila a la hoja RC_CHANGELOG del Gold Master.
    Crea la hoja si no existe. No toca hojas existentes.
    """
    if not _GM_PATH.exists():
        print(f"⚠ Gold Master no encontrado en {_GM_PATH}")
        return False

    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        wb   = load_workbook(str(_GM_PATH), keep_vba=False)
        sheetname = "RC_CHANGELOG"

        if sheetname not in wb.sheetnames:
            ws = wb.create_sheet(sheetname)
            # Header row
            headers = ["Fecha", "Versión", "Componente", "Tipo", "Descripción", "Archivos"]
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.font      = Font(bold=True, color="FFFFFF")
                cell.fill      = PatternFill("solid", start_color="1F4E79")
                cell.alignment = Alignment(horizontal="center")
            ws.column_dimensions["A"].width = 12
            ws.column_dimensions["B"].width = 18
            ws.column_dimensions["C"].width = 30
            ws.column_dimensions["D"].width = 14
            ws.column_dimensions["E"].width = 60
            ws.column_dimensions["F"].width = 40
        else:
            ws = wb[sheetname]

        # Agregar fila
        next_row = ws.max_row + 1
        ws.cell(row=next_row, column=1, value=fecha)
        ws.cell(row=next_row, column=2, value=version)
        ws.cell(row=next_row, column=3, value=name)
        ws.cell(row=next_row, column=4, value=comp_type)
        ws.cell(row=next_row, column=5, value=description[:200])
        ws.cell(row=next_row, column=6, value=" | ".join(files[:5]))

        wb.save(str(_GM_PATH))
        return True

    except Exception as e:
        print(f"⚠ No se pudo actualizar GM CHANGELOG: {e}")
        return False


# ── GIT PUSH ──────────────────────────────────────────────────────────────────

def _git_push_reminder(files: list[str], version: str) -> str:
    """Muestra los comandos git para completar el sync con GitHub."""
    files_str = " ".join(f'"{f}"' for f in files[:10])
    return (
        f"\n── GitHub Sync ──────────────────────────────────\n"
        f"git add {files_str}\n"
        f'git commit -m "sync({version}): diseminación canónica 4 fuentes"\n'
        f"git push origin main\n"
        f"──────────────────────────────────────────────────\n"
    )


# ── API PRINCIPAL ──────────────────────────────────────────────────────────────

def sync_delivery(
    component_name: str,
    component_type: str            = "code",
    description:    str            = "",
    version:        str            = "",
    files:          list[str]      = None,
    sat_codes:      list[str]      = None,
    dimensions:     list[str]      = None,
    auto_git:       bool           = False,
) -> dict:
    """
    Ejecuta el protocolo canónico de sync para un componente entregado.

    Args:
        component_name: nombre del componente (ej: "RC-7.3 Calibration Layer")
        component_type: 'code' | 'data' | 'arch' | 'calibration'
        description:    descripción para vault note y GM changelog
        version:        versión (ej: "RC-7.3")
        files:          lista de archivos creados/modificados
        sat_codes:      SAT codes implicados
        dimensions:     dimensiones TGI
        auto_git:       si True, ejecuta git add+commit+push automáticamente

    Returns:
        dict con resultado de cada sync target
    """
    fecha  = datetime.now().strftime("%Y-%m-%d")
    files  = files     or []
    sats   = sat_codes or []
    dims   = dimensions or ["D3_EJECUCION", "D2_PLANIFICACION"]

    results: dict = {
        "component": component_name,
        "version":   version,
        "fecha":     fecha,
        "vault":     None,
        "gm":        None,
        "github":    None,
    }

    # ── 1. Obsidian vault ────────────────────────────────────────────────────
    print(f"📒 Sync vault Obsidian → {_VAULT_MODULE.name}/...")
    try:
        note_path = _write_vault_note(
            name        = component_name,
            comp_type   = component_type,
            description = description,
            version     = version,
            files       = files,
            sat_codes   = sats,
            dimensions  = dims,
            fecha       = fecha,
        )
        print(f"   ✓ Nota creada: {note_path.name}")
        results["vault"] = str(note_path)
    except Exception as e:
        print(f"   ✗ Error vault: {e}")
        results["vault"] = f"ERROR: {e}"

    # ── 2. Gold Master changelog ──────────────────────────────────────────────
    print(f"📊 Sync Gold Master → RC_CHANGELOG...")
    ok = _update_gm_changelog(
        name        = component_name,
        comp_type   = component_type,
        description = description,
        version     = version,
        files       = files,
        fecha       = fecha,
    )
    results["gm"] = "OK" if ok else "SKIPPED"
    if ok:
        print(f"   ✓ Fila añadida en RC_CHANGELOG")

    # ── 3. GitHub ─────────────────────────────────────────────────────────────
    if auto_git:
        print(f"🐙 Sync GitHub → git push...")
        try:
            os.chdir(str(_ROOT))
            for f in files:
                subprocess.run(["git", "add", f], check=True, capture_output=True)
            msg = (
                f"sync({version}): diseminación canónica 4 fuentes — "
                f"{component_name}\n\n"
                f"Co-Authored-By: Claude Sonnet 4.6 <noreply@anthropic.com>"
            )
            subprocess.run(["git", "commit", "-m", msg], check=True, capture_output=True)
            subprocess.run(["git", "push", "origin", "main"], check=True, capture_output=True)
            print(f"   ✓ Push a GitHub OK")
            results["github"] = "PUSHED"
        except Exception as e:
            print(f"   ✗ Error git: {e}")
            results["github"] = f"ERROR: {e}"
    else:
        print(_git_push_reminder(files, version))
        results["github"] = "MANUAL_REQUIRED"

    # ── Resumen ───────────────────────────────────────────────────────────────
    print(
        f"\n✅ Sync {version} completado:\n"
        f"   Vault:  {results['vault'] or 'SKIP'}\n"
        f"   GM:     {results['gm']}\n"
        f"   GitHub: {results['github']}\n"
    )
    return results


# ── MODO CLI ──────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    """
    Uso:
        python scripts/sync_protocol.py <nombre> <tipo> <version> [archivos...]

    Ejemplo:
        python scripts/sync_protocol.py "RC-7.3 Calibration Layer" calibration RC-7.3 \\
            sentinel/calibration_layer.py data/rc72_calibration.json
    """
    args = sys.argv[1:]
    if len(args) < 3:
        print("Uso: sync_protocol.py <nombre> <tipo> <version> [archivos...]")
        sys.exit(1)

    name   = args[0]
    ctype  = args[1]
    ver    = args[2]
    ffiles = args[3:] if len(args) > 3 else []

    sync_delivery(
        component_name = name,
        component_type = ctype,
        version        = ver,
        files          = ffiles,
        description    = f"Componente {name} v{ver} entregado en QUIRA OS.",
        auto_git       = False,
    )
