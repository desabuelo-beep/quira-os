# -*- coding: utf-8 -*-
"""
update_gold_master_rc1.py
Update SIAP-ICPI Gold Master v5.5 with QUIRA OS RC-1.1 completion data.
Sheets: H84_SNAPSHOT_REGISTRY, H85_ALERTS_LOG, H02_GLOSARIO_QUIRA,
        H77_DATA_DICTIONARY, H36c_OBSIDIAN_MAP
Dylus Lab (c) 2026
"""

import sys
import shutil

try:
    from config import DATOS_DIR as _DATOS
except Exception:                                        # noqa: BLE001
    import os as _os
    from pathlib import Path as _P
    _DATOS = _P(_os.environ.get("QUIRA_DATOS", "."))

try:
    import openpyxl
except ImportError:
    sys.exit("ERROR: openpyxl not installed. Run: pip install openpyxl")

SRC = str(_DATOS / "SIAP-ICPI_GOLD_MASTER_v5.5_TGI.xlsx")
BAK = str(_DATOS / "SIAP-ICPI_GOLD_MASTER_v5.5_TGI_BACKUP_PRE_RC1.xlsx")

# ── 1. BACKUP ─────────────────────────────────────────────────────────────────
shutil.copy2(SRC, BAK)
print(f"[OK] Backup: {BAK}")

# ── 2. LOAD (data_only=False to preserve formulas) ────────────────────────────
wb = openpyxl.load_workbook(SRC, data_only=False)
print("[OK] Workbook loaded")

TS = "2026-05-18T00:00:00Z"

# ── H84_SNAPSHOT_REGISTRY ─────────────────────────────────────────────────────
# Columns: ARCHIVO_BACKUP | TIMESTAMP | CHECKSUM_SHA256 | TAMAÑO_BYTES | ESTADO
ws = wb["H84_SNAPSHOT_REGISTRY"]
ws.append([
    "SIAP-ICPI_GOLD_MASTER_v5.5_TGI_20260518.xlsx",
    TS,
    "1CB27504FD2FEC2A1C58E97BE95D765E53EC1096F53008DCD0A7309CA5F90564",
    "936916",
    "CERTIFICADO - v5.5 Gold Master TGI Framework QUIRA OS RC-1.1",
])
print(f"[OK] H84: 1 row added (max_row={ws.max_row})")

# ── H85_ALERTS_LOG ────────────────────────────────────────────────────────────
# Columns: TIMESTAMP | ALERTA | SEVERIDAD | ESTADO | DESCRIPCION | (x2)
ws = wb["H85_ALERTS_LOG"]
rc1_log = [
    [None] * 7,
    ["RC-1 COMPLETION LOG - QUIRA OS Sprints 2.8A to RC-1.1 - Dylus Lab 2026-05-18",
     None, None, None, None, None, None],
    [TS, "SPRINT_2_8A_COMPLETADO", "INFO", "RESUELTA",
     "Sprint 2.8A: Memoria Operativa - patrones de resolucion aprendidos", None, None],
    [TS, "SPRINT_2_8B_COMPLETADO", "INFO", "RESUELTA",
     "Sprint 2.8B: Antecedentes Comparables - matching tipo+3 entidad+2 severidad+2", None, None],
    [TS, "SPRINT_2_8C_COMPLETADO", "INFO", "RESUELTA",
     "Sprint 2.8C: Borrador Institucional - sugerencias pre-redactadas por contexto", None, None],
    [TS, "SPRINT_2_9A_COMPLETADO", "INFO", "RESUELTA",
     "Sprint 2.9A: Ruta de Atencion - maquina 7 estados bitacora append-only", None, None],
    [TS, "RC1_A_COMPLETADO", "INFO", "RESUELTA",
     "RC-1.A: 7 modulos QUIRA OS integrados - app.py wiring verificado", None, None],
    [TS, "RC1_B_COMPLETADO", "INFO", "RESUELTA",
     "RC-1.B: Vista Ejecutiva Alcalde - semaforos holding PDF exportacion", None, None],
    [TS, "RC1_C_COMPLETADO", "INFO", "RESUELTA",
     "RC-1.C: seed_demo.py - 10 alertas 12 KPIs 24 eventos 3 patrones Supabase", None, None],
    [TS, "RC1_D_COMPLETADO", "INFO", "RESUELTA",
     "RC-1.D: Schema audit Supabase - 8 tablas 18 indices columnas verificadas", None, None],
    [TS, "RC1_1_GLOSARIO", "INFO", "RESUELTA",
     "RC-1.1: GLOSARIO_INSTITUCIONAL.md - 30 equivalencias tecnico-institucional", None, None],
    [TS, "RC1_1_WALKTHROUGH", "INFO", "RESUELTA",
     "RC-1.1: RC1_1_WALKTHROUGH.md - 17 pasos 7 flujos criterios de aprobacion", None, None],
    [TS, "OBSIDIAN_SYNC_RC1_1", "INFO", "RESUELTA",
     "Obsidian: _Indice Sentinel + Glosario Institucional actualizados RC-1.1", None, None],
    [TS, "GOLD_MASTER_SYNC_RC1_1", "INFO", "RESUELTA",
     "Gold Master v5.5: H84 H85 H02 H77 H36c actualizados - QUIRA OS RC-1.1", None, None],
]
for row in rc1_log:
    ws.append(row)
print(f"[OK] H85: {len(rc1_log)} rows added (max_row={ws.max_row})")

# ── H02_GLOSARIO_QUIRA ────────────────────────────────────────────────────────
# Columns: Termino | Definicion | Referencia | (empty x3)
# Footer is at the last row - insert new terms before it
ws = wb["H02_GLOSARIO_QUIRA"]
footer_row = ws.max_row  # currently row 83

sentinel_terms = [
    # [Termino, Definicion, Referencia]
    ["Antecedentes Comparables",
     "Alertas similares del pasado recuperadas para contextualizar la situacion actual. "
     "Scoring: tipo+3 entidad+2 severidad+2 trimestre+1",
     "QUIRA OS RC-1.1 - SENTINEL-Aprendizaje Sprint 2.8B"],
    ["Borrador Institucional",
     "Texto pre-redactado sugerido al funcionario basado en antecedentes comparables "
     "y patrones de resolucion aprendidos por Sentinel",
     "QUIRA OS RC-1.1 - SENTINEL-Aprendizaje Sprint 2.8C"],
    ["Corte Institucional",
     "Estado registrado del sistema en un periodo cerrado (mes/anio). "
     "Equivalente institucional del termino tecnico snapshot",
     "QUIRA OS RC-1.1 - Glosario Institucional"],
    ["Holding Municipal",
     "Conjunto de 4 entidades del GAD Municipal de Montecristi: "
     "GADMCM + Cuerpo de Bomberos + EMAI-EP + Patronato Municipal",
     "QUIRA OS RC-1.1 - Estructura institucional Montecristi"],
    ["Memoria Operativa",
     "Modulo que aprende de resoluciones anteriores para sugerir soluciones "
     "contextualizadas. Sin ML - solo reglas deterministicas auditables",
     "QUIRA OS RC-1.1 - SENTINEL-Aprendizaje Sprint 2.8A"],
    ["QUIRA OS",
     "Sistema operativo de gobernanza institucional digital para el GAD Municipal "
     "de Montecristi. Desarrollado por Dylus Lab (c) 2026 - TGI Framework",
     "Dylus Lab - TGI Framework RC-1.1"],
    ["Registro de Estado",
     "Documento generado al cerrar un periodo con el estado completo del Holding "
     "Municipal. Exportable como PDF institucional",
     "QUIRA OS RC-1.1 - Vista Ejecutiva RC-1.B"],
    ["Ruta de Atencion",
     "Flujo formal de gestion de alertas con 7 estados institucionales y "
     "bitacora inmutable append-only. 7 estados: Abierta -> Archivada",
     "QUIRA OS RC-1.1 - SENTINEL-Ruta-Atencion Sprint 2.9A"],
    ["Sentinel",
     "Asistente de inteligencia artificial integrado a QUIRA OS. Copiloto "
     "institucional para el GADM Montecristi - interpreta SIAP-ICPI",
     "QUIRA OS RC-1.1 - _Indice Sentinel"],
    ["Ti (Tasa de Inversion)",
     "Indicador de ejecucion presupuestaria mensual por entidad. "
     "Semaforo Holding: Verde>=35% Amarillo 15-34.9% Rojo<15%",
     "QUIRA OS RC-1.1 - Semaforos Holding Municipal"],
    ["Validacion Documental",
     "Proceso de lectura verificacion e ingesta de cedulas presupuestarias eSIGEF. "
     "Equivalente institucional del termino tecnico parser",
     "QUIRA OS RC-1.1 - Sprint 2.5B Ingesta Mensual"],
    ["Vista Ejecutiva",
     "Panel simplificado para Alcalde y Directivos con semaforos del Holding "
     "Municipal alertas criticas y resumen institucional sin tecnicismos",
     "QUIRA OS RC-1.1 - p_ejecutivo.py RC-1.B"],
]

# Insert 13 rows before footer (1 section header + 12 terms)
ws.insert_rows(footer_row, amount=13)

# Section header at footer_row
ws.cell(row=footer_row, column=1,
        value="Modulo SENTINEL - QUIRA OS RC-1.1 (lenguaje institucional congelado)")

# Terms at footer_row+1 ... footer_row+12
for i, (term, defn, ref) in enumerate(sentinel_terms, 1):
    ws.cell(row=footer_row + i, column=1, value=term)
    ws.cell(row=footer_row + i, column=2, value=defn)
    ws.cell(row=footer_row + i, column=3, value=ref)

print(f"[OK] H02: 13 rows inserted before footer (footer now at row {footer_row + 13})")

# ── H77_DATA_DICTIONARY ───────────────────────────────────────────────────────
# Columns: MODULO/HOJA | TIPO | DEFINICION | RESTRICCIONES | VERSION
ws = wb["H77_DATA_DICTIONARY"]
sentinel_tables = [
    ["alert_timeline",
     "TRAZABILIDAD_SENTINEL",
     "Bitacora append-only de eventos por alerta. Registra los 7 estados "
     "institucionales: pendiente en_revision derivada resuelta observada validada archivada. "
     "Actor + timestamp + nota inmutable",
     "INSERT solo - NUNCA UPDATE. Integridad auditada. Sin borrado de historia.",
     "RC-1.1 - 2026"],
    ["resolution_patterns",
     "MEMORIA_OPERATIVA",
     "Patrones de resolucion aprendidos por Sentinel. "
     "Multi-criteria matching: tipo+3 entidad+2 severidad+2 trimestre+1. "
     "Alimenta Borrador Institucional y Antecedentes Comparables",
     "INSERT solo - agregar patrones nuevos al final. Sin borrado de historia.",
     "RC-1.1 - 2026"],
    ["alerts_history",
     "ALERTAS_HOLDING",
     "Historial Ti inversion 4 entidades Holding Municipal de Montecristi. "
     "Genera alertas semaforo QUIRA OS RC-1: Verde>=35% Amarillo 15-34.9% Rojo<15%",
     "Umbrales congelados RC-1.1. Unique constraint: (anio mes entidad tipo). "
     "ON CONFLICT DO UPDATE para re-ingesta idempotente.",
     "RC-1.1 - 2026"],
]
for row in sentinel_tables:
    ws.append(row)
print(f"[OK] H77: {len(sentinel_tables)} rows added (max_row={ws.max_row})")

# ── H36c_OBSIDIAN_MAP ─────────────────────────────────────────────────────────
# Columns: Variable_QUIRA | Descripcion | Valor_Excel | Nota_Obsidian |
#          Carpeta_Vault | Dimension_TGI | Prioridad | Tipo_Alerta | Estado
ws = wb["H36c_OBSIDIAN_MAP"]
# Insert before row 25 (the blank separator before "MAPA NOTAS" section)
insert_at = 25
sentinel_vars = [
    ["SENTINEL_GLOBAL_STATUS",
     "Estado global del Holding Municipal (ROJO/AMARILLO/VERDE) segun Ti inversion",
     "ROJO (Q1-2026 Ene-Feb)",
     "[[_Indice Sentinel]]",
     "06_Sentinel", "D3", "CRITICA", "STATUS", "ACTIVO"],
    ["SENTINEL_ALERTAS_CRITICAS",
     "Total alertas criticas activas en el Holding (Ti<15%)",
     "3 activas Q1-2026 (GAD+Bomberos+Patronato)",
     "[[ALERTA-Holding_Ti_Critico_2026]]",
     "06_Sentinel", "D3", "ALTA", "ALERTA", "ACTIVO"],
    ["SENTINEL_TI_GAD",
     "Ti inversion GAD Municipal de Montecristi - serie mensual Q1-2026",
     "8.5% Ene / 12.8% Feb / 18.2% Mar",
     "[[CEDULAS_HOLDING_ENE_MAR_2026]]",
     "06_Sentinel", "D3", "CRITICA", "STATUS", "ACTIVO"],
    ["SENTINEL_TI_BOMBEROS",
     "Ti inversion Cuerpo de Bomberos de Montecristi - serie mensual Q1-2026",
     "22.3% Ene / 28.7% Feb / 35.8% Mar",
     "[[CEDULAS_HOLDING_ENE_MAR_2026]]",
     "06_Sentinel", "D3", "ALTA", "STATUS", "ACTIVO"],
    ["SENTINEL_TI_EMAI_EP",
     "Ti inversion EMAI-EP Empresa Municipal de Aseo Integral - serie Q1-2026",
     "41.2% Ene / 38.5% Feb / 44.1% Mar",
     "[[CEDULAS_HOLDING_ENE_MAR_2026]]",
     "06_Sentinel", "D3", "MEDIA", "STATUS", "ACTIVO"],
    ["SENTINEL_TI_PATRONATO",
     "Ti inversion Patronato Municipal de Amparo Social - serie mensual Q1-2026",
     "12.1% Ene / 19.4% Feb / 28.3% Mar",
     "[[CEDULAS_HOLDING_ENE_MAR_2026]]",
     "06_Sentinel", "D3", "ALTA", "STATUS", "ACTIVO"],
    ["SENTINEL_VERSION",
     "Version RC activa de QUIRA OS Sentinel en produccion",
     "RC-1.1",
     "[[_Indice Sentinel]]",
     "06_Sentinel", "D5", "MEDIA", "INFO", "ACTIVO"],
    ["SENTINEL_PATRON_TOP",
     "Causa raiz mas frecuente de alertas segun Memoria Operativa (Sprint 2.8A)",
     "reforma_presupuestaria (freq=3)",
     "[[SENTINEL-Aprendizaje]]",
     "06_Sentinel", "D3", "MEDIA", "INFO", "ACTIVO"],
    ["SENTINEL_ALERTAS_ARCHIVADAS",
     "Total alertas archivadas (caso cerrado) en Supabase - Ruta de Atencion",
     "0 archivadas Q1-2026 (sistema nuevo RC-1.1)",
     "[[SENTINEL-Ruta-Atencion]]",
     "06_Sentinel", "D3", "MEDIA", "INFO", "ACTIVO"],
]
ws.insert_rows(insert_at, amount=len(sentinel_vars))
for i, row in enumerate(sentinel_vars):
    for j, val in enumerate(row, 1):
        ws.cell(row=insert_at + i, column=j, value=val)
print(f"[OK] H36c: {len(sentinel_vars)} rows inserted at row {insert_at} "
      f"(MAPA NOTAS section now at row {insert_at + len(sentinel_vars) + 1})")

# ── 3. SAVE ───────────────────────────────────────────────────────────────────
wb.save(SRC)
print(f"[OK] Saved: {SRC}")
print()
print("=" * 60)
print("Gold Master v5.5 updated - QUIRA OS RC-1.1 COMPLETE")
print("  H84: +1 snapshot (v5.5 Gold Master)")
print("  H85: +14 events (RC-1 completion log)")
print("  H02: +13 rows (12 Sentinel terms + section header)")
print("  H77: +3 rows (alert_timeline, resolution_patterns, alerts_history)")
print("  H36c: +9 rows (Sentinel variables QUIRA OS map)")
print("=" * 60)
