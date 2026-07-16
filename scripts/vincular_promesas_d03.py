"""
QUIRA OS — d03 · PROPUESTA de vinculación promesa CNE → meta PDOT
═══════════════════════════════════════════════════════════════════════════════
ADR-035 §5 (regla constitucional): la IA PROPONE, el humano VALIDA. Jamás al revés.
Este script NO decide el canon: escribe una PROPUESTA sobre la COPIA de trabajo, con
cada vinculación marcada `Propuesta IA · validar`, para que Javo la revise y subsane.

Método: razonamiento semántico promesa↔meta (NO similitud textual — ese método falló:
emparejaba "destino turístico" con "Concejo de Salud").

Escala de score (la del canon · H03 f16):
  1.00 Directa            — la promesa ES la meta
  0.75 Directa con matiz  — cae dentro de la meta, con alcance distinto
  0.50 Parcial            — la toca solo en parte / es un medio para ella
  0.00 Sin vinculación    — ninguna meta del PDOT la recoge

Uso:  python scripts/vincular_promesas_d03.py
Dylus Lab © 2026
"""
from __future__ import annotations

import re

import openpyxl

COPIA = r"C:\Users\DELL\Desktop\Javo\Dylus Lab\ProyecT\_TRABAJO_d03_promesas.xlsx"

# ── PROPUESTA · id_nuevo: (meta_pdot, score, tipo) ────────────────────────────
# Justificación = el "sistema" de la meta en H04 (Agua, Vialidad, Salud, Turismo…).
V = {
    # ── ECONÓMICO ────────────────────────────────────────────────────────────
    "EC-001": ("EP-L-X-01", 0.75, "Directa con matiz"),   # entorno empresarial → fortalecimiento productivo
    "EC-002": ("EP-L-X-01", 0.50, "Parcial"),             # exoneración tributaria → medio para lo productivo
    "EC-003": ("PI-I-G-01", 1.00, "Directa"),             # mercado → equipamientos (mercado) ✔ ya validada
    "EC-004": ("EP-L-X-01", 1.00, "Directa"),             # sector artesanal → productivo (artesanos)
    "EC-005": ("PI-I-G-01", 0.50, "Parcial"),             # vendedores ambulantes → ordenamiento de equipamientos
    "EC-006": ("PI-TUR-02", 1.00, "Directa"),             # festividades → eventos turísticos
    "EC-007": ("FA-C-X-01", 0.75, "Directa con matiz"),   # Paseo Lúdico → áreas verdes y parques
    "EC-008": ("EP-L-X-01", 1.00, "Directa"),             # promoción artesanías → productivo
    "EC-009": ("PI-TUR-01", 0.75, "Directa con matiz"),   # rutas de ciclismo → turismo cantonal
    "EC-010": ("PI-TUR-01", 0.75, "Directa con matiz"),   # gastronomía → turismo cantonal
    "EC-011": ("PI-TUR-01", 1.00, "Directa"),             # posicionamiento destino turístico → turismo
    "EC-012": ("PI-TUR-01", 1.00, "Directa"),             # ruta turística → turismo
    "EC-013": ("PI-TUR-01", 1.00, "Directa"),             # Playa San José / Isla de la Plata → turismo ✔
    "EC-014": ("PI-TUR-01", 1.00, "Directa"),             # Manglar → turismo
    # ── INSTITUCIONAL ────────────────────────────────────────────────────────
    "IN-001": ("PI-I-G-02", 0.75, "Directa con matiz"),   # plan de desarrollo territorial → PDOT/PUGS
    "IN-002": ("PI-I-G-02", 1.00, "Directa"),             # PDOT 2023 con ODS → PDOT/PUGS ✔
    "IN-003": ("AH-C-X-02", 1.00, "Directa"),             # catastro → información territorial ✔
    "IN-004": ("AH-C-X-01", 0.75, "Directa con matiz"),   # base de vulnerables → derechos sociales
    "IN-005": ("AH-C-X-02", 0.75, "Directa con matiz"),   # registro comunal → información territorial
    "IN-006": ("AH-C-X-02", 0.75, "Directa con matiz"),   # terrenos comunales → información territorial
    "IN-007": ("PI-TUR-01", 0.50, "Parcial"),             # marca ciudad → certificación Montecristi (turismo)
    "IN-008": ("AH-I-X-04", 1.00, "Directa"),             # equipos informáticos → modernización
    "IN-009": ("PI-I-G-02", 0.75, "Directa con matiz"),   # plataforma digital trámites → PDOT/PUGS (37 trámites)
    "IN-010": ("SC-I-N-03", 0.75, "Directa con matiz"),   # transparencia y rendición → participación
    "IN-011": ("AH-C-X-02", 0.75, "Directa con matiz"),   # registro de la propiedad → información territorial
    "IN-012": ("AH-I-X-01", 0.50, "Parcial"),             # alianzas con otros niveles → sostenibilidad financiera
    "IN-013": ("FA-I-X-01", 0.75, "Directa con matiz"),   # alarmas comunitarias → gestión del riesgo
    "IN-014": ("FA-I-X-01", 0.50, "Parcial"),             # UPC → riesgo (seguridad no es competencia exclusiva)
    "IN-015": ("FA-I-X-01", 0.75, "Directa con matiz"),   # cuarteles de bomberos → gestión del riesgo
    # ── SOCIAL ───────────────────────────────────────────────────────────────
    "SC-001": ("AH-I-X-03", 0.75, "Directa con matiz"),   # Concejo de Salud → salud integral
    "SC-002": ("AH-I-X-03", 1.00, "Directa"),             # Centro de Salud Tipo C → salud ✔
    "SC-003": ("AH-C-X-01", 1.00, "Directa"),             # Centro Geriátrico → grupos prioritarios
    "SC-004": ("SC-L-N-02", 0.50, "Parcial"),             # reorganizar instituciones → talento humano
    "SC-005": ("AH-I-X-03", 1.00, "Directa"),             # salud puerta a puerta → salud
    "SC-006": ("AH-I-X-03", 0.75, "Directa con matiz"),   # charlas educativas de salud → salud
    "SC-007": ("AH-I-X-03", 0.75, "Directa con matiz"),   # clubes de salud → salud
    "SC-008": ("PI-I-G-02", 0.50, "Parcial"),             # evaluación del plan → cumplimiento PDOT/PUGS
    "SC-009": ("AH-I-X-03", 0.50, "Parcial"),             # albergue animales → salud (zoonosis)
    "SC-010": ("EP-L-X-01", 0.75, "Directa con matiz"),   # plataforma de capacitación → productivo
    "SC-011": ("EP-L-X-01", 1.00, "Directa"),             # centro de emprendedores → productivo (mipymes)
    "SC-012": ("EP-L-X-01", 0.75, "Directa con matiz"),   # capacitación comerciantes → productivo
    "SC-013": ("SC-I-N-03", 1.00, "Directa"),             # liderazgo juvenil en política → participación
    "SC-014": ("AH-C-X-01", 0.75, "Directa con matiz"),   # becas universidad → derechos sociales
    "SC-015": ("FA-L-N-01", 1.00, "Directa"),             # lectura, pintura, cultura → cultura
    "SC-016": ("EP-L-X-01", 1.00, "Directa"),             # Centro Formación Artesanal → productivo (artesanos)
    "SC-017": ("AH-C-X-02", 0.50, "Parcial"),             # Infocentro → información/conectividad
    "SC-018": ("SC-I-N-03", 1.00, "Directa"),             # participación barrial → participación
    "SC-019": ("FA-I-X-02", 0.75, "Directa con matiz"),   # espacios deportivos → equipamiento urbano
    "SC-020": ("FA-L-N-01", 0.75, "Directa con matiz"),   # espacios culturales → cultura
    "SC-021": ("FA-C-X-01", 0.50, "Parcial"),             # wifi en parques → áreas verdes/parques
    # ── AMBIENTAL ────────────────────────────────────────────────────────────
    "AM-001": ("FA-C-X-01", 0.50, "Parcial"),             # flora y fauna / minería → áreas verdes
    "AM-002": ("FA-CC-01", 0.75, "Directa con matiz"),    # remediación ambiental → cambio climático
    "AM-003": ("FA-CC-01", 0.50, "Parcial"),              # calidad del aire → cambio climático
    "AM-004": ("AH-I-N-01", 1.00, "Directa"),             # residuos desde la fuente → desechos sólidos
    "AM-005": ("AH-I-N-01", 1.00, "Directa"),             # ampliación del basurero → desechos
    "AM-006": ("AH-I-N-01", 1.00, "Directa"),             # cobertura de recolección → desechos
    "AM-007": ("SC-L-G-01", 1.00, "Directa"),             # planta de oxidación → alcantarillado (PTAR)
    "AM-008": ("", 0.00, "Sin vinculación"),              # derogación de decreto nacional → NO es competencia del GAD
    "AM-009": ("FA-C-X-01", 1.00, "Directa"),             # arborización → áreas verdes (IVU)
    # ── TERRITORIAL ──────────────────────────────────────────────────────────
    "TE-001": ("AH-I-X-02", 0.75, "Directa con matiz"),   # equipo de pavimento → vialidad
    "TE-002": ("AH-I-X-02", 1.00, "Directa"),             # construcción/mantenimiento de vías → vialidad
    "TE-003": ("SC-L-G-01", 1.00, "Directa"),             # alcantarillado por sectores → alcantarillado
    "TE-004": ("AH-I-X-02", 0.75, "Directa con matiz"),   # aceras y bordillos → vialidad
    "TE-005": ("PI-I-G-01", 0.75, "Directa con matiz"),   # paradas de buses → equipamientos
    "TE-006": ("PI-L-G-01", 1.00, "Directa"),             # señalización de calles → señalización vial
    "TE-007": ("AH-I-X-02", 0.75, "Directa con matiz"),   # inventario de vías → vialidad
    "TE-008": ("PI-L-G-01", 0.75, "Directa con matiz"),   # reordenamiento vial → señalización/movilidad
    "TE-009": ("PI-I-G-01", 1.00, "Directa"),             # nuevo terminal → equipamientos (terminal terrestre)
    "TE-010": ("PI-I-G-01", 0.75, "Directa con matiz"),   # camal → equipamientos
    "TE-011": ("PI-I-G-01", 0.50, "Parcial"),             # cementerio parroquial → equipamientos
    "TE-012": ("PI-I-G-01", 0.50, "Parcial"),             # nuevo cementerio → equipamientos
    "TE-013": ("PI-I-G-01", 0.50, "Parcial"),             # mantenimiento cementerios → equipamientos
    "TE-014": ("SC-I-N-01", 1.00, "Directa"),             # cobertura de agua potable → agua potable
    "TE-015": ("SC-I-N-01", 0.75, "Directa con matiz"),   # reservorios planta potabilizadora → agua
    "TE-016": ("SC-I-N-01", 0.75, "Directa con matiz"),   # pozos acuíferos → agua
    "TE-017": ("AH-AP-04", 1.00, "Directa"),              # macro medidor / control → continuidad del servicio
    "TE-018": ("SC-I-N-01", 0.75, "Directa con matiz"),   # reservorios rurales → agua
}

_PID = re.compile(r"^(EC|IN|SC|AM|TE)-\d{3}$")


def main() -> None:
    wb = openpyxl.load_workbook(COPIA)
    ws = wb["H03_S1_ELECTORAL_CNE"]
    escritas, faltan, suma = 0, [], 0.0
    for r in range(17, 95):
        pid = ws.cell(row=r, column=1).value
        if not (pid and isinstance(pid, str) and _PID.match(pid.strip())):
            continue
        pid = pid.strip()
        if pid not in V:
            faltan.append(pid)
            continue
        meta, score, tipo = V[pid]
        ws.cell(row=r, column=4).value = meta or "—"
        ws.cell(row=r, column=5).value = score
        ws.cell(row=r, column=6).value = tipo
        ws.cell(row=r, column=7).value = "Propuesta IA · validar"
        suma += score
        escritas += 1
    if faltan:
        print("SIN PROPUESTA:", faltan)
    wb.save(COPIA)
    con_meta = sum(1 for m, _, _ in V.values() if m)
    print(f"OK - {escritas} vinculaciones propuestas escritas en la COPIA")
    print(f"   con meta        : {con_meta}/{len(V)}  ·  sin vinculación: {len(V)-con_meta}")
    print(f"   suma de score   : {suma:.2f}  ->  índice propuesto = {suma/len(V)*100:.2f}%")
    print(f"   TODAS marcadas  : 'Propuesta IA · validar' (ninguna se declara verificada)")


if __name__ == "__main__":
    main()
