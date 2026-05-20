"""
sentinel/cedula_parser.py
Parser de Cédula Presupuestaria Mensual — Sprint 2.5A

Procesa el Excel oficial que el GAD sube a eSIGEF/LOTAIP.
El mismo archivo que se reporta a Ministerio de Economía y Finanzas.

Guardrails de validación (G1-G5):
    G1  ¿Tiene hojas obligatorias?
    G2  ¿Tiene mínimo de filas (>5)?
    G3  ¿Tiene columnas obligatorias?
    G4  ¿No es duplicado por SHA256?
    G5  ¿El mes ya existe? → versionar si sí

Grupos de gasto relevantes:
    Grupos 7 y 8 = Inversión (los que miden D3 en TGI)
    Grupo 5      = Gasto corriente

Dylus Lab © 2026
"""
from __future__ import annotations

import hashlib
import io
import re
from dataclasses import dataclass, field
from typing import Optional

import openpyxl
import pandas as pd

# ── CONSTANTES ────────────────────────────────────────────────────────────────
# Columnas obligatorias mínimas que DEBE tener la cédula presupuestaria
REQUIRED_COLS_KEYWORDS = [
    ["codificado", "cod"],
    ["devengado", "deveng"],
    ["pagado", "pag"],
]

# Hojas aceptables (la cédula puede llamarse de varias formas)
CEDULA_SHEET_KEYWORDS = [
    "cedula", "cédula", "ejecucion", "ejecución",
    "presupuesto", "egresos", "sheet", "hoja"
]

# Grupos de gasto de inversión (grupos 7 y 8 del clasificador presupuestario)
GRUPOS_INVERSION = {"7", "8", "73", "75", "84"}

# Umbral mínimo de filas de datos (sin contar encabezados)
MIN_DATA_ROWS = 5


# ── RESULTADO DE VALIDACIÓN ───────────────────────────────────────────────────
@dataclass
class ValidationResult:
    """Resultado de cada guardrail de validación."""
    passed:    bool
    guardrail: str   # G1, G2, G3, G4, G5
    mensaje:   str
    detalle:   str = ""


@dataclass
class CedulaParseResult:
    """Resultado completo del parsing de una cédula."""
    # Metadata documental
    sha256:        str = ""
    size_bytes:    int = 0
    sheet_name:    str = ""
    total_rows:    int = 0
    # Datos presupuestarios
    df:            Optional[pd.DataFrame] = None
    # Indicadores clave
    codificado_total:  float = 0.0
    devengado_total:   float = 0.0
    pagado_total:      float = 0.0
    codificado_inv:    float = 0.0   # solo grupos 7+8
    devengado_inv:     float = 0.0   # solo grupos 7+8
    ti_mensual_pct:    float = 0.0   # Ti = devengado_inv / codificado_inv × 100
    # Validaciones
    validations:   list[ValidationResult] = field(default_factory=list)
    ok:            bool = False
    error:         str  = ""
    # Columnas mapeadas
    col_map:       dict = field(default_factory=dict)


# ── HELPERS ───────────────────────────────────────────────────────────────────
def _sha256(content: bytes) -> str:
    return hashlib.sha256(content).hexdigest()


def _find_sheet(wb: openpyxl.Workbook) -> str:
    """Encuentra la hoja principal de la cédula."""
    sheets = wb.sheetnames
    if len(sheets) == 1:
        return sheets[0]
    for s in sheets:
        sl = s.lower().strip()
        if any(kw in sl for kw in CEDULA_SHEET_KEYWORDS):
            return s
    # Si ninguna coincide, devolver la primera hoja no oculta
    return sheets[0]


def _find_header_row(df_raw: pd.DataFrame) -> int:
    """
    Encuentra la fila que contiene los encabezados de la cédula.
    Busca la fila con más columnas no-vacías o con keywords clave.
    """
    keywords = {"devengado", "codificado", "pagado", "comprometido",
                "vigente", "partida", "descripcion", "descripción",
                "saldo", "porcentaje", "programa"}
    best_row = 0
    best_score = 0
    for i in range(min(20, len(df_raw))):
        row_vals = [str(v).lower().strip() for v in df_raw.iloc[i] if pd.notna(v) and str(v).strip()]
        score = sum(1 for v in row_vals if any(kw in v for kw in keywords))
        if score > best_score:
            best_score = score
            best_row = i
    return best_row


def _looks_like_partida(series: pd.Series) -> bool:
    """
    Valida que una columna contiene códigos de partida presupuestaria.
    Una partida válida empieza con dígito (ej: '510105', '73.01.01', '7').
    Al menos el 60% de los valores no-nulos deben cumplir esto.
    """
    sample = series.dropna().astype(str).str.strip().head(20)
    if len(sample) == 0:
        return False
    numeric_first = sample.str[0].str.isdigit().sum()
    return numeric_first >= max(1, len(sample) * 0.6)


def _map_columns(df: pd.DataFrame) -> dict[str, str]:
    """
    Mapea columnas del DataFrame a nombres canónicos.
    Retorna dict: nombre_canonico → nombre_real_en_df

    Maneja formatos distintos de eSIGEF:
      - GAD / Bomberos: 'Cuenta'=partida numérica, 'Categoría'=descripción grupo
      - EMAI-EP: 'Cuenta'=nombre categoría, 'Categoría'=partida numérica (invertido)
    Detecta el caso invertido validando los valores reales de cada columna.
    """
    canon_map = {
        "codificado":    ["codificado", "cod_vigente", "codif"],
        "devengado":     ["devengado", "deveng"],
        "pagado":        ["pagado", "pag"],
        "comprometido":  ["comprometido", "comp"],
        "vigente":       ["vigente", "vig", "codificado_vigente"],
        "saldo":         ["saldo", "sal"],
        "pct_ejecucion": ["porcentaje", "pct", "ejecucion", "ejecución", "%"],
        "partida":       ["partida", "clasificador", "cod_partida", "cuenta"],
        "descripcion":   ["descripcion", "descripción", "nombre", "detalle"],
        "categoria":     ["categoría", "categoria", "cat"],
        "programa":      ["programa", "prog"],
        "proyecto":      ["proyecto", "proy"],
        "unidad":        ["unidad", "direcc", "responsable"],
    }
    col_lower = {c.lower().strip(): c for c in df.columns}
    result = {}
    for canon, keywords in canon_map.items():
        for kw in keywords:
            matches = [c_real for c_low, c_real in col_lower.items() if kw in c_low]
            if matches:
                result[canon] = matches[0]
                break

    # ── Validación de partida: ¿los valores son realmente códigos numéricos? ──
    # Si no, buscar otra columna candidata (caso EMAI-EP con columnas invertidas)
    if "partida" in result:
        if not _looks_like_partida(df[result["partida"]]):
            # La columna mapeada como 'partida' no tiene códigos numéricos.
            # Buscar la primera columna no usada que sí los tenga.
            used = set(result.values())
            for col in df.columns:
                if col not in used and _looks_like_partida(df[col]):
                    # Antes de reasignar, mover la columna actual a 'descripcion'
                    # si 'descripcion' aún no está mapeada (tiene texto descriptivo)
                    if "descripcion" not in result:
                        result["descripcion"] = result["partida"]
                    result["partida"] = col
                    break

    return result


def _to_float(val) -> float:
    """
    Convierte valor a float de forma segura.
    Maneja formatos:
      - Europeo (eSIGEF Ecuador): '1.234.567,89' → punto=miles, coma=decimal
      - Estándar:                 '1234567.89'
      - Guión como cero:          '-' → 0.0
    """
    if pd.isna(val):
        return 0.0
    try:
        s = str(val).replace("$", "").replace("%", "").strip()
        if not s or s in ("-", "—", "–", "N/A", "n/a"):
            return 0.0
        # Detectar formato europeo: la última coma está después del último punto
        last_dot   = s.rfind(".")
        last_comma = s.rfind(",")
        if last_comma > last_dot:
            # Formato europeo: 1.234.567,89 → quitar puntos, cambiar coma por punto
            s = s.replace(".", "").replace(",", ".")
        else:
            # Formato estándar o solo punto decimal: quitar comas eventuales
            s = s.replace(",", "")
        return float(s) if s else 0.0
    except (ValueError, TypeError):
        return 0.0


def _detect_grupo(partida: str) -> str:
    """Extrae el grupo de gasto de una partida presupuestaria (primer dígito)."""
    if not partida:
        return ""
    # La partida puede ser "7.3.01.01" o "730101" — extraer primer dígito
    clean = re.sub(r"[.\-\s]", "", str(partida))
    return clean[0] if clean else ""


# ── GUARDRAILS ────────────────────────────────────────────────────────────────
def _g1_has_data_sheet(wb: openpyxl.Workbook) -> ValidationResult:
    """G1: ¿Tiene al menos una hoja con datos?"""
    if not wb.sheetnames:
        return ValidationResult(False, "G1", "Sin hojas en el archivo",
                                "El archivo parece estar vacío o corrupto")
    sheet = _find_sheet(wb)
    ws = wb[sheet]
    rows = list(ws.iter_rows(values_only=True))
    has_data = any(any(c for c in r if c) for r in rows[:30])
    if not has_data:
        return ValidationResult(False, "G1", f"Hoja '{sheet}' sin datos",
                                "No se encontraron valores en las primeras 30 filas")
    return ValidationResult(True, "G1", f"Hoja detectada: '{sheet}'", f"Total hojas: {len(wb.sheetnames)}")


def _g2_min_rows(df: pd.DataFrame) -> ValidationResult:
    """G2: ¿Tiene mínimo de filas?"""
    n = len(df)
    if n < MIN_DATA_ROWS:
        return ValidationResult(False, "G2", f"Solo {n} filas de datos (mínimo {MIN_DATA_ROWS})",
                                "La cédula parece incompleta")
    return ValidationResult(True, "G2", f"{n} filas de datos", "")


def _g3_required_cols(col_map: dict) -> ValidationResult:
    """G3: ¿Tiene columnas obligatorias mínimas?"""
    required = {"codificado", "devengado"}
    missing  = required - set(col_map.keys())
    if missing:
        return ValidationResult(False, "G3", f"Faltan columnas: {missing}",
                                "El archivo no tiene la estructura de cédula presupuestaria estándar")
    return ValidationResult(True, "G3", f"Columnas encontradas: {set(col_map.keys())}", "")


def _g4_no_duplicate(sha256: str, existing_hashes: list[str]) -> ValidationResult:
    """G4: ¿No es un archivo duplicado?"""
    if sha256 in existing_hashes:
        return ValidationResult(False, "G4", "Archivo idéntico ya existe en la base de datos",
                                f"SHA256: {sha256[:12]}… ya fue procesado")
    return ValidationResult(True, "G4", "SHA256 único — archivo nuevo", "")


def _g5_version_check(year: int, month: int, existing_periods: list[tuple]) -> ValidationResult:
    """G5: ¿El período ya existe? Si sí, informa versionado."""
    exists = any(y == year and m == month for y, m in existing_periods)
    if exists:
        return ValidationResult(True, "G5", f"{year}-{month:02d} ya existe → se creará v2+",
                                "El período tiene datos previos. Se registrará como nueva versión.")
    return ValidationResult(True, "G5", f"Período {year}-{month:02d} nuevo", "")


# ── PARSER PRINCIPAL ──────────────────────────────────────────────────────────
def parse_cedula(
    content:          bytes,
    year:             int,
    month:            int,
    existing_hashes:  list[str]   = None,
    existing_periods: list[tuple] = None,
) -> CedulaParseResult:
    """
    Parsea una cédula presupuestaria mensual.

    Args:
        content          — bytes del archivo Excel
        year             — año del período
        month            — mes del período (1-12)
        existing_hashes  — lista de SHA256 ya en la DB (para G4)
        existing_periods — lista de (year, month) ya en la DB (para G5)

    Returns:
        CedulaParseResult con datos, indicadores y validaciones.
    """
    result = CedulaParseResult()
    result.sha256      = _sha256(content)
    result.size_bytes  = len(content)
    existing_hashes    = existing_hashes or []
    existing_periods   = existing_periods or []

    try:
        # ── Abrir con openpyxl para G1 ────────────────────────────────────────
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=False, data_only=True)

        # G1: hojas
        g1 = _g1_has_data_sheet(wb)
        result.validations.append(g1)
        if not g1.passed:
            result.error = g1.mensaje
            return result

        result.sheet_name = _find_sheet(wb)

        # ── Leer con pandas para análisis ─────────────────────────────────────
        df_raw = pd.read_excel(io.BytesIO(content), sheet_name=result.sheet_name, header=None)

        # Encontrar fila de encabezados
        header_row = _find_header_row(df_raw)
        df = pd.read_excel(io.BytesIO(content), sheet_name=result.sheet_name, header=header_row)

        # Limpiar columnas
        df.columns = [str(c).strip() if pd.notna(c) else f"col_{i}"
                      for i, c in enumerate(df.columns)]
        df = df.dropna(how="all")

        # G2: filas mínimas
        g2 = _g2_min_rows(df)
        result.validations.append(g2)
        if not g2.passed:
            result.error = g2.mensaje
            return result

        result.total_rows = len(df)

        # Mapear columnas
        col_map = _map_columns(df)
        result.col_map = col_map

        # G3: columnas obligatorias
        g3 = _g3_required_cols(col_map)
        result.validations.append(g3)
        if not g3.passed:
            result.error = g3.mensaje
            return result

        # G4: duplicados
        g4 = _g4_no_duplicate(result.sha256, existing_hashes)
        result.validations.append(g4)
        # G4 falla → no bloquear, solo informar (el usuario decide)

        # G5: versionado
        g5 = _g5_version_check(year, month, existing_periods)
        result.validations.append(g5)

        # ── Extraer valores numéricos ─────────────────────────────────────────
        cod_col  = col_map.get("codificado")
        dev_col  = col_map.get("devengado")
        pag_col  = col_map.get("pagado")
        par_col  = col_map.get("partida")
        vig_col  = col_map.get("vigente") or cod_col   # fallback

        if cod_col:
            df["_codificado"] = df[cod_col].apply(_to_float)
        else:
            df["_codificado"] = 0.0

        if dev_col:
            df["_devengado"] = df[dev_col].apply(_to_float)
        else:
            df["_devengado"] = 0.0

        if pag_col:
            df["_pagado"] = df[pag_col].apply(_to_float)
        else:
            df["_pagado"] = 0.0

        # Detectar grupos de inversión
        if par_col:
            df["_grupo"] = df[par_col].apply(lambda x: _detect_grupo(str(x)))
            df["_es_inversion"] = df["_grupo"].isin(GRUPOS_INVERSION)
        else:
            df["_grupo"] = ""
            df["_es_inversion"] = False

        # ── Totales ───────────────────────────────────────────────────────────
        result.codificado_total = float(df["_codificado"].sum())
        result.devengado_total  = float(df["_devengado"].sum())
        result.pagado_total     = float(df["_pagado"].sum())

        inv_mask = df["_es_inversion"]
        result.codificado_inv   = float(df.loc[inv_mask, "_codificado"].sum())
        result.devengado_inv    = float(df.loc[inv_mask, "_devengado"].sum())

        if result.codificado_inv > 0:
            result.ti_mensual_pct = round(
                (result.devengado_inv / result.codificado_inv) * 100, 2
            )

        result.df = df
        result.ok = all(v.passed for v in result.validations if v.guardrail != "G4")

    except Exception as e:
        result.error = f"Error al parsear: {e}"
        result.ok    = False

    return result


def format_result(r: CedulaParseResult) -> str:
    """Resumen legible del resultado del parsing."""
    lines = [
        f"SHA256   : {r.sha256[:16]}…",
        f"Tamaño   : {r.size_bytes/1024:.1f} KB",
        f"Hoja     : {r.sheet_name}",
        f"Filas    : {r.total_rows}",
        f"Ti mensual: {r.ti_mensual_pct:.2f}%",
        f"Dev inv  : ${r.devengado_inv:,.0f} / ${r.codificado_inv:,.0f}",
        "",
        "Validaciones:",
    ]
    for v in r.validations:
        icon = "✅" if v.passed else "❌"
        lines.append(f"  {icon} {v.guardrail}: {v.mensaje}")
    lines.append(f"\nResultado: {'OK' if r.ok else 'ERROR: ' + r.error}")
    return "\n".join(lines)
