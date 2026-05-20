"""
sentinel/snapshot_engine.py
Motor de Snapshots Institucionales Mensuales — Sprint 2.7

Genera y persiste un snapshot del estado completo del sistema al cierre de cada período:
  • Congruencia institucional (3 capas)
  • Integridad semántica (6 KPIs)
  • Alertas activas y resueltas
  • Métricas de respuesta institucional

Uso principal:
  • Rendición de cuentas
  • LOTAIP
  • Transición administrativa
  • Auditoría externa
  • Memoria histórica operativa

Dylus Lab © 2026
"""
from __future__ import annotations

import io
from datetime import datetime, date
from typing import Optional

from sentinel.db_config import get_connection, init_db

MESES_ES = {
    1: "Enero", 2: "Febrero", 3: "Marzo",    4: "Abril",
    5: "Mayo",  6: "Junio",   7: "Julio",    8: "Agosto",
    9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre",
}


# ── GENERACIÓN ────────────────────────────────────────────────────────────────
def generate_snapshot(
    period_year: int,
    period_month: int,
    triggered_by: str = "manual",
    notas: str = "",
) -> dict:
    """
    Consolida el estado institucional del período y lo persiste en monthly_snapshots.
    Si ya existe un snapshot para ese período, lo sobreescribe (INSERT OR REPLACE).

    Returns: dict con todos los campos del snapshot + "saved": bool
    """
    init_db()  # garantiza que monthly_snapshots existe

    # ── Congruencia ───────────────────────────────────────────────────────────
    try:
        from sentinel.parity_engine import get_congruencia_status
        cong = get_congruencia_status()
        congruencia_status = cong["global_status"]
        fuente_status      = cong["fuente"]["status"]
        memoria_status     = cong["memoria"]["status"]
        motor_status       = cong["motor"]["status"]
    except Exception:
        congruencia_status = fuente_status = memoria_status = motor_status = "error"

    # ── Integridad ────────────────────────────────────────────────────────────
    try:
        from sentinel.integrity_engine import run_integrity_check
        integ = run_integrity_check()
        integridad_status  = integ["status"]
        integridad_n_ok    = integ["n_ok"]
        integridad_n_total = integ["n_total"]
    except Exception:
        integridad_status = "error"; integridad_n_ok = 0; integridad_n_total = 0

    # ── Alertas ───────────────────────────────────────────────────────────────
    try:
        from sentinel.alert_engine import get_alert_kpis
        kpis = get_alert_kpis()
        alertas_criticas      = kpis["criticas_activas"]
        alertas_advertencias  = kpis["advertencias_activas"]
        alertas_resueltas_mes = kpis["resueltas_este_mes"]
        alertas_dias_max      = kpis["dias_abierta_max"]
    except Exception:
        alertas_criticas = alertas_advertencias = alertas_resueltas_mes = alertas_dias_max = 0

    now  = datetime.now().isoformat()
    snap = {
        "period_year":           period_year,
        "period_month":          period_month,
        "period_label":          f"{MESES_ES.get(period_month, str(period_month))} {period_year}",
        "generated_at":          now,
        "triggered_by":          triggered_by,
        "congruencia_status":    congruencia_status,
        "integridad_status":     integridad_status,
        "integridad_n_ok":       integridad_n_ok,
        "integridad_n_total":    integridad_n_total,
        "alertas_criticas":      alertas_criticas,
        "alertas_advertencias":  alertas_advertencias,
        "alertas_resueltas_mes": alertas_resueltas_mes,
        "alertas_dias_max":      alertas_dias_max,
        "fuente_status":         fuente_status,
        "memoria_status":        memoria_status,
        "motor_status":          motor_status,
        "notas":                 notas,
        "saved":                 False,
    }

    # ── Persistir ─────────────────────────────────────────────────────────────
    try:
        conn = get_connection()
        c    = conn.cursor()
        if conn.mode == "supabase":
            sql = """
                INSERT INTO monthly_snapshots
                    (period_year, period_month, generated_at, triggered_by,
                     congruencia_status, integridad_status, integridad_n_ok, integridad_n_total,
                     alertas_criticas, alertas_advertencias, alertas_resueltas_mes, alertas_dias_max,
                     fuente_status, memoria_status, motor_status, notas)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                ON CONFLICT (period_year, period_month) DO UPDATE SET
                    generated_at=EXCLUDED.generated_at, triggered_by=EXCLUDED.triggered_by,
                    congruencia_status=EXCLUDED.congruencia_status,
                    integridad_status=EXCLUDED.integridad_status,
                    integridad_n_ok=EXCLUDED.integridad_n_ok,
                    integridad_n_total=EXCLUDED.integridad_n_total,
                    alertas_criticas=EXCLUDED.alertas_criticas,
                    alertas_advertencias=EXCLUDED.alertas_advertencias,
                    alertas_resueltas_mes=EXCLUDED.alertas_resueltas_mes,
                    alertas_dias_max=EXCLUDED.alertas_dias_max,
                    fuente_status=EXCLUDED.fuente_status,
                    memoria_status=EXCLUDED.memoria_status,
                    motor_status=EXCLUDED.motor_status,
                    notas=EXCLUDED.notas
            """
        else:
            sql = """
                INSERT OR REPLACE INTO monthly_snapshots
                    (period_year, period_month, generated_at, triggered_by,
                     congruencia_status, integridad_status, integridad_n_ok, integridad_n_total,
                     alertas_criticas, alertas_advertencias, alertas_resueltas_mes, alertas_dias_max,
                     fuente_status, memoria_status, motor_status, notas)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            """
        c.execute(sql, (
            period_year, period_month, now, triggered_by,
            congruencia_status, integridad_status, integridad_n_ok, integridad_n_total,
            alertas_criticas, alertas_advertencias, alertas_resueltas_mes, alertas_dias_max,
            fuente_status, memoria_status, motor_status, notas,
        ))
        conn.commit()
        conn.close()
        snap["saved"] = True
    except Exception:
        pass

    return snap


def get_snapshots(limit: int = 24) -> list[dict]:
    """Retorna los últimos N snapshots ordenados por período descendente."""
    try:
        conn = get_connection()
        c    = conn.cursor()
        rows = c.execute(
            "SELECT * FROM monthly_snapshots ORDER BY period_year DESC, period_month DESC LIMIT ?",
            (limit,),
        ).fetchall()
        conn.close()
        result = []
        for r in rows:
            d = dict(r)
            d["period_label"] = (
                f"{MESES_ES.get(d.get('period_month', 0), '?')} {d.get('period_year', '?')}"
            )
            result.append(d)
        return result
    except Exception:
        return []


# ── EXPORT XLSX ───────────────────────────────────────────────────────────────
def export_alerts_xlsx(history: list[dict], kpis: dict, period_label: str = "") -> bytes:
    """
    Genera un XLSX institucional con:
      - Hoja 1: Resumen KPIs
      - Hoja 2: Alertas activas
      - Hoja 3: Historial completo
    Retorna bytes listos para st.download_button.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = openpyxl.Workbook()

        _H_FILL  = PatternFill("solid", fgColor="0D1F3C")
        _H_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        _BOLD    = Font(name="Arial", bold=True, size=10)
        _NORMAL  = Font(name="Arial", size=10)
        _MUTED   = Font(name="Arial", size=9, color="888888")
        _RED_F   = PatternFill("solid", fgColor="FFF0F3")
        _YEL_F   = PatternFill("solid", fgColor="FFFBEA")
        _GRN_F   = PatternFill("solid", fgColor="F0FFF8")
        _BORDER  = Border(
            bottom=Side(style="thin", color="DDDDDD"),
        )
        _CENTER  = Alignment(horizontal="center", vertical="center")

        def _hrow(ws, row: int, values: list, widths: list | None = None) -> None:
            for col, val in enumerate(values, 1):
                c = ws.cell(row=row, column=col, value=val)
                c.font = _H_FONT; c.fill = _H_FILL; c.alignment = _CENTER
            if widths:
                for col, w in enumerate(widths, 1):
                    ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w

        def _set_row_font(ws, row: int, ncols: int, font, fill=None, border=None) -> None:
            for col in range(1, ncols + 1):
                c = ws.cell(row=row, column=col)
                c.font  = font
                if fill:   c.fill   = fill
                if border: c.border = border

        # ── Hoja 1: Resumen ─────────────────────────────────────────────────
        ws1 = wb.active
        ws1.title = "Resumen"
        ws1["A1"] = f"QUIRA OS · Alertas de Cumplimiento{' · ' + period_label if period_label else ''}"
        ws1["A1"].font = Font(name="Arial", bold=True, size=13)
        ws1["A2"] = f"Generado: {datetime.now().strftime('%d %b %Y %H:%M')} · Dylus Lab · GAD Montecristi"
        ws1["A2"].font = _MUTED
        ws1.row_dimensions[1].height = 22

        kpi_rows = [
            ("Alertas críticas activas",   kpis.get("criticas_activas", 0)),
            ("Alertas advertencias activas", kpis.get("advertencias_activas", 0)),
            ("Resueltas este mes",          kpis.get("resueltas_este_mes", 0)),
            ("Días sin resolver (máximo)",  kpis.get("dias_abierta_max", 0)),
        ]
        ws1.append([])
        _hrow(ws1, 4, ["Indicador", "Valor"], [38, 15])
        for i, (label, val) in enumerate(kpi_rows, 5):
            ws1.cell(row=i, column=1, value=label).font = _NORMAL
            ws1.cell(row=i, column=2, value=val).font   = _BOLD
            ws1.cell(row=i, column=2).alignment          = _CENTER
            _set_row_font(ws1, i, 2, _NORMAL, border=_BORDER)

        # ── Hoja 2: Alertas activas ─────────────────────────────────────────
        ws2 = wb.create_sheet("Alertas Activas")
        _hrow(ws2, 1,
              ["Período", "Entidad", "Tipo", "Severidad", "Título", "Detalle", "Acción sugerida", "Estado"],
              [14, 20, 14, 14, 35, 45, 40, 14])
        activas = [h for h in history if h.get("estado", "") != "resuelta"]
        for i, a in enumerate(activas, 2):
            vals = [
                f"{a.get('period_month','')}/{a.get('period_year','')}",
                a.get("entidad", ""),
                a.get("tipo", "").title(),
                a.get("severidad", ""),
                a.get("titulo", ""),
                a.get("detalle", ""),
                a.get("accion", ""),
                "Abierta" if a.get("estado") == "pendiente" else "En Revisión",
            ]
            for col, val in enumerate(vals, 1):
                ws2.cell(row=i, column=col, value=val).font = _NORMAL
            fill = _RED_F if a.get("status") == "error" else _YEL_F
            _set_row_font(ws2, i, 8, _NORMAL, fill=fill, border=_BORDER)

        # ── Hoja 3: Historial completo ─────────────────────────────────────
        ws3 = wb.create_sheet("Historial Completo")
        _hrow(ws3, 1,
              ["Detectada", "Período", "Entidad", "Tipo", "Severidad",
               "Título", "Valor", "Umbral", "Estado", "Resuelta en", "Resolución institucional"],
              [20, 12, 14, 14, 14, 35, 10, 10, 14, 20, 45])
        for i, a in enumerate(history, 2):
            estado_lbl = {
                "pendiente":   "Abierta",
                "en_revision": "En Revisión",
                "resuelta":    "Resuelta",
            }.get(a.get("estado", ""), a.get("estado", ""))
            vals = [
                (a.get("detected_at") or "")[:16],
                f"{a.get('period_month','')}/{a.get('period_year','')}",
                a.get("entidad", ""),
                a.get("tipo", "").title(),
                a.get("severidad", ""),
                a.get("titulo", ""),
                a.get("valor"),
                a.get("umbral"),
                estado_lbl,
                (a.get("resuelta_en") or "")[:16],
                a.get("resolucion") or "",
            ]
            for col, val in enumerate(vals, 1):
                ws3.cell(row=i, column=col, value=val).font = _NORMAL
            fill = (_GRN_F if a.get("estado") == "resuelta"
                    else _RED_F if a.get("status") == "error" else _YEL_F)
            _set_row_font(ws3, i, 11, _NORMAL, fill=fill, border=_BORDER)

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    except Exception as e:
        # Si openpyxl no está disponible, retorna bytes vacíos
        return b""


def export_snapshot_xlsx(snapshot: dict, history: list[dict]) -> bytes:
    """
    XLSX de snapshot mensual para rendición de cuentas / LOTAIP.
    Hoja única con estado institucional completo del período.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        wb  = openpyxl.Workbook()
        ws  = wb.active
        ws.title = f"Snapshot {snapshot.get('period_label','')}"

        _ST_COLOR = {"ok": "00E096", "warning": "FFB800", "error": "FF4D6D"}
        _ST_LABEL = {"ok": "Operativo", "warning": "Requiere Atención", "error": "Revisión Pendiente"}
        _BOLD  = Font(name="Arial", bold=True, size=10)
        _NORM  = Font(name="Arial", size=10)
        _TITLE = Font(name="Arial", bold=True, size=14)
        _SUB   = Font(name="Arial", size=10, color="666666")

        ws.column_dimensions["A"].width = 32
        ws.column_dimensions["B"].width = 22
        ws.column_dimensions["C"].width = 22

        ws["A1"] = f"QUIRA OS · Snapshot Institucional"
        ws["A1"].font = _TITLE
        ws["A2"] = f"Período: {snapshot.get('period_label','')} · Generado: {datetime.now().strftime('%d %b %Y %H:%M')}"
        ws["A2"].font = _SUB
        ws["A3"] = f"GAD Municipal de Montecristi · Dylus Lab © 2026"
        ws["A3"].font = _SUB

        sections = [
            ("CONGRUENCIA INSTITUCIONAL", [
                ("Estado global", _ST_LABEL.get(snapshot.get("congruencia_status",""), "–")),
                ("Fuente Operativa", _ST_LABEL.get(snapshot.get("fuente_status",""), "–")),
                ("Memoria Institucional", _ST_LABEL.get(snapshot.get("memoria_status",""), "–")),
                ("Motor Analítico", _ST_LABEL.get(snapshot.get("motor_status",""), "–")),
            ]),
            ("INTEGRIDAD SEMÁNTICA", [
                ("Estado",               _ST_LABEL.get(snapshot.get("integridad_status",""), "–")),
                ("Indicadores correctos", f"{snapshot.get('integridad_n_ok',0)}/{snapshot.get('integridad_n_total',0)}"),
            ]),
            ("ALERTAS DE CUMPLIMIENTO", [
                ("Alertas críticas activas",   snapshot.get("alertas_criticas", 0)),
                ("Advertencias activas",       snapshot.get("alertas_advertencias", 0)),
                ("Resueltas en el período",    snapshot.get("alertas_resueltas_mes", 0)),
                ("Días sin resolver (máx.)",   snapshot.get("alertas_dias_max", 0)),
            ]),
        ]

        row = 5
        for sec_title, rows in sections:
            ws.cell(row=row, column=1, value=sec_title).font = Font(
                name="Arial", bold=True, size=11, color="0D1F3C"
            )
            ws.cell(row=row, column=1).fill = PatternFill("solid", fgColor="E8EDF5")
            row += 1
            for label, val in rows:
                ws.cell(row=row, column=1, value=label).font = _NORM
                ws.cell(row=row, column=2, value=val).font   = _BOLD
                row += 1
            row += 1

        if snapshot.get("notas"):
            ws.cell(row=row, column=1, value="Notas").font = _BOLD
            ws.cell(row=row+1, column=1, value=snapshot["notas"]).font = _NORM

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()
    except Exception:
        return b""
