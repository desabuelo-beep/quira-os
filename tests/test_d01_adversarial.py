# -*- coding: utf-8 -*-
"""
tests/test_d01_adversarial.py — atacar d01, no confirmarlo
════════════════════════════════════════════════════════════════════════════════
POR QUÉ EXISTE (2026-08-26 · cierre del piloto ADR-053). Tras migrar d01, el
inventario lo clasificó `protegido_sin_atacar`, y esa lectura es la honesta:

> *«Sin atacar: d01 — resistir no está demostrado, sólo no refutado.»*

El colega fijó las cuatro fronteras que hay que atacar, y la razón de fondo:

> *«Un sistema inexpugnable no es uno que declara que nunca falla. Es uno que,
> cuando intentan hacerle afirmar algo que no puede sostener, no inventa
> confianza.»*

    1 · IDENTIDAD     hacer que afirme sobre un sujeto que no es
    2 · PROCEDENCIA   hacer que sostenga con evidencia de otro Gold Master
    3 · GRADO         hacer que un peso suba sin evidencia nueva
    4 · EQUIVALENCIA  hacer que recalcule en vez de leer

En d07 estos ataques encontraron **dos agujeros que nadie había visto** —el
sujeto del sello y la huella de `dpe_entidad_id`—. El valor de atacar no es
confirmar que todo está bien: es que a veces no lo está.

Dylus Lab © 2026
"""
from __future__ import annotations

import json
import sys
from pathlib import Path

import pytest

RAIZ = Path(__file__).resolve().parents[1]
if str(RAIZ) not in sys.path:
    sys.path.insert(0, str(RAIZ))

from app.agents import procedencia as P                   # noqa: E402
from app.agents import sujeto as S                        # noqa: E402
from app.agents.d01 import motor as M                     # noqa: E402


def _hay_gm() -> bool:
    return M._GM_DEFAULT.exists()


def _gold_master_falso(tmp: Path, ipe: float) -> Path:
    """Un Gold Master mínimo con el IPE que se le diga. Sirve para comprobar que
    d01 devuelve **lo que está escrito**, no algo derivado."""
    import openpyxl
    tmp.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = M._HOJA
    ws[M._CELDA_IPE] = ipe
    ws[M._CELDA_COBERTURA] = 0.5
    ws[M._CELDA_INV_VINCULADA] = 1234.0
    p = tmp / "gm_falso.xlsx"
    wb.save(str(p))
    return p


# ── ATAQUE 1 · IDENTIDAD ──────────────────────────────────────────────────────
def test_ataque_1_alterar_el_sujeto_cambia_lo_que_d01_puede_afirmar():
    """Si se altera la identidad con la que QUIRA va a las fuentes, la
    afirmación de d01 **no puede seguir pareciendo la misma**.

    Es el ataque que en d07 (2026-08-19) dejó todo en verde midiendo a un GAD
    con la evidencia de otro. Aquí se ejerce sobre el dominio recién migrado."""
    if not _hay_gm():
        pytest.skip("Gold Master no accesible")

    perfil = S._SUJETOS / f"{S.POR_DEFECTO}.json"
    respaldo = perfil.read_bytes()
    antes = M.leer_metricas()["procedencia"]
    try:
        d = json.loads(respaldo.decode("utf-8"))
        d["identidad_en_fuentes"]["ruc"] = "9999999999999"   # otra entidad
        perfil.write_text(json.dumps(d, ensure_ascii=False, indent=1),
                          encoding="utf-8")
        S.cargar.cache_clear()
        despues = M.leer_metricas()["procedencia"]
    finally:
        perfil.write_bytes(respaldo)
        S.cargar.cache_clear()

    assert antes["sujeto_huella"] != despues["sujeto_huella"], (
        "cambiar la identidad con la que se va a la fuente NO alteró la huella "
        "estampada por d01: la afirmación seguiría pareciendo la misma sobre "
        "un sujeto distinto")
    assert M.leer_metricas()["procedencia"] == antes, (
        "la huella no volvió a su valor tras restaurar el perfil")


# ── ATAQUE 2 · PROCEDENCIA ────────────────────────────────────────────────────
def test_ataque_2_dos_gold_master_distintos_no_pueden_dar_la_misma_evidencia(tmp_path):
    """La evidencia identifica **qué versión del motor** produjo el número.

    Si dos Gold Master con contenido distinto produjeran la misma evidencia, la
    afirmación no podría distinguirlos — y con correcciones sobre copia (Regla
    de Oro 1) eso deja de ser hipotético."""
    a = _gold_master_falso(tmp_path / "a", 0.42)
    b = _gold_master_falso(tmp_path / "b", 0.99)

    ma, mb = M.leer_metricas(a), M.leer_metricas(b)
    assert ma["status"] == mb["status"] == "ok"
    assert ma["evidencia_sha256"] != mb["evidencia_sha256"], (
        "dos Gold Master con IPE distinto produjeron la misma evidencia: "
        "la afirmación no podría decir de cuál salió")
    assert ma["ipe_ejecutado"] != mb["ipe_ejecutado"]


def test_ataque_2b_la_evidencia_no_se_puede_declarar_a_mano(tmp_path):
    """Intento de falsificación directa: sostener el IPE de un Gold Master
    citando la evidencia de otro. La cadena debe seguir describiendo lo que
    realmente se leyó, no lo que se declare."""
    a = _gold_master_falso(tmp_path, 0.42)
    real = M.leer_metricas(a)["evidencia_sha256"]

    # Se construye a mano una procedencia con evidencia ajena.
    falsa = P.Procedencia(
        fuente=f"Gold Master · {M._HOJA}", captura="2026-08-26",
        estado_adquisicion="leido_del_motor",
        evidencia="0000000000000000",           # no corresponde a ningún GM leído
        verificador="d01.motor.leer_metricas",
        prueba_del_verificador="test_d01_lee_el_ipe_sin_recalcularlo",
        sujeto=f"{S.POR_DEFECTO} {S.nombre_corto()}")
    s = P.sostener("el IPE es 0.42", falsa)

    # ✅ ESCALÓN 4 CERRADO el 2026-08-30. Esta procedencia NO declara artefacto,
    # así que sigue acreditando por existencia — el residuo conocido.
    assert s.peso == P.HECHO_VERIFICABLE, (
        "sin artefacto declarado se acredita por existencia: residuo bajo "
        "trinquete en `test_escalon4_evidencia`")

    # Y AQUÍ LA ASERCIÓN INVERTIDA: en cuanto la procedencia declara con qué
    # artefacto se comprueba, el hash ajeno deja de acreditar.
    con_artefacto = P.Procedencia(**{**vars(falsa), "artefacto": str(a)})
    s2 = P.sostener("el IPE es 0.42", con_artefacto)
    assert s2.peso == P.HALLAZGO_DE_VERIFICABILIDAD, (
        "declarando el artefacto, un hash que no es el suyo debe degradar")
    assert real != "0000000000000000", "la evidencia real no puede ser el relleno"


# ── ATAQUE 3 · GRADO ──────────────────────────────────────────────────────────
def test_ataque_3_no_se_puede_subir_el_grado_de_una_lectura_fallida():
    """Cuando el motor no se puede leer, QUIRA no sabe nada del sujeto. Pedir
    `hecho_verificable` sobre esa cadena **no debe concederlo**."""
    s = M.sostener_ipe("/no/existe/gm.xlsx")
    assert s.peso != P.HECHO_VERIFICABLE
    assert s.habla_del_sujeto is False

    # Y forzando el peso pretendido tampoco sube.
    vacia = P.Procedencia(fuente=f"Gold Master · {M._HOJA}",
                          sujeto=f"{S.POR_DEFECTO} {S.nombre_corto()}")
    for pretendido in (P.HECHO_VERIFICABLE, P.HALLAZGO_DE_VERIFICABILIDAD):
        assert P.sostener("x", vacia, pretendido).peso == P.NO_DETERMINABLE, (
            "una cadena sin captura ni evidencia no mejora por pretender más")


def test_ataque_3b_quitar_la_prueba_degrada_la_afirmacion_de_d01():
    """Si la prueba que acredita al verificador desapareciera, d01 **debe dejar
    de afirmar sobre el sujeto** — no seguir como si nada.

    Se simula con un nombre de prueba inexistente: es lo que ocurriría si
    alguien renombrara `test_d01_lee_el_ipe_sin_recalcularlo` sin actualizar la
    referencia del motor."""
    if not _hay_gm():
        pytest.skip("Gold Master no accesible")
    m = M.leer_metricas()
    sin_prueba = P.Procedencia(
        fuente=f"Gold Master · {M._HOJA}", captura="2026-08-26",
        estado_adquisicion="leido_del_motor", evidencia=m["evidencia_sha256"],
        verificador="d01.motor.leer_metricas",
        prueba_del_verificador="test_que_alguien_renombro_y_ya_no_existe",
        sujeto=f"{S.POR_DEFECTO} {S.nombre_corto()}")
    s = P.sostener("el IPE es X", sin_prueba)
    assert s.peso == P.HALLAZGO_DE_VERIFICABILIDAD
    assert "prueba_del_verificador" in s.faltan


# ── ATAQUE 4 · EQUIVALENCIA · ¿lee o recalcula? ───────────────────────────────
def test_ataque_4_d01_devuelve_lo_ESCRITO_aunque_sea_absurdo(tmp_path):
    """EL ATAQUE QUE PROPUSO EL COLEGA, y el más importante de los cuatro.

    > *«Introducir deliberadamente una situación donde H16b!B15 = X y la fórmula
    > recalculada daría Y, y comprobar que d01 continúa devolviendo X.»*

    Se le da un Gold Master cuyo IPE es un valor que **ninguna fórmula sensata
    produciría** —0.0 y 1.0 exactos, y un valor fuera del rango habitual—. Si
    d01 devolviera algo distinto, estaría derivando en vez de leer, y violaría
    las Reglas de Oro 1 y 4.

    Esto convierte la equivalencia de una comparación estática en una
    **propiedad defendida**: no se compara con lo de ayer, se comprueba que la
    lectura es fiel a lo escrito, sea lo que sea."""
    for valor in (0.0, 1.0, 0.123456789, 42.0):
        gm = _gold_master_falso(tmp_path / f"v{valor}", valor)
        m = M.leer_metricas(gm)
        assert m["status"] == "ok"
        assert m["ipe_ejecutado"] == valor, (
            f"d01 devolvió {m['ipe_ejecutado']} donde la celda dice {valor}: "
            f"está derivando el número en vez de leerlo (Regla de Oro 1 y 4)")


def test_ataque_4b_si_falta_la_hoja_no_se_inventa_el_numero(tmp_path):
    """Un Gold Master sin `H16b_IPE` no autoriza a estimar el IPE por otra vía.
    Falta el dato: se dice, no se rellena."""
    import openpyxl
    wb = openpyxl.Workbook()
    wb.active.title = "OtraHoja"
    p = tmp_path / "sin_hoja.xlsx"
    wb.save(str(p))

    m = M.leer_metricas(p)
    assert m["status"] == "failed"
    assert "ipe_ejecutado" not in m, (
        "sin la hoja no puede aparecer un IPE: sería un número inventado")
    s = M.sostener_ipe(p)
    assert s.peso != P.HECHO_VERIFICABLE
