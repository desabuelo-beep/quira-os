# -*- coding: utf-8 -*-
"""
tests/test_recorrido_icpi.py — la cifra madre, recorrida hacia atrás
════════════════════════════════════════════════════════════════════════════════
Javo eligió la cifra: **el ICPI**. El colega fijó el método: no recalcular, no
corregir, no usar el propio artefacto como autoridad — sólo demostrar hasta
dónde puede reconstruirse.

EL RECORRIDO, medido:

    H12!B33  = B31/B32
      B31    = SUM(J6:J30)      Jᵢ = Pᵢ×Rᵢ×Vᵢ×Eᵢ×Tᵢ×Cᵢ
      B32    = SUM(K6:K30)      Kᵢ = Pᵢ×Rᵢ
        Pᵢ  ← H14_PONDERADORES!G      Rᵢ ← H14_PONDERADORES!F
        Vᵢ  ← VLOOKUP H13_VARIABLES_Vi
        Eᵢ  ← **literal, 25 de 25**
        Tᵢ  ← H07b!B20 ← H07_S5_FINANCIERO_eSIGEF!B23
              └─ Fuente declarada: «Cédula LOTAIP GAD Montecristi Abr-2026»
        Cᵢ  ← VLOOKUP H01_PARÁMETROS

EL HALLAZGO. `H04!B7` se llama **`Total_Metas_PDOT`** y vale **25**. Javo:

> *«para ICPI tomamos metas PDOT el total. Cuando empezamos la construcción
> empezamos con una **muestra de 25**, pero luego decidimos hacerlo completo,
> pero **nunca acoplamos el resto de metas**.»*

El parámetro etiquetado «total» es el tamaño de la muestra. Y BOOT ya tiene la
regla que lo nombra: **«etiqueta incorrecta = número falso»** (§6-sexies).

⚠️ EL MOTOR NO CALCULA MAL. Hace exactamente lo que dice hacer, y los
ponderadores suman 1.0000 sobre esas 25. Lo que no corresponde es el **alcance
declarado**: una cifra llamada `ICPI_GLOBAL_SISTEMA` calculada sobre un
subconjunto que se declara total. Es un problema de frontera, no de aritmética —
y esa distinción es la que esta sesión existe para no perder.

⛔ NO SE REPARA. El Gold Master es inmutable en su fórmula canónica y esto es una
decisión de gobernanza (Regla de Oro 1).

Dylus Lab © 2026
"""
from __future__ import annotations

import sys
from pathlib import Path

import pytest

RAIZ = Path(__file__).resolve().parents[1]
if str(RAIZ) not in sys.path:
    sys.path.insert(0, str(RAIZ))


def _gm():
    try:
        import openpyxl
        from config import DATOS_DIR
        p = Path(DATOS_DIR) / "SIAP-ICPI_GOLD_MASTER_v5.5_TGI.xlsx"
        if not p.exists():
            return None
        return openpyxl.load_workbook(p, data_only=False, read_only=True)
    except Exception:                                    # noqa: BLE001
        return None


def test_la_formula_del_ICPI_es_la_declarada():
    """El punto de partida, literal y sin evaluar. `H12!B33` es la fuente única
    del ICPI y su fórmula es inmutable (Regla de Oro 1)."""
    wb = _gm()
    if wb is None:
        pytest.skip("Gold Master no accesible")
    ws = wb["H12_MOTOR_ICPI_CANÓNICO"]
    assert ws["B33"].value == "=B31/B32"
    assert ws["B31"].value == "=SUM(J6:J30)"
    assert ws["B32"].value == "=SUM(K6:K30)"
    wb.close()


def test_el_parametro_llamado_total_es_el_tamano_de_la_muestra():
    """EL HALLAZGO DEL RECORRIDO, fijado con trinquete.

    `Total_Metas_PDOT = 25` — y las 25 fueron una muestra inicial que nunca se
    amplió. El día que se acoplen las demás, este número subirá y la prueba
    fallará: **y ese fallo será el progreso**, no una regresión."""
    wb = _gm()
    if wb is None:
        pytest.skip("Gold Master no accesible")
    ws = wb["H04_S2_PLANIFICACIÓN_PDOT"]
    etiqueta = ws.cell(row=7, column=1).value
    assert etiqueta == "Total_Metas_PDOT", f"cambió la etiqueta: {etiqueta}"
    wb.close()

    wb2 = _gm()
    import openpyxl
    from config import DATOS_DIR
    v = openpyxl.load_workbook(Path(DATOS_DIR) / "SIAP-ICPI_GOLD_MASTER_v5.5_TGI.xlsx",
                               data_only=True, read_only=True)
    total = v["H04_S2_PLANIFICACIÓN_PDOT"].cell(row=7, column=2).value
    v.close(); wb2.close()
    assert total == 25, (
        f"Total_Metas_PDOT pasó de 25 a {total}: si se acoplaron las metas que "
        f"faltaban, actualizar este hallazgo — dejó de ser una muestra")


def test_Ei_es_un_literal_sin_definicion_en_el_canon():
    """El segundo hallazgo, independiente del primero. `Eᵢ` participa en el
    numerador del ICPI y es un **valor escrito a mano en las 25 filas**: no tiene
    fórmula, no referencia ninguna hoja, y su única aparición en todo el libro es
    la cabecera de su propia columna.

    Mientras `H01_PARÁMETROS` declara: *«FUENTE ÚNICA de toda configuración.
    NINGUNA otra hoja tiene valores de configuración hardcodeados»*."""
    wb = _gm()
    if wb is None:
        pytest.skip("Gold Master no accesible")
    ws = wb["H12_MOTOR_ICPI_CANÓNICO"]
    literales = [r for r in range(6, 31)
                 if ws.cell(row=r, column=5).value is not None
                 and not str(ws.cell(row=r, column=5).value).startswith("=")]
    assert len(literales) == 25, (
        f"Eᵢ dejó de ser literal en las 25 filas ({len(literales)}): si ahora "
        f"deriva de alguna hoja, el hallazgo cambió")
    wb.close()


def test_el_eslabon_Ti_llega_a_un_documento_primario_declarado():
    """LO QUE SÍ SE RECONSTRUYE. De los seis insumos, `Tᵢ` es el único que
    retrocede hasta un documento nombrado: `H07_S5_FINANCIERO_eSIGEF` declara en
    su fila 11 la **Cédula LOTAIP GAD Montecristi Abr-2026**.

    No es trazabilidad completa —el documento se nombra en prosa, sin SHA ni
    ruta—, pero es más de lo que declaran los otros cinco insumos, que sólo
    apuntan a otras hojas del mismo libro."""
    wb = _gm()
    if wb is None:
        pytest.skip("Gold Master no accesible")
    ws = wb["H07_S5_FINANCIERO_eSIGEF"]
    assert ws.cell(row=11, column=1).value == "Fuente"
    doc = str(ws.cell(row=11, column=2).value)
    assert "Cédula" in doc and "2026" in doc, f"cambió la fuente declarada: {doc}"
    wb.close()


# ── EL REGISTRO DE DEUDA · una deuda no se declara sola ──────────────────────
def test_toda_deuda_declarada_tiene_su_ataque_localizado():
    """LA REGLA QUE HACE ÚTIL EL REGISTRO. Sin prueba asociada, una deuda es una
    nota que envejece — y este sistema ya sabe lo que pasa con las notas. Con
    ataque, el día que se subsane **la prueba falla**, y ese fallo es la señal.

    Se comprueba que el ataque EXISTE en disco, no que esté escrito en el
    registro: nombrar una prueba inexistente sería acreditar sin nada detrás, el
    mismo defecto que la escalera de apropiación cerró en su escalón 2."""
    from app.agents import deuda as D

    c = D.cobertura_de_deuda()
    assert c["deudas"], "el registro de deuda quedó vacío"
    assert not c["sin_ataque_localizado"], (
        f"deudas sin prueba que las fije: {c['sin_ataque_localizado']}")


def test_el_registro_no_pretende_ser_exhaustivo():
    """C0 · regla 1 aplicada a la deuda. Un registro de deudas **no se barre**:
    alguien las encuentra y las declara. Presentarlo como completo diría que el
    resto del sistema está limpio, y las capas C4–C7 ni siquiera se han mirado."""
    from app.agents import deuda as D

    u = D.cobertura_de_deuda()["universo"]
    assert u["mecanismo"]["tipo"] == "explicitamente_limitado"
    limites = " ".join(u["fuera_de_alcance"])
    assert "no barre" in limites and "C4" in limites


def test_el_motor_lee_la_version_que_el_canon_declara(gold_master):
    """D-002 · CERRADA, y la prueba se invirtió — que era la señal acordada.

    Decía «BOOT declara v5.7 y el código abre v5.5»; ahora exige lo contrario.
    La raíz estaba en `config.py`, la puerta única a los datos, que fijaba
    `GOLD_MASTER_VERSION = "v5.5_TGI"` a mano: **once archivos replicaban lo que
    ella declaraba** mientras `app/connectors/gold_master.py` ya resolvía bien.
    El sistema tenía dos respuestas a «¿cuál es mi Gold Master?».

    La regla de autoridad la dio Javo —*«debe terminar en TGI para ser tomada
    por el sistema, y lo que cambia es 5.6, 5.7…»*— y ahora se **resuelve** en un
    solo lugar: sufijo `_TGI`, se excluyen `_FREEZE`/`_`/`~$`, gana la versión
    numérica más alta.

    ⚠️ Verificado antes de migrar: las cifras NO se movieron. IPE
    0.9557408659866722, cobertura 0.96, ICPI 0.27458226534062735 — idénticas en
    v5.5 y v5.7. La reparación alineó la fuente sin tocar un solo resultado."""
    import re

    import config

    boot = (RAIZ / "governance" / "BOOT.md").read_text(encoding="utf-8", errors="replace")
    m = re.search(r"v(\d\.\d)_TGI", boot)
    assert m, "BOOT dejó de declarar la versión del Gold Master"
    assert config.GOLD_MASTER_VERSION == f"v{m.group(1)}_TGI", (
        f"el canon declara v{m.group(1)}_TGI y config resuelve "
        f"{config.GOLD_MASTER_VERSION}: el desfase volvió")


def test_la_version_se_resuelve_y_no_se_escribe():
    """El trinquete de D-002. Si alguien vuelve a fijar la versión a mano, el
    sistema tendrá otra vez dos respuestas a la misma pregunta."""
    # ⚠️ SE COMPRUEBA LA ASIGNACIÓN, NO EL TEXTO. La primera versión buscaba
    # el literal `GOLD_MASTER_VERSION = "v5` en todo el archivo y lo encontró
    # **en el comentario que documenta la reparación**: el texto no es el
    # código, otra vez. Se mira la línea que asigna de verdad.
    fuente = (RAIZ / "config.py").read_text(encoding="utf-8")
    assert "_resolver_gold_master_vigente" in fuente
    asignaciones = [ln.strip() for ln in fuente.splitlines()
                    if ln.startswith("GOLD_MASTER_VERSION")]
    assert asignaciones, "config dejó de exponer GOLD_MASTER_VERSION"
    assert all("_resolver_gold_master_vigente()" in a for a in asignaciones), (
        f"la versión volvió a escribirse a mano: {asignaciones}")
    literales = [f for f in (RAIZ / "scripts").glob("enrich_*.py")
                 if "GOLD_MASTER_v5.5_TGI.xlsx" in f.read_text(encoding="utf-8")
                 and "_gold_master_vigente" not in f.read_text(encoding="utf-8")]
    assert not literales, f"enrichers que volvieron al literal: {literales}"
